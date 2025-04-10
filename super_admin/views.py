from django.http.response import HttpResponse,JsonResponse
from django.shortcuts import render,HttpResponse,redirect   ,HttpResponseRedirect
from django.contrib.auth import authenticate,get_user_model,login,logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q , F
from .models import *
from rest_framework.response import Response
from django.core.mail import send_mail as sm,EmailMessage
from csv import reader
from csv import DictReader
from .serializers import *
import json
from datetime import date,datetime,timedelta
from hashlib import sha512
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
# from paywix.payu import Payu
from django.core.mail import EmailMessage
import hashlib
# import pandas as pd 
import math
# from pandas import ExcelWriter
# from pandas import ExcelFile 
import csv
import os
#pdf import
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
# from xhtml2pdf import pisa
import xlwt
import logging,traceback
logger = logging.getLogger('django')
from .functions import *
from django.db.models import Min, Max
from openpyxl import load_workbook
import random
import string
from django.contrib.auth.hashers import make_password
import threading
from threading import Thread


class EmailThread(threading.Thread):
    def __init__(self, subject, body, recipient_list):
        self.subject = subject
        self.recipient_list = recipient_list
        self.body = body
        threading.Thread.__init__(self)
        
    def run(self):
        # print("mail sent through threading. !!")
        sm(
            self.subject,
            self.body,
            'testmail@erp.ciisindia.in',
            self.recipient_list,
            fail_silently=False
        )

def send_html_email(subject, body, recipient_list):
    EmailThread(subject, body, recipient_list).start()


def Test(request):
    f = open("demofile2.txt", "a")
    f.write(f"{str(datetime.now())} :\n")
    f.close()
    return HttpResponse("Hello")


@login_required(login_url='/login/')
def home(request):
    if request.user.is_superuser or request.user.is_data_entry:

        params = {
            "total_student":Student.objects.filter(Q(archive=False) & Q(is_cancelled=False)).count(),
            "pending_student":Student.objects.filter(Q(archive=False) & Q(is_cancelled=False) & Q(enrolled=False)).count(),
            "registered_student":Student.objects.filter(Q(archive=False) & Q(is_cancelled=False) & Q(enrolled=True)).count(),
            "courses":Course.objects.all()[:15]
        }
        return render(request,"super_admin/homepage.html",params)

def index(request):
    if request.user.is_authenticated:
        user_id = request.user.id
        
        if request.user.is_superuser:
            return redirect('overview')
        elif request.user.is_student:
            return redirect('profile')
        elif request.user.is_data_entry:
            return redirect('overview')
    else:
        return redirect('login')
    
@login_required(login_url='/login/')
def Overview(request):
    # logger.info('>>>>>>>>>>>>>> Something Debug wrong!')
    # print("System 25")

    student_count= ''
    student_list = []
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_data_entry:
            if request.method == "POST":
                payment_reciept_id = request.POST.get('payment_reciept_id')
                payment_reciept_status = request.POST.get('payment_reciept_status')
                if payment_reciept_id and payment_reciept_status:
                    get_payment_reciept = PaymentReciept.objects.get(id = payment_reciept_id)
                    get_payment_reciept.status = payment_reciept_status
                    total_fees = get_payment_reciept.semyearfees
                    paid_fees = get_payment_reciept.uncleared_amount
                    print(total_fees , paid_fees)
                    if int(paid_fees) > int(total_fees):
                        get_payment_reciept.advanceamount = int(paid_fees) - int(total_fees)
                        get_payment_reciept.paidamount = int(paid_fees)
                        get_payment_reciept.pendingamount = 0
                        get_payment_reciept.uncleared_amount = 0
                    elif int(paid_fees) < int(total_fees):
                        get_payment_reciept.advanceamount = 0
                        get_payment_reciept.paidamount = int(paid_fees)
                        get_payment_reciept.pendingamount = int(total_fees) - int(paid_fees)
                        get_payment_reciept.uncleared_amount = 0
                    elif int(paid_fees) == int(total_fees):
                        get_payment_reciept.advanceamount = 0
                        get_payment_reciept.paidamount = int(paid_fees)
                        get_payment_reciept.pendingamount = int(total_fees) - int(paid_fees)
                        get_payment_reciept.uncleared_amount = 0
                    # print("total fees :",get_payment_reciept.semyearfees)
                    # print("paid fees :",get_payment_reciept.paidamount)
                    # print("pending fees :",get_payment_reciept.pendingamount)
                    # print("uncleared amount :",get_payment_reciept.uncleared_amount)
                    get_payment_reciept.save()
                    updated_payment_data = []
                    all_pending_realisation = PaymentReciept.objects.filter(status = "Not Realised")
                    for i in all_pending_realisation:
                        try:
                            getstudent = Student.objects.get(id = i.student)
                            enrolled = Enrolled.objects.get(student = i.student)
                            course_name = Course.objects.get(id = enrolled.course)
                            stream_name = Stream.objects.get(id = enrolled.stream)
                            obj = {
                                "student_id":getstudent.id,
                                "name":getstudent.name,
                                "email":getstudent.email,
                                "mobile":getstudent.mobile,
                                "course":course_name.name,
                                "stream":stream_name.name,
                                "current_semester":enrolled.current_semyear,
                                "payment_reciept_id":i.id,                        
                                "payment_status":i.status,
                                "transactionID":i.transactionID,
                                "uncleared_amount":i.uncleared_amount
                            }
                            updated_payment_data.append(obj)
                        except Student.DoesNotExist:
                            pass
                    return JsonResponse({'data': updated_payment_data})

            all_pending_realisation = PaymentReciept.objects.filter(status = "Not Realised")
            for i in all_pending_realisation:
                try:
                    getstudent = Student.objects.get(id = i.student.id)
                    enrolled = Enrolled.objects.get(student = i.student)
                    course_name = Course.objects.get(id = enrolled.course.id)
                    stream_name = Stream.objects.get(id = enrolled.stream.id)
                    obj = {
                        "student_id":getstudent.id,
                        "name":getstudent.name,
                        "email":getstudent.email,
                        "mobile":getstudent.mobile,
                        "course":course_name.name,
                        "stream":stream_name.name,
                        "current_semester":enrolled.current_semyear,
                        "payment_reciept_id":i.id,                        
                        "payment_status":i.status,
                        "transactionID":i.transactionID,
                        "uncleared_amount":i.uncleared_amount
                    }
                    student_list.append(obj)
                except Student.DoesNotExist:
                    pass
            student_count = Student.objects.all().count()
            params= {
                "allstudents":student_count,
                "students":Student.objects.filter(~Q(is_cancelled = True) & Q(is_quick_register=False)).order_by('-id'), ## modified by Avani
                "payment_student":student_list,   
            }
            return render(request,"super_admin/overview.html",params)

def ForgotPassword(request):
    if request.method == "POST":
        email = request.POST.get('forgot_password')
        if email:
            try:
                get_user = User.objects.get(email=email)
                # create_password_session = ForgotPasswordSession(
                #     email=email,
                #     temp_password="kjghjgsaduhyghjgjhsdgd"
                # )
                res = ''.join(random.choices(string.ascii_uppercase +
                                            string.digits, k=20))
                print("The generated random string : " + str(res))
                # create_password_session.save()
                get_user.password = make_password(res)
                get_user.save()
                print("Sending Email")

                message = f"""
        Dear User,

        We hope this email finds you well. We recently received a request to reset your password for CIIS. Your security is our top priority, and we are here to assist you in regaining access to your account.

        Please find below your temporary password that will grant you entry to the system:

        Temporary Password: {res}

    """

                subject = "Retrieve Access: Your Temporary Password | CIIS"

                # Send the email
                send_html_email(subject, message, [get_user.email])
                # subject = "Retrieve Access: Your Temporary Password | CABACCOUNTS"

                # send_html_email(subject,msg,[get_user.email])
                # print(send_mail("Retrieve Access: Your Temporary Password | CABACCOUNTS", msg, 'info@cabsaveaccounts.cab-accounts.co.uk',[get_user.email], fail_silently=False))

                print("Email Sent")
                # password = make_password(res)
                
                return JsonResponse({'user':'yes'})
            except User.DoesNotExist:
                return JsonResponse({'user':'no'})
    return JsonResponse({'data':'data'})


def ULogin(request):
    showsignup = ""
    usernamepresent = ''
    loggedinuser = ''
    if not request.user.is_authenticated:
        if request.method == "POST":
            usignup = request.POST.get('signup')
            if usignup == "signup":
                showsignup = "yes"
            else:
                showsignup = "no"
                
            # Remail = request.POST.get('Remail')
            # Rpassword = request.POST.get('Rpassword')
            # if  Remail  and Rpassword:
            #     try:
            #         checkuser = User.objects.get(email=Remail)
            #         print("checkuser :",checkuser)
            #     except User.DoesNotExist:
            #         createuser = User.objects.create_user(email=Remail,password=Rpassword)
            #         createuser.save()
            #         usercreated = "yes"
            #     if checkuser:
            #         usernamepresent = "yes"
            email = request.POST.get('email')
            password = request.POST.get('password')
            if email and password:
                print(email,password)
                authen = authenticate(email = email,password = password)
                print("authen :",authen)
                if authen:
                    if authen.is_superuser:
                        login(request,authen)
                        print("super admin logged in")
                        return redirect('overview')
                        
                    elif authen.is_student:
                        login(request,authen)
                        print("student logged in")
                        return redirect('profile')
                    elif authen.is_data_entry:
                        login(request,authen)
                        print("data entry logged in")
                        return redirect('overview')
                else:
                    return redirect('login')
    params = {
        "signup":showsignup,
        "usernamepresent":usernamepresent,
        # "usercreated":usercreated,
        "loggedinuser":loggedinuser
    }
    return render(request,'includes/login.html',params)


def ULogout(request):
    print(request.user, " in logout process")
    url = ''
    if not request.user.is_authenticated:
         url = 'examination_login'
    else:
        url = 'login'
    logout(request)
    print(url)
    #return redirect('login')
    return redirect(url)

@login_required(login_url='/login/')
def ChangePassword(request):
    msg = ""
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        if password == confirm_password:
            user = request.user
            user.set_password(password)
            user.save()
            msg = "Password Changed Successfully !!"
            
            # print("password changed for user :",user)
    params = {
        "msg":msg
    }
    return render(request,'includes/change_password.html',params)

@login_required(login_url='/login/')
def AddCourse(request):
    main = []
    university = University.objects.all()
    for i in university:
        courses = Course.objects.filter(university = i.id)
        obj = {
            "university_name":i.university_name,
            "course_name":courses
        }
        main.append(obj)
    # print("main : ",main)

    
    
    if request.method == "POST":
        print(request.POST)
        add_coursename = request.POST.get('add_coursename')
        add_university = request.POST.get('add_university')
        course_year = request.POST.get('course_year')
        importdata = request.FILES.get('importdata')
        if importdata:
            df=pd.read_excel(importdata,sheet_name='Sheet1')
            for i in df.index:
                course_name = df['courseName'][i]
                stream_name = df['streamName'][i]
                stream_duration = df['streamDurationYear'][i]
                

                # try:
                #     getcourse = Course.objects.get(name = course_name)
                # except Course.DoesNotExist:
                #     getcourse = Course(name=course_name)
                #     getcourse.save()
                # getcourse = Course.objects.get(name = course_name)
                # try:
                #     getstream = Stream.objects.get(Q(name = stream_name) & Q(course = getcourse))
                # except Stream.DoesNotExist:
                #     getstream = Stream(name = stream_name,course=getcourse,sem=stream_duration)
                #     getstream.save()
                

        university_name = request.POST.get('university_name')
        if university_name:
            try:
                getuniversity = University.objects.get(university_name = university_name)
                getcourse = Course.objects.filter(university = getuniversity.id)
                course_list = []
                for i in getcourse:
                    courseserializer = {
                        "id":i.id,
                        "name":i.name,
                        'year':i.year
                    }
                    course_list.append(courseserializer)
                return JsonResponse({'course':course_list})
            except University.DoesNotExist:
                print("University does not exist")        


        update_course_id = request.POST.get('update_course_id')
        update_course_name = request.POST.get('update_course_name')     
        update_course_year = request.POST.get('update_course_year')       
        if update_course_id and update_course_name and update_course_year:
            updatecourse = Course.objects.get(id=update_course_id)
            if updatecourse.name == update_course_name:
                pass
            else:
                updatecourse.name = update_course_name
            if updatecourse.year == update_course_year:
                pass
            else:
                updatecourse.year = update_course_year
            
            updatecourse.save()
            return JsonResponse({'updated':'yes'})

        delete_course_id = request.POST.get('delete_course_id')
        if delete_course_id:
            deletecourse = Course.objects.get(id=delete_course_id)
            deletecourse.delete()
            deletestream = Stream.objects.filter(course=delete_course_id)
            deletestream.delete()
            return JsonResponse({'deleted':'yes'})
        
        if add_coursename and add_university != "Select University" and course_year:
            addcourse = Course(name = add_coursename,university = University.objects.get(id=add_university),year=course_year)
            addcourse.save()
            getlatestcourse = Course.objects.latest('id')
            allcoursecount = Course.objects.all().count()
            obj = {
                "id":getlatestcourse.id,
                "name":getlatestcourse.name,
                "count":allcoursecount
            }
            return JsonResponse({'course':obj})
        else:
            return JsonResponse({'course':'no','redirect':'yes'})
        
    params = {
        "main":main,
        "courses":Course.objects.all(),
        "university":University.objects.all()
    }
    return render(request,"super_admin/addcourse.html",params)

@login_required(login_url='/login/')
def AddUniversity(request):
    display = ""
    level_of_user = ""
    universitysame = ""
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_data_entry:
            if request.method == "POST":
                edit_university_id = request.POST.get('edit_university_id')
                edit_university_name = request.POST.get('edit_university_name')
                edit_university_address = request.POST.get('edit_university_address')
                edit_university_state = request.POST.get('edit_university_state')
                edit_university_city = request.POST.get('edit_university_city')
                edit_university_pincode = request.POST.get('edit_university_pincode')
                edit_university_logo = request.FILES.get('edit_university_logo', False)
                if edit_university_id:
                    print('inside editttttttttttttttt')
                    print(edit_university_id,edit_university_name,edit_university_address,edit_university_state,edit_university_city,edit_university_pincode,edit_university_logo)
                    try:
                        university = University.objects.get(id = edit_university_id)
                        university.university_name = edit_university_name
                        university.university_address = edit_university_address
                        university.university_state = edit_university_state
                        university.university_city = edit_university_city
                        university.university_pincode = edit_university_pincode
                        university.university_logo = edit_university_logo
                        university.save()
                        return JsonResponse({'updated':'yes'})
                    except University.DoesNotExist:
                        pass




                university_name = request.POST.get('university_name')
                university_address = request.POST.get('university_address')
                university_city = request.POST.get('university_city')
                university_state = request.POST.get('university_state')
                university_pincode = request.POST.get('university_pincode')
                try:
                    university_logo = request.FILES['university_logo']
                except KeyError:
                    university_logo = ''
                if university_name:
                    try:
                        checkuniversity = University.objects.get(university_name = university_name)
                        if checkuniversity:
                            universitysame = "yes"
                    except University.DoesNotExist:
                        removespaces = university_name.replace(' ','')
                        lowercase = removespaces.lower()
                        today = str(date.today())
                        today = today.replace('-','')
                        registrationID = str(lowercase) + str(today) + str('UNI')
                        adduniversity = University(university_name=university_name,university_address=university_address,university_city=university_city,university_state=university_state,university_pincode=university_pincode,university_logo=university_logo,registrationID=registrationID)
                        adduniversity.save()
                        return redirect('adduniversity')

                    
                
        params = {
            "display":display,
            "level_of_user":level_of_user,
            "universitysame":universitysame,
            "universities":University.objects.all()
        }
        return render(request,'super_admin/adduniversity.html',params)

@login_required(login_url='/login/')
def AddStream(request):
    print("in this function")
    display = ""
    level_of_user = ""
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_data_entry:
            cdata = ''
            special = ''
            allcourse = Course.objects.all()
            mainlist = []
            
            

            for i in allcourse:
                id = i.id
                
                coursename = i.name
                
                allstream = Stream.objects.filter(course=id)
                specializationlist = []
                for j in allstream:
                    obj = {
                        "type":j.name,
                        "duration":j.sem
                    }
                    specializationlist.append(obj)
                main = {
                    "id":i.id,
                    "coursename":coursename,
                    "specialization":specializationlist
                }
                mainlist.append(main)
                

            main = []
            university = University.objects.all()
            for i in university:
                courses = Course.objects.filter(university = i.id)
                obj = {
                    "university_name":i.university_name,
                    "course_name":courses
                }
                main.append(obj)
            
            if request.method == "POST":
                get_course = request.POST.get('get_course')
                if get_course:
                    try:
                        get_all_course = Course.objects.filter(Q(university=get_course))
                        print(get_all_course)
                        courseserializer = CourseSerializer(get_all_course,many=True)
                    except:
                        getcourse = ''
                    return JsonResponse({'course':courseserializer.data})
                    # distinctyear = []
                    # distinct_years = Course.objects.order_by('year').values_list('year', flat=True).distinct()
                    # for i in distinct_years:
                    #     distinctyear.append(i)
                    # try:
                    #     getcourse = Course.objects.filter(university = get_course)
                    #     courseserializer = CourseSerializer(getcourse,many=True)
                        
                    #     return JsonResponse({'course':courseserializer.data,'distinct_years':distinctyear})
                    # except Course.DoesNotExist:
                    #     pass
                get_course_university = request.POST.get('get_course_university')
                get_course_name = request.POST.get('get_course_name')
                if get_course_university and get_course_name:
                    # print(get_course_name,get_course_university)
                    get_all_course = Course.objects.filter(Q(university=get_course_university) & Q(year=get_course_name))
                    print(get_all_course)
                    courseserializer = CourseSerializer(get_all_course,many=True)
                    return JsonResponse({'course':courseserializer.data})
                stream_delete_id = request.POST.get('stream_delete_id')
                if stream_delete_id:
                    try:
                        getdeletestream = Stream.objects.get(id=stream_delete_id)
                        getdeletestream.delete()
                        return JsonResponse({'deleted':'yes'})
                    except Stream.DoesNotExist:
                        pass

                streamchange = request.POST.get('streamchange')
                if streamchange:
                    jsondata = json.loads(streamchange)
                    data = jsondata['data']
                    for i in data:
                        stream_id = i['stream_id']
                        stream_name = i['stream_name']
                        stream_year = i['stream_year']
                        try:
                            updatestream = Stream.objects.get(id=stream_id)
                            if updatestream.name == stream_name:
                                pass
                            else:
                                updatestream.name = stream_name
                            if updatestream.year == stream_year:
                                pass
                            else:
                                updatestream.year = stream_year
                            updatestream.save()
                        except Stream.DoesNotExist:
                            print("stream not available")
                    return JsonResponse({'updated':'yes'})
                

                university_name = request.POST.get('university_name')
                if university_name:
                    try:
                        getuniversity = University.objects.get(university_name = university_name)
                        getcourse = Course.objects.filter(university = getuniversity.id)
                        course_list = []
                        for i in getcourse:
                            courseserializer = {
                                "id":i.id,
                                "name":i.name
                            }
                            course_list.append(courseserializer)
                        return JsonResponse({'course':course_list})
                    except University.DoesNotExist:
                        print("University does not exist") 

                course_id = request.POST.get('course_id')
                if course_id:
                    print(course_id)
                    try:
                        getcourse = Course.objects.get(id = course_id)
                        getstream = Stream.objects.filter(course = getcourse.id)
                        streamserializer = StreamSerializer(getstream,many=True)
                        if len(getstream) != 0:
                            streamcount = len(getstream)
                            return JsonResponse({'data':streamserializer.data,'stream':'yes','streamcount':streamcount})
                        else:
                            return JsonResponse({'data':streamserializer.data,'stream':'no'})
                    except Course.DoesNotExist:
                        print("no course")
                
                coursename = request.POST.get('add_selectcourse')
                print('course', coursename)
                specialization = request.POST.get('add_specialization')
                duration = request.POST.get('add_duration')
                add_year = request.POST.get('add_year')
                if coursename and specialization and add_year:
                    print("gotcoursename and specialization")
                    getcourseid = Course.objects.get(id = coursename)
                    save_specialization = Stream(name=specialization,course=getcourseid,sem=duration,year=add_year)
                    save_specialization.save()
                    return JsonResponse({'added':'yes'})
        
            params = {
                "maindata":main,
                "courses":Course.objects.all(),
                "university":University.objects.all(),
                "main":mainlist,
                "display":display,
                "level_of_user":level_of_user
            }
            return render(request,"super_admin/addstream.html",params)

@login_required(login_url='/login/')
def AddStreamFees(request):
    
    if request.method == "POST":
        get_course = request.POST.get('get_course')
        data = request.POST.get('data')
        if get_course:
            try:
                fetch_filter_course_serialized(get_course)
                return JsonResponse({'course':fetch_filter_course_serialized(get_course)})
            except:
                pass
        if data:
            try:
                fetch_filter_stream_serialized(data)
                return JsonResponse({'stream':fetch_filter_stream_serialized(data)})
            except:
                pass

        courseid = request.POST.get('coursename')
        streamid = request.POST.get('streamname')
        semesters = request.POST.get('semesters')
        substream = request.POST.get('substreamname') ##added by Avani 14/08
        # print("data",courseid , streamid , semesters)
        if courseid and streamid and semesters:
            streamfees = Stream.objects.get(id = streamid)
            all_fees_data = []
            for i in range(1,int(semesters)+1):
                obj = {
                    "semester":i,
                    f"tution{i}":request.POST.get(f'tution{i}'),
                    f"exam{i}":request.POST.get(f'exam{i}'),
                    f"book{i}":request.POST.get(f'book{i}'),
                    f"resitting{i}":request.POST.get(f'resitting{i}'),
                    f"entrance{i}":request.POST.get(f'entrance{i}'),
                    f"extra{i}":request.POST.get(f'extra{i}'),
                    f"discount{i}":request.POST.get(f'discount{i}'),
                    f"total{i}":request.POST.get(f'total{i}')
                }
                all_fees_data.append(obj)
            create_new_semester_fees(request,streamid,substream,semesters,all_fees_data) ##modified by Avani 14/08
            return JsonResponse({'added':'yes'})
            
        year_coursename = request.POST.get('year_coursename')
        year_streamname = request.POST.get('year_streamname')
        year_substream = request.POST.get('year_substreamname') ##added by Avani 14/08
        year_semesters = request.POST.get('year_semesters')

        if year_coursename and year_streamname and year_semesters:
            all_fees_data = []
            for i in range(1,int(year_semesters)+1):
                obj = {
                    "year":i,
                    f"tution{i}":request.POST.get(f'year_tution{i}'),
                    f"exam{i}":request.POST.get(f'year_exam{i}'),
                    f"book{i}":request.POST.get(f'year_book{i}'),
                    f"resitting{i}":request.POST.get(f'year_resitting{i}'),
                    f"entrance{i}":request.POST.get(f'year_entrance{i}'),
                    f"extra{i}":request.POST.get(f'year_extra{i}'),
                    f"discount{i}":request.POST.get(f'year_discount{i}'),
                    f"total{i}":request.POST.get(f'year_total{i}')
                }
                all_fees_data.append(obj)
            create_new_year_fees(request,year_streamname,year_substream, year_semesters,all_fees_data)##modified by Avani 14/08
            
            return JsonResponse({'added':'yes'})
                


        year_c_name = request.POST.get('year_c_name')
        year_s_name = request.POST.get('year_s_name')
        year_ss_name = request.POST.get('year_ss_name')
        if year_c_name and year_s_name:
            print('in here')
            stream = Stream.objects.get(id =year_s_name)
            streamserializer = StreamSerializer(stream,many=False)
            try:
                if year_ss_name != '':
                    getfeesdetails = YearFees.objects.filter(stream = year_s_name, substream=year_ss_name)##modified by Avani 14/08
                else:
                    getfeesdetails = YearFees.objects.filter(stream = year_s_name, substream__isnull=True)
                feesdetails = YearFeesSerializer(getfeesdetails,many=True)
                return JsonResponse({'stream':streamserializer.data,'feesdetails':feesdetails.data})
                
            except YearFees.DoesNotExist:
                return JsonResponse({'stream':streamserializer.data})
            
            
                
            

        coursename = request.POST.get('c_name')
        streamname = request.POST.get('s_name')
        substreamname= request.POST.get('ss_name')##added by Avani 14/08
        
        if coursename and streamname:
            stream = Stream.objects.get(id =streamname)
            print("fgfgfg", stream)
            streamserializer = StreamSerializer(stream,many=False)
            try:
                if substreamname != '':
                    getfeesdetails = SemesterFees.objects.filter(stream = streamname, substream=substreamname)##modified by Avani 14/08
                else:
                     getfeesdetails = SemesterFees.objects.filter(stream = streamname, substream__isnull=True)
                feesdetails = SemesterFeesSerializer(getfeesdetails,many=True)
                return JsonResponse({'stream':streamserializer.data,'feesdetails':feesdetails.data})
            except SemesterFees.DoesNotExist:
                return JsonResponse({'stream':streamserializer.data})
            
                
    params = {
        "university":University.objects.all(),
        "course":Course.objects.all(),
    }
    return render(request,"super_admin/addstreamfees.html",params)

@login_required(login_url='/login/')    
def AddStudent(request): 
    if request.user.is_superuser or request.user.is_data_entry:
        roll= ''
        mobile_unique = ''
        get_course = request.GET.get('get_course')
        print("the course:  ", get_course)
        if get_course:
            #distinctyear = []
            #distinct_years = Course.objects.order_by('year').values_list('year', flat=True).distinct()
            #distinct_years = Stream.objects.order_by('sem').values_list('sem', flat=True).distinct()
            # for i in distinct_years:
            #     print(i)
            #     distinctyear.append(i)
            try:
                print("getting course")
                getcourse = Course.objects.filter(university = University.objects.get(registrationID = get_course))
                print("coursess", getcourse)
                courseserializer = CourseSerializer(getcourse,many=True)
                return JsonResponse({'course':courseserializer.data,})
            except Course.DoesNotExist:
                pass
        get_country_id = request.GET.get('get_country_id')
        if get_country_id:
            states = States.objects.filter(country_id = get_country_id)
            states_serializer = StatesSerializer(states,many=True)
            return JsonResponse({'states':states_serializer.data})

        get_cities_id = request.GET.get('get_cities_id')
        if get_cities_id:
            states = Cities.objects.filter( state_id = get_cities_id )
            states_serializer = CitiesSerializer(states,many=True)
            return JsonResponse({'cities':states_serializer.data})
        print("bbbbbb")
        selected_course = request.GET.get('selected_course')
        print(selected_course)
        if selected_course:
            get_course_id = Course.objects.get(id = selected_course)
            get_specialization = Stream.objects.filter(course = get_course_id.id)
            serializer = StreamSerializer(get_specialization,many=True)
            return JsonResponse({'data':serializer.data})
        if request.method == "POST":
            get_course_university = request.POST.get('get_course_university')
            get_course_name = request.POST.get('get_course_name')
            print("get_course_university", get_course_university)
            print("get_course_name", get_course_name)
            if get_course_university and get_course_name:
                # print(get_course_name,get_course_university)
                #get_all_course = Course.objects.filter(Q(university=University.objects.get(registrationID = get_course_university)) & Q(year=get_course_name))
                courses = Course.objects.filter(university=University.objects.get(registrationID = get_course_university), stream__sem=get_course_name).distinct()
                print("courses:", courses)
                courseserializer = CourseSerializer(courses,many=True)
                return JsonResponse({'course':courseserializer.data})
            course_name = request.POST.get('course_name')
            stream_name = request.POST.get('stream_name')
            
            study_pattern = request.POST.get('study_pattern')
            if course_name and stream_name and study_pattern:
                if study_pattern == "Semester":
                    try:
                        getstream = Stream.objects.get(id=stream_name)
                        obj = {
                            "semyear":int(getstream.sem) * 2,
                            "study_pattern":"Semester"
                        }
                        print(obj)
                        return JsonResponse({'obj':obj})
                    except Stream.DoesNotExist:
                        pass
                elif study_pattern == "Annual":
                    try:
                        getstream = Stream.objects.get(id=stream_name)
                        obj = {
                            "semyear":int(getstream.sem),
                            "study_pattern":"Annual"
                        }
                        print(obj)
                        return JsonResponse({'obj':obj})
                    except Stream.DoesNotExist:
                        pass
            get_c_id = request.POST.get('c_id')
            get_s_id = request.POST.get('s_id')
            get_ss_id = request.POST.get('ss_id')
            get_studypattern = request.POST.get('studypattern')
            get_semyear = request.POST.get('semyear')
            print("sssss ", get_ss_id)
            if get_c_id and get_s_id and get_studypattern and get_semyear:
                if get_studypattern == "Semester":
                    try:
                        if get_ss_id == '':
                            getsemesterfees = SemesterFees.objects.get(Q(stream=get_s_id) & Q(sem = get_semyear)& Q(substream__isnull=True))
                        else:
                            getsemesterfees = SemesterFees.objects.get(Q(stream=get_s_id) & Q(sem = get_semyear) & Q(substream=get_ss_id))##modified by Avani 14/08
                        getsemesterfeesserializer = SemesterFeesSerializer(getsemesterfees,many=False)
                        return JsonResponse({'data':getsemesterfeesserializer.data})
                    except SemesterFees.DoesNotExist:
                        pass
                elif get_studypattern == "Annual":
                    try:
                        if get_ss_id == '':
                            getyearfees = YearFees.objects.get(Q(stream=get_s_id) & Q(year = get_semyear)& Q(substream__isnull=True))
                        else:
                            getyearfees = YearFees.objects.get(Q(stream=get_s_id) & Q(year = get_semyear) & Q(substream=get_ss_id))##modified by Avani 14/08
                        getyearfeesserializer = YearFeesSerializer(getyearfees,many=False)
                        return JsonResponse({'data':getyearfeesserializer.data})
                    except YearFees.DoesNotExist:
                        pass
        #     xuniversity = request.POST.get('xuniversity')
        #     if xuniversity:
        #         print("university :",xuniversity)


            # course_name = request.POST.get('course_name')
            # stream_name = request.POST.get('stream_name')
            # study_pattern = request.POST.get('study_pattern')
            
        #     student_image = request.FILES.get('student_image', False)
        #     name = request.POST.get('name')
        #     dob = request.POST.get('dateofbirth')
        #     fathername = request.POST.get('fathers_name')
        #     mothername = request.POST.get('mothers_name')
        #     email = request.POST.get('email')
        #     alternateemail = request.POST.get('alternateemail')
        #     mobile = request.POST.get('mobile')
        #     alternatemobile1 = request.POST.get('alternatemobile1')
        #     gender = request.POST.get('gender')
        #     category = request.POST.get('category')
        
        #     address = request.POST.get('address')
        #     alternateaddress = request.POST.get('alternateaddress')
        #     nationality = request.POST.get('nationality')
        #     country = request.POST.get('countryId')
        #     state = request.POST.get('stateId')
        #     city = request.POST.get('cityId')
        #     pincode = request.POST.get('Pincode')

        #     counselor_name = request.POST.get('counselor_name')
        #     reference_name = request.POST.get('reference_name')
        #     university_enrollment_number = request.POST.get('university_enrollment_number')
        #     student_remarks = request.POST.get('student_remarks')
            


        #     fees = request.POST.get('fees')
        #     total_fees = request.POST.get('total_fees')
        #     course = request.POST.get('course')
        #     streamID = request.POST.get('Stream')
        #     studypattern = request.POST.get('studypattern')
        #     semyear = request.POST.get('semyear')
        #     university = request.POST.get('university')
        #     session = request.POST.get('session')
        #     entry_mode = request.POST.get('entry_mode')

        #     discount = request.POST.get('discount')
        #     # Documents
            
        #     totaldocuments = request.POST.get('totaldocuments')
        #     document1 = request.POST.get('document1')
        #     DocumentName1 = request.POST.get('DocumentName1')
        #     DocumentID1 = request.POST.get('DocumentID1')
        #     DocumentFront1 = request.FILES.get('DocumentFront1', False)
        #     DocumentBack1 = request.FILES.get('DocumentBack1', False)

        #     document2 = request.POST.get('document2')
        #     DocumentName2 = request.POST.get('DocumentName2')
        #     DocumentID2 = request.POST.get('DocumentID2')
        #     DocumentFront2 = request.FILES.get('DocumentFront2', False)
        #     DocumentBack2 = request.FILES.get('DocumentBack2', False)
            
        #     document3 = request.POST.get('document3')
        #     DocumentName3 = request.POST.get('DocumentName3')
        #     DocumentID3 = request.POST.get('DocumentID3')
        #     DocumentFront3 = request.FILES.get('DocumentFront3', False)
        #     DocumentBack3 = request.FILES.get('DocumentBack3', False)
            
        #     document4 = request.POST.get('document4')
        #     DocumentName4 = request.POST.get('DocumentName4')
        #     DocumentID4 = request.POST.get('DocumentID4')
        #     DocumentFront4 = request.FILES.get('DocumentFront4', False)
        #     DocumentBack4 = request.FILES.get('DocumentBack4', False)
            
        #     document5 = request.POST.get('document5')
        #     DocumentName5 = request.POST.get('DocumentName5')
        #     DocumentID5 = request.POST.get('DocumentID5')
        #     DocumentFront5 = request.FILES.get('DocumentFront5', False)
        #     DocumentBack5 = request.FILES.get('DocumentBack5', False)
            
            
            
            

            
        #     # Qualifications
        #     secondary_year = request.POST.get('secondary_year')
        #     secondary_board = request.POST.get('secondary_board')
        #     secondary_percentage = request.POST.get('secondary_percentage')
        #     secondary_document = request.FILES.get('secondary_document', False)
            
        #     sr_year = request.POST.get('sr_year')
        #     sr_board = request.POST.get('sr_board')
        #     sr_percentage = request.POST.get('sr_percentage')
        #     sr_document = request.FILES.get('sr_document', False)
            
        #     under_year = request.POST.get('under_year')
        #     under_board = request.POST.get('under_board')
        #     under_percentage = request.POST.get('under_percentage')
        #     under_document = request.FILES.get('under_document', False)
            
        #     post_year = request.POST.get('post_year')
        #     post_board = request.POST.get('post_board')
        #     post_percentage = request.POST.get('post_percentage')
        #     post_document = request.FILES.get('post_document', False)
            
        #     mphil_year = request.POST.get('mphil_year')
        #     mphil_board = request.POST.get('mphil_board')
        #     mphil_percentage = request.POST.get('mphil_percentage')
        #     mphil_document = request.FILES.get('mphil_document', False)
            
        #     others_year = request.POST.get('other_year')
        #     others_board = request.POST.get('other_board')
        #     others_percentage = request.POST.get('other_percentage')
        #     others_document = request.FILES.get('other_document', False)


        #     fee_reciept_type = request.POST.get('fee_reciept_type')
        #     other_data = request.POST.get('other_data')
        #     transaction_date = request.POST.get('transaction_date')
        #     payment_mode = request.POST.get('payment_mode')
        #     cheque_no = request.POST.get('cheque_no')
        #     bank_name = request.POST.get('bank_name')
        #     other_bank = request.POST.get('other_bank')
        #     remarks = request.POST.get('remarks')
        #     random_number = request.POST.get('random_number')
        #     print("random number :",random_number)
        #     print(name, dob , fathername , mothername , email , mobile , gender , category ,  address , nationality , country , state , city , pincode)
        #     print(secondary_year , secondary_board , secondary_percentage , secondary_document , sr_year , sr_board , sr_percentage , sr_document , under_year , under_board , under_percentage , under_document , post_year , post_board , post_percentage , post_document , mphil_year , mphil_board , mphil_percentage , mphil_document , others_year , others_board , others_percentage , others_document)
        #     print(course , streamID , studypattern , semyear , session , university)
        #     print(fees , fee_reciept_type , other_data , transaction_date , payment_mode , cheque_no , bank_name , remarks)

        #     if payment_mode == "Cheque":
        #         payment_status = "Not Realised"
        #     else:
        #         payment_status = "Realised"
                
            
            # if course_name and stream_name and study_pattern:
            #     print(course_name, stream_name , study_pattern)
            #     if study_pattern == "Semester":
            #         try:
            #             getstream = Stream.objects.get(id=stream_name)
            #             obj = {
            #                 "semyear":int(getstream.sem) * 2,
            #                 "study_pattern":"Semester"
            #             }
            #             print(obj)
            #             return JsonResponse({'obj':obj})
            #         except Stream.DoesNotExist:
            #             pass
            #     elif study_pattern == "Annual":
            #         try:
            #             getstream = Stream.objects.get(id=stream_name)
            #             obj = {
            #                 "semyear":int(getstream.sem),
            #                 "study_pattern":"Annual"
            #             }
            #             print(obj)
            #             return JsonResponse({'obj':obj})
            #         except Stream.DoesNotExist:
            #             pass

            # get_c_id = request.POST.get('c_id')
            # get_s_id = request.POST.get('s_id')
            # get_studypattern = request.POST.get('studypattern')
            # get_semyear = request.POST.get('semyear')
            # if get_c_id and get_s_id and get_studypattern and get_semyear:
            #     if get_studypattern == "Semester":
            #         try:
            #             getsemesterfees = SemesterFees.objects.get(Q(stream=get_s_id) & Q(sem = get_semyear))
            #             getsemesterfeesserializer = SemesterFeesSerializer(getsemesterfees,many=False)
            #             return JsonResponse({'data':getsemesterfeesserializer.data})
            #         except SemesterFees.DoesNotExist:
            #             pass
            #     elif get_studypattern == "Annual":
            #         try:
            #             getyearfees = YearFees.objects.get(Q(stream=get_s_id) & Q(year = get_semyear))
            #             getyearfeesserializer = YearFeesSerializer(getyearfees,many=False)
            #             return JsonResponse({'data':getyearfeesserializer.data})
            #         except YearFees.DoesNotExist:
            #             pass
            

            
        #     email_unique = ''
        #     try:
        #         random = Student.objects.latest('id')
        #     except Student.DoesNotExist:
        #         random = 1
        #     if random==1:
        #         enroll = 50000
        #         registration_id = 250000
        #     else:
        #         enroll = int(random.enrollment_id)+ 1
        #         registration_id = int(random.registration_id) +1
            
        #     if mobile:
        #         print("mobile : ",mobile)
        #         try:
        #             check_mobile = Student.objects.get(mobile = mobile)
        #             print("check_mobile ; ",check_mobile)
        #             mobile_unique = "no"
        #         except Student.DoesNotExist:
        #             mobile_unique = "yes"
        #         print("mobile_unique ; ",mobile_unique)
        #     if email:
        #         try:
        #             check_email = Student.objects.get(email = email)
        #             email_unique = "no"
        #         except Student.DoesNotExist:
        #             email_unique = "yes"
        #     if mobile_unique == "no" or email_unique == "no":
        #         obj = {
        #             'mobile_unique':mobile_unique,
        #             'email_unique':email_unique
        #         }
        #         print("data :",obj)
        #         return JsonResponse({'data':obj})
            
        #     print("student Image :",student_image)
        #     if mobile_unique == "yes" and email_unique == "yes":       
        #         if name or dob or fathername or mothername or email or mobile or gender or category  or address or nationality or country or state or city or pincode or (course and streamID and studypattern and semyear and session and entry_mode):
        #             create_student = Student(name=name,father_name=fathername,mother_name=mothername,dateofbirth=dob,mobile=mobile,alternate_mobile1=alternatemobile1,email=email,alternateemail=alternateemail,gender=gender,category=category,address=address,alternateaddress=alternateaddress,nationality=nationality,country=country,state=state,city=city,pincode=pincode,registration_id=registration_id,enrollment_id=enroll,image=student_image,verified=False,university=university,created_by=user_id,modified_by=user_id,student_remarks=student_remarks)
        #             create_student.save()
        #             print("student saved")
        #             print(name , dob , fathername , mothername , email , mobile , gender , category ,  address , nationality , country , state , city , pincode)
        #             print(secondary_year , secondary_board , secondary_percentage , secondary_document , sr_year , sr_board , sr_percentage , sr_document , under_year , under_board , under_percentage , under_document , post_year , post_board , post_percentage , post_document , mphil_year , mphil_board , mphil_percentage , mphil_document , others_year , others_board , others_percentage , others_document)
        #             print(course , streamID , studypattern , semyear , session , university)
        #             print(total_fees,fees , fee_reciept_type , other_data , transaction_date , payment_mode , cheque_no , bank_name , remarks)
                    
        #             # return JsonResponse({'added':'yes'})
                    
        #             # print("Total Fees : ",total_fees)
        #             # print("Paid Fees : ",fees)
        #             # pendingfees = abs(int(total_fees) - int(fees))
        #             # print("Pending Fees : ",pendingfees)

                        
                    
                    
                        
                        
            
        #             latest_student = Student.objects.latest('id')
        #             archive_post = StudentArchive()
        #             for field in latest_student._meta.fields:
        #                 setattr(archive_post, field.name, getattr(latest_student, field.name))
        #             archive_post.pk = None
        #             archive_post.save()
        #             if course and streamID and studypattern and semyear and session and entry_mode:
        #                 try:
        #                     stream = Stream.objects.get(id = streamID)
        #                 except Stream.DoesNotExist:
        #                     pass
        #                 totalsem = ""
        #                 if studypattern == "Semester":
        #                     totalsem = int(stream.sem) * 2
        #                     add_enrollmentdetails = Enrolled(student=student,course=course,stream=streamID,course_pattern=studypattern,session=session,entry_mode=entry_mode,total_semyear=totalsem,current_semyear=semyear)
        #                     add_enrollmentdetails.save()
        #                 elif studypattern == "Annual":
        #                     totalsem = int(stream.sem)
        #                     add_enrollmentdetails = Enrolled(student=latest_student.id,course=course,stream=streamID,course_pattern=studypattern,session=session,entry_mode=entry_mode,total_semyear=totalsem,current_semyear=semyear)
        #                     add_enrollmentdetails.save()
        #                 elif studypattern == "Full Course":
        #                     totalsem = int(stream.sem)
        #                     add_enrollmentdetails = Enrolled(student=latest_student.id,course=course,stream=streamID,course_pattern=studypattern,session=session,entry_mode=entry_mode,total_semyear=totalsem,current_semyear="1")
        #                     add_enrollmentdetails.save()
        #                 print("enrolled student")

        #             print(bank_name , other_bank)
        #             if document1 == "Select":
        #                 print("document not selected")
        #             else:
        #                 print(document1,DocumentName1,DocumentID1,DocumentFront1,DocumentBack1)
        #                 add_student_document = StudentDocuments(document=document1,document_name=DocumentName1,document_ID_no=DocumentID1,document_image_front= DocumentFront1,document_image_back = DocumentBack1,student=latest_student.id)
        #                 add_student_document.save()
                    
        #             if document2:
        #                 print(document2,DocumentName2,DocumentID2,DocumentFront2,DocumentBack2)
        #                 add_student_document = StudentDocuments(document=document2,document_name=DocumentName2,document_ID_no=DocumentID2,document_image_front= DocumentFront2,document_image_back = DocumentBack2,student=latest_student.id)
        #                 add_student_document.save()
                    
        #             if document3:
        #                 print(document3,DocumentName3,DocumentID3,DocumentFront3,DocumentBack3)
        #                 add_student_document = StudentDocuments(document=document3,document_name=DocumentName3,document_ID_no=DocumentID3,document_image_front= DocumentFront3,document_image_back = DocumentBack3,student=latest_student.id)
        #                 add_student_document.save()
                    
        #             if document4:
        #                 print(document4,DocumentName4,DocumentID4,DocumentFront4,DocumentBack4)
        #                 add_student_document = StudentDocuments(document=document4,document_name=DocumentName4,document_ID_no=DocumentID4,document_image_front= DocumentFront4,document_image_back = DocumentBack4,student=latest_student.id)
        #                 add_student_document.save()
                    
        #             if document5:
        #                 print(document5,DocumentName5,DocumentID5,DocumentFront5,DocumentBack5)
        #                 add_student_document = StudentDocuments(document=document5,document_name=DocumentName5,document_ID_no=DocumentID5,document_image_front= DocumentFront5,document_image_back = DocumentBack5,student=latest_student.id)
        #                 add_student_document.save()
                        
                    

                
        #             add_additional_details = AdditionalEnrollmentDetails(counselor_name=counselor_name,reference_name=reference_name,university_enrollment_id=university_enrollment_number,student=latest_student.id)
        #             add_additional_details.save()

        #             if studypattern == "Semester":
        #                 try:
        #                     getsemesterfees = SemesterFees.objects.filter(stream=streamID)
        #                     for i in getsemesterfees:
        #                         addstudentfees = StudentFees(student = latest_student.id,studypattern="Semester",stream=streamID,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.sem)
        #                         addstudentfees.save()
        #                 except SemesterFees.DoesNotExist:
        #                     pass
        #             elif studypattern == "Annual":
        #                 try:
        #                     print("Annual try reached")
        #                     getyearfees = YearFees.objects.filter(stream=streamID)
        #                     for i in getyearfees:
        #                         addstudentfees = StudentFees(student = latest_student.id,studypattern="Annual",stream=streamID,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.year)
        #                         addstudentfees.save()
        #                 except YearFees.DoesNotExist:
        #                     print("annual not reached")
                    
        #             if secondary_year or secondary_board or secondary_percentage or secondary_document or sr_year or sr_board or sr_percentage or sr_document or under_year or under_board or under_percentage or under_document or post_year or post_board or post_percentage or post_document or mphil_year or mphil_board or mphil_percentage or mphil_document or others_year or others_board or others_percentage or others_document:
        #                 setqualification = Qualification(student=latest_student.id,secondary_year = secondary_year,sr_year = sr_year,under_year=under_year,post_year=post_year,mphil_year=mphil_year,others_year = others_year,secondary_board=secondary_board,sr_board=sr_board,under_board=under_board,post_board=post_board,mphil_board=mphil_board,others_board=others_board,secondary_percentage=secondary_percentage,sr_percentage=sr_percentage,under_percentage=under_percentage,post_percentage=post_percentage,mphil_percentage=mphil_percentage,others_percentage=others_percentage,secondary_document=secondary_document,sr_document=sr_document,under_document=under_document,post_document=post_document,mphil_document=mphil_document,others_document=others_document)
        #                 setqualification.save()
        #                 print("qualification saved")
                    
        #             try:
        #                 getlatestreciept = PaymentReciept.objects.latest('id')
        #             except PaymentReciept.DoesNotExist:
        #                 getlatestreciept = "none"
        #             if getlatestreciept == "none":
        #                 transactionID = "TXT445FE101"
        #             else:
        #                 tid = getlatestreciept.transactionID
        #                 tranx = tid.replace("TXT445FE",'')
        #                 transactionID =  str("TXT445FE") + str(int(tranx) + 1)
        #             if fee_reciept_type == "Others":
        #                 reciept_type = other_data
        #             else:
        #                 reciept_type = fee_reciept_type
        #             if payment_mode == "Cheque":
        #                 if studypattern == "Full Course":
        #                     if fees and fee_reciept_type and transaction_date and payment_mode:
        #                         pending_fees = int(total_fees) - int(fees)
        #                         if pending_fees > 0:
        #                             print("positive")
        #                             pending_amount = pending_fees
        #                             advance_amount = 0
        #                         elif pending_fees == 0:
        #                             print("no pending semyear clear")
        #                             pending_amount = 0
        #                             advance_amount = 0
        #                         elif pending_fees < 0:
        #                             print("negative hence advance payment")
        #                             pending_amount = 0
        #                             advance_amount = abs(pending_fees)
        #                         if pending_fees == 0 | pending_fees < 0:
        #                             paymenttype = "Full Payment"
        #                         else:
        #                             paymenttype = "Part Payment"
        #                         if bank_name == "Others":
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=other_bank,
        #                                 semyearfees=total_fees,
        #                                 paidamount=0,
        #                                 pendingamount=total_fees,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear="1",
        #                                 uncleared_amount=fees,
        #                                 status=payment_status)
        #                         else:
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=bank_name,
        #                                 semyearfees=total_fees,
        #                                 paidamount=0,
        #                                 pendingamount=total_fees,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear="1",
        #                                 uncleared_amount=fees,
        #                                 status=payment_status)
        #                             add_payment_reciept.save()
        #                 else:
        #                     if fees and fee_reciept_type and transaction_date and payment_mode:
        #                         pending_fees = int(total_fees) - int(fees)
        #                         if pending_fees > 0:
        #                             print("positive")
        #                             pending_amount = pending_fees
        #                             advance_amount = 0
        #                         elif pending_fees == 0:
        #                             print("no pending semyear clear")
        #                             pending_amount = 0
        #                             advance_amount = 0
        #                         elif pending_fees < 0:
        #                             print("negative hence advance payment")
        #                             pending_amount = 0
        #                             advance_amount = abs(pending_fees)
                                    
        #                         if pending_fees == 0:
        #                             paymenttype = "Full Payment"
        #                         else:
        #                             paymenttype = "Part Payment"
        #                         if bank_name == "Others":
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=other_bank,
        #                                 semyearfees=total_fees,
        #                                 paidamount=0,
        #                                 pendingamount=total_fees,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear=semyear,
        #                                 uncleared_amount=fees,
        #                                 status=payment_status)
        #                             add_payment_reciept.save()
        #                         else:
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=bank_name,
        #                                 semyearfees=total_fees,
        #                                 paidamount=0,
        #                                 pendingamount=total_fees,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear=semyear,
        #                                 uncleared_amount=fees,
        #                                 status=payment_status)
        #                             add_payment_reciept.save()
        #             else:
        #                 if studypattern == "Full Course":
        #                     if fees and fee_reciept_type and transaction_date and payment_mode:
        #                         pending_fees = int(total_fees) - int(fees)
        #                         if pending_fees > 0:
        #                             print("positive")
        #                             pending_amount = pending_fees
        #                             advance_amount = 0
        #                         elif pending_fees == 0:
        #                             print("no pending semyear clear")
        #                             pending_amount = 0
        #                             advance_amount = 0
        #                         elif pending_fees < 0:
        #                             print("negative hence advance payment")
        #                             pending_amount = 0
        #                             advance_amount = abs(pending_fees)
        #                         if pending_fees == 0 | pending_fees < 0:
        #                             paymenttype = "Full Payment"
        #                         else:
        #                             paymenttype = "Part Payment"
        #                         if bank_name == "Others":
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=other_bank,
        #                                 semyearfees=total_fees,
        #                                 paidamount=fees,
        #                                 pendingamount=pending_amount,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear="1",
        #                                 status=payment_status)
        #                         else:
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=bank_name,
        #                                 semyearfees=total_fees,
        #                                 paidamount=fees,
        #                                 pendingamount=pending_amount,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear="1",
        #                                 status=payment_status)
        #                             add_payment_reciept.save()
        #                 else:
        #                     if fees and fee_reciept_type and transaction_date and payment_mode:
        #                         pending_fees = int(total_fees) - int(fees)
        #                         if pending_fees > 0:
        #                             print("positive")
        #                             pending_amount = pending_fees
        #                             advance_amount = 0
        #                         elif pending_fees == 0:
        #                             print("no pending semyear clear")
        #                             pending_amount = 0
        #                             advance_amount = 0
        #                         elif pending_fees < 0:
        #                             print("negative hence advance payment")
        #                             pending_amount = 0
        #                             advance_amount = abs(pending_fees)
                                    
        #                         if pending_fees == 0:
        #                             paymenttype = "Full Payment"
        #                         else:
        #                             paymenttype = "Part Payment"
        #                         if bank_name == "Others":
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=other_bank,
        #                                 semyearfees=total_fees,
        #                                 paidamount=fees,
        #                                 pendingamount=pending_amount,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear=semyear,
        #                                 status=payment_status)
        #                             add_payment_reciept.save()
        #                         else:
        #                             add_payment_reciept = PaymentReciept(
        #                                 student = latest_student.id,
        #                                 payment_for="Course Fees",
        #                                 payment_categories = "New",
        #                                 payment_type=paymenttype,
        #                                 fee_reciept_type=reciept_type,
        #                                 transaction_date= transaction_date,
        #                                 cheque_no=cheque_no,
        #                                 bank_name=bank_name,
        #                                 semyearfees=total_fees,
        #                                 paidamount=fees,
        #                                 pendingamount=pending_amount,
        #                                 advanceamount = advance_amount,
        #                                 transactionID = transactionID,
        #                                 paymentmode=payment_mode,
        #                                 remarks=remarks,
        #                                 session=session,
        #                                 semyear=semyear,
        #                                 status=payment_status)
        #                             add_payment_reciept.save()


                
        #             return JsonResponse({'added':'yes'})
            

            
    
        params = {
            "university":University.objects.all(),
            "mobile_unique":mobile_unique,
            "countries":Countries.objects.all(),
            "paymentmodes":PaymentModes.objects.filter(status=True), ## added by Avani on 09/08 -pick data from mastertables
            "feereceiptoptions":FeeReceiptOptions.objects.filter(status=True),## added by Avani on 09/08 -pick data from mastertables
            "banknames":BankNames.objects.filter(status=True),## added by Avani on 09/08 -pick data from mastertables
            "sessionnames":SessionNames.objects.filter(status=True),## added by Avani on 12/08 -pick data from mastertables
        }
        return render(request,"super_admin/addstudent.html",params)

def SaveStudent(request):
    print("inside save student function")
    if request.method == "POST":
        print("inside save student function post method")
        university = request.POST.get('university')
        student_image = request.FILES.get('student_image', False)
        name = request.POST.get('firstname')
        dob = request.POST.get('dob')
        fathername = request.POST.get('fathername')
        mothername = request.POST.get('mothername')
        email = request.POST.get('email')
        alternateemail = request.POST.get('alt_email')
        mobile = request.POST.get('mobile')
        alternatemobile = request.POST.get('alternatemobile1')
        gender = request.POST.get('gender')
        category = request.POST.get('category')
        address = request.POST.get('address')
        alternateaddress = request.POST.get('alternate_address')
        nationality = request.POST.get('nationality')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        pincode = request.POST.get('pincode')

        counselor_name = request.POST.get('counselor_name')
        reference_name = request.POST.get('reference_name')
        university_enrollment_number = request.POST.get('university_enroll_number')
        student_remarks = request.POST.get('student_remarks')
        course = request.POST.get('course')
        streamID = request.POST.get('Stream')
        substream = request.POST.get('SubStream')##added by Avani 14/08
        studypattern = request.POST.get('studypattern')
        semyear = request.POST.get('semyear')
        
        session = request.POST.get('session')
        entry_mode = request.POST.get('entry_mode')

        totaldocuments = request.POST.get('no_of_documents')

        secondary_year = request.POST.get('secondary_year')
        secondary_board = request.POST.get('secondary_board')
        secondary_percentage = request.POST.get('secondary_percentage')
        secondary_document = request.FILES.get('secondary_document', False)
        
        sr_year = request.POST.get('sr_year')
        sr_board = request.POST.get('sr_board')
        sr_percentage = request.POST.get('sr_percentage')
        sr_document = request.FILES.get('sr_document', False)
        
        under_year = request.POST.get('under_year')
        under_board = request.POST.get('under_board')
        under_percentage = request.POST.get('under_percentage')
        under_document = request.FILES.get('under_document', False)
        
        post_year = request.POST.get('post_year')
        post_board = request.POST.get('post_board')
        post_percentage = request.POST.get('post_percentage')
        post_document = request.FILES.get('post_document', False)
        
        mphil_year = request.POST.get('mphil_year')
        mphil_board = request.POST.get('mphil_board')
        mphil_percentage = request.POST.get('mphil_percentage')
        mphil_document = request.FILES.get('mphil_document', False)
      
        print(request.POST.get('qual_other_counter'), "cooooo")
        others_counter = request.POST.get('qual_other_counter')
        
        files_data = []
        for i in range(1,int(others_counter)+1):
            doctype = request.POST.get(f'other_{i}')
            year = request.POST.get(f'other_year_{i}')
            board = request.POST.get(f'other_board_{i}')
            docfile= request.FILES.get(f'other_document_{i}')
            file_data = {
                'doctype': doctype,
                'year': year,
                'board': board,
                'file_path': None
            }

            if docfile:
                # Save the file and get its path
                upload_dir = os.path.join(settings.MEDIA_ROOT, 'University_Documents')
                file_path = os.path.join(upload_dir, docfile.name)
                save_path = f'University_Documents/{docfile.name}'
                with open(file_path, 'wb+') as destination:
                    for chunk in docfile.chunks():
                        destination.write(chunk)
                file_data['file_path'] = save_path

            files_data.append(file_data)

            print("files_data", files_data)            

        fee_reciept_type = request.POST.get('fees_reciept')
        total_fees = request.POST.get('total_fees')
        fees = request.POST.get('fees')
        transaction_date = request.POST.get('date_of_transaction')
        payment_mode = request.POST.get('payment_mode')
        cheque_no = request.POST.get('cheque_no')
        bank_name = request.POST.get('bank_name')
        other_bank = request.POST.get('other_bank')
        other_data = request.POST.get('other_data')
        remarks = request.POST.get('remarks')

        try:
            random = Student.objects.latest('id')
            enroll = int(random.enrollment_id)+ 1
            registration_id = int(random.registration_id) +1
        except Student.DoesNotExist:
            enroll = 50000
            registration_id = 250000

        create_student = Student(
            name=name,
            father_name=fathername,
            mother_name=mothername,
            dateofbirth=dob,
            mobile=mobile,
            alternate_mobile1=alternatemobile,
            email=email,
            alternateemail=alternateemail,
            gender=gender,
            category=category,
            address=address,
            alternateaddress=alternateaddress,
            nationality=nationality,
            country=Countries.objects.get(id=country),
            state=States.objects.get(id=state),
            city=Cities.objects.get(id=city),
            pincode=pincode,
            registration_id=registration_id,
            enrollment_id=enroll,
            image=student_image,
            verified=True, #modified by Avani 18/07 client said all students are verified
            university=University.objects.get(registrationID=university),
            created_by=request.user.id,
            student_remarks=student_remarks,
            user_id= request.user.id, ## added by Avani for showing who enrolled the student
            is_enrolled = True,  ## added by Avani on 23/08 bypassing the verification process
            )
        create_student.save()

        ## added by Avani for creating user for student- bypassing the verification process
        student = Student.objects.get(enrollment_id = enroll)
        create_user = User(
                        email = student.email,
                        is_student = True,
                        password = make_password(student.email)
                    )
        create_user.save()
        print("student saved successfully :",create_user.id)
        student.user = User.objects.get(id=create_user.id)
        student.save()

        try:
            stream = Stream.objects.get(id = streamID)
        except Stream.DoesNotExist:
            pass
        ##added by Avani 14/08
        try:
            substream = SubStream.objects.get(id = substream)
        except:
            substream = None
        ##end of add
        totalsem = ""
        if studypattern == "Semester":
            totalsem = int(stream.sem) * 2
            add_enrollmentdetails = Enrolled(
                student=student,
                course=Course.objects.get(id=course),
                stream=stream,
                course_pattern=studypattern,
                session=session,
                entry_mode=entry_mode,
                total_semyear=totalsem,
                current_semyear=semyear,
                substream=substream##added by Avani 14/08
                )
        elif studypattern == "Annual":
            totalsem = int(stream.sem)
            add_enrollmentdetails = Enrolled(
                student=student,
                course=Course.objects.get(id=course),
                stream=stream,
                course_pattern=studypattern,
                session=session,
                entry_mode=entry_mode,
                total_semyear=totalsem,
                current_semyear=semyear,
                substream=substream##added by Avani 14/08
                )
        else:
            totalsem = int(stream.sem)
            add_enrollmentdetails = Enrolled(
                student=student,
                course=Course.objects.get(id=course),
                stream=stream,
                course_pattern=studypattern,
                session=session,
                entry_mode=entry_mode,
                total_semyear=totalsem,
                current_semyear="1",
                substream=substream##added by Avani 14/08
                )
        add_enrollmentdetails.save()    
        print("enrolled student")
        print(bank_name , other_bank)
        for i in range(1,int(totaldocuments)+1):
            
            document = request.POST.get(f'document{i}')
            DocumentName = request.POST.get(f'DocumentName{i}')
            ## added by Avani - if Document type is Other, store the name from the field other{i}
            if document == 'Other':
                document = request.POST.get(f'other{i}')
                print('other document',document)
            DocumentID = request.POST.get(f'DocumentID{i}')
            DocumentFront = request.FILES.get(f'DocumentFront{i}', False)
            DocumentBack = request.FILES.get(f'DocumentBack{i}', False)
            # print(document,DocumentName,DocumentID,DocumentFront,DocumentBack)
            add_student_document = StudentDocuments(document=document,document_name=DocumentName,document_ID_no=DocumentID,document_image_front= DocumentFront,document_image_back = DocumentBack,student=student)
            add_student_document.save()
            print('documents saveeeeeeeeeee in StudentDocuments -----------------')
    
        add_additional_details = AdditionalEnrollmentDetails(
            counselor_name=counselor_name,
            reference_name=reference_name,
            university_enrollment_id=university_enrollment_number,
            student=student)
        add_additional_details.save()

        if studypattern == "Semester":
            try:
                getsemesterfees = SemesterFees.objects.filter(stream=streamID, substream=substream)##modified by Avani 14/08
                for i in getsemesterfees:
                    addstudentfees = StudentFees(
                        student = student,
                        studypattern="Semester",
                        stream=Stream.objects.get(id=streamID),
                        tutionfees=i.tutionfees,
                        examinationfees=i.examinationfees,
                        bookfees=i.bookfees,
                        resittingfees=i.resittingfees,
                        entrancefees=i.entrancefees,
                        extrafees=i.extrafees,
                        discount=i.discount,
                        totalfees=i.totalfees,
                        sem=i.sem,
                        substream=i.substream,##added by Avani 14/08
                        )
                    addstudentfees.save()
            except SemesterFees.DoesNotExist:
                pass
        elif studypattern == "Annual":
            try:
                print("Annual try reached")
                getyearfees = YearFees.objects.filter(stream=streamID, substream=substream)##modified by Avani 14/08
                for i in getyearfees:
                    addstudentfees = StudentFees(
                        student = student,
                        studypattern="Annual",
                        stream=Stream.objects.get(id=streamID),
                        tutionfees=i.tutionfees,
                        examinationfees=i.examinationfees,
                        bookfees=i.bookfees,
                        resittingfees=i.resittingfees,
                        entrancefees=i.entrancefees,
                        extrafees=i.extrafees,
                        discount=i.discount,
                        totalfees=i.totalfees,
                        sem=i.year,
                        substream = i.substream##added by Avani 14/08
                        )
                    addstudentfees.save()
            except YearFees.DoesNotExist:
                print("annual not reached")
        print("inside save data above Qualification")
        if secondary_year or secondary_board or secondary_percentage or secondary_document or sr_year or sr_board or sr_percentage or sr_document or under_year or under_board or under_percentage or under_document or post_year or post_board or post_percentage or post_document or mphil_year or mphil_board or mphil_percentage or mphil_document or files_data:
            print("inside save data above Qualification ifffffff")
            setqualification = Qualification(
                others=files_data,
                student=student,
                secondary_year = secondary_year,
                sr_year = sr_year,
                under_year=under_year,
                post_year=post_year,
                mphil_year=mphil_year,
                #others_year = other_year,
                secondary_board=secondary_board,
                sr_board=sr_board,
                under_board=under_board,
                post_board=post_board,
                mphil_board=mphil_board,
                #others_board=other_board,
                secondary_percentage=secondary_percentage,
                sr_percentage=sr_percentage,
                under_percentage=under_percentage,
                post_percentage=post_percentage,
                mphil_percentage=mphil_percentage,
                #others_percentage=other_percentage,
                secondary_document=secondary_document,
                sr_document=sr_document,
                under_document=under_document,
                post_document=post_document,
                mphil_document=mphil_document,
                #others_document=other_document
                )
            setqualification.save()
            print("qualification saved")
        
        try:
            getlatestreciept = PaymentReciept.objects.latest('id')
            tid = getlatestreciept.transactionID
            tranx = tid.replace("TXT445FE",'')
            transactionID =  str("TXT445FE") + str(int(tranx) + 1)
        except PaymentReciept.DoesNotExist:
            transactionID = "TXT445FE101"
        if payment_mode == "Cheque":
            payment_status = "Not Realised"
        else:
            payment_status = "Realised"
            
        if fee_reciept_type == "Others":
            reciept_type = other_data
        else:
            reciept_type = fee_reciept_type
        obj = {

        }
        if payment_mode == "Cheque":
            if studypattern == "Full Course":
                if fees and fee_reciept_type and transaction_date and payment_mode:
                    pending_fees = int(total_fees) - int(fees)
                    if pending_fees > 0:
                        print("positive")
                        pending_amount = pending_fees
                        advance_amount = 0
                    elif pending_fees == 0:
                        print("no pending semyear clear")
                        pending_amount = 0
                        advance_amount = 0
                    elif pending_fees < 0:
                        print("negative hence advance payment")
                        pending_amount = 0
                        advance_amount = abs(pending_fees)
                    if pending_fees == 0 | pending_fees < 0:
                        paymenttype = "Full Payment"
                    else:
                        paymenttype = "Part Payment"
                    if bank_name == "Others":
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=other_bank,
                            semyearfees=total_fees,
                            paidamount=0,
                            pendingamount=total_fees,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear="1",
                            uncleared_amount=fees,
                            status=payment_status)
                    else:
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=bank_name,
                            semyearfees=total_fees,
                            paidamount=0,
                            pendingamount=total_fees,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear="1",
                            uncleared_amount=fees,
                            status=payment_status)
                    add_payment_reciept.save()
            else:
                if fees and fee_reciept_type and transaction_date and payment_mode:
                    pending_fees = int(total_fees) - int(fees)
                    if pending_fees > 0:
                        print("positive")
                        pending_amount = pending_fees
                        advance_amount = 0
                    elif pending_fees == 0:
                        print("no pending semyear clear")
                        pending_amount = 0
                        advance_amount = 0
                    elif pending_fees < 0:
                        print("negative hence advance payment")
                        pending_amount = 0
                        advance_amount = abs(pending_fees)
                        
                    if pending_fees == 0:
                        paymenttype = "Full Payment"
                    else:
                        paymenttype = "Part Payment"
                    if bank_name == "Others":
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=other_bank,
                            semyearfees=total_fees,
                            paidamount=0,
                            pendingamount=total_fees,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear=semyear,
                            uncleared_amount=fees,
                            status=payment_status)
                        add_payment_reciept.save()
                    else:
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=bank_name,
                            semyearfees=total_fees,
                            paidamount=0,
                            pendingamount=total_fees,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear=semyear,
                            uncleared_amount=fees,
                            status=payment_status)
                        add_payment_reciept.save()
        else:
            if studypattern == "Full Course":
                if fees and fee_reciept_type and transaction_date and payment_mode:
                    pending_fees = int(total_fees) - int(fees)
                    if pending_fees > 0:
                        print("positive")
                        pending_amount = pending_fees
                        advance_amount = 0
                    elif pending_fees == 0:
                        print("no pending semyear clear")
                        pending_amount = 0
                        advance_amount = 0
                    elif pending_fees < 0:
                        print("negative hence advance payment")
                        pending_amount = 0
                        advance_amount = abs(pending_fees)
                    if pending_fees == 0 | pending_fees < 0:
                        paymenttype = "Full Payment"
                    else:
                        paymenttype = "Part Payment"
                    if bank_name == "Others":
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=other_bank,
                            semyearfees=total_fees,
                            paidamount=fees,
                            pendingamount=pending_amount,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear="1",
                            status=payment_status)
                    else:
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=bank_name,
                            semyearfees=total_fees,
                            paidamount=fees,
                            pendingamount=pending_amount,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear="1",
                            status=payment_status)
                        add_payment_reciept.save()
            else:
                if fees and fee_reciept_type and transaction_date and payment_mode:
                    pending_fees = int(total_fees) - int(fees)
                    if pending_fees > 0:
                        print("positive")
                        pending_amount = pending_fees
                        advance_amount = 0
                    elif pending_fees == 0:
                        print("no pending semyear clear")
                        pending_amount = 0
                        advance_amount = 0
                    elif pending_fees < 0:
                        print("negative hence advance payment")
                        pending_amount = 0
                        advance_amount = abs(pending_fees)
                        
                    if pending_fees == 0:
                        paymenttype = "Full Payment"
                    else:
                        paymenttype = "Part Payment"
                    if bank_name == "Others":
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=other_bank,
                            semyearfees=total_fees,
                            paidamount=fees,
                            pendingamount=pending_amount,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear=semyear,
                            status=payment_status)
                        add_payment_reciept.save()
                    else:
                        add_payment_reciept = PaymentReciept(
                            student = student,
                            payment_for="Course Fees",
                            payment_categories = "New",
                            payment_type=paymenttype,
                            fee_reciept_type=reciept_type,
                            transaction_date= transaction_date,
                            cheque_no=cheque_no,
                            bank_name=bank_name,
                            semyearfees=total_fees,
                            paidamount=fees,
                            pendingamount=pending_amount,
                            advanceamount = advance_amount,
                            transactionID = transactionID,
                            paymentmode=payment_mode,
                            remarks=remarks,
                            session=session,
                            semyear=semyear,
                            status=payment_status)
                        add_payment_reciept.save()

        
        print(course , streamID , studypattern , semyear , session ,entry_mode, university)

        # commented



        # print(total_fees,fees , fee_reciept_type  , transaction_date , payment_mode , cheque_no , bank_name ,other_bank, remarks)
    ## modified by Avani on 23/07 - redirect to registerd students page. Client does not want pending verification.
    #return redirect('pending_verification')
    return redirect('viewregister')

def check_mobile_unique(request):
    if request.method == "POST":
        student = Student.objects.filter(Q(mobile__contains=request.POST.get('mobile_unique')) | Q(alternate_mobile1__contains=request.POST.get('mobile_unique')))
        if student:
             ## added by Avani on 19/07 - need to show that another student with same mobile has been registered
            user_id = student[0].created_by
            username = User.objects.get(id = user_id)
            return JsonResponse({'unique':'false', 'studentname': student[0].name, 'createdby': username.email })
        else:
            return JsonResponse({'unique':'true'})

def check_email_unique(request):
    if request.method == "POST":
        student = Student.objects.filter(Q(email__contains=request.POST.get('email_unique')) | Q(alternateemail__contains=request.POST.get('email_unique')))
        if student:
            ## added by Avani on 19/07 - need to show that another student with same email has been registered
            user_id = student[0].created_by
            username = User.objects.get(id = user_id)
            return JsonResponse({'unique':'false', 'studentname': student[0].name, 'createdby': username.email })
        else:
            return JsonResponse({'unique':'true'})

@login_required(login_url='/login/')
def PendingStudentVerification(request):
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            verified = request.POST.get('verified')
            if verified:
                try:
                    getstudent = Student.objects.get(enrollment_id = verified)
                    getstudent.verified = True
                    getstudent.is_enrolled = True
                    create_user = User(
                        email = getstudent.email,
                        is_student = True,
                        password = make_password(getstudent.email)
                    )
                    create_user.save()
                    print("student saved successfully :",create_user.id)
                    getstudent.user = User.objects.get(id=create_user.id)
                    getstudent.save()
                except Student.DoesNotExist:
                    pass
            delete_student = request.POST.get('delete_student')
            if delete_student:
                # print("delete_student : ",delete_student)
                getstudent = Student.objects.get(enrollment_id = delete_student)
                # print("getstudent : ",getstudent)
                # print("Delete Student : ",getstudent)
                getenroll = Enrolled.objects.get(student = getstudent)
                studentdocuments = StudentDocuments.objects.filter(student = getstudent)
                studentfees = StudentFees.objects.filter(student = getstudent)
                qualification = Qualification.objects.filter(student = getstudent)
                additionalenrollmentdetails = AdditionalEnrollmentDetails.objects.filter(student = getstudent)
                courier = Courier.objects.filter(student = getstudent)
                paymentreciept = PaymentReciept.objects.filter(student = getstudent)
                studentsyllabus = StudentSyllabus.objects.filter(student = getstudent)
                getstudent.delete()
                getenroll.delete()
                studentdocuments.delete()
                studentfees.delete()
                qualification.delete()
                additionalenrollmentdetails.delete()
                courier.delete()
                paymentreciept.delete()
                studentsyllabus.delete()
                return JsonResponse({'deleted':'yes'})
        params = {
            "students":Student.objects.filter(verified=False)
        }

    return render(request,"super_admin/pending_verification.html",params)

def testing_api(request):
    allstudent = Student.objects.filter(verified = False).order_by('-id')
    studentlist = []
    
    for i in allstudent:
        student_id = i.id
        university = University.objects.get(id=i.university.id)
        try:
            enrolled = Enrolled.objects.get(student = i)
            # print(enrolled)
        except Enrolled.DoesNotExist:
            # print("error beacuse of student id :",student_id)
            enrolled = ''
        course = Course.objects.get(id = enrolled.course.id)
        stream = Stream.objects.get(id = enrolled.stream.id)
        # logger.info(enrolled)
        # print("print enrolled error:",enrolled)
        # print(enrolled.course_pattern)
        obj = {
            "name":str(i.name),
            "father_name":i.father_name,
            "mother_name":i.mother_name,
            "dateofbirth":datetime.strptime(str(i.dateofbirth), '%Y-%m-%d').strftime('%d/%m/%Y'),
            "mobile":i.mobile,
            "email":i.email,
            "gender":i.gender,
            "category":i.category,
            "enrollment_id":i.enrollment_id,
            "enrollment_date":str(i.enrollment_date),
            "registration_id":i.registration_id,
            "course":course.name,
            "stream":stream.name,
            "semyear":enrolled.current_semyear,
            "course_pattern":enrolled.course_pattern,
            "session":enrolled.session,
            "entry_mode":enrolled.entry_mode,
            "university":university.university_name
        }
        studentlist.append(obj)
    
    return JsonResponse({'data':studentlist})


def PendingStudentVerification_fees(request):
    if request.method == "POST":
        fees = request.POST.get('fees')
        if fees:
            try:
                getstudent = Student.objects.get(enrollment_id = fees)
                try:
                    getfeespaid = PaymentReciept.objects.filter(student = getstudent)
                    feeserializer = PaymentRecieptSerializer(getfeespaid,many=True)
                    return JsonResponse({'fees':feeserializer.data})
                except PaymentReciept.DoesNotExist:
                    getfeespaid = "none"
            except Student.DoesNotExist:
                pass
        return JsonResponse({'fees':"fees"})

def PendingStudentVerification_fees_viewreciept(request):
    if request.method == "POST":
        getpayment_reciept = request.POST.get('getpayment_reciept')
        if getpayment_reciept:
            # print(getpayment_reciept)
            get_reciept = PaymentReciept.objects.get(id = getpayment_reciept)
            paymentrecieptserializer = PaymentRecieptSerializer(get_reciept,many=False)
            view_student = Student.objects.get(id=get_reciept.student.id)
            view_enroll = Enrolled.objects.get(student=view_student.id)
            view_course = Course.objects.get(id = view_enroll.course.id)
            view_stream = Stream.objects.get(id = view_enroll.stream.id)
            temp = {
                "name":view_student.name,
                "email":view_student.email,
                "mobile":view_student.mobile,
                "enrollment_id":view_student.enrollment_id,
                "address":view_student.address,
                "country":view_student.country.name,
                "state":view_student.state.name,
                "city":view_student.city.name,
                "pincode":view_student.pincode,
                "course":view_course.name,
                "stream":view_stream.name
            }
            return JsonResponse({'view_reciept':paymentrecieptserializer.data,'personal':temp})

def PendingStudentVerification_get_documents(request):
    if request.method == "POST":
        enrollid = request.POST.get('enrollid')
        if enrollid:
            try:
                getstudent = Student.objects.get(enrollment_id = enrollid)
                getqualification = Qualification.objects.get(student = getstudent.id)
                qualificationserializer = QualificationSerializer(getqualification,many=False)
                return JsonResponse({'student':'yes','qualification':qualificationserializer.data})
            except Student.DoesNotExist:
                return JsonResponse({'student':'no'})

@login_required(login_url='/login/')    
def AddStudentQuick(request):
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            xuniversity = request.POST.get('xuniversity')
            if xuniversity:
                print("university :",xuniversity)



            
            university = request.POST.get('university')
            student_image = request.FILES.get('student_image', False)
            name = request.POST.get('name')
            father_name = request.POST.get('father_name')
            dob = request.POST.get('dateofbirth')
            
            email = request.POST.get('email')
            mobile = request.POST.get('mobile')
            counselor_name = request.POST.get('counselor_name')
            university_enrollment_number = request.POST.get('university_enrollment_number')
            student_remarks = request.POST.get('student_remarks')

            fees = request.POST.get('fees')
            total_fees = request.POST.get('total_fees')
            course = request.POST.get('course')
            streamID = request.POST.get('Stream')
            studypattern = request.POST.get('studypattern')
            semyear = request.POST.get('semyear')
            session = request.POST.get('session')
            entry_mode = request.POST.get('entry_mode')
            substream= request.POST.get('SubStream') ## added by Avani on 14/08 
            

            fee_reciept_type = request.POST.get('fee_reciept_type')
            other_data = request.POST.get('other_data')
            transaction_date = request.POST.get('transaction_date')
            payment_mode = request.POST.get('payment_mode')
            cheque_no = request.POST.get('cheque_no')
            bank_name = request.POST.get('bank_name')
            other_bank = request.POST.get('other_bank')
            remarks = request.POST.get('remarks')
            random_number = request.POST.get('random_number')
            is_quick_register= request.POST.get('is_quick_register') #added by Avani 12/07
            print("is_quick_register: " , is_quick_register)
            # print("random number :",random_number)
            # print( university , student_image , name ,dob  , email , mobile ,university_enrollment_number , student_remarks )
            # print(course , streamID , studypattern , semyear , session , entry_mode)
            # print(fees , fee_reciept_type , other_data , transaction_date , payment_mode , cheque_no , bank_name , remarks)

            
            
            

            
            

            
            email_unique = ''
            try:
                random = Student.objects.latest('id')
            except Student.DoesNotExist:
                random = 1
            if random==1:
                enroll = 50000
                registration_id = 250000
            else:
                enroll = int(random.enrollment_id)+ 1
                registration_id = int(random.registration_id) +1
            
            if mobile:
                print("mobile : ",mobile)
                try:
                    check_mobile = Student.objects.get(Q(mobile = mobile) | Q(alternate_mobile1 = mobile))
                    print("check_mobile ; ",check_mobile)
                    mobile_unique = "no"
                    ## added by Avani on 19/07 to get student name and created by details if not unique
                    user_id = check_mobile.created_by
                    studentname= check_mobile.name
                    username = User.objects.get(id = user_id)
                except Student.DoesNotExist:
                    mobile_unique = "yes"
                print("mobile_unique ; ",mobile_unique)
            if email:
                try:
                    check_email = Student.objects.get(email = email)
                    email_unique = "no"
                    ## added by Avani on 19/07 to get student name and created by details if not unique
                    user_id = check_email.created_by
                    studentname= check_email.name
                    username = User.objects.get(id = user_id)
                except Student.DoesNotExist:
                    email_unique = "yes"
            if mobile_unique == "no" or email_unique == "no":
                obj = {
                    'mobile_unique':mobile_unique,
                    'email_unique':email_unique,
                    'studentname': studentname,
                    'createdby':username.email
                }
                print("data :",obj)
                return JsonResponse({'data':obj})
            

            if mobile_unique == "yes" and email_unique == "yes":
                if name or dob  or email or mobile  or (course and streamID and studypattern and semyear and session and entry_mode):
                    ##modified by Avani to not save none, undefined etc in database.
                    #create_student = Student(name=name,father_name=father_name,dateofbirth=dob,mobile=mobile,email=email,country=Countries.objects.get(id=101),state=States.objects.get(id=22),city=Cities.objects.get(id=2707),address="Mumbai",nationality="Indian",registration_id=registration_id,enrollment_id=enroll,image=student_image,verified=False,university=University.objects.get(registrationID=university),created_by=request.user.id,modified_by=request.user.id,student_remarks=student_remarks, is_quick_register=is_quick_register)
                    create_student = Student(name=name,father_name=father_name,mother_name="",dateofbirth=dob,mobile=mobile,email=email,country=Countries.objects.get(id=101),state=States.objects.get(id=22),city=Cities.objects.get(id=2707),address="",nationality="",registration_id=registration_id,enrollment_id=enroll,image=student_image,verified=True,university=University.objects.get(registrationID=university),created_by=request.user.id,modified_by=request.user.id,student_remarks=student_remarks, is_quick_register=is_quick_register, alternate_mobile1='', alternateemail='', gender='', pincode='', category='',alternateaddress='',registration_number='', is_enrolled=True)
                    create_student.save()
                    print("student saved")
                    ## added by Avani for creating user for student- bypassing the verification process
                    student = Student.objects.get(enrollment_id = enroll)
                    create_user = User(
                        email = student.email,
                        is_student = True,
                        password = make_password(student.email)
                    )
                    create_user.save()
                    print("student saved successfully :",create_user.id)
                    student.user = User.objects.get(id=create_user.id)
                    student.save()

                    # print(name ,father_name, dob , email , mobile )
                    
                    # print(course , streamID , studypattern , semyear , session , university)
                    # print(fees , fee_reciept_type , other_data , transaction_date , payment_mode , cheque_no , bank_name , remarks)
                    
                    if payment_mode == "Cheque":
                        payment_status = "Not Realised"
                        uncleared_amount = fees
                        paidfees = 0
                    else:
                        payment_status = "Realised"
                        uncleared_amount = 0
                        paidfees = fees            
              
                    latest_student = Student.objects.latest('id')
                    if course and streamID and studypattern and semyear and session and entry_mode:
                        try:
                            stream = Stream.objects.get(id = streamID)
                        except Stream.DoesNotExist:
                            pass
                        ##added by Avani 14/08
                        try:
                                substream = SubStream.objects.get(id=substream)
                        except:
                                substream = None
                        #end of add
                        totalsem = ""
                        if studypattern == "Semester":
                            totalsem = int(stream.sem) * 2
                            
                            add_enrollmentdetails = Enrolled(student=latest_student,course=Course.objects.get(id=course),stream=Stream.objects.get(id=streamID),course_pattern=studypattern,session=session,entry_mode=entry_mode,total_semyear=totalsem,current_semyear=semyear, substream=substream)##modified by Avani 14/08
                            add_enrollmentdetails.save()
                        elif studypattern == "Annual":
                            totalsem = int(stream.sem)
                            add_enrollmentdetails = Enrolled(student=latest_student,course=Course.objects.get(id=course),stream=Stream.objects.get(id=streamID),course_pattern=studypattern,session=session,entry_mode=entry_mode,total_semyear=totalsem,current_semyear=semyear, substream=substream)##modified by Avani 14/08
                            add_enrollmentdetails.save()
                        elif studypattern == "Full Course":
                            totalsem = int(stream.sem)
                            add_enrollmentdetails = Enrolled(student=latest_student,course=Course.objects.get(id=course),stream=Stream.objects.get(id=streamID),course_pattern=studypattern,session=session,entry_mode=entry_mode,total_semyear=totalsem,current_semyear="1", substream=substream)##modified by Avani 14/08
                            add_enrollmentdetails.save()
                        print("enrolled student")

                    print(bank_name , other_bank)
                    
                        
                    

                    ##modified to not pass none to database - Avani 15/07
                    #add_additional_details = AdditionalEnrollmentDetails(student=latest_student,counselor_name=counselor_name,university_enrollment_id=university_enrollment_number)
                    add_additional_details = AdditionalEnrollmentDetails(student=latest_student,counselor_name=counselor_name,university_enrollment_id=university_enrollment_number, reference_name='')
                    add_additional_details.save()

                    if studypattern == "Semester":
                        try:
                            getsemesterfees = SemesterFees.objects.filter(stream=streamID, substream=substream) ## modified by Avani 14/08
                            for i in getsemesterfees:
                                addstudentfees = StudentFees(student = latest_student,studypattern="Semester",stream=Stream.objects.get(id=streamID),tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.sem, substream=i.substream)##modified by Avani 14/08
                                addstudentfees.save()
                        except SemesterFees.DoesNotExist:
                            pass
                    elif studypattern == "Annual":
                        try:
                            print("Annual try reached")
                            getyearfees = YearFees.objects.filter(stream=streamID, substream=substream)
                            for i in getyearfees:
                                addstudentfees = StudentFees(student = latest_student,studypattern="Annual",stream=Stream.objects.get(id=streamID),tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.year, substream=i.substream)##modified by Avani 14/08
                                addstudentfees.save()
                        except YearFees.DoesNotExist:
                            print("annual not reached")

                    try:
                        getlatestreciept = PaymentReciept.objects.latest('id')
                    except PaymentReciept.DoesNotExist:
                        getlatestreciept = "none"
                    if getlatestreciept == "none":
                        transactionID = "TXT445FE101"
                    else:
                        tid = getlatestreciept.transactionID
                        tranx = tid.replace("TXT445FE",'')
                        transactionID =  str("TXT445FE") + str(int(tranx) + 1)
                    if fee_reciept_type == "Others":
                        reciept_type = other_data
                    else:
                        reciept_type = fee_reciept_type
                    if payment_mode == "Cheque":
                        if studypattern == "Full Course":
                            if fees and fee_reciept_type and transaction_date and payment_mode:
                                pending_fees = int(total_fees) - int(fees)
                                if pending_fees > 0:
                                    print("positive")
                                    pending_amount = pending_fees
                                    advance_amount = 0
                                elif pending_fees == 0:
                                    print("no pending semyear clear")
                                    pending_amount = 0
                                    advance_amount = 0
                                elif pending_fees < 0:
                                    print("negative hence advance payment")
                                    pending_amount = 0
                                    advance_amount = abs(pending_fees)
                                if pending_fees == 0 | pending_fees < 0:
                                    paymenttype = "Full Payment"
                                else:
                                    paymenttype = "Part Payment"
                                if bank_name == "Others":
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=other_bank,
                                        semyearfees=total_fees,
                                        paidamount=0,
                                        pendingamount=total_fees,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear="1",
                                        uncleared_amount=fees,
                                        status=payment_status)
                                    add_payment_reciept.save()
                                else:
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=bank_name,
                                        semyearfees=total_fees,
                                        paidamount=0,
                                        pendingamount=total_fees,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear="1",
                                        uncleared_amount=fees,
                                        status=payment_status)
                                    add_payment_reciept.save()
                        else:
                            if fees and fee_reciept_type and transaction_date and payment_mode:
                                pending_fees = int(total_fees) - int(fees)
                                if pending_fees > 0:
                                    print("positive")
                                    pending_amount = pending_fees
                                    advance_amount = 0
                                elif pending_fees == 0:
                                    print("no pending semyear clear")
                                    pending_amount = 0
                                    advance_amount = 0
                                elif pending_fees < 0:
                                    print("negative hence advance payment")
                                    pending_amount = 0
                                    advance_amount = abs(pending_fees)
                                    
                                if pending_fees == 0:
                                    paymenttype = "Full Payment"
                                else:
                                    paymenttype = "Part Payment"
                                if bank_name == "Others":
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=other_bank,
                                        semyearfees=total_fees,
                                        paidamount=0,
                                        pendingamount=total_fees,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear=semyear,
                                        uncleared_amount=fees,
                                        status=payment_status)
                                    add_payment_reciept.save()
                                else:
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=bank_name,
                                        semyearfees=total_fees,
                                        paidamount=0,
                                        pendingamount=total_fees,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear=semyear,
                                        uncleared_amount=fees,
                                        status=payment_status)
                                    add_payment_reciept.save()
                    else:
                        if studypattern == "Full Course":
                            if fees and fee_reciept_type and transaction_date and payment_mode:
                                pending_fees = int(total_fees) - int(fees)
                                if pending_fees > 0:
                                    print("positive")
                                    pending_amount = pending_fees
                                    advance_amount = 0
                                elif pending_fees == 0:
                                    print("no pending semyear clear")
                                    pending_amount = 0
                                    advance_amount = 0
                                elif pending_fees < 0:
                                    print("negative hence advance payment")
                                    pending_amount = 0
                                    advance_amount = abs(pending_fees)
                                if pending_fees == 0 | pending_fees < 0:
                                    paymenttype = "Full Payment"
                                else:
                                    paymenttype = "Part Payment"
                                if bank_name == "Others":
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=other_bank,
                                        semyearfees=total_fees,
                                        paidamount=fees,
                                        pendingamount=pending_amount,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear="1",
                                        status=payment_status)
                                    add_payment_reciept.save()
                                else:
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=bank_name,
                                        semyearfees=total_fees,
                                        paidamount=fees,
                                        pendingamount=pending_amount,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear="1",
                                        status=payment_status)
                                    add_payment_reciept.save()
                        else:
                            if fees and fee_reciept_type and transaction_date and payment_mode:
                                pending_fees = int(total_fees) - int(fees)
                                if pending_fees > 0:
                                    print("positive")
                                    pending_amount = pending_fees
                                    advance_amount = 0
                                elif pending_fees == 0:
                                    print("no pending semyear clear")
                                    pending_amount = 0
                                    advance_amount = 0
                                elif pending_fees < 0:
                                    print("negative hence advance payment")
                                    pending_amount = 0
                                    advance_amount = abs(pending_fees)
                                    
                                if pending_fees == 0:
                                    paymenttype = "Full Payment"
                                else:
                                    paymenttype = "Part Payment"
                                if bank_name == "Others":
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=other_bank,
                                        semyearfees=total_fees,
                                        paidamount=fees,
                                        pendingamount=pending_amount,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear=semyear,
                                        status=payment_status)
                                    add_payment_reciept.save()
                                else:
                                    add_payment_reciept = PaymentReciept(
                                        student = latest_student,
                                        payment_for="Course Fees",
                                        payment_categories = "New",
                                        payment_type=paymenttype,
                                        fee_reciept_type=reciept_type,
                                        transaction_date= transaction_date,
                                        cheque_no=cheque_no,
                                        bank_name=bank_name,
                                        semyearfees=total_fees,
                                        paidamount=fees,
                                        pendingamount=pending_amount,
                                        advanceamount = advance_amount,
                                        transactionID = transactionID,
                                        paymentmode=payment_mode,
                                        remarks=remarks,
                                        session=session,
                                        semyear=semyear,
                                        status=payment_status)
                                    add_payment_reciept.save()
                                


                
                    return JsonResponse({'added':'yes'})
        params = {
            "university":University.objects.all(),
            "paymentmodes":PaymentModes.objects.filter(status=True), ## added by Avani on 09/08 -pick data from mastertables
            "feereceiptoptions":FeeReceiptOptions.objects.filter(status=True),## added by Avani on 09/08 -pick data from mastertables
            "banknames":BankNames.objects.filter(status=True),## added by Avani on 09/08 -pick data from mastertables
            "sessionnames":SessionNames.objects.filter(status=True),## added by Avani on 12/08 -pick data from mastertables
        }
        return render(request,"super_admin/quick_addstudent.html",params)

@login_required(login_url='/login/')
def AdvanceSearch(request):
    temp = []
    if request.user.is_superuser or request.user.is_data_entry:
        
        if 'term' in request.GET:
            q = request.GET.get('term', '')
            getstudents = Student.objects.filter(enrollment_date__icontains = q).order_by('-id')
            count = getstudents.count()
            print(count)
            studentserializer = StudentSerializer(getstudents,many=True)
            return JsonResponse({'student':studentserializer.data})
        if request.method == "POST":
            enroll_id = request.POST.get('search_enroll_id')
            university = request.POST.get('search_university')
            course = request.POST.get('search_course')
            stream = request.POST.get('search_stream')
            name = request.POST.get('search_name')
            registration_year = request.POST.get('search_registration_year')
            paid_fees_switch = request.POST.get('search_paid_fees_switch')
            pending_fees_switch = request.POST.get('search_pending_fees_switch')
            if enroll_id or university or course or stream or name or registration_year or paid_fees_switch=="true" or pending_fees_switch=="true":
                print(enroll_id , university , course , stream , name , registration_year , paid_fees_switch , pending_fees_switch)
                getstudents = []
                if enroll_id:
                    getstudents = Student.objects.filter(
                        Q(enrollment_id__icontains = enroll_id) |
                        Q(email__icontains = enroll_id) |
                        Q(mobile__icontains = enroll_id)
                    )
                elif name:
                    getstudents = Student.objects.filter(
                        Q(name__icontains = name) 
                    )
                elif registration_year:
                    if getstudents:
                        getstudents = getstudents.filter(enrollment_date__icontains = registration_year)
                    else:
                        getstudents = Student.objects.filter(enrollment_date__icontains = registration_year)
                    print(getstudents)
                
                elif course and stream:
                    enrolled = Enrolled.objects.filter(Q(course = course) & Q(stream = stream))
                    for i in enrolled:
                        students = Student.objects.get(id = i.student.id)
                        getstudents.append(students)
                elif paid_fees_switch == "true":
                    getallfees = PaymentReciept.objects.filter(pendingamount = 0)
                    for i in getallfees:
                        stud = Student.objects.get(id = i.student)
                        if stud in getstudents:
                            pass
                        else:
                            getstudents.append(stud)
                elif pending_fees_switch == "true":
                    getallfees = PaymentReciept.objects.filter(~Q(pendingamount = 0))
                    print(getallfees)
                    for i in getallfees:
                        stud = Student.objects.get(id = i.student)
                        if stud in getstudents:
                            pass
                        else:
                            getstudents.append(stud)
                    print(getallfees)
                # print("students :",getstudents)
                for i in getstudents:
                    enrolled = Enrolled.objects.get(student= i)
                    course = Course.objects.get(id=enrolled.course.id)
                    stream = Stream.objects.get(id = enrolled.stream.id)
                    paymentreciept = PaymentReciept.objects.filter(student=i)
                    payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                    obj = {
                        "student_id":i.id,
                        "name":i.name,
                        "email":i.email,
                        "mobile":i.mobile,
                        "enrollment_id":i.enrollment_id,
                        "course_pattern":enrolled.course_pattern,
                        "total_semyear":enrolled.total_semyear,
                        "current_semyear":enrolled.current_semyear,
                        "course":course.name,
                        "stream":stream.name,
                        "paymentreciept":payment_reciept_serializer.data
                    }
                    temp.append(obj)
            
                return JsonResponse({'student_data':temp})

            name = request.POST.get('name')
            if name:
                getstudent = Student.objects.filter(Q(name = name) |  Q(name__icontains = name) |  Q(email__icontains = name))
                for i in getstudent:
                    enrolled = Enrolled.objects.get(student= i.id)
                    course = Course.objects.get(id=enrolled.course)
                    stream = Stream.objects.get(id = enrolled.stream)
                    paymentreciept = PaymentReciept.objects.filter(student=i.id)
                    payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                    obj = {
                        "student_id":i.id,
                        "name":i.name,
                        "email":i.email,
                        "mobile":i.mobile,
                        "enrollment_id":i.enrollment_id,
                        "course_pattern":enrolled.course_pattern,
                        "total_semyear":enrolled.total_semyear,
                        "current_semyear":enrolled.current_semyear,
                        "course":course.name,
                        "stream":stream.name,
                        "paymentreciept":payment_reciept_serializer.data
                    }
                    temp.append(obj)
                return JsonResponse({'student_data':temp})

            course_id = request.POST.get('course_id')
            stream_id = request.POST.get('stream_id')
            # if course_id and stream_id:
            #     enroll = Enrolled.objects.filter(Q(course = course_id) & Q(stream = stream_id))
                # for i in enroll:
                #     student = Student.objects.get(id=i.student)
                #     enrolled = Enrolled.objects.get(student=student.id)
                #     course = Course.objects.get(id=enrolled.course)
                #     stream = Stream.objects.get(id = enrolled.stream)
                #     paymentreciept = PaymentReciept.objects.filter(student=student.id)
                #     payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                #     obj = {
                #         "name":student.name,
                #         "email":student.email,
                #         "mobile":student.mobile,
                #         "enrollment_id":student.enrollment_id,
                #         "course_pattern":enrolled.course_pattern,
                #         "total_semyear":enrolled.total_semyear,
                #         "current_semyear":enrolled.current_semyear,
                #         "course":course.name,
                #         "stream":stream.name,
                #         "paymentreciept":payment_reciept_serializer.data
                #     }
                #     temp.append(obj)
                # return JsonResponse({'student_data':temp})
            
            courseid = request.POST.get('courseid')
            if courseid:
                getcourse = Course.objects.get(id=courseid)
                getstream = Stream.objects.filter(course=getcourse)
                streamserializer = StreamSerializer(getstream,many=True)
                return JsonResponse({'stream':streamserializer.data})
            university_registration_id = request.POST.get('university_registration_id')
            if university_registration_id:
                getuniversity = University.objects.get(registrationID = university_registration_id)
                getcourse = Course.objects.filter(university = getuniversity.id)
                courseserializer = CourseSerializer(getcourse,many=True)
                return JsonResponse({'course':courseserializer.data})

            # enroll_id = request.POST.get('enroll_id')
            # if enroll_id:
            #     checkstudent = Student.objects.filter(Q(enrollment_id=enroll_id) | Q(mobile__icontains=enroll_id) | Q(email__icontains=enroll_id) | Q(name__icontains=enroll_id))
                # for i in checkstudent:
                #     enrolled = Enrolled.objects.get(student= i.id)
                #     course = Course.objects.get(id=enrolled.course)
                #     stream = Stream.objects.get(id = enrolled.stream)
                #     paymentreciept = PaymentReciept.objects.filter(student=i.id)
                #     payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                #     obj = {
                #         "name":i.name,
                #         "email":i.email,
                #         "mobile":i.mobile,
                #         "enrollment_id":i.enrollment_id,
                #         "course_pattern":enrolled.course_pattern,
                #         "total_semyear":enrolled.total_semyear,
                #         "current_semyear":enrolled.current_semyear,
                #         "course":course.name,
                #         "stream":stream.name,
                #         "paymentreciept":payment_reciept_serializer.data
                #     }
                #     temp.append(obj)
            
                # return JsonResponse({'student_data':temp})


            # search_enrollment_id = request.POST.get('search_enrollment_id')
            # if search_enrollment_id:
            #     getstudent = Student.objects.get(enrollment_id = search_enrollment_id)
                


        params = {
            "university":University.objects.all(),
        }
        return render(request,"super_admin/advancesearch.html",params)


@login_required(login_url='/login/')
def CancelledStudent(request):
    students = []
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            student_active_id = request.POST.get('student_id')
            if student_active_id:
                print("Active student",student_active_id)
                getstudent = Student.objects.get(id = student_active_id)
                getstudent.is_cancelled = False
                getstudent.archive = False
                getstudent.enrolled = True
                getstudent.save()
                return JsonResponse({'success':'yes'})
        allstudents = Student.objects.filter(Q(is_cancelled = True))
        for i in allstudents:
            student_id = i.id
            enrolled = Enrolled.objects.get(student = student_id)
            course = Course.objects.get(id = enrolled.course.id)
            stream = Stream.objects.get(id = enrolled.stream.id)
            university = University.objects.get(id = i.university.id)
            # datetimedata = datetime.strptime(str(i.enrollment_date), '%Y-%m-%d').strftime('%d-%m-%Y')
            # print(i.name,datetimedata)
            obj = {
                "id":i.id,
                "name":str(i.name),
                "father_name":i.father_name,
                "mother_name":i.mother_name,
                "dateofbirth":datetime.strptime(str(i.dateofbirth), '%Y-%m-%d').strftime('%d-%m-%Y'),
                "mobile":i.mobile,
                "email":i.email,
                "gender":i.gender,
                "category":i.category,
                "enrollment_id":i.enrollment_id,
                "enrollment_date":datetime.strptime(str(i.enrollment_date), '%Y-%m-%d').strftime('%d-%m-%Y'),
                "registration_id":i.registration_id,
                "course_name":course.name,
                "stream_name":stream.name,
                "course_pattern":enrolled.course_pattern,
                "current_semester":enrolled.current_semyear,
                "session":enrolled.session,
                'entry_mode':enrolled.entry_mode,
                'university':university.university_name
            }
            students.append(obj)

            
        params = {
            "students":students
        }
        return render(request,"super_admin/cancelled_students.html",params)


@login_required(login_url='/login/')
def UserCreation(request):
    get_all_users = []
    if request.user.is_superuser:
        get_all_users = User.objects.filter(Q(is_student=False) & Q(is_superuser = False))
        # level = UserLevel.objects.get(user = request.user.id)
        # if level.level == "4":
        #     level_of_user = level.level
        #     display = "yes"
        #     alluserlevel = UserLevel.objects.filter(~Q(level = '1') & ~Q(level='4'))
            
        #     for i in alluserlevel:
        #         try:
        #             alluser = User.objects.get(id=i.user_id)
        #             print(alluser)
        #             obj = {
        #                 "email":alluser.email,
        #                 "level":i.level
        #             }
        #             main.append(obj)
        #         except User.DoesNotExist:
        #             print("user doesnt exist")
        #     print(main)
        if request.method == "POST":
            useremail = request.POST.get('email')
            userpassword = request.POST.get('password')
            userlevel = request.POST.get('level')
            print(useremail,userpassword,userlevel)
            if useremail and userpassword and userlevel:
                try:
                    get_user = User.objects.get(email=useremail)
                    return JsonResponse({'created':'no'})
                except User.DoesNotExist:
                    if userlevel == "2":
                        createuser = User.objects.create_user(email=useremail,password=userpassword,is_data_entry=True)
                        createuser.save()
                    elif userlevel == "3":
                        createuser = User.objects.create_user(email=useremail,password=userpassword,is_fee_clerk=True)
                        createuser.save()
                    getlatestuser = User.objects.get(email=useremail)
                    # obj1 = {
                    #     "email":
                    # }
                
                    return JsonResponse({'created':'sucessfully'})
        params = {
            "main":get_all_users,
        }
        return render(request,"super_admin/user_creation.html",params)

@login_required(login_url='/login/')
def PrintAddress(request):
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            student_id = request.POST.get('student_id')
            if student_id:
                try:
                    getstudent = Student.objects.get(id = student_id)
                    obj = {
                        "address":getstudent.address,
                        "country":getstudent.country.name,
                        "state":getstudent.state.name,
                        "city":getstudent.city.name,
                        "pincode":getstudent.pincode
                    }
                    return JsonResponse({'address':'yes','data':obj})
                except Student.DoesNotExist:
                    print("student not avaliable")
        params = {
            "students":Student.objects.all()
        }
        return render(request,"super_admin/print_address.html",params)

@login_required
def SetExamination(request):
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            student_course_id = request.POST.get('student_course_id')
            student_stream_id = request.POST.get('student_stream_id')
            if student_course_id and student_stream_id:
                print(student_course_id, student_stream_id)
            university = request.POST.get("university")
            course = request.POST.get("course")
            stream = request.POST.get('stream')
            exam_date = request.POST.get('exam_date')
            exam_time = request.POST.get('exam_time')
            exam_name = request.POST.get('exam_name')
            total_marks = request.POST.get('total_marks')
            array_data = request.POST.get('array_data')
            file =  request.FILES.get('upload_file')
            ##added by Avani  on 15/10 to save additional data as per change requests.
            exam_end_time = request.POST.get('exam_end_time')
            exam_duration = request.POST.get('exam_duration')
            studypattern = request.POST.get('studypattern')
            semyear =  request.POST.get('semyear')
            if university and file:
                is_file = False
                try:
                    upload_file = request.FILES['upload_file']
                    is_file = True
                except:
                    print("no")
                    pass
                if is_file == True:
                # print(university , file)
                    # df=pd.read_excel(file,sheet_name='Sheet1')	
                    workbook = load_workbook(upload_file)
                    sheet = workbook['Sheet1']	
                    count = 0
                    ## modified by Avanin on 15/10 - save additional data in the table
                    #add_examination = Examination(university=University.objects.get(id=university),course = Course.objects.get(id=course), stream = Stream.objects.get(id=stream) ,examname= exam_name,examdate=exam_date,examtime=exam_time,totalmarks=total_marks)
                    add_examination = Examination(university=University.objects.get(id=university),course = Course.objects.get(id=course), stream = Stream.objects.get(id=stream) ,examname= exam_name,examdate=exam_date,examtime=exam_time,totalmarks=total_marks, examendtime = exam_end_time, examduration= exam_duration, studypattern = studypattern, semyear = semyear)
                    add_examination.save()
                    latest_examination = Examination.objects.latest('id')
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        
                        if all(value is None for value in row):
                            # print("null values hence skip")
                            continue  # Skip the row if all values are None
                        # print(row)
                        count = count + 1
                        question = row[0]
                        image = ""
                        question_type = row[1]
                        marks = row[2]
                        option1 = row[3]
                        option2 = row[4]
                        option3 = row[5]
                        option4 = row[6]
                        option5 = ""
                        option6 =""
                        shortanswer = ""
                        answer =  row[7]
                        #print(option1,option2,option3,option4,answer)
                        # print(row[0])
                        # count += 1
                        # print(count)
                        # name = str(row[0])
                        # loop over the lines and save them in db. If error , store as string and then display
                        
                        
                        
                        add_questions = Questions(
                            exam=latest_examination,
                            question=question,
                            image=image,
                            type=question_type,
                            marks=marks,
                            option1=option1,
                            option2=option2,
                            option3=option3,
                            option4=option4,
                            option5=option5,
                            option6=option6,
                            shortanswer=shortanswer,
                            answer=answer)
                        add_questions.save()
                    latest_examination.totalquestions = count
                    latest_examination.save()
                    # print("Count :",count)
                    # print("data len",len(data))
                    if array_data:
                        # print("exam time :",exam_time)
                        # d = datetime.strptime(exam_time, "%H:%M")
                        # now = datetime.now()
                        # current_time = now.strftime("%I:%M %p")
                        # print("Current Time =", current_time)
                        # if current_time > d.strftime("%I:%M %p"):
                        #     print("current time greater")
                        # else:
                        #     print("current time is less")
                        # print("Exam time is :",d.strftime("%I:%M %p"))


                        x = array_data.split(',')
                        for i in x:
                            print("added student :",i, " for exam :",latest_examination.id)
                            add_StudentAppearingExam = StudentAppearingExam(
                                exam = latest_examination,
                                student = Student.objects.get(id=i)
                                )
                            add_StudentAppearingExam.save()
                            stud = Student.objects.get(id=i)
                            try:
                                additional_enroll_detail = AdditionalEnrollmentDetails.objects.get(student=stud)
                            except:
                                print("error")
                            print("Student email sent to",stud.email)
                            print(additional_enroll_detail.university_enrollment_id)
                            ## modified by Avani on 15/10.. user's enrollment id will be the userd id and password for the exam login - change request
                            res = sm(
                                subject = "CIIS Examination",
                                message = '''Enter Email ID & Password in below link \n
                                Email :{} \n
                                Password :{} \n
                                Click on the link below to Login \n
                                https://erp.ciisindia.in/examination_login/'''.format(stud.enrollment_id,stud.enrollment_id),
                                from_email = 'testmail@erp.ciisindia.in',
                                recipient_list = [stud.email],
                                fail_silently=False,
                                    )
                            print('''Enter Email ID & Password in below link \n
                                Email :{} \n
                                Password :{} \n
                                Click on the link below to Login \n
                                https://erp.ciisindia.in/examination_login/'''.format(stud.enrollment_id,stud.enrollment_id))
                            
                    return JsonResponse({'saved':'yes'})
                


                        
                        
        params = {
            "university":University.objects.all(),
            "sessionnames":SessionNames.objects.filter(status=True),## added by Avani on 12/08 -pick data from mastertables
        }
        return render(request,"super_admin/setexamination.html",params)

@login_required(login_url='/login/')
def SetExam(request):
    get_course = request.GET.get('get_course')
    if get_course:
        #distinctyear = []
        #distinct_years = Course.objects.order_by('year').values_list('year', flat=True).distinct()
        #for i in distinct_years:
        #    distinctyear.append(i)

        
        try:
            getcourse = Course.objects.filter(university = University.objects.get(id = get_course))
            courseserializer = CourseSerializer(getcourse,many=True)
            
            #return JsonResponse({'course':courseserializer.data,'distinct_years':distinctyear})
            return JsonResponse({'course':courseserializer.data,})
        except Course.DoesNotExist:
            pass
    universitysame = ""
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            get_course_university = request.POST.get('get_course_university')
            get_course_name = request.POST.get('get_course_name')
            if get_course_university and get_course_name:
                # print(get_course_name,get_course_university)
                get_all_course = Course.objects.filter(Q(university=University.objects.get(id = get_course_university)) & Q(year=get_course_name))
                courseserializer = CourseSerializer(get_all_course,many=True)
                return JsonResponse({'course':courseserializer.data})
            student_course_id = request.POST.get('student_course_id')
            student_stream_id = request.POST.get('student_stream_id')
            if student_course_id and student_stream_id:
                getstream = Stream.objects.get(Q(id = student_stream_id) & Q(course = student_course_id))
                return JsonResponse({'total_semester': getstream.sem})
            university = request.POST.get('university')
            course = request.POST.get('course')
            stream = request.POST.get('stream')
            exam_name = request.POST.get('exam_name')
            exam_date = request.POST.get('exam_date')
            total_marks = request.POST.get('total_marks')
            total_questions = request.POST.get('total_questions')
            # print(university,course,stream,exam_name,exam_date,total_marks,total_questions)
            data = request.POST
            # print(data)
            

            course_id = request.POST.get('course_id')
            # print(course_id)
            if course_id:
                try:
                    getstream = Stream.objects.filter(course=course_id)
                    # print(getstream)
                except Stream.DoesNotExist:
                    getstream = "yes"
                if getstream != "yes":
                    streamserializer = StreamSerializer(getstream,many=True)
                    return JsonResponse({'stream':streamserializer.data})

            get_students_course_id = request.POST.get('get_students_course_id')
            get_students_stream_id = request.POST.get('get_students_stream_id')
            get_students_study_pattern = request.POST.get('get_students_study_pattern')
            get_students_semyear = request.POST.get('get_students_semyear')
            ## added by Avani on 14/10
            get_students_entry_mode = request.POST.get('get_students_entry_mode')
            get_students_session = request.POST.get('get_students_session')
            examination_all_student = []
            if get_students_course_id and get_students_stream_id and get_students_study_pattern and get_students_semyear:
                print(get_students_course_id , get_students_stream_id , get_students_study_pattern , get_students_semyear)
                ## modified by Avani on 14/10 - to include other variables in query and also if any is null, that variable should not be considered.
                # Initialize the base query
                query = Q()

                # Conditionally add filters if the variables are not empty
                if get_students_course_id:
                    query &= Q(course=get_students_course_id)

                if get_students_stream_id:
                    query &= Q(stream=get_students_stream_id)

                if get_students_study_pattern:
                    query &= Q(course_pattern=get_students_study_pattern)

                if get_students_semyear:
                    query &= Q(current_semyear=get_students_semyear)

                if get_students_session:
                    query &= Q(session=get_students_session)

                if get_students_entry_mode:
                    query &= Q(entry_mode=get_students_entry_mode)

                print("the query is :", query)
                # Apply the query
                get_total_enrolled = Enrolled.objects.filter(query)

                
                # get_total_enrolled = Enrolled.objects.filter(
                #     course = get_students_course_id,
                #     stream = get_students_stream_id,
                #     course_pattern = get_students_study_pattern,
                #     current_semyear = get_students_semyear
                # )
                for i in get_total_enrolled:
                    try:
                        get_student = Student.objects.get(id = i.student.id)
                        examination_all_student.append(get_student)
                        
                    except Student.DoesNotExist:
                        pass
                student_serializer = StudentSerializer(examination_all_student,many=True)
                return JsonResponse({'data': student_serializer.data})
        params = {
            "university":University.objects.all()
        }
        return render(request,"super_admin/setexam.html",params)

@login_required(login_url='/login/')
def CheckResult(request):
    if request.user.is_superuser or request.user.is_data_entry:
        if request.method == "POST":
            course_id = request.POST.get('course_id')
            stream_id = request.POST.get('stream_id')
            if course_id and stream_id:
                getexamination = Examination.objects.filter(Q(course=course_id) & Q(stream=stream_id))
                examinationserializer = ExaminationSerializer(getexamination,many=True)
                return JsonResponse({'data':examinationserializer.data})

            exam_id = request.POST.get('exam_id')
            if exam_id:
                all_data = []
                get_all_results = Result.objects.filter(exam= exam_id)
                for i in get_all_results:
                    getstudent = Student.objects.get(id = i.student.id)
                    percentage_of_marks_obtained = (int(i.score) / int(i.total_marks) )* 100
                    print("total marks :",i.total_marks , "marks obained :",i.score)
                    percentage_of_marks_obtained = round(percentage_of_marks_obtained,2)
                    print("percentage_of_marks_obtained :",percentage_of_marks_obtained)
                    obj = {
                        "exam_id":i.exam.id,
                        "name":getstudent.name,
                        "email":getstudent.email,
                        "enrollment_id":getstudent.enrollment_id,
                        "total_question":i.total_question,
                        "total_marks":i.total_marks,
                        "score":i.score,
                        "result":i.result,
                        "percentage":percentage_of_marks_obtained
                    }
                    all_data.append(obj)
                return JsonResponse({'data':all_data})
                
            student_id = request.POST.get('check_result_student_id')
            check_result_exam_id = request.POST.get('check_result_exam_id')
            
            if student_id and check_result_exam_id:
                print("student id :",student_id , "exam id :",check_result_exam_id)
                try:
                    all_questions_data = []
                    getstudent = Student.objects.get(enrollment_id = student_id)
                    print(getstudent.id)
                    get_submitted_questions = SubmittedExamination.objects.filter(Q(student=getstudent.id) & Q(exam = check_result_exam_id))
                    print(get_submitted_questions)
                    for i in get_submitted_questions:
                        getquestion = Questions.objects.get(id=i.question)
                        # print("selected answer : ",i.answer)
                        question_answer = ''
                        if i.type == "radio" or i.type == "mcq(single)":
                            # print("submitted_answer :",selected_ans)
                            if i.submitted_answer == "option1":
                                selected_answer = getquestion.option1
                            elif i.submitted_answer == "option2":
                                selected_answer = getquestion.option2
                            elif i.submitted_answer == "option3":
                                selected_answer = getquestion.option3
                            elif i.submitted_answer == "option4":
                                selected_answer = getquestion.option4
                            if i.answer == "option1":
                                question_answer = getquestion.option1
                            elif i.answer == "option2":
                                question_answer = getquestion.option2
                            elif i.answer == "option3":
                                question_answer = getquestion.option3
                            elif i.answer == "option4":
                                question_answer = getquestion.option4
                        
                            
                        obj = {
                            "question":getquestion.question,
                            "answer":question_answer,
                            "submitted_answer":selected_answer,
                            "marks":i.marks,
                            "marks_obtained":i.marks_obtained,
                            "test":"test"
                        }
                        all_questions_data.append(obj)
                    print(all_questions_data)
                    
                    return JsonResponse({'data':all_questions_data})
                except Student.DoesNotExist:
                    print("no student with that enrollment")


    params = {
        "university":University.objects.all(),
    }
    return render(request,"examination/check_result.html",params)

# import pandas as pd

def verify_examination_import_upload(request):
    column_error = ""
    type_error_list = []
    answer_error_list = []
    if request.method == 'POST':
        is_file = False
        try:
            upload_file = request.FILES['upload_file']
            print("upload_file " + upload_file)
            is_file = True
        except:
            print("no")
            pass
        if is_file == True:
            expected_cols = ['questions','type','marks','option1','option2','option3','option4','answer']
            workbook = load_workbook(upload_file)
            sheet = workbook['Sheet1']
            column_names = [cell.value for cell in sheet[1]]
            if column_names != expected_cols:
                # print("Error: Column names do not match expected columns :",column_names)
                column_error = f"Error: Column names do not match expected columns : {column_names}"
            else:
                column_error = ""
                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    count += 1
                    type = row[1]
                    if type != 'mcq(single)':
                        type_error = f"Please Check Spelling mistake in Line no {count} at {type}"
                        type_error_list.append(type_error)
                    answer = row[7]
                    if answer == "option1" or answer == "option2" or answer == "option3" or answer == "option4":
                        pass
                    else:
                        ans_error = f"Please Check Spelling mistake in Line no {count} at {answer}"
                        answer_error_list.append(ans_error)
            # Print the column names
            # for column_name in column_names:
            #     print(column_name)
            # for row in sheet.iter_rows(values_only=True):
            #     print(row[0])
        return JsonResponse({'data':'data','column_error':column_error,'type_error_list':type_error_list,'answer_error_list':answer_error_list})
        

def examination_login(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        if email and password:
            try:
                #Modified by Avani on 15/10 - login  and password for exam module will be student enrollment id 
                student = Student.objects.get(enrollment_id = email)
                #student = Student.objects.get(email = email)
                #additional = AdditionalEnrollmentDetails.objects.get(student = student.id)
                if password == student.enrollment_id:
                    studentserializer = StudentSerializer(student,many=False)
                    request.session['logged_in_user_object'] = studentserializer.data
                    return redirect('giveexamination')
                else:
                    print("not matched")
            except Student.DoesNotExist:
                pass
    return render(request,"examination/examination_login.html")

def ExaminationThroughLink(request):
    # del request.session['is_logged_in']
    # del request.session['student_id']
    temp = []
    exam_id = 0
    hours = 0
    minutes = 0
    show_exam_list = []
    if 'is_logged_in' in request.session and 'student_id' in request.session: 
        student_id = request.session['student_id']
        student_appearing_exam = StudentAppearingExam.objects.filter(student=student_id)
        # print("student_appearing_exam:",student_appearing_exam)
        for i in student_appearing_exam:
            exam = Examination.objects.get(id=i.exam)
            if exam.examdate == str(date.today()):
                time_obj = datetime.strptime(datetime.now().strftime("%H:%M:%S"), '%H:%M:%S').time()
                # print(exam.examtime , time_obj)
                if time_obj >= exam.examtime:
                    show_exam_list.append(exam)
                    # print("can give exam")
                else:
                    pass
            else:
                pass
    print(show_exam_list)
        # try:
        
    today = date.today()
    now = datetime.now()
    current_time = now.strftime("%I:%M %p")
    # del request.session['started_exam']
    # print("session deleted")
    if request.method == "POST":
        print(request.POST)
        email = request.POST.get("email")
        password = request.POST.get("password")
        # 75764567476
        if email and password:
            try:
                stud = Student.objects.get(email=email)
                additional = AdditionalEnrollmentDetails.objects.get(student=stud.id)
                if password == additional.university_enrollment_id:
                    request.session['is_logged_in'] = "yes"
                    request.session['student_id'] = stud.id
                    return redirect('examination_through_link')
                else:
                    if 'is_logged_in' in request.session:
                        del request.session['is_logged_in']

            except Student.DoesNotExist:
                pass
        check_timing = request.POST.get("check_timing")
        if check_timing:
            print("check timing :",check_timing)
            try:
                exam = Examination.objects.get(id = check_timing)
                h, m, s = str(exam.examtime).split(':')
                exam_time = datetime.strptime(str(h)+":"+str(m), "%H:%M").strftime("%I:%M %p")
                # print("exam time: ",exam_time)
                # print("current time :",current_time , ">" , " exam time :", exam_time)
                if current_time >= exam_time:
                    try:
                        result = Result.objects.get(Q(exam = exam.id) & Q(student=student_id))
                        return JsonResponse({'error':"Examination already Given. Please Wait Until Result is Announced"})
                    except Result.DoesNotExist:
                        
                        request.session['started_exam'] = "yes"
                        request.session['started_exam_id'] = exam.id
                        print("yupp no result hence can give exam")
                        return JsonResponse({'exam_id':exam.id,'exam_started':'yes'})
                else:
                    print("cannot give exam")
                    error = "Please Wait Till EXAM TIME. Exam Starts At : {}".format(exam_time)
                    return JsonResponse({'error':error})





            except Examination.DoesNotExist:
                print("no examination found")
    if 'started_exam' in request.session and 'started_exam_id' in request.session:
        print("exam started")
        try:
            exam_id = request.session['started_exam_id']
            hours = 0
            minutes = 0
            try:
                getexaminationtime = StudentExaminationTime.objects.get(Q(student = student_id ) & Q(exam = exam_id))
                # print("time data available",getexaminationtime)
                get_time_in_minutes = int(getexaminationtime.time_in_minutes)
                time_string = str(timedelta(minutes=get_time_in_minutes))[:-3]
                data_list = time_string.split(":")
                hours = data_list[0]
                minutes = data_list[1]
                print(hours , minutes)

                # getexaminationtime.time_in_minutes = (int(hours) * 60 ) + int(minutes)
                # getexaminationtime.save()
                # print("examination time found",getexaminationtime,"time left:",(int(hours) * 60 ) + int(minutes))
            except StudentExaminationTime.DoesNotExist:
                print("time data not available")
                # print("timer not found",(int(hours) * 60 ) + int(minutes))
                # create_timer = StudentExaminationTime(
                #     student = student.id,
                #     exam = exam_id_timer,
                #     time_in_minutes = (int(hours) * 60 ) + int(minutes)
                # )
                # create_timer.save()
            getexamination = Examination.objects.get(id=exam_id)
            getquestions = Questions.objects.filter(exam = getexamination.id)
            questionsserializer = QuestionsSerializer(getquestions, many=True)
            temp = []
            for i in getquestions:
                # print(i)
                try:
                    get_submitted_questions = SubmittedExamination.objects.get(Q(student = student_id) & Q(exam = exam_id) & Q(question=i.id))
                    print("get submitted questions :",get_submitted_questions)
                    submitted_answer = get_submitted_questions.submitted_answer
                except SubmittedExamination.DoesNotExist:
                    submitted_answer = ""
                
                obj = {
                    "id":i.id,
                    "exam":i.exam,
                    "question":i.question,
                    # "image":i.image,
                    "type":i.type,
                    "marks":i.marks,
                    "option1":i.option1,
                    "option2":i.option2,
                    "option3":i.option3,
                    "option4":i.option4,
                    # "option5":i.option5,
                    # "option6":i.option6,
                    # "shortanswer":i.shortanswer,
                    "answer":i.answer,
                    "submitted_answer":submitted_answer
                }
                temp.append(obj)
            # print(temp)
            
            # return JsonResponse({'data':temp,'exam_id':getexamination.id,'hours':hours,'minutes':minutes})
        except:
            print("invalid url")
            
    else:
        print("no data")
    paginator = Paginator(temp, 1)  # Show 25 contacts per page.


    page_number = request.GET.get("page")
    if page_number:
        page_number = page_number
    else:
        page_number = 1
    page_obj = paginator.get_page(page_number)
    data = temp[int(page_number) - 1]
    params = {
        "exam_list":show_exam_list,
        "data":data,
        "page_obj": page_obj,
        # 'exam_id':exam_id,
        # 'hours':hours,
        # 'minutes':minutes
    }
    return render(request,"examination/examination_through_link.html",params)


def ImportStudent(request):
    display = ""
    inserted = ""
    level_of_user = 0
    if request.user.is_superuser:
        if request.method == "POST":
            is_file = False
            try:
                upload_file = request.FILES['import_student']
                is_file = True
            except:
                print("no")
                pass
            if is_file == True:
                expected_cols = ['name','dob','mobile','email','university_enrollment_id','university','course','stream','pattern','semyear','session','entry_mode','fees_type','fees','amount_paid','payment_mode','payment_remarks']
                workbook = load_workbook(upload_file)
                sheet = workbook['Sheet1']
                column_names = [cell.value for cell in sheet[1]]
                if column_names != expected_cols:
                    # print("Error: Column names do not match expected columns :",column_names)
                    column_error = f"Error: Column names do not match expected columns : {column_names}"
                    print(column_error)
                else:
                    column_error = ""
                    count = 0
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if all(value is None for value in row):
                            # print("null values hence skip")
                            continue  # Skip the row if all values are None
                        # print(row)
                        # print(row[0])
                        # count += 1
                        # print(count)
                        name = str(row[0])
                        dob = str(row[1])
                        mobile = str(row[2])
                        email = str(row[3])
                        university_enrollment_id = str(row[4])
                        university = str(row[5])
                        course = str(row[6])
                        stream = str(row[7])
                        pattern = str(row[8])
                        semyear = str(row[9])
                        session = str(row[10])
                        entry_mode = str(row[11])
                        fees_type = str(row[12])
                        fees = str(row[13])
                        amount_paid = str(row[14])
                        payment_mode = str(row[15])
                        payment_remarks = str(row[16])
                        date_obj = datetime.strptime(dob, "%d-%m-%Y")
                        formatted_date = datetime.strftime(date_obj, "%Y-%m-%d")
                        print(formatted_date)
                        try:
                            random = Student.objects.latest('id')
                        except Student.DoesNotExist:
                            random = 1
                        if random==1:
                            enroll = 50000
                            registration_id = 250000
                        else:
                            enroll = int(random.enrollment_id)+ 1
                            registration_id = int(random.registration_id) +1
                        university_obj = University.objects.get(id=university)
                        add_student = Student(
                            name = name,
                            dateofbirth=formatted_date,
                            mobile=mobile,
                            email=email,
                            university=university_obj,
                            registration_id=registration_id,
                            enrollment_id=enroll,
                            enrolled=True,
                            verified=True
                        )
                        try:
                            add_student.save()
                            get_latest_student = Student.objects.get(Q(email=email) & Q(mobile=mobile))
                            stream_id = Stream.objects.get(id = stream)
                            
                            if pattern == "Semester":
                                total_sem_year = int(stream_id.sem) * 2
                            else:
                                total_sem_year = int(stream_id.sem)
                            add_enrollment_details = Enrolled(
                                student=get_latest_student,
                                course=Course.objects.get(id=course),
                                stream=Stream.objects.get(id=stream),
                                course_pattern=pattern,
                                session=session,
                                entry_mode=entry_mode,
                                total_semyear=total_sem_year,
                                current_semyear=semyear
                            )
                            add_enrollment_details.save()

                            add_additional_details = AdditionalEnrollmentDetails(
                                student = get_latest_student,
                                university_enrollment_id = university_enrollment_id
                            )
                            add_additional_details.save()


                            if pattern == "Semester":
                                try:
                                    getsemesterfees = SemesterFees.objects.filter(stream=stream)
                                    for i in getsemesterfees:
                                        addstudentfees = StudentFees(student = get_latest_student,studypattern="Semester",stream=stream,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.sem)
                                        addstudentfees.save()
                                except SemesterFees.DoesNotExist:
                                    pass
                            elif pattern == "Annual":
                                try:
                                    print("Annual try reached")
                                    getyearfees = YearFees.objects.filter(stream=stream)
                                    for i in getyearfees:
                                        addstudentfees = StudentFees(student = get_latest_student,studypattern="Annual",stream=stream,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.year)
                                        addstudentfees.save()
                                except YearFees.DoesNotExist:
                                    print("annual not reached")
                            

                            try:
                                getlatestreciept = PaymentReciept.objects.latest('id')
                            except PaymentReciept.DoesNotExist:
                                getlatestreciept = "none"
                            if getlatestreciept == "none":
                                transactionID = "TXT445FE101"
                            else:
                                tid = getlatestreciept.transactionID
                                tranx = tid.replace("TXT445FE",'')
                                transactionID =  str("TXT445FE") + str(int(tranx) + 1)
                            
                            add_payment_reciept = PaymentReciept(
                                student = get_latest_student,
                                payment_for = "Course Fees",
                                payment_categories = "New",
                                payment_type = "Full Payment",
                                fee_reciept_type = fees_type,
                                transaction_date = date.today(),
                                cheque_no = "",
                                semyearfees = fees,
                                paidamount = amount_paid,
                                pendingamount = 0,
                                advanceamount = 0,
                                transactionID = transactionID,
                                paymentmode = payment_mode,
                                remarks = payment_remarks,
                                semyear = semyear,
                                uncleared_amount = "",
                                status = "Realised",

                            )
                            add_payment_reciept.save()
                                
                                    
                                        
                        except Exception as e:
                            print("error :",e)
            # if is_file == True:
            
            #     df = pd.read_excel(upload_file)

            #     for i in df.index:
            #         name = str(df['name'][i])
            #         dob = str(df['dob'][i])
            #         mobile = str(df['mobile'][i])
            #         email = str(df['email'][i])
            #         university_enrollment_id = str(df['university_enrollment_id'][i])
            #         university = str(df['university'][i])
            #         course = str(df['course'][i])
            #         stream = str(df['stream'][i])
            #         pattern = str(df['pattern'][i])
            #         semyear = str(df['semyear'][i])
            #         session = str(df['session'][i])
            #         entry_mode = str(df['entry_mode'][i])
            #         fees_type = str(df['fees_type'][i])
            #         fees = str(df['fees'][i])
            #         amount_paid = str(df['amount_paid'][i])
            #         payment_mode = str(df['payment_mode'][i])
            #         payment_remarks = str(df['payment_remarks'][i])
            #         # print(name,dob,mobile,email,university_enrollment_id,university,course,stream,pattern,semyear,session,entry_mode,fees_type,amount_paid,payment_mode,payment_remarks)
            #         date_obj = datetime.strptime(dob, "%Y-%m-%d  %H:%M:%S")
            #         formatted_date = datetime.strftime(date_obj, "%Y-%m-%d")
            #         print(formatted_date)
            #         try:
            #             random = Student.objects.latest('id')
            #         except Student.DoesNotExist:
            #             random = 1
            #         if random==1:
            #             enroll = 50000
            #             registration_id = 250000
            #         else:
            #             enroll = int(random.enrollment_id)+ 1
            #             registration_id = int(random.registration_id) +1
            #         university_obj = University.objects.get(id=university)
            #         add_student = Student(
            #             name = name,
            #             dateofbirth=formatted_date,
            #             mobile=mobile,
            #             email=email,
            #             university=university_obj,
            #             registration_id=registration_id,
            #             enrollment_id=enroll,
            #             enrolled=True,
            #             verified=True
            #         )
            #         try:
            #             add_student.save()
            #             get_latest_student = Student.objects.get(Q(email=email) & Q(mobile=mobile))
            #             stream_id = Stream.objects.get(id = stream)
                        
            #             if pattern == "Semester":
            #                 total_sem_year = int(stream_id.sem) * 2
            #             else:
            #                 total_sem_year = int(stream_id.sem)
            #             add_enrollment_details = Enrolled(
            #                 student=get_latest_student,
            #                 course=Course.objects.get(id=course),
            #                 stream=Stream.objects.get(id=stream),
            #                 course_pattern=pattern,
            #                 session=session,
            #                 entry_mode=entry_mode,
            #                 total_semyear=total_sem_year,
            #                 current_semyear=semyear
            #             )
            #             add_enrollment_details.save()

            #             add_additional_details = AdditionalEnrollmentDetails(
            #                 student = get_latest_student,
            #                 university_enrollment_id = university_enrollment_id
            #             )
            #             add_additional_details.save()


            #             if pattern == "Semester":
            #                 try:
            #                     getsemesterfees = SemesterFees.objects.filter(stream=stream)
            #                     for i in getsemesterfees:
            #                         addstudentfees = StudentFees(student = get_latest_student,studypattern="Semester",stream=stream,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.sem)
            #                         addstudentfees.save()
            #                 except SemesterFees.DoesNotExist:
            #                     pass
            #             elif pattern == "Annual":
            #                 try:
            #                     print("Annual try reached")
            #                     getyearfees = YearFees.objects.filter(stream=stream)
            #                     for i in getyearfees:
            #                         addstudentfees = StudentFees(student = get_latest_student,studypattern="Annual",stream=stream,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.year)
            #                         addstudentfees.save()
            #                 except YearFees.DoesNotExist:
            #                     print("annual not reached")
                        

            #             try:
            #                 getlatestreciept = PaymentReciept.objects.latest('id')
            #             except PaymentReciept.DoesNotExist:
            #                 getlatestreciept = "none"
            #             if getlatestreciept == "none":
            #                 transactionID = "TXT445FE101"
            #             else:
            #                 tid = getlatestreciept.transactionID
            #                 tranx = tid.replace("TXT445FE",'')
            #                 transactionID =  str("TXT445FE") + str(int(tranx) + 1)
                        
            #             add_payment_reciept = PaymentReciept(
            #                 student = get_latest_student,
            #                 payment_for = "Course Fees",
            #                 payment_categories = "New",
            #                 payment_type = "Full Payment",
            #                 fee_reciept_type = fees_type,
            #                 transaction_date = date.today(),
            #                 cheque_no = "",
            #                 semyearfees = fees,
            #                 paidamount = amount_paid,
            #                 pendingamount = 0,
            #                 advanceamount = 0,
            #                 transactionID = transactionID,
            #                 paymentmode = payment_mode,
            #                 remarks = payment_remarks,
            #                 semyear = semyear,
            #                 uncleared_amount = "",
            #                 status = "Realised",

            #             )
            #             add_payment_reciept.save()
                            
                                
                                    
            #         except Exception as e:
            #             print("error :",e)
                        
    params = {
        "inserted":inserted
    }
    return render(request,"super_admin/import_student.html",params)


def ValidateStudentImportFile(request):
    if request.method == "POST":
        column_error = ""
        missing_details = []
        mobile_matched_list = []
        email_matched_list = []
        university_enrollment_id_matched_list = []

        university_error_list = []
        course_error_list = []
        stream_error_list = []
        
        if request.method == 'POST':
            is_file = False
            try:
                upload_file = request.FILES['upload_file']
                is_file = True
            except:
                print("no")
                pass
            if is_file == True:
                expected_cols = ['name','dob','mobile','email','university_enrollment_id','university','course','stream','pattern','semyear','session','entry_mode','fees_type','fees','amount_paid','payment_mode','payment_remarks']
                workbook = load_workbook(upload_file)
                sheet = workbook['Sheet1']
                column_names = [cell.value for cell in sheet[1]]
                if column_names != expected_cols:
                    # print("Error: Column names do not match expected columns :",column_names)
                    column_error = f"Error: Column names do not match expected columns : {column_names}"
                    print(column_error)
                else:
                    column_error = ""
                    count = 0
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if all(value is None for value in row):
                            # print("null values hence skip")
                            continue  # Skip the row if all values are None
                        # print(row)
                        # print(row[0])
                        # count += 1
                        # print(count)
                        name = str(row[0])
                        dob = str(row[1])
                        mobile = str(row[2])
                        email = str(row[3])
                        university_enrollment_id = str(row[4])
                        university = str(row[5])
                        course = str(row[6])
                        stream = str(row[7])
                        pattern = str(row[8])
                        semyear = str(row[9])
                        session = str(row[10])
                        entry_mode = str(row[11])
                        fees_type = str(row[12])
                        fees = str(row[13])
                        amount_paid = str(row[14])
                        payment_mode = str(row[15])
                        payment_remarks = str(row[16])
                        
                        
                        if name != "None" and dob != "None" and mobile != "None" and email != "None" and university_enrollment_id != "None" and university != "None" and course != "None" and stream != "None" and pattern != "None" and semyear != "None" and session != "None" and entry_mode != "None"  and fees_type != "None" and fees != "None" and amount_paid != "None" and payment_mode != "None" and payment_remarks:
                            # print(name,dob,mobile,email,university_enrollment_id,university,course,stream,pattern,semyear,session,entry_mode,fees_type,fees,amount_paid,payment_mode,payment_remarks)
                            
                            mobile_matched = Student.objects.filter(Q(mobile=mobile) | Q(alternate_mobile1=mobile))
                            if mobile_matched:
                                mobile_error = f"Mobile Number Already Exist in system for Student Name : {name}"
                                mobile_matched_list.append(mobile_error)
                            
                            email_matched = Student.objects.filter(Q(email=email) | Q(alternateemail=email))
                            if email_matched:
                                email_error = f"Email Already Exist in system for Student Name : {name}"
                                email_matched_list.append(email_error)
                            
                            university_enrollment_id_matched = AdditionalEnrollmentDetails.objects.filter(university_enrollment_id = university_enrollment_id)
                            if university_enrollment_id_matched:
                                university_enrollment_id_error = f"University Enrollment Number Already Exist in system for Student Name : {name}"
                                university_enrollment_id_matched_list.append(university_enrollment_id_error)

                            university_matched = University.objects.filter(id=university)
                            if university_matched:
                                pass
                            else:
                                university_error = f"University Does Not Exist for ID : {university} for Student Name : {name}"
                                university_error_list.append(university_error)
                            
                            stream_matched = Stream.objects.filter(id=stream)
                            if stream_matched:
                                pass
                            else:
                                stream_error = f"Stream Does Not Exist for ID : {stream} for Student Name : {name}"
                                stream_error_list.append(stream_error)
                            
                            course_matched = Course.objects.filter(id=course)
                            if course_matched:
                                pass
                            else:
                                course_error = f"Course Does Not Exist for ID : {course} for Student Name : {name}"
                                course_error_list.append(course_error)

                        else:
                            missing_error_string = f"Data is missing at student name : {name}"
                            # print(missing_error_string)
                            missing_details.append(missing_error_string)




                    # for row in sheet.iter_rows(min_row=2, values_only=True):
                    #     count += 1
                    #     type = row[1]
                    #     if type != 'mcq(single)':
                    #         type_error = f"Please Check Spelling mistake in Line no {count} at {type}"
                    #         type_error_list.append(type_error)
                    #     answer = row[7]
                    #     if answer == "option1" or answer == "option2" or answer == "option3" or answer == "option4":
                    #         pass
                    #     else:
                    #         ans_error = f"Please Check Spelling mistake in Line no {count} at {answer}"
                    #         answer_error_list.append(ans_error)
            # Print the column names
            # is_file = False
            # try:
            #     upload_file = request.FILES['upload_file']
            #     is_file = True
            # except:
            #     print("no")
            #     pass
            # if is_file == True:
            #     expected_cols = ['name','dob','mobile','email','university_enrollment_id','university','course','stream','pattern','semyear','session','entry_mode','fees_type','fees','amount_paid','payment_mode','payment_remarks']

            #     df = pd.read_excel(upload_file)

            #     # Validate column names
            #     if list(df.columns) != expected_cols:
            #         print("Error: Column names do not match expected columns :",list(df.columns))
            #         column_error = f"Error: Column names do not match expected columns : {list(df.columns)}"
            #     else:
            #         column_error = ""
                    # for i in df.index:
                    #     name = str(df['name'][i])
                    #     dob = str(df['dob'][i])
                    #     mobile = str(df['mobile'][i])
                    #     email = str(df['email'][i])
                    #     university_enrollment_id = str(df['university_enrollment_id'][i])
                    #     university = str(df['university'][i])
                    #     course = str(df['course'][i])
                    #     stream = str(df['stream'][i])
                    #     pattern = str(df['pattern'][i])
                    #     semyear = str(df['semyear'][i])
                    #     session = str(df['session'][i])
                    #     entry_mode = str(df['entry_mode'][i])
                    #     fees_type = str(df['fees_type'][i])
                    #     fees = str(df['fees'][i])
                    #     amount_paid = str(df['amount_paid'][i])
                    #     payment_mode = str(df['payment_mode'][i])
                    #     payment_remarks = str(df['payment_remarks'][i])
                    #     # print(university_enrollment_id == None,str(university_enrollment_id) == "None")
                    #     # print(course)
                    #     print(name,dob,mobile,email,university_enrollment_id,university,course,stream,pattern,semyear,session,entry_mode,fees_type,fees,amount_paid,payment_mode,payment_remarks)
                    #     if name != "nan" and dob != "nan" and mobile != "nan" and email != "nan" and university_enrollment_id != "nan" and university != "nan" and course != "nan" and stream != "nan" and pattern != "nan" and semyear != "nan" and session != "nan" and entry_mode != "nan"  and fees_type != "nan" and fees != "nan" and amount_paid != "nan" and payment_mode != "nan" and payment_remarks:
                    #         # print("all details given")
                            
                    #         mobile_matched = Student.objects.filter(Q(mobile=mobile) | Q(alternate_mobile1=mobile))
                    #         if mobile_matched:
                    #             mobile_error = f"Mobile Number Already Exist in system for Student Name : {name}"
                    #             mobile_matched_list.append(mobile_error)
                            
                    #         email_matched = Student.objects.filter(Q(email=email) | Q(alternateemail=email))
                    #         if email_matched:
                    #             email_error = f"Email Already Exist in system for Student Name : {name}"
                    #             email_matched_list.append(email_error)
                            
                    #         university_enrollment_id_matched = AdditionalEnrollmentDetails.objects.filter(university_enrollment_id = university_enrollment_id)
                    #         if university_enrollment_id_matched:
                    #             university_enrollment_id_error = f"University Enrollment Number Already Exist in system for Student Name : {name}"
                    #             university_enrollment_id_matched_list.append(university_enrollment_id_error)

                    #         university_matched = University.objects.filter(id=university)
                    #         if university_matched:
                    #             pass
                    #         else:
                    #             university_error = f"University Does Not Exist for ID : {university} for Student Name : {name}"
                    #             university_error_list.append(university_error)
                            
                    #         stream_matched = Stream.objects.filter(id=stream)
                    #         if stream_matched:
                    #             pass
                    #         else:
                    #             stream_error = f"Stream Does Not Exist for ID : {stream} for Student Name : {name}"
                    #             stream_error_list.append(stream_error)
                            
                    #         course_matched = Course.objects.filter(id=course)
                    #         if course_matched:
                    #             pass
                    #         else:
                    #             course_error = f"Course Does Not Exist for ID : {course} for Student Name : {name}"
                    #             course_error_list.append(course_error)

                    #     else:
                    #         missing_error_string = f"Data is missing at student name : {name}"
                    #         # print(missing_error_string)
                    #         missing_details.append(missing_error_string)
                        
                        
                        

                       
        # print("Missing Details :",missing_details)
        # print("mobile_matched_list :",mobile_matched_list)
        # print("email_matched_list :",email_matched_list)
        # print("university_enrollment_id_matched_list :",university_enrollment_id_matched_list)
        # print("university_error_list :",university_error_list)
        # print("course_error_list :",course_error_list)
        # print("stream_error_list :",stream_error_list)
        params = {
            'column_error':column_error,
            "missing_details":missing_details,
            "mobile_matched_list":mobile_matched_list,
            "email_matched_list":email_matched_list,
            "university_enrollment_id_matched_list":university_enrollment_id_matched_list,
            "university_error_list":university_error_list,
            "course_error_list":course_error_list,
            "stream_error_list":stream_error_list,
        }
        
        
        return JsonResponse(params)

def GiveExamination(request):
    display = ""
    error = ""
    exams = []
    givenexams = [] ## added by Avani on 16/10
    if request.user.is_superuser:
        user_id = request.user.id
        student = Student.objects.get(user=user_id)
        print(student.id)
        today = date.today()
        now = datetime.now()
        current_time = now.strftime("%I:%M %p")
        print("Current date:", today,"Current time :",current_time)
        
        can_give_exam = StudentAppearingExam.objects.filter(student = student.id)
        if can_give_exam:
            enrolled = Enrolled.objects.get(student = student.id)
            for i in can_give_exam:
                exam = Examination.objects.filter(Q(stream = enrolled.stream) & Q(id = i.exam))
                for j in exam:
                    if str(j.examdate) == str(today):
                        exams.append(j)
        else:
            print("cant give exam")
        
        enrolled = Enrolled.objects.get(student = student.id)
        exam = Examination.objects.filter(stream = enrolled.stream)
        for i in exam:
            h, m, s = str(i.examtime).split(':')
            exam_time = datetime.strptime(str(h)+":"+str(m), "%H:%M").strftime("%I:%M %p")
            print("exam time: ",exam_time)

            now = datetime.now()
            current_time = now.strftime("%I:%M %p")
            print("current time :",current_time)
            # print("Current Time =", current_time)
            if current_time >= exam_time:
                print("You can give exam")
            else:
                print("There is time to give exam")
            print("Exam time is :",d.strftime("%I:%M %p"))
            try:
                result = Result.objects.get(Q(exam = i.id) & Q(student = user_id))
                # print("Exam ID :",i.id,type(i.id) , "Result Exam :",result.exam)
                if int(i.id) == int(result.exam):
                    print("found")
                else:
                    tester = 0
                    # print("not found")
            except Result.DoesNotExist:
                # print("not found")
                exams.append(i)
        if request.method == "POST":
            print("in post")
            check_timing = request.POST.get("check_timing")
            if check_timing:
                print("check timing :",check_timing)
                try:
                    exam = Examination.objects.get(id = check_timing)
                    h, m, s = str(exam.examtime).split(':')
                    exam_time = datetime.strptime(str(h)+":"+str(m), "%H:%M").strftime("%I:%M %p")
                    # print("exam time: ",exam_time)
                    # print("current time :",current_time , ">" , " exam time :", exam_time)
                    if current_time >= exam_time:
                        try:
                            result = Result.objects.get(Q(exam = exam.id) & Q(student=student.id))
                            return JsonResponse({'error':"Examination already Given. Please Wait Until Result is Announced"})
                        except Result.DoesNotExist:
                            print("yupp no result hence can give exam")
                            return JsonResponse({'exam_id':exam.id})
                    else:
                        print("cannot give exam")
                        error = "Please Wait Till EXAM TIME. Exam Starts At : {}".format(exam_time)
                        return JsonResponse({'error':error})





                except Examination.DoesNotExist:
                    print("no examination found")
            save_question_no = request.POST.get('save_question_no')
            save_question_id = request.POST.get('save_question_id')
            save_question_type = request.POST.get('save_question_type')
            save_question = request.POST.get('save_question')
            save_value = request.POST.get('save_value')
            
            uploaded_files = request.FILES.getlist('uploaded_files')
            if uploaded_files:
                # print("uploaded_files :",uploaded_files)
                print(save_question,save_question_type,save_question_id,save_value)
            if save_question_no and save_question_id and save_question_type and save_question and save_value:
                print("got everything")
                try:
                    getsubmittedanswer = SubmittedExamination.objects.get(Q(student = user_id) & Q(question = save_question_id))
                    print(getsubmittedanswer)
                    print("Avani ........ try is used")
                    getquestion = Questions.objects.get(id = getsubmittedanswer.question)
                    print(len(getquestion.answer))
                    #if save_question_type == "radio" or save_question_type == "mcq(single)":
                    if save_question_type == "radio" or save_question_type == "mcq(single)" or save_question_type == "multi":
                        # print("question answer :",getquestion.answer , len(getquestion.answer))
                        # print("submitted answer :",submitted_answer , len(submitted_answer))
                        if str(getquestion.answer) == str(save_value):
                            print("right answer")
                            print("getsubmittedanswer:",getsubmittedanswer)
                            getsubmittedanswer.marks_obtained=getquestion.marks
                            getsubmittedanswer.submitted_answer=save_value
                            getsubmittedanswer.result="right"
                            getsubmittedanswer.save()
                            print("data saved")
                            
                        else:
                            print("wrong answer")
                            getsubmittedanswer.marks_obtained=0
                            getsubmittedanswer.submitted_answer=save_value
                            getsubmittedanswer.result="wrong"
                            getsubmittedanswer.save()
                    elif save_question_type == "mcq(multiple)":
                        answer = getquestion.answer.split(",")
                        # print(len(answer),answer)
                        count = 0
                        save_values_list = save_value.split(",")
                        # print(len(save_values_list),save_values_list)
                        for j in answer:
                            for k in save_values_list:
                                if k == j:
                                    count = count +1
                        length_of_answer = len(answer)
                        marks_per_answer = len(answer)/int(getquestion.marks)
                        # print("marks per answer:",math.trunc(marks_per_answer))
                        set_marks = int(count) * math.trunc(marks_per_answer)
                        # marks_obtained = marks_obtained + int(set_marks)
                        # print("set_marks:",set_marks)


                        getsubmittedanswer.marks_obtained=set_marks
                        getsubmittedanswer.submitted_answer=save_value
                        getsubmittedanswer.result="right"
                        getsubmittedanswer.save()
                        
                    elif save_question_type == "shortanswer":
                        
                        getsubmittedanswer.submitted_answer=save_value
                        getsubmittedanswer.save()
                    elif save_question_type == "long_desc":
                        print("try used and long desc") 
                        getsubmittedanswer.submitted_answer = save_value
                        getsubmittedanswer.save()
                        get_all_uploads = Descriptive_Answer.objects.filter(Q(student=user_id) & Q(exam = getquestion.exam) & Q(question = getquestion.id))
                        print("all uploads :",get_all_uploads)
                        if uploaded_files:
                            get_all_uploads.delete()
                            for i in uploaded_files:
                                save_descriptive = Descriptive_Answer(
                                    student = user_id,
                                    exam = getquestion.exam,
                                    question = getquestion.id,
                                    upload = i
                                )
                                save_descriptive.save()

                except SubmittedExamination.DoesNotExist:
                    print("except is used")
                    getquestion = Questions.objects.get(id = save_question_id)
                    if save_question_type == "radio" or save_question_type == "mcq(single)":
                        if getquestion.answer == save_value:
                            # print("right answer")
                            # marks_obtained = marks_obtained + int(getquestion.marks)
                            submitanswer = SubmittedExamination(
                                student=user_id,
                                exam=getquestion.exam,
                                question=getquestion.id,
                                type=getquestion.type,
                                marks=getquestion.marks,
                                marks_obtained=getquestion.marks,
                                submitted_answer=save_value,
                                answer=getquestion.answer,
                                result="right"
                                )
                            submitanswer.save()
                        else:
                            # print("wrong answer")
                            submitanswer = SubmittedExamination(
                                student=user_id,
                                exam=getquestion.exam,
                                question=getquestion.id,
                                type=getquestion.type,
                                marks=getquestion.marks,
                                marks_obtained=0,
                                submitted_answer=save_value,
                                answer=getquestion.answer,
                                result="wrong"
                                )
                            submitanswer.save()
                    elif save_question_type == "mcq(multiple)":
                        answer = getquestion.answer.split(",")
                        # print(len(answer),answer)
                        count = 0
                        save_values_list = save_value.split(",")
                        # print(len(save_values_list),save_values_list)
                        for j in answer:
                            for k in save_values_list:
                                if k == j:
                                    count = count +1
                        length_of_answer = len(answer)
                        marks_per_answer = len(answer)/int(getquestion.marks)
                        # print("marks per answer:",math.trunc(marks_per_answer))
                        set_marks = int(count) * math.trunc(marks_per_answer)
                        # marks_obtained = marks_obtained + int(set_marks)
                        # print("set_marks:",set_marks)
                        submitanswer = SubmittedExamination(
                            student=user_id,
                            exam=getquestion.exam,
                            question=getquestion.id,
                            type=getquestion.type,
                            marks=getquestion.marks,
                            marks_obtained=set_marks,
                            submitted_answer=save_value,
                            answer=getquestion.answer,
                            result="right"
                            )
                        submitanswer.save()
                    elif save_question_type == "shortanswer":
                        
                        submitanswer = SubmittedExamination(
                            student=user_id,
                            exam=getquestion.exam,
                            question=getquestion.id,
                            type=getquestion.type,
                            marks=getquestion.marks,
                            marks_obtained=0,
                            submitted_answer=save_value,
                            answer=getquestion.answer,
                            result="pending"
                            )
                        submitanswer.save()
                    elif save_question_type == "long_desc":
                        print("saved question long description")
                        submitanswer = SubmittedExamination(
                            student=user_id,
                            exam=getquestion.exam,
                            question=getquestion.id,
                            type=getquestion.type,
                            marks=getquestion.marks,
                            marks_obtained=0,
                            submitted_answer=save_value,
                            answer=getquestion.answer,
                            result="pending"
                            )
                        submitanswer.save()
                        if uploaded_files:
                            for i in uploaded_files:
                                print(i)
                                save_descriptive = Descriptive_Answer(
                                    student = user_id,
                                    exam = getquestion.exam,
                                    question = getquestion.id,
                                    upload = i
                                )
                                save_descriptive.save()
            exam_id_timer = request.POST.get('exam_id_timer')
            exam_in_minutes = request.POST.get('exam_in_minutes')
            if exam_id_timer and exam_in_minutes:
                # print("Time :", exam_in_minutes,exam_id_timer)
                try:
                    getexaminationtime = StudentExaminationTime.objects.get(Q(student = student.id ) & Q(exam = exam_id_timer))
                    getexaminationtime.time_in_minutes = exam_in_minutes
                    getexaminationtime.save()
                    # print("examination time found",getexaminationtime,"time left:",exam_in_minutes)
                except StudentExaminationTime.DoesNotExist:
                    # print("timer not found")
                    create_timer = StudentExaminationTime(
                        student = student.id,
                        exam = exam_id_timer,
                        time_in_minutes = 180
                    )
                    create_timer.save()
            # print(request.POST)
            # print("1 16 radio what is 10+10? option4")
            # print("2 17 radio what is a sun? option3")
            # print("3 18 mcq(single) what is a rose option1")
            # print("4 19 mcq(single) what is 1+1? option2")
            # print("5 20 mcq(multiple) select animal from following option1")
            # print("5 20 mcq(multiple) select animal from following option2")
            # print("6 21 shortanswer write in brief what is a star star")
            # print("_____________________________________________________________________________________________")
            result_exam_id = request.POST.get('result_exam_id')
            question_count = request.POST.get("count")
            if question_count and result_exam_id:
                question_count = int(question_count)
                marks_obtained = 0
                print(result_exam_id)
                getexamination = Examination.objects.get(id = result_exam_id)
                get_submitted_questions = SubmittedExamination.objects.filter(Q(student = user_id) & Q(exam = result_exam_id))
                for i in get_submitted_questions:
                    marks_obtained = int(i.marks_obtained) + marks_obtained
                print("total_marks : ",getexamination.totalmarks , "Marks Obtained : ",marks_obtained)
                total_marks_of_exam = getexamination.totalmarks
                percentage_of_marks_obtained = (int(marks_obtained) / int(total_marks_of_exam) )* 100
                print("total marks :",total_marks_of_exam , "marks obained :",marks_obtained)
                print("percentage_of_marks_obtained :",percentage_of_marks_obtained)
                if(percentage_of_marks_obtained >= 35):
                    result = "Pass"
                else:
                    result = "Fail"
                set_result = Result(
                    student=student.id,
                    exam=result_exam_id,
                    total_question=getexamination.totalquestions,
                    attempted=question_count,
                    total_marks=getexamination.totalmarks,
                    score=marks_obtained,
                    result=result
                    )
                set_result.save()
                return JsonResponse({'added':'yes'})
                    #sun is an astonomical object made up of hydrogen and helium which generates plasma and since it has huge mass it creates a huge gravity.
            exam_id = request.POST.get('exam_id')
            if exam_id:
                hours = 0
                minutes = 0
                try:
                    getexaminationtime = StudentExaminationTime.objects.get(Q(student = student.id ) & Q(exam = exam_id))
                    print("time data available",getexaminationtime)
                    get_time_in_minutes = int(getexaminationtime.time_in_minutes)
                    time_string = str(timedelta(minutes=get_time_in_minutes))[:-3]
                    data_list = time_string.split(":")
                    hours = data_list[0]
                    minutes = data_list[1]
                    print(hours , minutes)

                    # getexaminationtime.time_in_minutes = (int(hours) * 60 ) + int(minutes)
                    # getexaminationtime.save()
                    # print("examination time found",getexaminationtime,"time left:",(int(hours) * 60 ) + int(minutes))
                except StudentExaminationTime.DoesNotExist:
                    print("time data not available")
                    # print("timer not found",(int(hours) * 60 ) + int(minutes))
                    # create_timer = StudentExaminationTime(
                    #     student = student.id,
                    #     exam = exam_id_timer,
                    #     time_in_minutes = (int(hours) * 60 ) + int(minutes)
                    # )
                    # create_timer.save()
                getexamination = Examination.objects.get(id=exam_id)
                getquestions = Questions.objects.filter(exam = getexamination.id)
                questionsserializer = QuestionsSerializer(getquestions, many=True)
                temp = []
                for i in getquestions:
                    # print(i)
                    try:
                        get_submitted_questions = SubmittedExamination.objects.get(Q(student = user_id) & Q(exam = exam_id) & Q(question=i.id))
                        print("get submitted questions :",get_submitted_questions)
                        submitted_answer = get_submitted_questions.submitted_answer
                    except SubmittedExamination.DoesNotExist:
                        submitted_answer = ""
                    
                    obj = {
                        "id":i.id,
                        "exam":i.exam,
                        "question":i.question,
                        "image":i.image,
                        "type":i.type,
                        "marks":i.marks,
                        "option1":i.option1,
                        "option2":i.option2,
                        "option3":i.option3,
                        "option4":i.option4,
                        "option5":i.option5,
                        "option6":i.option6,
                        "shortanswer":i.shortanswer,
                        "answer":i.answer,
                        "submitted_answer":submitted_answer
                    }
                    temp.append(obj)
                print(temp)
                return JsonResponse({'data':temp,'exam_id':getexamination.id,'hours':hours,'minutes':minutes})
            # show = request.POST.get('show')
            # if show == "yes":
            #     getexamination = Examination.objects.latest('id')
            #     getquestions = Questions.objects.filter(exam = getexamination.id)
            #     questionsserializer = QuestionsSerializer(getquestions, many=True)
            #     return JsonResponse({'data':questionsserializer.data})
        # print(exams)
    else:
        if 'logged_in_user_object' in request.session:
            print("yes can give exam")
            display = "yes"
            user_id = request.session['logged_in_user_object']['id']
            student = Student.objects.get(id=user_id)
            print(student.id)
            today = date.today()
            now = datetime.now()
            current_time = now.strftime("%I:%M %p")
            print("Current date:", today,"Current time :",current_time)
            
            can_give_exam = StudentAppearingExam.objects.filter(student = student.id)
            if can_give_exam:
                enrolled = Enrolled.objects.get(student = student.id)
                for i in can_give_exam:
                    exam = Examination.objects.filter(Q(stream = enrolled.stream) & Q(id = i.exam.id))
                    for j in exam:
                        if str(j.examdate) == str(today):
                            time_obj = datetime.strptime(datetime.now().strftime("%H:%M:%S"), '%H:%M:%S').time()
                            print(j.examtime , time_obj, j.examendtime)
                            ## modified by Avani on 15/10  - check if current time lies between start and end time for an exam
                            if j.examendtime != None:
                                
                                if time_obj >= j.examtime and time_obj <= j.examendtime:

                                    try:
                                        result = Result.objects.get(Q(exam = j.id) & Q(student=student.id))
                                        givenexams.append(j)
                                    except Result.DoesNotExist:
                                        print("yupp no result hence can give exam")
                                        exams.append(j)
                        
                                
                                
                            else:
                                if time_obj >= j.examtime:
                                    try:
                                        result = Result.objects.get(Q(exam = exam.id) & Q(student=student.id))
                                        givenexams.append(j)
                                    except Result.DoesNotExist:
                                        print("yupp no result hence can give exam")
                                        exams.append(j)
            else:
                print("cant give exam")
            
            # enrolled = Enrolled.objects.get(student = student.id)
            # exam = Examination.objects.filter(stream = enrolled.stream)
            # for i in exam:
                # h, m, s = str(i.examtime).split(':')
                # exam_time = datetime.strptime(str(h)+":"+str(m), "%H:%M").strftime("%I:%M %p")
                # print("exam time: ",exam_time)

                # now = datetime.now()
                # current_time = now.strftime("%I:%M %p")
                # print("current time :",current_time)
                # # print("Current Time =", current_time)
            #     if current_time >= exam_time:
            #         print("You can give exam")
            #     else:
            #         print("There is time to give exam")
                # print("Exam time is :",d.strftime("%I:%M %p"))
                # try:
                #     result = Result.objects.get(Q(exam = i.id) & Q(student = user_id))
                #     # print("Exam ID :",i.id,type(i.id) , "Result Exam :",result.exam)
                #     if int(i.id) == int(result.exam):
                #         print("found")
                #     else:
                #         tester = 0
                #         # print("not found")
                # except Result.DoesNotExist:
                #     # print("not found")
                #     exams.append(i)
            if request.method == "POST":
                marks_obtained = 0
                print("ffffff")
                # check_timing = request.POST.get("check_timing")
                # if check_timing:
                #     print("check timing :",check_timing)
                #     try:
                #         exam = Examination.objects.get(id = check_timing)
                #         h, m, s = str(exam.examtime).split(':')
                #         exam_time = datetime.strptime(str(h)+":"+str(m), "%H:%M").strftime("%I:%M %p")
                #         print("exam time: ",exam_time)
                #         print("current time :",current_time , ">" , " exam time :", exam_time)
                #         exam_date_time = datetime.combine(datetime.now().date(), datetime.strptime(exam_time, "%I:%M %p").time())
                #         current_date_time = datetime.combine(datetime.now().date(), datetime.strptime(current_time, "%I:%M %p").time())
                #         if current_date_time >= exam_date_time:
                #             try:
                #                 result = Result.objects.get(Q(exam = exam.id) & Q(student=student.id))
                #                 return JsonResponse({'error':"Examination already Given. Please Wait Until Result is Announced"})
                #             except Result.DoesNotExist:
                #                 print("yupp no result hence can give exam")
                #                 return JsonResponse({'exam_id':exam.id})
                #         else:
                #             print("cannot give exam")
                #             error = "Please Wait Till EXAM TIME. Exam Starts At : {}".format(exam_time)
                #             return JsonResponse({'error':error})





                #     except Examination.DoesNotExist:
                #         print("no examination found")
                save_question_no = request.POST.get('save_question_no')
                save_question_id = request.POST.get('save_question_id')
                save_question_type = request.POST.get('save_question_type')
                save_question = request.POST.get('save_question')
                save_value = request.POST.get('save_value')
                
                uploaded_files = request.FILES.getlist('uploaded_files')
                if uploaded_files:
                    print(save_question,save_question_type,save_question_id,save_value)
                if save_question_no and save_question_id and save_question_type and save_question and save_value:
                    print("AVANI ...... got everything")
                    try:
                        getsubmittedanswer = SubmittedExamination.objects.get(Q(student = student) & Q(question = save_question_id))
                        print("try is used")
                        getquestion = Questions.objects.get(id = getsubmittedanswer.question)
                        print(len(getquestion.answer))
                        #if save_question_type == "radio" or save_question_type == "mcq(single)":
                        if save_question_type == "radio" or save_question_type == "mcq(single)" or save_question_type == "multi":
                            print("question answer :",getquestion.answer , len(getquestion.answer))
                            #print("submitted answer :",submitted_answer , len(submitted_answer))
                            if str(getquestion.answer) == str(save_value):
                                print("right answer")
                                print("getsubmittedanswer:",getsubmittedanswer)
                                getsubmittedanswer.marks_obtained=getquestion.marks
                                getsubmittedanswer.submitted_answer=save_value
                                getsubmittedanswer.result="right"
                                getsubmittedanswer.save()
                                print("data saved")
                                
                            else:
                                print("wrong answer")
                                getsubmittedanswer.marks_obtained=0
                                getsubmittedanswer.submitted_answer=save_value
                                getsubmittedanswer.result="wrong"
                                getsubmittedanswer.save()
                        elif save_question_type == "mcq(multiple)":
                            answer = getquestion.answer.split(",")
                            print(len(answer),answer)
                            count = 0
                            save_values_list = save_value.split(",")
                            print(len(save_values_list),save_values_list)
                            for j in answer:
                                for k in save_values_list:
                                    if k == j:
                                        count = count +1
                            length_of_answer = len(answer)
                            marks_per_answer = len(answer)/int(getquestion.marks)
                            print("marks per answer:",math.trunc(marks_per_answer))
                            set_marks = int(count) * math.trunc(marks_per_answer)
                            marks_obtained = marks_obtained + int(set_marks)
                            print("set_marks:",set_marks)


                            getsubmittedanswer.marks_obtained=set_marks
                            getsubmittedanswer.submitted_answer=save_value
                            getsubmittedanswer.result="right"
                            getsubmittedanswer.save()
                            
                        elif save_question_type == "shortanswer":
                            
                            getsubmittedanswer.submitted_answer=save_value
                            getsubmittedanswer.save()
                        elif save_question_type == "long_desc":
                            print("try used and long desc") 
                            getsubmittedanswer.submitted_answer = save_value
                            getsubmittedanswer.save()
                            get_all_uploads = Descriptive_Answer.objects.filter(Q(student=user_id) & Q(exam = getquestion.exam) & Q(question = getquestion.id))
                            print("all uploads :",get_all_uploads)
                            if uploaded_files:
                                get_all_uploads.delete()
                                for i in uploaded_files:
                                    save_descriptive = Descriptive_Answer(
                                        student = user_id,
                                        exam = getquestion.exam,
                                        question = getquestion.id,
                                        upload = i
                                    )
                                    save_descriptive.save()

                    except SubmittedExamination.DoesNotExist:
                        print("except is used")
                        getquestion = Questions.objects.get(id = save_question_id)
                        #if save_question_type == "radio" or save_question_type == "mcq(single)":
                        if save_question_type == "radio" or save_question_type == "mcq(single)" or save_question_type == "multi":
                            if getquestion.answer == save_value:
                                print("right answer")
                                marks_obtained = marks_obtained + int(getquestion.marks)
                                submitanswer = SubmittedExamination(
                                    student=student,
                                    exam=getquestion.exam,
                                    question=getquestion.id,
                                    type=getquestion.type,
                                    marks=getquestion.marks,
                                    marks_obtained=getquestion.marks,
                                    submitted_answer=save_value,
                                    answer=getquestion.answer,
                                    result="right"
                                    )
                                submitanswer.save()
                            else:
                                print("wrong answer")
                                submitanswer = SubmittedExamination(
                                    student=student,
                                    exam=getquestion.exam,
                                    question=getquestion.id,
                                    type=getquestion.type,
                                    marks=getquestion.marks,
                                    marks_obtained=0,
                                    submitted_answer=save_value,
                                    answer=getquestion.answer,
                                    result="wrong"
                                    )
                                submitanswer.save()
                        elif save_question_type == "mcq(multiple)":
                            answer = getquestion.answer.split(",")
                            print(len(answer),answer)
                            count = 0
                            save_values_list = save_value.split(",")
                            print(len(save_values_list),save_values_list)
                            for j in answer:
                                for k in save_values_list:
                                    if k == j:
                                        count = count +1
                            length_of_answer = len(answer)
                            marks_per_answer = len(answer)/int(getquestion.marks)
                            print("marks per answer:",math.trunc(marks_per_answer))
                            set_marks = int(count) * math.trunc(marks_per_answer)
                            marks_obtained = marks_obtained + int(set_marks)
                            print("set_marks:",set_marks)
                            submitanswer = SubmittedExamination(
                                student=user_id,
                                exam=getquestion.exam,
                                question=getquestion.id,
                                type=getquestion.type,
                                marks=getquestion.marks,
                                marks_obtained=set_marks,
                                submitted_answer=save_value,
                                answer=getquestion.answer,
                                result="right"
                                )
                            submitanswer.save()
                        elif save_question_type == "shortanswer":
                            
                            submitanswer = SubmittedExamination(
                                student=user_id,
                                exam=getquestion.exam,
                                question=getquestion.id,
                                type=getquestion.type,
                                marks=getquestion.marks,
                                marks_obtained=0,
                                submitted_answer=save_value,
                                answer=getquestion.answer,
                                result="pending"
                                )
                            submitanswer.save()
                        elif save_question_type == "long_desc":
                            print("saved question long description")
                            submitanswer = SubmittedExamination(
                                student=user_id,
                                exam=getquestion.exam,
                                question=getquestion.id,
                                type=getquestion.type,
                                marks=getquestion.marks,
                                marks_obtained=0,
                                submitted_answer=save_value,
                                answer=getquestion.answer,
                                result="pending"
                                )
                            submitanswer.save()
                            if uploaded_files:
                                for i in uploaded_files:
                                    print(i)
                                    save_descriptive = Descriptive_Answer(
                                        student = user_id,
                                        exam = getquestion.exam,
                                        question = getquestion.id,
                                        upload = i
                                    )
                                    save_descriptive.save()
                exam_id_timer = request.POST.get('exam_id_timer')
                exam_in_minutes = request.POST.get('exam_in_minutes')
                if exam_id_timer and exam_in_minutes:
                    print("Avani Time :", exam_in_minutes,exam_id_timer)
                    try:
                        getexaminationtime = StudentExaminationTime.objects.get(Q(student = student ) & Q(exam = exam_id_timer))
                        getexaminationtime.time_in_minutes = exam_in_minutes
                        getexaminationtime.save()
                        print("examination time found",getexaminationtime,"time left:",exam_in_minutes)
                    except StudentExaminationTime.DoesNotExist:
                        print("timer not found")
                        create_timer = StudentExaminationTime(
                            student = student,
                            exam = exam_id_timer,
                            time_in_minutes = 180
                        )
                        create_timer.save()
                # print(request.POST)
                # print("1 16 radio what is 10+10? option4")
                # print("2 17 radio what is a sun? option3")
                # print("3 18 mcq(single) what is a rose option1")
                # print("4 19 mcq(single) what is 1+1? option2")
                # print("5 20 mcq(multiple) select animal from following option1")
                # print("5 20 mcq(multiple) select animal from following option2")
                # print("6 21 shortanswer write in brief what is a star star")
                # print("_____________________________________________________________________________________________")
                result_exam_id = request.POST.get('result_exam_id')
                question_count = request.POST.get("count")
                if question_count and result_exam_id:
                    question_count = int(question_count)
                    marks_obtained = 0
                    print(result_exam_id)
                    getexamination = Examination.objects.get(id = result_exam_id)
                    get_submitted_questions = SubmittedExamination.objects.filter(Q(student = user_id) & Q(exam = result_exam_id))
                    for i in get_submitted_questions:
                        marks_obtained = int(i.marks_obtained) + marks_obtained
                    print("total_marks : ",getexamination.totalmarks , "Marks Obtained : ",marks_obtained)
                    total_marks_of_exam = getexamination.totalmarks
                    percentage_of_marks_obtained = (int(marks_obtained) / int(total_marks_of_exam) )* 100
                    print("total marks :",total_marks_of_exam , "marks obained :",marks_obtained)
                    print("percentage_of_marks_obtained :",percentage_of_marks_obtained)
                    if(percentage_of_marks_obtained >= 35):
                        result = "Pass"
                    else:
                        result = "Fail"
                    set_result = Result(
                        student=student,
                        exam=getexamination,
                        total_question=getexamination.totalquestions,
                        attempted=question_count,
                        total_marks=getexamination.totalmarks,
                        score=marks_obtained,
                        result=result
                        )
                    set_result.save()
                    res = sm(
                        subject = "Examination Submitted",
                        message = '''Examination has been submitted \n
                        CIIS INDIA \n
                        Content\n
                        https://erp.ciisindia.in/''',
                        from_email = 'testmail@erp.ciisindia.in',
                        recipient_list = [student.email],
                        fail_silently=False,
                        )
                    print("email sent to student")
                    return JsonResponse({'added':'yes'})
                        #sun is an astonomical object made up of hydrogen and helium which generates plasma and since it has huge mass it creates a huge gravity.
                exam_id = request.POST.get('exam_id')
                if exam_id:
                    hours = 0
                    minutes = 0
                    #added by Avani on 15/10 - pick hours and minutes for timer from database.
                    getexamination = Examination.objects.get(id=exam_id)
                    duration = getexamination.examduration
                    if duration != 0:
                       hours = int(duration) // 60 
                       minutes = int(duration) % 60
                    
                    try:
                        getexaminationtime = StudentExaminationTime.objects.get(Q(student = student.id ) & Q(exam = exam_id))
                        print("time data available",getexaminationtime)
                        get_time_in_minutes = int(getexaminationtime.time_in_minutes)
                        time_string = str(timedelta(minutes=get_time_in_minutes))[:-3]
                        data_list = time_string.split(":")
                        hours = data_list[0]
                        minutes = data_list[1]
                        print(hours , minutes)

                        # getexaminationtime.time_in_minutes = (int(hours) * 60 ) + int(minutes)
                        # getexaminationtime.save()
                        # print("examination time found",getexaminationtime,"time left:",(int(hours) * 60 ) + int(minutes))
                    except StudentExaminationTime.DoesNotExist:
                        print("time data not available")
                        # print("timer not found",(int(hours) * 60 ) + int(minutes))
                        # create_timer = StudentExaminationTime(
                        #     student = student.id,
                        #     exam = exam_id_timer,
                        #     time_in_minutes = (int(hours) * 60 ) + int(minutes)
                        # )
                        # create_timer.save()
                    
                    getquestions = Questions.objects.filter(exam = getexamination.id)
                    questionsserializer = QuestionsSerializer(getquestions, many=True)
                    temp = []
                    for i in getquestions:
                        
                        try:
                            get_submitted_questions = SubmittedExamination.objects.get(Q(student = user_id) & Q(exam = exam_id) & Q(question=i.id))
                            print("get submitted questions :",get_submitted_questions)
                            submitted_answer = get_submitted_questions.submitted_answer
                        except SubmittedExamination.DoesNotExist:
                            submitted_answer = ""
                        
                        obj = {
                            "id":i.id,
                            "exam":i.exam.id,
                            "question":i.question,
                            "image":i.image,
                            "type":i.type,
                            "marks":i.marks,
                            "option1":i.option1,
                            "option2":i.option2,
                            "option3":i.option3,
                            "option4":i.option4,
                            "option5":i.option5,
                            "option6":i.option6,
                            "shortanswer":i.shortanswer,
                            "answer":i.answer,
                            "submitted_answer":submitted_answer
                        }
                        temp.append(obj)
                    #print(temp)
                    print("AAAAAAA", {'data':temp,'exam_id':getexamination.id,'hours':hours,'minutes':minutes})
                    return JsonResponse({'data':temp,'exam_id':getexamination.id,'hours':hours,'minutes':minutes})
    params = {
        "exams":exams,
        "error":error,
        "givenexams": givenexams
    }
    return render(request,"examination/give_examination.html",params)

def ExamSubmitted(request):
    return render(request,"examination/exam_submitted.html")


@login_required(login_url='/login/')
def EditOldStudentFees(request):
    temp = []
    if request.user.is_superuser:
        if request.method == "POST":
            tution1 = request.POST.get('tution1')
            exam1 = request.POST.get('exam1')
            book1 = request.POST.get('book1')
            resitting1 = request.POST.get('resitting1')
            entrance1 = request.POST.get('entrance1')
            extra1 = request.POST.get('extra1')
            discount1 = request.POST.get('discount1')
            total1 = request.POST.get('total1')

            tution2 = request.POST.get('tution2')
            exam2 = request.POST.get('exam2')
            book2 = request.POST.get('book2')
            resitting2 = request.POST.get('resitting2')
            entrance2 = request.POST.get('entrance2')
            extra2 = request.POST.get('extra2')
            discount2 = request.POST.get('discount2')
            total2 = request.POST.get('total2')

            tution3 = request.POST.get('tution3')
            exam3 = request.POST.get('exam3')
            book3 = request.POST.get('book3')
            resitting3 = request.POST.get('resitting3')
            entrance3 = request.POST.get('entrance3')
            extra3 = request.POST.get('extra3')
            discount3 = request.POST.get('discount3')
            total3 = request.POST.get('total3')

            tution4 = request.POST.get('tution4')
            exam4 = request.POST.get('exam4')
            book4 = request.POST.get('book4')
            resitting4 = request.POST.get('resitting4')
            entrance4 = request.POST.get('entrance4')
            extra4 = request.POST.get('extra4')
            discount4 = request.POST.get('discount4')
            total4 = request.POST.get('total4')

            tution5 = request.POST.get('tution5')
            exam5 = request.POST.get('exam5')
            book5 = request.POST.get('book5')
            resitting5 = request.POST.get('resitting5')
            entrance5 = request.POST.get('entrance5')
            extra5 = request.POST.get('extra5')
            discount5 = request.POST.get('discount5')
            total5 = request.POST.get('total5')

            tution6 = request.POST.get('tution6')
            exam6 = request.POST.get('exam6')
            book6 = request.POST.get('book6')
            resitting6 = request.POST.get('resitting6')
            entrance6 = request.POST.get('entrance6')
            extra6 = request.POST.get('extra6')
            discount6 = request.POST.get('discount6')
            total6 = request.POST.get('total6')

            tution7 = request.POST.get('tution7')
            exam7 = request.POST.get('exam7')
            book7 = request.POST.get('book7')
            resitting7 = request.POST.get('resitting7')
            entrance7 = request.POST.get('entrance7')
            extra7 = request.POST.get('extra7')
            discount7 = request.POST.get('discount7')
            total7 = request.POST.get('total7')

            tution8 = request.POST.get('tution8')
            exam8 = request.POST.get('exam8')
            book8 = request.POST.get('book8')
            resitting8 = request.POST.get('resitting8')
            entrance8 = request.POST.get('entrance8')
            extra8 = request.POST.get('extra8')
            discount8 = request.POST.get('discount8')
            total8 = request.POST.get('total8')
            print("data :",tution1,exam1,book1, resitting1, entrance1, extra1,discount1, total1)
            print("data :",tution2,exam2,book2, resitting2, entrance2, extra2,discount2, total2)
            print("data :",tution3,exam3,book3, resitting3, entrance3, extra3,discount3, total3)
            print("data :",tution4,exam4,book4, resitting4, entrance4, extra4,discount4, total4)
            print("data :",tution5,exam5,book5, resitting5, entrance5, extra5,discount5, total5)
            print("data :",tution6,exam6,book6, resitting6, entrance6, extra6,discount6, total6)
            print("data :",tution7,exam7,book7, resitting7, entrance7, extra7,discount7, total7)
            print("data :",tution8,exam8,book8, resitting8, entrance8, extra8,discount8, total8)
            
            
            



            student_id = request.POST.get('student')
            no_of_semester = request.POST.get('no_of_semester')
            stream_semester = request.POST.get('stream_semester')
            
            
            if student_id and no_of_semester:
                enrolled = Enrolled.objects.get(student = student_id)
                print(enrolled)
                if no_of_semester == '4':
                    
                    try:
                        semesterfees1 = StudentFees.objects.get(Q(student = student_id) & Q(sem="1"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution1,exam1,book1, resitting1, entrance1, extra1,discount1, total1)
                        if semesterfees1.tutionfees == tution1:
                            pass
                        else:
                            semesterfees1.tutionfees = tution1
                        if semesterfees1.examinationfees == exam1:
                            pass
                        else:
                            semesterfees1.examinationfees = exam1
                        if semesterfees1.bookfees == book1:
                            pass
                        else:
                            semesterfees1.bookfees = book1
                        if semesterfees1.resittingfees == resitting1:
                            pass
                        else:
                            semesterfees1.resittingfees = resitting1
                        if semesterfees1.entrancefees == entrance1:
                            pass
                        else:
                            semesterfees1.entrancefees = entrance1
                        if semesterfees1.extrafees == extra1:
                            pass
                        else:
                            semesterfees1.extrafees = extra1
                        if semesterfees1.discount == discount1:
                            pass
                        else:
                            semesterfees1.discount = discount1
                        if semesterfees1.totalfees == total1:
                            pass
                        else:
                            tempfees = semesterfees1.totalfees
                            semesterfees1.totalfees = total1
                            if mainreciept:
                                mainreciept.pendingamount = int(total1) - int(mainreciept.paidamount)
                                diff = int(total1) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")                                                      
                        semesterfees1.save()
                        
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution1,examinationfees=exam1,bookfees=book1,resittingfees=resitting1,entrancefees=entrance1,extrafees=extra1,discount=discount1,totalfees=total1,sem="1")
                        addsemesterfees.save()
                    try:
                        semesterfees2 = StudentFees.objects.get(Q(student = student_id) & Q(sem="2"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution2,exam2,book2, resitting2, entrance2, extra2,discount2, total2)
                        if semesterfees2.tutionfees == tution2:
                            pass
                        else:
                            semesterfees2.tutionfees = tution2
                        if semesterfees2.examinationfees == exam2:
                            pass
                        else:
                            semesterfees2.examinationfees = exam2
                        if semesterfees2.bookfees == book2:
                            pass
                        else:
                            semesterfees2.bookfees = book2
                        if semesterfees2.resittingfees == resitting2:
                            pass
                        else:
                            semesterfees2.resittingfees = resitting2
                        if semesterfees2.entrancefees == entrance2:
                            pass
                        else:
                            semesterfees2.entrancefees = entrance2
                        if semesterfees2.extrafees == extra2:
                            pass
                        else:
                            semesterfees2.extrafees = extra2
                        if semesterfees2.discount == discount2:
                            pass
                        else:
                            semesterfees2.discount = discount2
                        if semesterfees2.totalfees == total2:
                            pass
                        else:
                            tempfees = semesterfees2.totalfees
                            semesterfees2.totalfees = total2
                            if mainreciept:
                                mainreciept.pendingamount = int(total2) - int(mainreciept.paidamount)
                                diff = int(total2) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees2.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution2,examinationfees=exam2,bookfees=book2,resittingfees=resitting2,entrancefees=entrance2,extrafees=extra2,discount=discount2,totalfees=total2,sem="2")
                        addsemesterfees.save()
                    try:
                        semesterfees3 = StudentFees.objects.get(Q(student = student_id) & Q(sem="3"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution3,exam3,book3, resitting3, entrance3, extra3,discount3, total3)
                        if semesterfees3.tutionfees == tution3:
                            pass
                        else:
                            semesterfees3.tutionfees = tution3
                        if semesterfees3.examinationfees == exam3:
                            pass
                        else:
                            semesterfees3.examinationfees = exam3
                        if semesterfees3.bookfees == book3:
                            pass
                        else:
                            semesterfees3.bookfees = book3
                        if semesterfees3.resittingfees == resitting3:
                            pass
                        else:
                            semesterfees3.resittingfees = resitting3
                        if semesterfees3.entrancefees == entrance3:
                            pass
                        else:
                            semesterfees3.entrancefees = entrance3
                        if semesterfees3.extrafees == extra3:
                            pass
                        else:
                            semesterfees3.extrafees = extra3
                        if semesterfees3.discount == discount3:
                            pass
                        else:
                            semesterfees3.discount = discount3
                        if semesterfees3.totalfees == total3:
                            pass
                        else:
                            tempfees = semesterfees3.totalfees
                            semesterfees3.totalfees = total3
                            if mainreciept:
                                mainreciept.pendingamount = int(total3) - int(mainreciept.paidamount)
                                diff = int(total3) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees3.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution3,examinationfees=exam3,bookfees=book3,resittingfees=resitting3,entrancefees=entrance3,extrafees=extra3,discount=discount3,totalfees=total3,sem="3")
                        addsemesterfees.save()
                    try:
                        semesterfees4 = StudentFees.objects.get(Q(student = student_id) & Q(sem="4"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution4,exam4,book4, resitting4, entrance4, extra4,discount4, total4)
                        if semesterfees4.tutionfees == tution4:
                            pass
                        else:
                            semesterfees4.tutionfees = tution4
                        if semesterfees4.examinationfees == exam4:
                            pass
                        else:
                            semesterfees4.examinationfees = exam4
                        if semesterfees4.bookfees == book4:
                            pass
                        else:
                            semesterfees4.bookfees = book4
                        if semesterfees4.resittingfees == resitting4:
                            pass
                        else:
                            semesterfees4.resittingfees = resitting4
                        if semesterfees4.entrancefees == entrance4:
                            pass
                        else:
                            semesterfees4.entrancefees = entrance4
                        if semesterfees4.extrafees == extra4:
                            pass
                        else:
                            semesterfees4.extrafees = extra4
                        if semesterfees4.discount == discount4:
                            pass
                        else:
                            semesterfees4.discount = discount4
                        if semesterfees4.totalfees == total4:
                            pass
                        else:
                            tempfees = semesterfees4.totalfees
                            semesterfees4.totalfees = total4
                            if mainreciept:
                                mainreciept.pendingamount = int(total4) - int(mainreciept.paidamount)
                                diff = int(total4) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees4.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution4,examinationfees=exam4,bookfees=book4,resittingfees=resitting4,entrancefees=entrance4,extrafees=extra4,discount=discount4,totalfees=total4,sem="4")
                        addsemesterfees.save()
                    return JsonResponse({'added':'yes'})

                elif no_of_semester == '6':
                    try:
                        semesterfees1 = StudentFees.objects.get(Q(student = student_id)  & Q(sem="1"))
                        
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                            print("data :",tution1,exam1,book1, resitting1, entrance1, extra1,discount1, total1)
                        if semesterfees1.tutionfees == tution1:
                            pass
                        else:
                            semesterfees1.tutionfees = tution1
                        if semesterfees1.examinationfees == exam1:
                            pass
                        else:
                            semesterfees1.examinationfees = exam1
                        if semesterfees1.bookfees == book1:
                            pass
                        else:
                            semesterfees1.bookfees = book1
                        if semesterfees1.resittingfees == resitting1:
                            pass
                        else:
                            semesterfees1.resittingfees = resitting1
                        if semesterfees1.entrancefees == entrance1:
                            pass
                        else:
                            semesterfees1.entrancefees = entrance1
                        if semesterfees1.extrafees == extra1:
                            pass
                        else:
                            semesterfees1.extrafees = extra1
                        if semesterfees1.discount == discount1:
                            pass
                        else:
                            semesterfees1.discount = discount1
                        if semesterfees1.totalfees == total1:
                            pass
                        else:
                            tempfees = semesterfees1.totalfees
                            semesterfees1.totalfees = total1
                            if mainreciept:
                                mainreciept.pendingamount = int(total1) - int(mainreciept.paidamount)
                                diff = int(total1) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                            

                            
                        
                        semesterfees1.save()
                        
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution1,examinationfees=exam1,bookfees=book1,resittingfees=resitting1,entrancefees=entrance1,extrafees=extra1,discount=discount1,totalfees=total1,sem="1")
                        addsemesterfees.save()
                    try:
                        semesterfees2 = StudentFees.objects.get(Q(student = student_id) & Q(sem="2"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution2,exam2,book2, resitting2, entrance2, extra2,discount2, total2)
                        if semesterfees2.tutionfees == tution2:
                            pass
                        else:
                            semesterfees2.tutionfees = tution2
                        if semesterfees2.examinationfees == exam2:
                            pass
                        else:
                            semesterfees2.examinationfees = exam2
                        if semesterfees2.bookfees == book2:
                            pass
                        else:
                            semesterfees2.bookfees = book2
                        if semesterfees2.resittingfees == resitting2:
                            pass
                        else:
                            semesterfees2.resittingfees = resitting2
                        if semesterfees2.entrancefees == entrance2:
                            pass
                        else:
                            semesterfees2.entrancefees = entrance2
                        if semesterfees2.extrafees == extra2:
                            pass
                        else:
                            semesterfees2.extrafees = extra2
                        if semesterfees2.discount == discount2:
                            pass
                        else:
                            semesterfees2.discount = discount2
                        if semesterfees2.totalfees == total2:
                            pass
                        else:
                            
                            tempfees = semesterfees2.totalfees
                            semesterfees2.totalfees = total2
                            if mainreciept:
                                mainreciept.pendingamount = int(total2) - int(mainreciept.paidamount)
                                diff = int(total2) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees2.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution2,examinationfees=exam2,bookfees=book2,resittingfees=resitting2,entrancefees=entrance2,extrafees=extra2,discount=discount2,totalfees=total2,sem="2")
                        addsemesterfees.save()
                    try:
                        semesterfees3 = StudentFees.objects.get(Q(student = student_id) & Q(sem="3"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution3,exam3,book3, resitting3, entrance3, extra3,discount3, total3)
                        if semesterfees3.tutionfees == tution3:
                            pass
                        else:
                            semesterfees3.tutionfees = tution3
                        if semesterfees3.examinationfees == exam3:
                            pass
                        else:
                            semesterfees3.examinationfees = exam3
                        if semesterfees3.bookfees == book3:
                            pass
                        else:
                            semesterfees3.bookfees = book3
                        if semesterfees3.resittingfees == resitting3:
                            pass
                        else:
                            semesterfees3.resittingfees = resitting3
                        if semesterfees3.entrancefees == entrance3:
                            pass
                        else:
                            semesterfees3.entrancefees = entrance3
                        if semesterfees3.extrafees == extra3:
                            pass
                        else:
                            semesterfees3.extrafees = extra3
                        if semesterfees3.discount == discount3:
                            pass
                        else:
                            semesterfees3.discount = discount3
                        if semesterfees3.totalfees == total3:
                            pass
                        else:
                            tempfees = semesterfees3.totalfees
                            semesterfees3.totalfees = total3
                            if mainreciept:
                                mainreciept.pendingamount = int(total3) - int(mainreciept.paidamount)
                                diff = int(total3) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees3.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution3,examinationfees=exam3,bookfees=book3,resittingfees=resitting3,entrancefees=entrance3,extrafees=extra3,discount=discount3,totalfees=total3,sem="3")
                        addsemesterfees.save()
                    try:
                        semesterfees4 = StudentFees.objects.get(Q(student = student_id) & Q(sem="4"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution4,exam4,book4, resitting4, entrance4, extra4,discount4, total4)
                        if semesterfees4.tutionfees == tution4:
                            pass
                        else:
                            semesterfees4.tutionfees = tution4
                        if semesterfees4.examinationfees == exam4:
                            pass
                        else:
                            semesterfees4.examinationfees = exam4
                        if semesterfees4.bookfees == book4:
                            pass
                        else:
                            semesterfees4.bookfees = book4
                        if semesterfees4.resittingfees == resitting4:
                            pass
                        else:
                            semesterfees4.resittingfees = resitting4
                        if semesterfees4.entrancefees == entrance4:
                            pass
                        else:
                            semesterfees4.entrancefees = entrance4
                        if semesterfees4.extrafees == extra4:
                            pass
                        else:
                            semesterfees4.extrafees = extra4
                        if semesterfees4.discount == discount4:
                            pass
                        else:
                            semesterfees4.discount = discount4
                        if semesterfees4.totalfees == total4:
                            pass
                        else:
                            tempfees = semesterfees4.totalfees
                            semesterfees4.totalfees = total4
                            if mainreciept:
                                mainreciept.pendingamount = int(total4) - int(mainreciept.paidamount)
                                diff = int(total4) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees4.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution4,examinationfees=exam4,bookfees=book4,resittingfees=resitting4,entrancefees=entrance4,extrafees=extra4,discount=discount4,totalfees=total4,sem="4")
                        addsemesterfees.save()
                    try:
                        semesterfees5 = StudentFees.objects.get(Q(student = student_id) & Q(sem="5"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "5")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "5")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution5,exam5,book5, resitting5, entrance5, extra5,discount5, total5)
                        if semesterfees5.tutionfees == tution5:
                            pass
                        else:
                            semesterfees5.tutionfees = tution5
                        if semesterfees5.examinationfees == exam5:
                            pass
                        else:
                            semesterfees5.examinationfees = exam5
                        if semesterfees5.bookfees == book5:
                            pass
                        else:
                            semesterfees5.bookfees = book5
                        if semesterfees5.resittingfees == resitting5:
                            pass
                        else:
                            semesterfees5.resittingfees = resitting5
                        if semesterfees5.entrancefees == entrance5:
                            pass
                        else:
                            semesterfees5.entrancefees = entrance5
                        if semesterfees5.extrafees == extra5:
                            pass
                        else:
                            semesterfees5.extrafees = extra5
                        if semesterfees5.discount == discount5:
                            pass
                        else:
                            semesterfees5.discount = discount5
                        if semesterfees5.totalfees == total5:
                            pass
                        else:
                            tempfees = semesterfees5.totalfees
                            semesterfees5.totalfees = total5
                            if mainreciept:
                                mainreciept.pendingamount = int(total5) - int(mainreciept.paidamount)
                                diff = int(total5) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees5.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution5,examinationfees=exam5,bookfees=book5,resittingfees=resitting5,entrancefees=entrance5,extrafees=extra5,discount=discount5,totalfees=total5,sem="5")
                        addsemesterfees.save()
                    try:
                        semesterfees6 = StudentFees.objects.get(Q(student = student_id) & Q(sem="6"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "6")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "6")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution6,exam6,book6, resitting6, entrance6, extra6,discount6, total6)
                        if semesterfees6.tutionfees == tution6:
                            pass
                        else:
                            semesterfees6.tutionfees = tution6
                        if semesterfees6.examinationfees == exam6:
                            pass
                        else:
                            semesterfees6.examinationfees = exam6
                        if semesterfees6.bookfees == book6:
                            pass
                        else:
                            semesterfees6.bookfees = book6
                        if semesterfees6.resittingfees == resitting6:
                            pass
                        else:
                            semesterfees6.resittingfees = resitting6
                        if semesterfees6.entrancefees == entrance6:
                            pass
                        else:
                            semesterfees6.entrancefees = entrance6
                        if semesterfees6.extrafees == extra6:
                            pass
                        else:
                            semesterfees6.extrafees = extra6
                        if semesterfees6.discount == discount6:
                            pass
                        else:
                            semesterfees6.discount = discount6
                        if semesterfees6.totalfees == total6:
                            pass
                        else:
                            tempfees = semesterfees6.totalfees
                            semesterfees6.totalfees = total6
                            if mainreciept:
                                mainreciept.pendingamount = int(total6) - int(mainreciept.paidamount)
                                diff = int(total6) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees6.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution6,examinationfees=exam6,bookfees=book6,resittingfees=resitting6,entrancefees=entrance6,extrafees=extra6,discount=discount6,totalfees=total6,sem="6")
                        addsemesterfees.save()
                    return JsonResponse({'added':'yes'})
                    
                elif no_of_semester == '8':
                    try:
                        semesterfees1 = StudentFees.objects.get(Q(student = student_id) & Q(sem="1"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution1,exam1,book1, resitting1, entrance1, extra1,discount1, total1)
                        if semesterfees1.tutionfees == tution1:
                            pass
                        else:
                            semesterfees1.tutionfees = tution1
                        if semesterfees1.examinationfees == exam1:
                            pass
                        else:
                            semesterfees1.examinationfees = exam1
                        if semesterfees1.bookfees == book1:
                            pass
                        else:
                            semesterfees1.bookfees = book1
                        if semesterfees1.resittingfees == resitting1:
                            pass
                        else:
                            semesterfees1.resittingfees = resitting1
                        if semesterfees1.entrancefees == entrance1:
                            pass
                        else:
                            semesterfees1.entrancefees = entrance1
                        if semesterfees1.extrafees == extra1:
                            pass
                        else:
                            semesterfees1.extrafees = extra1
                        if semesterfees1.discount == discount1:
                            pass
                        else:
                            semesterfees1.discount = discount1
                        if semesterfees1.totalfees == total1:
                            pass
                        else:
                            tempfees = semesterfees1.totalfees
                            semesterfees1.totalfees = total1
                            if mainreciept:
                                mainreciept.pendingamount = int(total1) - int(mainreciept.paidamount)
                                diff = int(total1) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        
                        semesterfees1.save()
                        
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution1,examinationfees=exam1,bookfees=book1,resittingfees=resitting1,entrancefees=entrance1,extrafees=extra1,discount=discount1,totalfees=total1,sem="1")
                        addsemesterfees.save()
                    try:
                        semesterfees2 = StudentFees.objects.get(Q(student = student_id) & Q(sem="2"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution2,exam2,book2, resitting2, entrance2, extra2,discount2, total2)
                        if semesterfees2.tutionfees == tution2:
                            pass
                        else:
                            semesterfees2.tutionfees = tution2
                        if semesterfees2.examinationfees == exam2:
                            pass
                        else:
                            semesterfees2.examinationfees = exam2
                        if semesterfees2.bookfees == book2:
                            pass
                        else:
                            semesterfees2.bookfees = book2
                        if semesterfees2.resittingfees == resitting2:
                            pass
                        else:
                            semesterfees2.resittingfees = resitting2
                        if semesterfees2.entrancefees == entrance2:
                            pass
                        else:
                            semesterfees2.entrancefees = entrance2
                        if semesterfees2.extrafees == extra2:
                            pass
                        else:
                            semesterfees2.extrafees = extra2
                        if semesterfees2.discount == discount2:
                            pass
                        else:
                            semesterfees2.discount = discount2
                        if semesterfees2.totalfees == total2:
                            pass
                        else:
                            tempfees = semesterfees2.totalfees
                            semesterfees2.totalfees = total2
                            if mainreciept:
                                mainreciept.pendingamount = int(total2) - int(mainreciept.paidamount)
                                diff = int(total2) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees2.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution2,examinationfees=exam2,bookfees=book2,resittingfees=resitting2,entrancefees=entrance2,extrafees=extra2,discount=discount2,totalfees=total2,sem="2")
                        addsemesterfees.save()
                    try:
                        semesterfees3 = StudentFees.objects.get(Q(student = student_id) & Q(sem="3"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution3,exam3,book3, resitting3, entrance3, extra3,discount3, total3)
                        if semesterfees3.tutionfees == tution3:
                            pass
                        else:
                            semesterfees3.tutionfees = tution3
                        if semesterfees3.examinationfees == exam3:
                            pass
                        else:
                            semesterfees3.examinationfees = exam3
                        if semesterfees3.bookfees == book3:
                            pass
                        else:
                            semesterfees3.bookfees = book3
                        if semesterfees3.resittingfees == resitting3:
                            pass
                        else:
                            semesterfees3.resittingfees = resitting3
                        if semesterfees3.entrancefees == entrance3:
                            pass
                        else:
                            semesterfees3.entrancefees = entrance3
                        if semesterfees3.extrafees == extra3:
                            pass
                        else:
                            semesterfees3.extrafees = extra3
                        if semesterfees3.discount == discount3:
                            pass
                        else:
                            semesterfees3.discount = discount3
                        if semesterfees3.totalfees == total3:
                            pass
                        else:
                            tempfees = semesterfees3.totalfees
                            semesterfees3.totalfees = total3
                            if mainreciept:
                                mainreciept.pendingamount = int(total3) - int(mainreciept.paidamount)
                                diff = int(total3) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees3.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution3,examinationfees=exam3,bookfees=book3,resittingfees=resitting3,entrancefees=entrance3,extrafees=extra3,discount=discount3,totalfees=total3,sem="3")
                        addsemesterfees.save()
                    try:
                        semesterfees4 = StudentFees.objects.get(Q(student = student_id) & Q(sem="4"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution4,exam4,book4, resitting4, entrance4, extra4,discount4, total4)
                        if semesterfees4.tutionfees == tution4:
                            pass
                        else:
                            semesterfees4.tutionfees = tution4
                        if semesterfees4.examinationfees == exam4:
                            pass
                        else:
                            semesterfees4.examinationfees = exam4
                        if semesterfees4.bookfees == book4:
                            pass
                        else:
                            semesterfees4.bookfees = book4
                        if semesterfees4.resittingfees == resitting4:
                            pass
                        else:
                            semesterfees4.resittingfees = resitting4
                        if semesterfees4.entrancefees == entrance4:
                            pass
                        else:
                            semesterfees4.entrancefees = entrance4
                        if semesterfees4.extrafees == extra4:
                            pass
                        else:
                            semesterfees4.extrafees = extra4
                        if semesterfees4.discount == discount4:
                            pass
                        else:
                            semesterfees4.discount = discount4
                        if semesterfees4.totalfees == total4:
                            pass
                        else:
                            tempfees = semesterfees4.totalfees
                            semesterfees4.totalfees = total4
                            if mainreciept:
                                mainreciept.pendingamount = int(total4) - int(mainreciept.paidamount)
                                diff = int(total4) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees4.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution4,examinationfees=exam4,bookfees=book4,resittingfees=resitting4,entrancefees=entrance4,extrafees=extra4,discount=discount4,totalfees=total4,sem="4")
                        addsemesterfees.save()
                    try:
                        semesterfees5 = StudentFees.objects.get(Q(student = student_id) & Q(sem="5"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "5")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "5")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution5,exam5,book5, resitting5, entrance5, extra5,discount5, total5)
                        if semesterfees5.tutionfees == tution5:
                            pass
                        else:
                            semesterfees5.tutionfees = tution5
                        if semesterfees5.examinationfees == exam5:
                            pass
                        else:
                            semesterfees5.examinationfees = exam5
                        if semesterfees5.bookfees == book5:
                            pass
                        else:
                            semesterfees5.bookfees = book5
                        if semesterfees5.resittingfees == resitting5:
                            pass
                        else:
                            semesterfees5.resittingfees = resitting5
                        if semesterfees5.entrancefees == entrance5:
                            pass
                        else:
                            semesterfees5.entrancefees = entrance5
                        if semesterfees5.extrafees == extra5:
                            pass
                        else:
                            semesterfees5.extrafees = extra5
                        if semesterfees5.discount == discount5:
                            pass
                        else:
                            semesterfees5.discount = discount5
                        if semesterfees5.totalfees == total5:
                            pass
                        else:
                            tempfees = semesterfees5.totalfees
                            semesterfees5.totalfees = total5
                            if mainreciept:
                                mainreciept.pendingamount = int(total5) - int(mainreciept.paidamount)
                                diff = int(total5) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees5.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution5,examinationfees=exam5,bookfees=book5,resittingfees=resitting5,entrancefees=entrance5,extrafees=extra5,discount=discount5,totalfees=total5,sem="5")
                        addsemesterfees.save()
                    try:
                        semesterfees6 = StudentFees.objects.get(Q(student = student_id) & Q(sem="6"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "6")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "6")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution6,exam6,book6, resitting6, entrance6, extra6,discount6, total6)
                        if semesterfees6.tutionfees == tution6:
                            pass
                        else:
                            semesterfees6.tutionfees = tution6
                        if semesterfees6.examinationfees == exam6:
                            pass
                        else:
                            semesterfees6.examinationfees = exam6
                        if semesterfees6.bookfees == book6:
                            pass
                        else:
                            semesterfees6.bookfees = book6
                        if semesterfees6.resittingfees == resitting6:
                            pass
                        else:
                            semesterfees6.resittingfees = resitting6
                        if semesterfees6.entrancefees == entrance6:
                            pass
                        else:
                            semesterfees6.entrancefees = entrance6
                        if semesterfees6.extrafees == extra6:
                            pass
                        else:
                            semesterfees6.extrafees = extra6
                        if semesterfees6.discount == discount6:
                            pass
                        else:
                            semesterfees6.discount = discount6
                        if semesterfees6.totalfees == total6:
                            pass
                        else:
                            tempfees = semesterfees6.totalfees
                            semesterfees6.totalfees = total6
                            if mainreciept:
                                mainreciept.pendingamount = int(total6) - int(mainreciept.paidamount)
                                diff = int(total6) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees6.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution6,examinationfees=exam6,bookfees=book6,resittingfees=resitting6,entrancefees=entrance6,extrafees=extra6,discount=discount6,totalfees=total6,sem="6")
                        addsemesterfees.save()
                    try:
                        semesterfees7 = StudentFees.objects.get(Q(student = student_id) & Q(sem="7"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "7")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "7")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution7,exam7,book7, resitting7, entrance7, extra7,discount7, total7)
                        if semesterfees7.tutionfees == tution7:
                            pass
                        else:
                            semesterfees7.tutionfees = tution7
                        if semesterfees7.examinationfees == exam7:
                            pass
                        else:
                            semesterfees7.examinationfees = exam7
                        if semesterfees7.bookfees == book7:
                            pass
                        else:
                            semesterfees7.bookfees = book7
                        if semesterfees7.resittingfees == resitting7:
                            pass
                        else:
                            semesterfees7.resittingfees = resitting7
                        if semesterfees7.entrancefees == entrance7:
                            pass
                        else:
                            semesterfees7.entrancefees = entrance7
                        if semesterfees7.extrafees == extra7:
                            pass
                        else:
                            semesterfees7.extrafees = extra7
                        if semesterfees7.discount == discount7:
                            pass
                        else:
                            semesterfees7.discount = discount7
                        if semesterfees7.totalfees == total7:
                            pass
                        else:
                            tempfees = semesterfees7.totalfees
                            semesterfees7.totalfees = total7
                            if mainreciept:
                                mainreciept.pendingamount = int(total7) - int(mainreciept.paidamount)
                                diff = int(total7) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees7.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution7,examinationfees=exam7,bookfees=book7,resittingfees=resitting7,entrancefees=entrance7,extrafees=extra7,discount=discount7,totalfees=total7,sem="7")
                        addsemesterfees.save()
                    try:
                        semesterfees8 = StudentFees.objects.get(Q(student = student_id) & Q(sem="8"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "8")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "8")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data :",tution8,exam8,book8, resitting8, entrance8, extra8,discount8, total8)
                        if semesterfees8.tutionfees == tution8:
                            pass
                        else:
                            semesterfees8.tutionfees = tution8
                        if semesterfees8.examinationfees == exam8:
                            pass
                        else:
                            semesterfees8.examinationfees = exam8
                        if semesterfees8.bookfees == book8:
                            pass
                        else:
                            semesterfees8.bookfees = book8
                        if semesterfees8.resittingfees == resitting8:
                            pass
                        else:
                            semesterfees8.resittingfees = resitting8
                        if semesterfees8.entrancefees == entrance8:
                            pass
                        else:
                            semesterfees8.entrancefees = entrance8
                        if semesterfees8.extrafees == extra8:
                            pass
                        else:
                            semesterfees8.extrafees = extra8
                        if semesterfees8.discount == discount8:
                            pass
                        else:
                            semesterfees8.discount = discount8
                        if semesterfees8.totalfees == total8:
                            pass
                        else:
                            tempfees = semesterfees8.totalfees
                            semesterfees8.totalfees = total8
                            if mainreciept:
                                mainreciept.pendingamount = int(total8) - int(mainreciept.paidamount)
                                diff = int(total8) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")
                            
                        semesterfees8.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Semester",stream=stream_semester, tutionfees=tution8,examinationfees=exam8,bookfees=book8,resittingfees=resitting8,entrancefees=entrance8,extrafees=extra8,discount=discount8,totalfees=total8,sem="8")
                        addsemesterfees.save()
                    return JsonResponse({'added':'yes'})

            no_of_year = request.POST.get('no_of_year')
            stream_year = request.POST.get('stream_year')
            year_tution1 = request.POST.get('year_tution1')
            year_exam1 = request.POST.get('year_exam1')
            year_book1 = request.POST.get('year_book1')
            year_resitting1 = request.POST.get('year_resitting1')
            year_entrance1 = request.POST.get('year_entrance1')
            year_extra1 = request.POST.get('year_extra1')
            year_discount1 = request.POST.get('year_discount1')
            year_total1 = request.POST.get('year_total1')
            
            year_tution2 = request.POST.get('year_tution2')
            year_exam2 = request.POST.get('year_exam2')
            year_book2 = request.POST.get('year_book2')
            year_resitting2 = request.POST.get('year_resitting2')
            year_entrance2 = request.POST.get('year_entrance2')
            year_extra2 = request.POST.get('year_extra2')
            year_discount2 = request.POST.get('year_discount2')
            year_total2 = request.POST.get('year_total2')
            
            year_tution3 = request.POST.get('year_tution3')
            year_exam3 = request.POST.get('year_exam3')
            year_book3 = request.POST.get('year_book3')
            year_resitting3 = request.POST.get('year_resitting3')
            year_entrance3 = request.POST.get('year_entrance3')
            year_extra3 = request.POST.get('year_extra3')
            year_discount3 = request.POST.get('year_discount3')
            year_total3 = request.POST.get('year_total3')
            
            year_tution4 = request.POST.get('year_tution4')
            year_exam4 = request.POST.get('year_exam4')
            year_book4 = request.POST.get('year_book4')
            year_resitting4 = request.POST.get('year_resitting4')
            year_entrance4 = request.POST.get('year_entrance4')
            year_extra4 = request.POST.get('year_extra4')
            year_discount4 = request.POST.get('year_discount4')
            year_total4 = request.POST.get('year_total4')


            if student_id and no_of_year:
                if no_of_year == '2':
                    
                    try:
                        semesterfees1 = StudentFees.objects.get(Q(student = student_id) & Q(sem="1"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            print("all payment reciept : ",allpaymentreciept)
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution1, year_exam1, year_book1, year_resitting1, year_entrance1, year_extra1, year_discount1, year_total1)
                        if semesterfees1.tutionfees == year_tution1:
                            pass
                        else:
                            semesterfees1.tutionfees = year_tution1
                        if semesterfees1.examinationfees == year_exam1:
                            pass
                        else:
                            semesterfees1.examinationfees = year_exam1
                        if semesterfees1.bookfees == year_book1:
                            pass
                        else:
                            semesterfees1.bookfees = year_book1
                        if semesterfees1.resittingfees == year_resitting1:
                            pass
                        else:
                            semesterfees1.resittingfees = year_resitting1
                        if semesterfees1.entrancefees == year_entrance1:
                            pass
                        else:
                            semesterfees1.entrancefees = year_entrance1
                        if semesterfees1.extrafees == year_extra1:
                            pass
                        else:
                            semesterfees1.extrafees = year_extra1
                        if semesterfees1.discount == year_discount1:
                            pass
                        else:
                            semesterfees1.discount = year_discount1
                        if semesterfees1.totalfees == year_total1:
                            pass
                        else:
                            tempfees = semesterfees1.totalfees
                            semesterfees1.totalfees = year_total1
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total1) - int(mainreciept.paidamount)
                                diff = int(year_total1) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")                          
                        semesterfees1.save()
                        
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution1,examinationfees=year_exam1,bookfees=year_book1,resittingfees=year_resitting1,entrancefees=year_entrance1,extrafees=year_extra1,discount=year_discount1,totalfees=year_total1,sem="1")
                        addsemesterfees.save()
                        print("added year 2 year 1")
                    try:
                        semesterfees2 = StudentFees.objects.get(Q(student = student_id) & Q(sem="2"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution2, year_exam2, year_book2, year_resitting2, year_entrance2, year_extra2, year_discount2, year_total2)
                        if semesterfees2.tutionfees == year_tution2:
                            pass
                        else:
                            semesterfees2.tutionfees = year_tution2
                        if semesterfees2.examinationfees == year_exam2:
                            pass
                        else:
                            semesterfees2.examinationfees = year_exam2
                        if semesterfees2.bookfees == year_book2:
                            pass
                        else:
                            semesterfees2.bookfees = year_book2
                        if semesterfees2.resittingfees == year_resitting2:
                            pass
                        else:
                            semesterfees2.resittingfees = year_resitting2
                        if semesterfees2.entrancefees == year_entrance2:
                            pass
                        else:
                            semesterfees2.entrancefees = year_entrance2
                        if semesterfees2.extrafees == year_extra2:
                            pass
                        else:
                            semesterfees2.extrafees = year_extra2
                        if semesterfees2.discount == year_discount2:
                            pass
                        else:
                            semesterfees2.discount = year_discount2
                        if semesterfees2.totalfees == year_total2:
                            pass
                        else:
                            tempfees = semesterfees2.totalfees
                            semesterfees2.totalfees = year_total2
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total2) - int(mainreciept.paidamount)
                                diff = int(year_total2) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        semesterfees2.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id, studypattern="Annual",stream=stream_year, tutionfees=year_tution2,examinationfees=year_exam2,bookfees=year_book2,resittingfees=year_resitting2,entrancefees=year_entrance2,extrafees=year_extra2,discount=year_discount2,totalfees=year_total2,sem="2")
                        addsemesterfees.save()
                        print("added year 2 year 2")
                    
                    return JsonResponse({'added':'yes'})

                elif no_of_year == '3':
                    try:
                        semesterfees1 = StudentFees.objects.get(Q(student = student_id) & Q(sem="1"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution1, year_exam1, year_book1, year_resitting1, year_entrance1, year_extra1, year_discount1, year_total1)
                        if semesterfees1.tutionfees == year_tution1:
                            pass
                        else:
                            semesterfees1.tutionfees = year_tution1
                        if semesterfees1.examinationfees == year_exam1:
                            pass
                        else:
                            semesterfees1.examinationfees = year_exam1
                        if semesterfees1.bookfees == year_book1:
                            pass
                        else:
                            semesterfees1.bookfees = year_book1
                        if semesterfees1.resittingfees == year_resitting1:
                            pass
                        else:
                            semesterfees1.resittingfees = year_resitting1
                        if semesterfees1.entrancefees == year_entrance1:
                            pass
                        else:
                            semesterfees1.entrancefees = year_entrance1
                        if semesterfees1.extrafees == year_extra1:
                            pass
                        else:
                            semesterfees1.extrafees = year_extra1
                        if semesterfees1.discount == year_discount1:
                            pass
                        else:
                            semesterfees1.discount = year_discount1
                        if semesterfees1.totalfees == year_total1:
                            pass
                        else:
                            tempfees = semesterfees1.totalfees
                            semesterfees1.totalfees = year_total1
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total1) - int(mainreciept.paidamount)
                                diff = int(year_total1) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        
                        semesterfees1.save()
                        
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution1,examinationfees=year_exam1,bookfees=year_book1,resittingfees=year_resitting1,entrancefees=year_entrance1,extrafees=year_extra1,discount=year_discount1,totalfees=year_total1,sem="1")
                        addsemesterfees.save()
                        print("added year 3 year 1")
                    try:
                        semesterfees2 = StudentFees.objects.get(Q(student = student_id) & Q(sem="2"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution2, year_exam2, year_book2, year_resitting2, year_entrance2, year_extra2, year_discount2, year_total2)
                        if semesterfees2.tutionfees == year_tution2:
                            pass
                        else:
                            semesterfees2.tutionfees = year_tution2
                        if semesterfees2.examinationfees == year_exam2:
                            pass
                        else:
                            semesterfees2.examinationfees = year_exam2
                        if semesterfees2.bookfees == year_book2:
                            pass
                        else:
                            semesterfees2.bookfees = year_book2
                        if semesterfees2.resittingfees == year_resitting2:
                            pass
                        else:
                            semesterfees2.resittingfees = year_resitting2
                        if semesterfees2.entrancefees == year_entrance2:
                            pass
                        else:
                            semesterfees2.entrancefees = year_entrance2
                        if semesterfees2.extrafees == year_extra2:
                            pass
                        else:
                            semesterfees2.extrafees = year_extra2
                        if semesterfees2.discount == year_discount2:
                            pass
                        else:
                            semesterfees2.discount = year_discount2
                        if semesterfees2.totalfees == year_total2:
                            pass
                        else:
                            tempfees = semesterfees2.totalfees
                            semesterfees2.totalfees = year_total2
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total2) - int(mainreciept.paidamount)
                                diff = int(year_total2) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        semesterfees2.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution2,examinationfees=year_exam2,bookfees=year_book2,resittingfees=year_resitting2,entrancefees=year_entrance2,extrafees=year_extra2,discount=year_discount2,totalfees=year_total2,sem="2")
                        addsemesterfees.save()
                        print("added year 3 year 2")
                    try:
                        semesterfees3 = StudentFees.objects.get(Q(student = student_id) & Q(sem="3"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution3, year_exam3, year_book3, year_resitting3, year_entrance3, year_extra3, year_discount3, year_total3)
                        if semesterfees3.tutionfees == year_tution3:
                            pass
                        else:
                            semesterfees3.tutionfees = year_tution3
                        if semesterfees3.examinationfees == year_exam3:
                            pass
                        else:
                            semesterfees3.examinationfees = year_exam3
                        if semesterfees3.bookfees == year_book3:
                            pass
                        else:
                            semesterfees3.bookfees = year_book3
                        if semesterfees3.resittingfees == year_resitting3:
                            pass
                        else:
                            semesterfees3.resittingfees = year_resitting3
                        if semesterfees3.entrancefees == year_entrance3:
                            pass
                        else:
                            semesterfees3.entrancefees = year_entrance3
                        if semesterfees3.extrafees == year_extra3:
                            pass
                        else:
                            semesterfees3.extrafees = year_extra3
                        if semesterfees3.discount == year_discount3:
                            pass
                        else:
                            semesterfees3.discount = year_discount3
                        if semesterfees3.totalfees == year_total3:
                            pass
                        else:
                            tempfees = semesterfees3.totalfees
                            semesterfees3.totalfees = year_total3
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total3) - int(mainreciept.paidamount)
                                diff = int(year_total3) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        semesterfees3.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution3,examinationfees=year_exam3,bookfees=year_book3,resittingfees=year_resitting3,entrancefees=year_entrance3,extrafees=year_extra3,discount=year_discount3,totalfees=year_total3,sem="3")
                        addsemesterfees.save()
                        print("added year 3 year 3")
                    
                    return JsonResponse({'added':'yes'})
                    
                elif no_of_year == '4':
                    try:
                        semesterfees1 = StudentFees.objects.get(Q(student = student_id) & Q(sem="1"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "1")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution1, year_exam1, year_book1, year_resitting1, year_entrance1, year_extra1, year_discount1, year_total1)
                        if semesterfees1.tutionfees == year_tution1:
                            pass
                        else:
                            semesterfees1.tutionfees = year_tution1
                        if semesterfees1.examinationfees == year_exam1:
                            pass
                        else:
                            semesterfees1.examinationfees = year_exam1
                        if semesterfees1.bookfees == year_book1:
                            pass
                        else:
                            semesterfees1.bookfees = year_book1
                        if semesterfees1.resittingfees == year_resitting1:
                            pass
                        else:
                            semesterfees1.resittingfees = year_resitting1
                        if semesterfees1.entrancefees == year_entrance1:
                            pass
                        else:
                            semesterfees1.entrancefees = year_entrance1
                        if semesterfees1.extrafees == year_extra1:
                            pass
                        else:
                            semesterfees1.extrafees = year_extra1
                        if semesterfees1.discount == year_discount1:
                            pass
                        else:
                            semesterfees1.discount = year_discount1
                        if semesterfees1.totalfees == year_total1:
                            pass
                        else:
                            tempfees = semesterfees1.totalfees
                            semesterfees1.totalfees = year_total1
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total1) - int(mainreciept.paidamount)
                                diff = int(year_total1) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        
                        semesterfees1.save()
                        
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution1,examinationfees=year_exam1,bookfees=year_book1,resittingfees=year_resitting1,entrancefees=year_entrance1,extrafees=year_extra1,discount=year_discount1,totalfees=year_total1,sem="1")
                        addsemesterfees.save()
                        print("added year 4 year 1")
                    try:
                        semesterfees2 = StudentFees.objects.get(Q(student = student_id) & Q(sem="2"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "2")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution2, year_exam2, year_book2, year_resitting2, year_entrance2, year_extra2, year_discount2, year_total2)
                        if semesterfees2.tutionfees == year_tution2:
                            pass
                        else:
                            semesterfees2.tutionfees = year_tution2
                        if semesterfees2.examinationfees == year_exam2:
                            pass
                        else:
                            semesterfees2.examinationfees = year_exam2
                        if semesterfees2.bookfees == year_book2:
                            pass
                        else:
                            semesterfees2.bookfees = year_book2
                        if semesterfees2.resittingfees == year_resitting2:
                            pass
                        else:
                            semesterfees2.resittingfees = year_resitting2
                        if semesterfees2.entrancefees == year_entrance2:
                            pass
                        else:
                            semesterfees2.entrancefees = year_entrance2
                        if semesterfees2.extrafees == year_extra2:
                            pass
                        else:
                            semesterfees2.extrafees = year_extra2
                        if semesterfees2.discount == year_discount2:
                            pass
                        else:
                            semesterfees2.discount = year_discount2
                        if semesterfees2.totalfees == year_total2:
                            pass
                        else:
                            tempfees = semesterfees2.totalfees
                            semesterfees2.totalfees = year_total2
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total2) - int(mainreciept.paidamount)
                                diff = int(year_total2) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        semesterfees2.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution2,examinationfees=year_exam2,bookfees=year_book2,resittingfees=year_resitting2,entrancefees=year_entrance2,extrafees=year_extra2,discount=year_discount2,totalfees=year_total2,sem="2")
                        addsemesterfees.save()
                        print("added year 4 year 2")
                    try:
                        semesterfees3 = StudentFees.objects.get(Q(student = student_id) & Q(sem="3"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "3")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution3, year_exam3, year_book3, year_resitting3, year_entrance3, year_extra3, year_discount3, year_total3)
                        if semesterfees3.tutionfees == year_tution3:
                            pass
                        else:
                            semesterfees3.tutionfees = year_tution3
                        if semesterfees3.examinationfees == year_exam3:
                            pass
                        else:
                            semesterfees3.examinationfees = year_exam3
                        if semesterfees3.bookfees == year_book3:
                            pass
                        else:
                            semesterfees3.bookfees = year_book3
                        if semesterfees3.resittingfees == year_resitting3:
                            pass
                        else:
                            semesterfees3.resittingfees = year_resitting3
                        if semesterfees3.entrancefees == year_entrance3:
                            pass
                        else:
                            semesterfees3.entrancefees = year_entrance3
                        if semesterfees3.extrafees == year_extra3:
                            pass
                        else:
                            semesterfees3.extrafees = year_extra3
                        if semesterfees3.discount == year_discount3:
                            pass
                        else:
                            semesterfees3.discount = year_discount3
                        if semesterfees3.totalfees == year_total3:
                            pass
                        else:
                            tempfees = semesterfees3.totalfees
                            semesterfees3.totalfees = year_total3
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total3) - int(mainreciept.paidamount)
                                diff = int(year_total3) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        semesterfees3.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution3,examinationfees=year_exam3,bookfees=year_book3,resittingfees=year_resitting3,entrancefees=year_entrance3,extrafees=year_extra3,discount=year_discount3,totalfees=year_total3,sem="3")
                        addsemesterfees.save()
                        print("added year 4 year 3")
                    try:
                        semesterfees4 = StudentFees.objects.get(Q(student = student_id) & Q(sem="4"))
                        allpaymentreciept = PaymentReciept.objects.filter(Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                        mainreciept = ""
                        if allpaymentreciept:
                            mainreciept = allpaymentreciept[0]
                            print("First Main : ",mainreciept)
                            remainingpaymentreciept = PaymentReciept.objects.filter(~Q(id = mainreciept.id) & Q(student = student_id) & Q(payment_for = "Course Fees") & Q(semyear = "4")).order_by('id')
                            print("Remaining All : ",remainingpaymentreciept)
                        print("data year:", year_tution4, year_exam4, year_book4, year_resitting4, year_entrance4, year_extra4, year_discount4, year_total4)
                        if semesterfees4.tutionfees == year_tution4:
                            pass
                        else:
                            semesterfees4.tutionfees = year_tution4
                        if semesterfees4.examinationfees == year_exam4:
                            pass
                        else:
                            semesterfees4.examinationfees = year_exam4
                        if semesterfees4.bookfees == year_book4:
                            pass
                        else:
                            semesterfees4.bookfees = year_book4
                        if semesterfees4.resittingfees == year_resitting4:
                            pass
                        else:
                            semesterfees4.resittingfees = year_resitting4
                        if semesterfees4.entrancefees == year_entrance4:
                            pass
                        else:
                            semesterfees4.entrancefees = year_entrance4
                        if semesterfees4.extrafees == year_extra4:
                            pass
                        else:
                            semesterfees4.extrafees = year_extra4
                        if semesterfees4.discount == year_discount4:
                            pass
                        else:
                            semesterfees4.discount = year_discount4
                        if semesterfees4.totalfees == year_total4:
                            pass
                        else:
                            tempfees = semesterfees4.totalfees
                            semesterfees4.totalfees = year_total4
                            if mainreciept:
                                mainreciept.pendingamount = int(year_total4) - int(mainreciept.paidamount)
                                diff = int(year_total4) - int(tempfees)
                                print(diff)
                                print((int(mainreciept.paidamount) + int(mainreciept.pendingamount)) , int(mainreciept.paidamount) , int(mainreciept.pendingamount))
                                mainreciept.save()
                                for i in remainingpaymentreciept:
                                    i.pendingamount = int(i.pendingamount) + int(diff)
                                    i.save()
                            else:
                                print("not working")  
                            
                        semesterfees4.save()
                    except StudentFees.DoesNotExist:
                        addsemesterfees = StudentFees(student=student_id,studypattern="Annual",stream=stream_year, tutionfees=year_tution4,examinationfees=year_exam4,bookfees=year_book4,resittingfees=year_resitting4,entrancefees=year_entrance4,extrafees=year_extra4,discount=year_discount4,totalfees=year_total4,sem="4")
                        addsemesterfees.save()
                        print("added year 4 year 4")
                    
                    return JsonResponse({'added':'yes'})




            getoldstudentfees_enroll_id = request.POST.get('getoldstudentfees_enroll_id')
            if getoldstudentfees_enroll_id:
                try:
                    getstudent = Student.objects.get(enrollment_id = getoldstudentfees_enroll_id)
                    getenroll = Enrolled.objects.get(student=getstudent.id)
                    stream = Stream.objects.get(id=getenroll.stream.id)
                    if getstudent:
                        try:
                            getoldstudentfees = StudentFees.objects.filter(student=getstudent.id)
                            getoldstudentfeescount = StudentFees.objects.filter(student=getstudent.id).count() 
                            getstudentfeesserializer = StudentFeesSerializer(getoldstudentfees,many=True)
                            return JsonResponse({'data':getstudentfeesserializer.data,'count':getoldstudentfeescount,'studypattern':getenroll.course_pattern,'sem':stream.sem,'student':getstudent.id})
                        except StudentFees.DoesNotExist:
                            print("no student fees found")
                except Student.DoesNotExist:
                    print("no student Found")
            name = request.POST.get('name')
            if name:
                getstudent = Student.objects.filter(Q(name = name) | Q(name__icontains = name)  | Q(email__icontains = name))
                for i in getstudent:
                    enrolled = Enrolled.objects.get(student= i.id)
                    course = Course.objects.get(id=enrolled.course.id)
                    stream = Stream.objects.get(id = enrolled.stream.id)
                    ##added by Avani 14/08
                    try:
                        substream = SubStream.objects.get(id = enrolled.substream.id)
                        substreamname = substream.name
                    except:
                        substreamname = '-'
                    #end of add
                    paymentreciept = PaymentReciept.objects.filter(student=i.id)
                    payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                    obj = {
                        "name":i.name,
                        "email":i.email,
                        "mobile":i.mobile,
                        "enrollment_id":i.enrollment_id,
                        "course_pattern":enrolled.course_pattern,
                        "total_semyear":enrolled.total_semyear,
                        "current_semyear":enrolled.current_semyear,
                        "course":course.name,
                        "stream":stream.name,
                        "substream":substreamname,##added by Avani 14/08
                        "paymentreciept":payment_reciept_serializer.data
                    }
                    temp.append(obj)
                return JsonResponse({'student_data':temp})

            course_id = request.POST.get('course_id')
            stream_id = request.POST.get('stream_id')
            substream_id=request.POST.get('substream_id')##added by Avani 14/08
            if course_id and stream_id:
                print("inside course and stream")
                print("course_id", course_id, stream_id)
                ##added by Avani 14/08
                try:
                    if substream_id != None and substream_id != '':
                        print("ddfdd")
                        enroll = Enrolled.objects.filter(Q(course = course_id) & Q(stream = stream_id) & Q(substream = substream_id))
                    else:    
                        enroll = Enrolled.objects.filter(Q(course = course_id) & Q(stream = stream_id) & Q(substream__isnull=True))
                except:
                    enroll =''
                ## end of add
                print("enroll", enroll)
                for i in enroll:
                    print(i)
                    student = Student.objects.get(id=i.student_id)
                    enrolled = Enrolled.objects.get(student=student.id)
                    course = Course.objects.get(id=enrolled.course_id)
                    stream = Stream.objects.get(id = enrolled.stream_id)
                    ##added by Avani 14/08
                    if substream_id != None and substream_id != '':
                        substream = SubStream.objects.get(id =enrolled.substream_id)
                        substreamname = substream.name
                    else:
                        substreamname = '-'
                    ##end of add
                    paymentreciept = PaymentReciept.objects.filter(student=student.id)
                    payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                    obj = {
                        "name":student.name,
                        "email":student.email,
                        "mobile":student.mobile,
                        "enrollment_id":student.enrollment_id,
                        "course_pattern":enrolled.course_pattern,
                        "total_semyear":enrolled.total_semyear,
                        "current_semyear":enrolled.current_semyear,
                        "course":course.name,
                        "stream":stream.name,
                        "substream": substreamname,##added by Avani 14/08
                        "paymentreciept":payment_reciept_serializer.data
                    }
                    temp.append(obj)
                return JsonResponse({'student_data':temp})
            ##added by Avani 14/08
            streamid = request.POST.get('streamid')
            if streamid:
                getSubStream = SubStream.objects.filter(stream_id=streamid)
                substreamserializer = SubStreamSerializer(getSubStream,many=True)
                return JsonResponse({'substream':substreamserializer.data})
            ## end of add
            courseid = request.POST.get('courseid')
            if courseid:
                getcourse = Course.objects.get(id=courseid)
                getstream = Stream.objects.filter(course=getcourse)
                streamserializer = StreamSerializer(getstream,many=True)
                return JsonResponse({'stream':streamserializer.data})
            enroll_id = request.POST.get('enroll_id')
            if enroll_id:
                try:
                    checkstudent = Student.objects.get(Q(enrollment_id=enroll_id) | Q(mobile=enroll_id) | Q(email=enroll_id))
                    enrolled = Enrolled.objects.get(student= checkstudent.id)
                    course = Course.objects.get(id=enrolled.course.id)
                    stream = Stream.objects.get(id = enrolled.stream.id)
                    ##added by Avani 14/08
                    try:
                        substream = SubStream.objects.get(id =enrolled.substream.id)
                        substreamname = substream.name
                    except SubStream.DoesNotExist:
                        substreamname = '-'
                    ## end of add
                    paymentreciept = PaymentReciept.objects.filter(student=checkstudent.id)
                    payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
                    obj = {
                        "name":checkstudent.name,
                        "email":checkstudent.email,
                        "mobile":checkstudent.mobile,
                        "enrollment_id":checkstudent.enrollment_id,
                        "course_pattern":enrolled.course_pattern,
                        "total_semyear":enrolled.total_semyear,
                        "current_semyear":enrolled.current_semyear,
                        "course":course.name,
                        "stream":stream.name,
                        "substream":substreamname,##added by Avani 14/08
                        "paymentreciept":payment_reciept_serializer.data
                    }
                    temp.append(obj)
                except Student.DoesNotExist:
                    pass
                return JsonResponse({'student_data':temp})

    params = {
        "course":Course.objects.all()
    }
    return render(request,"super_admin/editoldstudentfees.html",params)

@login_required(login_url='/login/')
def saveEditStudent(request):
    print("inside this function editstudent")
    if request.user.is_superuser or request.user.is_data_entry:
        
        if request.method == "POST":
            print("in here")
            university = request.POST.get('university')
            enroll_id = request.POST.get('student_enrollment_id')
            print("enroll_id: ", enroll_id)
            student_image = request.FILES.get('student_image')
            print("student_image:",student_image)
            name = request.POST.get('name')
            dob = request.POST.get('dateofbirth')
            print("dateofbirth", dob)
            fathername = request.POST.get('fathers_name')
            mothername = request.POST.get('mothers_name')
            email = request.POST.get('email')
            alternateemail = request.POST.get('alternateemail')
            mobile = request.POST.get('mobile')
            alternatemobile1 = request.POST.get('alternatemobile1')
            gender = request.POST.get('gender')
            category = request.POST.get('category')
            address = request.POST.get('address')
            alternateaddress = request.POST.get('alternateaddress')
            nationality = request.POST.get('nationality')
            country = request.POST.get('country')
            state = request.POST.get('stateId')
            city = request.POST.get('cityId')
            pincode = request.POST.get('Pincode')
            student_remarks = request.POST.get('student_remarks')
            counselor_name = request.POST.get('counselor_name')
            reference_name = request.POST.get('reference_name')
            university_enrollment_number = request.POST.get('university_enrollment_number')
            if country:
                change_country = Countries.objects.get(id=country)
            else:
                change_country = Countries.objects.get(id=101)
            if state:
                change_state = States.objects.get(id=state)
            else:
                change_state = States.objects.get(id=22)
            if city:
                change_city = Cities.objects.get(id=city)
            else:
                change_city = Cities.objects.get(id=2707)
            try:
                getstudent = Student.objects.get(enrollment_id = enroll_id)
                getstudent.name = name
                getstudent.dateofbirth = dob
                getstudent.father_name = fathername
                getstudent.mother_name = mothername
                getstudent.gender = gender
                getstudent.category = category

                getstudent.email = email
                getstudent.alternateemail = alternateemail
                getstudent.mobile = mobile
                getstudent.alternate_mobile1 = alternatemobile1
                    
                getstudent.address = address
                getstudent.alternateaddress = alternateaddress
                getstudent.nationality = nationality
                getstudent.country = change_country
                getstudent.state = change_state
                getstudent.city = change_city
                getstudent.pincode = pincode
                if student_image != None:
                    print("inside if student image")
                    getstudent.image = student_image
                #getstudent.image = student_image
                getstudent.university = University.objects.get(registrationID = university)
                getstudent.student_remarks = student_remarks
                getstudent.save()
                    
            except Student.DoesNotExist:
                print("No student found")

            #Personal Documents
            getstudent = Student.objects.get(enrollment_id = enroll_id)
            id = getstudent.id
            no_of_docs = request.POST.get('personaldocs_counter')
            print("no of docus" , no_of_docs)
            for i in range(1,int(no_of_docs)+1):
                docid= request.POST.get(f'docid{i}')
                print("docid" , docid)
                document = request.POST.get(f'document{i}')
                DocumentName = request.POST.get(f'DocumentName{i}')
                ## added by Avani - if Document type is Other, store the name from the field other{i}
                if document == 'Other':
                    document = request.POST.get(f'other{i}')
                DocumentID = request.POST.get(f'DocumentID{i}')
                DocumentFront = request.FILES.get(f'DocumentFront{i}')
                DocumentBack = request.FILES.get(f'DocumentBack{i}')
                print("DocumentFront",  request.FILES.get(f'DocumentFront{i}'))
                # print(document,DocumentName,DocumentID,DocumentFront,DocumentBack)
                if docid != None:
                    getrecord = StudentDocuments.objects.get(id = docid)
                    getrecord.document = document
                    getrecord.document_name = DocumentName
                    getrecord.document_ID_no = DocumentID
                    if DocumentFront != None:
                            getrecord.document_image_front = DocumentFront
                    if DocumentBack != None:
                            getrecord.document_image_back = DocumentBack
                    getrecord.save()
                else:
                    add_student_document = StudentDocuments(document=document,document_name=DocumentName,document_ID_no=DocumentID,document_image_front= DocumentFront,document_image_back = DocumentBack,student_id=id)
                    add_student_document.save()

           

            # Qualifications
            secondary_year = request.POST.get('secondary_year')
            secondary_board = request.POST.get('secondary_board')
            secondary_percentage = request.POST.get('secondary_percentage')
            secondary_document = request.FILES.get('secondary_document')## modified by Avani, removed the False part 15/07
            
            sr_year = request.POST.get('sr_year')
            sr_board = request.POST.get('sr_board')
            sr_percentage = request.POST.get('sr_percentage')
            sr_document = request.FILES.get('sr_document')## modified by Avani, removed the False part 15/07
            
            under_year = request.POST.get('under_year')
            under_board = request.POST.get('under_board')
            under_percentage = request.POST.get('under_percentage')
            under_document = request.FILES.get('under_document') ## modified by Avani, removed the False part 15/07
            
            post_year = request.POST.get('post_year')
            post_board = request.POST.get('post_board')
            post_percentage = request.POST.get('post_percentage')
            post_document = request.FILES.get('post_document')## modified by Avani, removed the False part 15/07
            
            mphil_year = request.POST.get('mphil_year')
            print("mphil_year", mphil_year)
            mphil_board = request.POST.get('mphil_board')
            mphil_percentage = request.POST.get('mphil_percentage')
            mphil_document = request.FILES.get('mphil_document')## modified by Avani, removed the False part 15/07
            
            # others_year = request.POST.get('other_year')
            # others_board = request.POST.get('other_board')
            # others_percentage = request.POST.get('other_percentage')
            # others_document = request.FILES.get('other_document', False)
            others_counter = request.POST.get('qual_other_counter')
            print(others_counter, "others_counter")
            files_data = []
            for i in range(1,int(others_counter)+1):
                print("inside for")
                doctype = request.POST.get(f'other_{i}')
                print("doctype", doctype)
                year = request.POST.get(f'other_year_{i}')
                board = request.POST.get(f'other_board_{i}')
                percentage = request.POST.get(f'other_percentage_{i}')
                docfile= request.FILES.get(f'other_document_{i}')
                uploadeddoc_other = request.POST.get(f'olddoc_{i}')
               
                print("uploaded-doc-other_", uploadeddoc_other)
                print("docfile", docfile)

                
                if docfile:
                    # Save the file and get its path
                    upload_dir = os.path.join(settings.MEDIA_ROOT, 'University_Documents')
                    file_path = os.path.join(upload_dir, docfile.name)
                    save_path = f'University_Documents/{docfile.name}'
                    with open(file_path, 'wb+') as destination:
                        for chunk in docfile.chunks():
                            destination.write(chunk)
                        docfile = save_path
                else:
                    docfile = uploadeddoc_other
                file_data = {
                    'doctype': doctype,
                    'year': year,
                    'board': board,
                    'percentage': percentage,
                    'file_path': docfile
                }
                files_data.append(file_data)
                #files_data[f'other_{i}'] = file_data
                print("files_data", files_data)

            if secondary_year or secondary_board or secondary_percentage or secondary_document or sr_year or sr_board or sr_percentage or sr_document or under_year or under_board or under_percentage or under_document or post_year or post_board or post_percentage or post_document or mphil_year or mphil_board or mphil_percentage or mphil_document or files_data:
                try:

                    getqualification = Qualification.objects.get(student = getstudent.id)
                    print("sr_document", sr_document)
                    getqualification.secondary_year = secondary_year
                    getqualification.sr_year = sr_year
                    getqualification.under_year = under_year
                    getqualification.post_year = post_year
                    getqualification.mphil_year = mphil_year
                        
                    getqualification.secondary_board = secondary_board
                    getqualification.sr_board = sr_board
                    getqualification.under_board = under_board
                    getqualification.post_board = post_board
                    getqualification.mphil_board = mphil_board
                        
                    getqualification.secondary_percentage = secondary_percentage
                    getqualification.sr_percentage = sr_percentage
                    getqualification.under_percentage = under_percentage
                    getqualification.post_percentage = post_percentage
                    getqualification.mphil_percentage = mphil_percentage
                        
                    if secondary_document != None:
                            getqualification.secondary_document = secondary_document
                    if sr_document != None:
                            getqualification.sr_document = sr_document
                    if under_document != None:
                            getqualification.under_document = under_document
                    if post_document != None:
                            getqualification.post_document = post_document
                    if mphil_document != None:
                            getqualification.mphil_document = mphil_document
                   
                    getqualification.others = files_data
                    getqualification.save()

                except Qualification.DoesNotExist:
                    print("no record for qualification" , sr_document)
                    create_student_qualification = Qualification(
                            student_id = getstudent.id,
                            secondary_year = secondary_year,
                            sr_year = sr_year,
                            under_year = under_year,
                            post_year = post_year,
                            mphil_year = mphil_year,
                            others_year = others_year,
                        
                            secondary_board = secondary_board,
                            sr_board = sr_board,
                            under_board = under_board,
                            post_board = post_board,
                            mphil_board = mphil_board,
                            others_board = others_board,
                            
                            secondary_percentage = secondary_percentage,
                            sr_percentage = sr_percentage,
                            under_percentage = under_percentage,
                            post_percentage = post_percentage,
                            mphil_percentage = mphil_percentage,
                            others_percentage = others_percentage,
                            
                            secondary_document = secondary_document,
                            sr_document = sr_document,
                            under_document = under_document,
                            post_document = post_document,
                            mphil_document = mphil_document,
                            others_document = others_document,
                        )
                    create_student_qualification.save()
                if counselor_name or reference_name or university_enrollment_number:
                    try:
                        getadditionaldetails = AdditionalEnrollmentDetails.objects.get(student=getstudent.id)
                        getadditionaldetails.counselor_name=counselor_name
                        getadditionaldetails.reference_name=reference_name
                        getadditionaldetails.university_enrollment_id=university_enrollment_number
                        getadditionaldetails.save()
                    
                    except AdditionalEnrollmentDetails.DoesNotExist:
                        pass
                ##save any changes made  to Payment details (added by Avani on 15/7)
                try:
                    getPaymentDetails = PaymentReciept.objects.get(student = getstudent.id)
                    getPaymentDetails.fee_reciept_type =  request.POST.get('fees_reciept')
                    getPaymentDetails.transaction_date =  request.POST.get('transaction_date')
                    getPaymentDetails.cheque_no =  request.POST.get('cheque_no')
                    getPaymentDetails.bank_name =  request.POST.get('bank_name')
                    getPaymentDetails.paymentmode =  request.POST.get('payment_mode')
                    getPaymentDetails.remarks =  request.POST.get('remarks')
                    getPaymentDetails.semyearfees =  request.POST.get('total_fees')
                    getPaymentDetails.paidamount =  request.POST.get('fees')
                    getPaymentDetails.pendingamount =  request.POST.get('dues')
                    getPaymentDetails.save()
                except PaymentReciept.DoesNotExist:
                    pass
    enroll_id = request.POST.get('student_enrollment_id')
    type = False
    try:
        getstudent = Student.objects.get(enrollment_id = enroll_id)  
        type = getstudent.is_quick_register
    except:
        pass          
   
    if type == True:
                params= {
                        "students":Student.objects.filter(~Q(is_cancelled = True) & Q(is_quick_register=True)).order_by('-id'),
                        "title" : 'List of Quick Registered Students',
                        "msg": 'Student details updated successfully'
                        }
    else:
                params= {
                        "students":Student.objects.filter(~Q(is_cancelled = True) & Q(is_quick_register=False)).order_by('-id'),
                        "title" : 'List of  Registered Students',
                        "msg": 'Student details updated successfully'
                        
                        }
            
    return render(request,"super_admin/viewstudents.html", params)

@login_required(login_url='/login/')
def EditStudent(request,enroll_id):
    print("Jammy is hitted............................")
    print("Jammy is hitted............................")
    display = ""
    level_of_user = ""
    found_student = ""
    student = []
    selected_course = request.GET.get('selected_course')
    print(selected_course,'course')
    print("Jammy is hitted............................")
    if selected_course:
        get_course_id = Course.objects.get(id = selected_course)
        get_specialization = Stream.objects.filter(course = get_course_id.id)
        serializer = StreamSerializer(get_specialization,many=True)
        return JsonResponse({'data':serializer.data})
    if request.user.is_superuser or request.user.is_data_entry:
        
        if request.method == "POST":
            student_image = request.FILES.get('student_image', False)
            print("student_image:",student_image)
            name = request.POST.get('name')
            dob = request.POST.get('dateofbirth')
            fathername = request.POST.get('fathers_name')
            mothername = request.POST.get('mothers_name')
            email = request.POST.get('email')
            alternateemail = request.POST.get('alternateemail')
            mobile = request.POST.get('mobile')
            alternatemobile1 = request.POST.get('alternatemobile1')
            gender = request.POST.get('gender')
            category = request.POST.get('category')
            address = request.POST.get('address')
            alternateaddress = request.POST.get('alternateaddress')
            nationality = request.POST.get('nationality')
            country = request.POST.get('country')
            state = request.POST.get('stateId')
            city = request.POST.get('cityId')
            pincode = request.POST.get('Pincode')
            student_remarks = request.POST.get('student_remarks')
            counselor_name = request.POST.get('counselor_name')
            reference_name = request.POST.get('reference_name')
            university_enrollment_number = request.POST.get('university_enrollment_number')
            
            course = request.POST.get('course')
            streamID = request.POST.get('Stream')
            studypattern = request.POST.get('studypattern')
            semyear = request.POST.get('semyear')
            university = request.POST.get('university')
            print("univ", university)
            session = request.POST.get('session')
            entry_mode = request.POST.get('entry_mode')


            
            totaldocuments = request.POST.get('totaldocuments')
            document1 = request.POST.get('document1')
            DocumentName1 = request.POST.get('DocumentName1')
            DocumentID1 = request.POST.get('DocumentID1')
            DocumentUpload1 = request.FILES.getlist('DocumentUpload1')

            document2 = request.POST.get('document2')
            DocumentName2 = request.POST.get('DocumentName2')
            DocumentID2 = request.POST.get('DocumentID2')
            DocumentUpload2 = request.FILES.getlist('DocumentUpload2')

            document3 = request.POST.get('document3')
            DocumentName3 = request.POST.get('DocumentName3')
            DocumentID3 = request.POST.get('DocumentID3')
            DocumentUpload3 = request.FILES.getlist('DocumentUpload3')

            document4 = request.POST.get('document4')
            DocumentName4 = request.POST.get('DocumentName4')
            DocumentID4 = request.POST.get('DocumentID4')
            DocumentUpload4 = request.FILES.getlist('DocumentUpload4')
            
            # print("totaldocuments :",totaldocuments)
            # print("document1 :",document1)
            # print("DocumentName1 :",DocumentName1)
            # print("DocumentID1 :",DocumentID1)
            # print("___________________")
            # for i in DocumentUpload1:
            #     print(i)
            #     print(type(i))

            # print("document2 :",document2)
            # print("DocumentName2 :",DocumentName2)
            # print("DocumentID2 :",DocumentID2)
            # for i in DocumentUpload2:
            #     print(i)
            #     print(type(i))

            # print("document3 :",document3)
            # print("DocumentName3 :",DocumentName3)
            # print("DocumentID3 :",DocumentID3)
            # for i in DocumentUpload3:
            #     print(i)
            #     print(type(i))

            # print("document4 :",document4)
            # print("DocumentName4 :",DocumentName4)
            # print("DocumentID4 :",DocumentID4)
            # for i in DocumentUpload4:
            #     print(i)
            #     print(type(i))
            personaldocs_counter = request.POST.get('personaldocs')
            print("in post" , personaldocs_counter)
            
            # Qualifications
            secondary_year = request.POST.get('secondary_year')
            secondary_board = request.POST.get('secondary_board')
            secondary_percentage = request.POST.get('secondary_percentage')
            secondary_document = request.FILES.get('secondary_document')## modified by Avani, removed the False part 15/07
            
            sr_year = request.POST.get('sr_year')
            sr_board = request.POST.get('sr_board')
            sr_percentage = request.POST.get('sr_percentage')
            sr_document = request.FILES.get('sr_document')## modified by Avani, removed the False part 15/07
            
            under_year = request.POST.get('under_year')
            under_board = request.POST.get('under_board')
            under_percentage = request.POST.get('under_percentage')
            under_document = request.FILES.get('under_document') ## modified by Avani, removed the False part 15/07
            
            post_year = request.POST.get('post_year')
            post_board = request.POST.get('post_board')
            post_percentage = request.POST.get('post_percentage')
            post_document = request.FILES.get('post_document')## modified by Avani, removed the False part 15/07
            
            mphil_year = request.POST.get('mphil_year')
            mphil_board = request.POST.get('mphil_board')
            mphil_percentage = request.POST.get('mphil_percentage')
            mphil_document = request.FILES.get('mphil_document')## modified by Avani, removed the False part 15/07
            
            others_year = request.POST.get('other_year')
            others_board = request.POST.get('other_board')
            others_percentage = request.POST.get('other_percentage')
            others_document = request.FILES.get('other_document', False)
            print("country",country , "state",state , "city",city , "pincode",pincode)
            # print(firstname , middlename , lastname , dob , fathername , mothername , email , mobile , gender , category , idtype , idnumber , address , nationality , country , state , city , pincode)
            # print(secondary_year , secondary_board , secondary_percentage , secondary_document , sr_year , sr_board , sr_percentage , sr_document , under_year , under_board , under_percentage , under_document , post_year , post_board , post_percentage , post_document , mphil_year , mphil_board , mphil_percentage , mphil_document , others_year , others_board , others_percentage , others_document)
            # print(course , streamID , studypattern , semyear , session , entry_mode , university)
            # print(counselor_name , reference_name , university_enrollment_number)
            if name or dob or fathername or mothername or email or mobile or gender or category  or address or nationality or country or state or city or pincode:
                try:
                    getstudent = Student.objects.get(enrollment_id = enroll_id)
                    
                    if getstudent.email == email  or email == "None":
                        print("same email")
                    else:
                        try:
                            match = Student.objects.get(Q(email = email) | Q(alternateemail = email))
                            return JsonResponse({'match':'email','enroll_id':enroll_id})
                        except Student.DoesNotExist:
                            pass
                    
                    if getstudent.alternateemail == alternateemail or alternateaddress == "None":
                        print("same alternateemail")
                    else:
                        try:
                            match = Student.objects.get(Q(email = alternateemail) | Q(alternateemail = alternateemail))
                            return JsonResponse({'match':'alternateemail','enroll_id':enroll_id})
                        except Student.DoesNotExist:
                            pass
                    
                    if getstudent.mobile == mobile or mobile == "None":
                        print("same mobile")
                    else:
                        try:
                            match = Student.objects.get(Q(mobile = mobile) | Q(alternate_mobile1 = mobile))
                            return JsonResponse({'match':'mobile','enroll_id':enroll_id})
                        except Student.DoesNotExist:
                            pass
                    
                    if getstudent.alternate_mobile1 == alternatemobile1 or alternatemobile1 == "None":
                        print("same alternatemobile")
                    else:
                        try:
                            match = Student.objects.get(Q(mobile = alternatemobile1) | Q(alternate_mobile1 = alternatemobile1))
                            return JsonResponse({'match':'alternatemobile','enroll_id':enroll_id})
                        except Student.DoesNotExist:
                            pass
                    
                    

                    print("student found")
                    if country:
                        change_country = Countries.objects.get(id=country)
                    else:
                        change_country = Countries.objects.get(id=101)
                    if state:
                        change_state = States.objects.get(id=state)
                    else:
                        change_state = States.objects.get(id=22)
                    if city:
                        change_city = Cities.objects.get(id=city)
                    else:
                        change_city = Cities.objects.get(id=2707)
                    
                    getstudent.name = name
                    getstudent.dateofbirth = dob
                    getstudent.father_name = fathername
                    getstudent.mother_name = mothername
                    getstudent.gender = gender
                    getstudent.category = category

                    getstudent.email = email
                    getstudent.alternateemail = alternateemail
                    getstudent.mobile = mobile
                    getstudent.alternate_mobile1 = alternatemobile1
                    
                    getstudent.address = address
                    getstudent.alternateaddress = alternateaddress
                    getstudent.nationality = nationality
                    getstudent.country = change_country
                    getstudent.state = change_state
                    getstudent.city = change_city
                    getstudent.pincode = pincode
                    getstudent.image = student_image
                    getstudent.university = University.objects.get(registrationID = university)
                    getstudent.student_remarks = student_remarks
                    getstudent.save()
                    
                    
                except Student.DoesNotExist:
                    print("student not found ..!!")
            if secondary_year or secondary_board or secondary_percentage or secondary_document or sr_year or sr_board or sr_percentage or sr_document or under_year or under_board or under_percentage or under_document or post_year or post_board or post_percentage or post_document or mphil_year or mphil_board or mphil_percentage or mphil_document or others_year or others_board or others_percentage or others_document:
                try:
                    getstudent = Student.objects.get(enrollment_id = enroll_id)
                    ##modified the code, if no record in qualifications is found, create a new record. Avani 15/07
                    try:
                        getqualification = Qualification.objects.get(student = getstudent.id)
                        print("sr_document", sr_document)
                        getqualification.secondary_year = secondary_year
                        getqualification.sr_year = sr_year
                        getqualification.under_year = under_year
                        getqualification.post_year = post_year
                        getqualification.mphil_year = mphil_year
                        getqualification.others_year = others_year
                        
                        getqualification.secondary_board = secondary_board
                        getqualification.sr_board = sr_board
                        getqualification.under_board = under_board
                        getqualification.post_board = post_board
                        getqualification.mphil_board = mphil_board
                        getqualification.others_board = others_board
                        
                        getqualification.secondary_percentage = secondary_percentage
                        getqualification.sr_percentage = sr_percentage
                        getqualification.under_percentage = under_percentage
                        getqualification.post_percentage = post_percentage
                        getqualification.mphil_percentage = mphil_percentage
                        getqualification.others_percentage = others_percentage
                        
                        if secondary_document != None:
                            getqualification.secondary_document = secondary_document
                        if sr_document != None:
                            getqualification.sr_document = sr_document
                        if under_document != None:
                            getqualification.under_document = under_document
                        if post_document != None:
                            getqualification.post_document = post_document
                        if mphil_document != None:
                            getqualification.mphil_document = mphil_document
                        if others_document != None:
                            getqualification.others_document = others_document
                        getqualification.save()

                    except Qualification.DoesNotExist:
                        print("no record for qualification" , sr_document)
                        create_student_qualification = Qualification(
                            student_id = getstudent.id,
                            secondary_year = secondary_year,
                            sr_year = sr_year,
                            under_year = under_year,
                            post_year = post_year,
                            mphil_year = mphil_year,
                            others_year = others_year,
                        
                            secondary_board = secondary_board,
                            sr_board = sr_board,
                            under_board = under_board,
                            post_board = post_board,
                            mphil_board = mphil_board,
                            others_board = others_board,
                            
                            secondary_percentage = secondary_percentage,
                            sr_percentage = sr_percentage,
                            under_percentage = under_percentage,
                            post_percentage = post_percentage,
                            mphil_percentage = mphil_percentage,
                            others_percentage = others_percentage,
                            
                            secondary_document = secondary_document,
                            sr_document = sr_document,
                            under_document = under_document,
                            post_document = post_document,
                            mphil_document = mphil_document,
                            others_document = others_document,
                        )
                        create_student_qualification.save()

                except Student.DoesNotExist:
                    pass

            if counselor_name or reference_name or university_enrollment_number:
                try:
                    getstudent = Student.objects.get(enrollment_id = enroll_id)
                    getadditionaldetails = AdditionalEnrollmentDetails.objects.get(student=getstudent.id)
                    getadditionaldetails.counselor_name=counselor_name
                    getadditionaldetails.reference_name=reference_name
                    getadditionaldetails.university_enrollment_id=university_enrollment_number
                    getadditionaldetails.save()
                    
                except Student.DoesNotExist:
                    pass
            ##save any changes made  to Payment details (added by Avani on 15/7)
            getPaymentDetails = PaymentReciept.objects.get(student = getstudent.id)
            getPaymentDetails.fee_reciept_type =  request.POST.get('fees_reciept')
            getPaymentDetails.transaction_date =  request.POST.get('transaction_date')
            getPaymentDetails.cheque_no =  request.POST.get('cheque_no')
            getPaymentDetails.bank_name =  request.POST.get('bank_name')
            getPaymentDetails.paymentmode =  request.POST.get('payment_mode')
            getPaymentDetails.remarks =  request.POST.get('remarks')
            getPaymentDetails.semyearfees =  request.POST.get('total_fees')
            getPaymentDetails.paidamount =  request.POST.get('fees')
            getPaymentDetails.pendingamount =  request.POST.get('dues')
            getPaymentDetails.save()


            return JsonResponse({'saved':'yes','enroll_id':enroll_id})

        try:
            getstudent = Student.objects.get(enrollment_id = enroll_id)
            print("student found")
            found_student = "yes"
            try:
                getenroll = Enrolled.objects.get(student=getstudent.id)
                print(getenroll)
                course = Course.objects.get(id = getenroll.course.id)
                stream = Stream.objects.get(id = getenroll.stream.id)
                ##added by Avani 14/08
                try:
                    substream= SubStream.objects.get(id = getenroll.substream.id)
                except:
                    substream = ''
                ## end of add
                print(stream.id)
                # print(course)
                # print(stream)
                getdocuments = StudentDocuments.objects.filter(student=getstudent.id)
                # print(getdocuments)

                print(course,stream,substream,substream.id )
                try:
                    getqualification = Qualification.objects.get(student=getstudent.id)
                    # print(getqualification)
                except Qualification.DoesNotExist:
                    getqualification = "none"
                try:
                    getuniversity = University.objects.get(id=getstudent.university.id)
                    # print(getuniversity)
                except University.DoesNotExist:
                    getuniversity = "none"

                try:
                    getadditionalenrollmentdetails = AdditionalEnrollmentDetails.objects.get(student = getstudent.id)
                    # print(getadditionalenrollmentdetails)
                except AdditionalEnrollmentDetails.DoesNotExist:
                    getadditionalenrollmentdetails = "none"
                ## Fetch payment details to show in fronten added by Avani 15/07
                try:
                    paymentdetails = PaymentReciept.objects.filter(student=getstudent.id).order_by('transactiontime').first()
                except PaymentReciept.DoesNotExist:
                    paymentdetails = "none"
                try:
                    personaldocuments = StudentDocuments.objects.filter(student_id=getstudent.id)
                    print("personaldoc", personaldocuments)
                except StudentDocuments.DoesNotExist:
                    personaldocuments = 'none'
                personaldoccounter = personaldocuments.count()
                print('personaldoccounter', personaldoccounter)
                obj = {
                    "enrollment_id":getstudent.enrollment_id,
                    "image":getstudent.image,
                    "name":getstudent.name,
                    "father_name":getstudent.father_name,
                    "mother_name":getstudent.mother_name,
                    "dateofbirth":str(getstudent.dateofbirth),
                    "mobile":getstudent.mobile,
                    "alternate_mobile1":getstudent.alternate_mobile1,
                    "email":getstudent.email,
                    "alternateemail":getstudent.alternateemail,
                    "gender":getstudent.gender,
                    "category":getstudent.category,
                    "address":getstudent.address,
                    "alternateaddress":getstudent.alternateaddress,
                    "nationality":getstudent.nationality,
                    "country":getstudent.country,
                    "state":getstudent.state,
                    "city":getstudent.city,
                    "pincode":getstudent.pincode,
                    "university":getstudent.university,
                    "qualification":getqualification,
                    "course":course,
                    "stream":stream,
                    "substream":substream,##added by Avani 14/08
                    "course_pattern":getenroll.course_pattern,
                    "session":getenroll.session,
                    "entry_mode":getenroll.entry_mode,
                    "current_semyear":getenroll.current_semyear,
                    "university":getuniversity,
                    "student_remarks":getstudent.student_remarks,
                    "additionalenrollmentdetails":getadditionalenrollmentdetails,
                    "documents":getdocuments, # show docoument related details in frontend added by Avani 15/07
                    "personaldocs":personaldocuments,
                    
                }
                student.append(obj)
            except Enrolled.DoesNotExist:
                getenroll = "none"

        except Student.DoesNotExist:
            # print("student not found")
            found_student = "no"
        
            
            
            
            




    params = {
        "course":Course.objects.all(),
        "university":University.objects.all(),
        "found_student":found_student,
        "student":student,
        "country":Countries.objects.all(),
        "all_course_year":Course.objects.order_by('year').values_list('year', flat=True).distinct(),
        "paymentdetails":paymentdetails, ## pass paymentdetails object added by Avani 15/07
        'personaldoccounter':personaldoccounter,
        "mobile":getstudent.mobile,
        "alternate_mobile1":getstudent.alternate_mobile1,
        "email":getstudent.email,
        "alternateemail":getstudent.alternateemail,
        "paymentmodes":PaymentModes.objects.filter(status=True), ## added by Avani on 09/08 -pick data from mastertables
        "feereceiptoptions":FeeReceiptOptions.objects.filter(status=True),## added by Avani on 09/08 -pick data from mastertables
        "banknames":BankNames.objects.filter(status=True),## added by Avani on 09/08 -pick data from mastertables
        "sessionnames":SessionNames.objects.filter(status=True),## added by Avani on 12/08 -pick data from mastertables

    }
    return render(request,"super_admin/editstudent.html",params)


def changeStudentCourse(request):
    if request.method == "POST":
        student_enrollment_id = request.POST.get('student_enrollment_id')
        show_modal_course = request.POST.get('show_modal_course')
        show_modal_stream = request.POST.get('show_modal_stream')
        show_modal_study_pattern = request.POST.get('show_modal_study_pattern')
        show_modal_semyear = request.POST.get('show_modal_semyear')
        show_modal_substream = request.POST.get('show_modal_substream')##added by Avani 14/08
        try:
            total_fees_paid = 0
            student = Student.objects.get(enrollment_id=student_enrollment_id)
            get_all_payment_reciepts = PaymentReciept.objects.filter(student=student)
            for reciept in get_all_payment_reciepts:
                if reciept.payment_for == "Course Fees":
                    total_fees_paid += int(reciept.paidamount)
            payment_serializer = PaymentRecieptSerializer(get_all_payment_reciepts,many=True)
            if show_modal_semyear:
                semyear = show_modal_semyear
            else:
                semyear = 0
            pending_amount = 0
            advance_amount = 0
            total_year_fees = 0
            if show_modal_study_pattern == "Full Course":
                print("Fees of full Course")
            else:
                streamyearfees = CalculateFees(total_fees_paid,show_modal_study_pattern,show_modal_stream,show_modal_substream,semyear)##modified by Avani 14/08
                if streamyearfees != '':
                    calc = total_fees_paid - int(streamyearfees.totalfees)
                    total_year_fees = streamyearfees.totalfees
                    if calc >= 0:
                        pending_amount = 0
                        advance_amount = calc
                    else:
                        print("total fees is less than streamyear fees hence pending")
                        pending_amount = int(streamyearfees.totalfees) - total_fees_paid
                        advance_amount = 0
                else:
                    pending_amount = 0
                    advance_amount = calc
                    
            obj = {
                "study_pattern":show_modal_study_pattern,
                "semyear":semyear,
                "semyear_fees":total_year_fees,
                "total_fees_paid": total_fees_paid,
                "reciepts":payment_serializer.data,
                'pending_fees':pending_amount,
                'advance_amount':advance_amount,
            }
            return JsonResponse({'data':obj})
        except Student.DoesNotExist:
            pass
        

    return JsonResponse({'data':'data'})


def CalculateFees(paid_fees,study_pattern,stream,substream, semyear): ##modified by Avani 14/08
    # print(paid_fees,study_pattern,stream,semyear)
    if study_pattern == "Semester":
        ##modified by Avani 14/08
        if substream != None and substream != '':
            try:
                    get_semester_fees = SemesterFees.objects.get(Q(stream=Stream.objects.get(id=stream))  & Q(substream=SubStream.objects.get(id=substream)) & Q(sem=semyear))
            except:
                    get_semester_fees = []
        else:
            try:
                get_semester_fees = SemesterFees.objects.get(Q(stream=Stream.objects.get(id=stream)) & Q(sem=semyear))
            except:
                get_semester_fees = []
        ##end of modification
    elif study_pattern == "Annual":
        ##modified by Avani 14/08
        if substream != None and substream != '':
            try:
                get_semester_fees = YearFees.objects.get(Q(stream=Stream.objects.get(id=stream)) & Q(substream=SubStream.objects.get(id=substream)) & Q(year=semyear))
            except:
                get_semester_fees = []
        else:
            try:
                get_semester_fees = YearFees.objects.get(Q(stream=Stream.objects.get(id=stream)) & Q(year=semyear))
            except:
                get_semester_fees = []
        ## end of modification
    else:
        get_semester_fees = []
    return get_semester_fees


@login_required(login_url='/login/')  
def DeleteArchive(request,enroll_id):
    display = ""
    found_student = ""
    student = []
    if request.user.is_superuser:

        if request.method == "POST":
            delete_student_id = request.POST.get('delete_student_id')
            delete = request.POST.get('delete')
            if delete_student_id:
                print("student got")
                try:
                    getstudent = Student.objects.get(id = delete_student_id)
                    print("Delete Student : ",getstudent)
                    getenroll = Enrolled.objects.filter(student = getstudent.id)
                    studentdocuments = StudentDocuments.objects.filter(student = getstudent.id)
                    studentfees = StudentFees.objects.filter(student = getstudent.id)
                    qualification = Qualification.objects.filter(student = getstudent.id)
                    additionalenrollmentdetails = AdditionalEnrollmentDetails.objects.filter(student = getstudent.id)
                    courier = Courier.objects.filter(student = getstudent.id)
                    paymentreciept = PaymentReciept.objects.filter(student = getstudent.id)
                    studentsyllabus = StudentSyllabus.objects.filter(student = getstudent.id)

                    getstudent.delete()
                    getenroll.delete()
                    studentdocuments.delete()
                    studentfees.delete()
                    qualification.delete()
                    additionalenrollmentdetails.delete()
                    courier.delete()
                    paymentreciept.delete()
                    studentsyllabus.delete()
                    

                    return JsonResponse({'deleted':'yes'})

                except Student.DoesNotExist:
                    pass
            archive = request.POST.get('archive')
            student_id = request.POST.get('student_id')
            if student_id:
                try:
                    getstudent = Student.objects.get(id = student_id)
                    enrolled = Enrolled.objects.get(student = student_id)
                    studentfees = StudentFees.objects.filter(student = student_id)
                    try:
                        qualification = Qualification.objects.get(student = student_id)
                        create_archive_qualification = Qualification(
                            id=qualification.id,
                            student= qualification.student,
                            secondary_year = qualification.secondary_year,
                            sr_year = qualification.sr_year,
                            under_year= qualification.under_year,
                            post_year= qualification.post_year,
                            mphil_year= qualification.mphil_year,
                            others_year = qualification.others_year,
                            secondary_board= qualification.secondary_board,
                            sr_board= qualification.sr_board,
                            under_board= qualification.under_board,
                            post_board= qualification.post_board,
                            mphil_board= qualification.mphil_board,
                            others_board= qualification.others_board,
                            secondary_percentage= qualification.secondary_percentage,
                            sr_percentage= qualification.sr_percentage,
                            under_percentage= qualification.under_percentage,
                            post_percentage= qualification.post_percentage,
                            mphil_percentage= qualification.mphil_percentage,
                            others_percentage= qualification.others_percentage,
                            secondary_document= qualification.secondary_document,
                            sr_document= qualification.sr_document,
                            under_document= qualification.under_document,
                            post_document= qualification.post_document,
                            mphil_document= qualification.mphil_document,
                            others_document= qualification.others_document
                            )
                        create_archive_qualification.save(using="archive")
                        qualification.delete()

                    except Qualification.DoesNotExist:
                        pass
                    additionalenrollmentdetails = AdditionalEnrollmentDetails.objects.get(student = student_id)
                    university_enrollment = UniversityEnrollment.objects.filter(student = student_id)
                    courier = Courier.objects.filter(student = student_id)
                    paymentreciept = PaymentReciept.objects.filter(student = student_id)
                    submitted_answer = SubmittedExamination.objects.filter(student = getstudent.user)
                    results = Result.objects.filter(student = student_id)
                    personal_documents = PersonalDocuments.objects.filter(student = student_id)
                    student_documents = StudentDocuments.objects.filter(student = student_id)
                    university_examination = UniversityExamination.objects.filter(student = student_id)
                    result_uploaded = ResultUploaded.objects.filter(student = student_id)
                    if archive == "yes":
                        # print("yes")
                        getstudent.archive = True
                        getstudent.save()
                        create_archive_student = Student(
                            id = getstudent.id,
                            name = getstudent.name, 
                            father_name = getstudent.father_name, 
                            mother_name = getstudent.mother_name, 
                            dateofbirth = getstudent.dateofbirth, 
                            mobile = getstudent.mobile, 
                            alternate_mobile1 = getstudent.alternate_mobile1, 
                            email = getstudent.email,
                            gender = getstudent.gender,
                            category = getstudent.category,
                            address = getstudent.address,
                            nationality = getstudent.nationality,
                            country = getstudent.country,
                            state = getstudent.state,
                            city = getstudent.city,
                            pincode = getstudent.pincode,
                            registration_id = getstudent.registration_id,
                            old_university_enrollment_id = getstudent.old_university_enrollment_id,
                            new_university_enrollment_id = getstudent.new_university_enrollment_id,
                            enrollment_id = getstudent.enrollment_id,
                            enrollment_date = getstudent.enrollment_date,
                            university = getstudent.university,
                            image = getstudent.image,
                            verified = getstudent.verified,
                            user = getstudent.user,
                            enrolled = getstudent.enrolled,
                            archive = getstudent.archive,
                            is_cancelled = getstudent.is_cancelled,
                            created_by = getstudent.created_by,
                            modified_by = getstudent.modified_by
                            )
                        create_archive_enrolled = Enrolled(
                            id=enrolled.id,
                            student=enrolled.student,
                            course=enrolled.course,
                            stream=enrolled.stream,
                            course_pattern=enrolled.course_pattern,
                            session=enrolled.session,
                            entry_mode=enrolled.entry_mode,
                            total_semyear=enrolled.total_semyear,
                            current_semyear=enrolled.current_semyear
                            )
                        create_archive_student.save(using="archive")
                        create_archive_enrolled.save(using="archive")
                        
                        
                        for i in studentfees:
                            create_archive_studntfees = StudentFees(
                                id=i.id,
                                student=i.student,
                                studypattern=i.studypattern,
                                stream=i.stream,
                                tutionfees=i.tutionfees,
                                examinationfees=i.examinationfees,
                                bookfees=i.bookfees,
                                resittingfees=i.resittingfees,
                                entrancefees=i.entrancefees,
                                extrafees=i.extrafees,
                                discount=i.discount,
                                totalfees=i.totalfees,
                                sem=i.sem
                                )
                            create_archive_studntfees.save(using="archive")
                            i.delete()

                        for i in university_enrollment:
                            create_archive_universityenrollment = UniversityEnrollment(
                                id=i.id,
                                student = i.student,
                                type = i.type,
                                course_id = i.course_id,
                                course_name = i.course_name,
                                enrollment_id = i.enrollment_id
                                )
                            create_archive_universityenrollment.save(using="archive")
                            i.delete()
                        for i in courier:
                            create_archive_courier = Courier(
                                id=i.id,
                                student=i.student, 
                                article_name = i.article_name , 
                                courier_from = i.courier_from , 
                                courier_to = i.courier_to, 
                                booking_date = i.booking_date, 
                                courier_company = i.courier_company , 
                                tracking_id = i.tracking_id , 
                                remarks = i.remarks 
                                )
                            create_archive_courier.save(using="archive")
                            i.delete()
                        for i in paymentreciept:
                            create_archive_paymentreciept = PaymentReciept(
                                id=i.id,
                                student = i.student,
                                payment_for= i.payment_for,
                                payment_type= i.payment_type,
                                fee_reciept_type= i.fee_reciept_type,
                                transaction_date= i.transaction_date,
                                cheque_no= i.cheque_no,
                                bank_name= i.bank_name,
                                paidamount= i.paidamount,
                                pendingamount= i.pendingamount,
                                transactionID = i.transactionID,
                                paymentmode= i.paymentmode,
                                remarks= i.remarks,
                                session= i.session,
                                semyear= i.semyear
                                )
                            create_archive_paymentreciept.save(using="archive")
                            i.delete()
                        for i in submitted_answer:
                            create_archive_submitted_answer = SubmittedExamination(
                                id=i.id,
                                student=i.student,
                                exam=i.exam,
                                question=i.question,
                                type=i.type,
                                marks=i.marks,
                                marks_obtained=i.marks_obtained,
                                submitted_answer=i.submitted_answer,
                                answer=i.answer,
                                result=i.result
                                )
                            create_archive_submitted_answer.save(using="archive")
                            i.delete()



                        for i in results:
                            create_archive_result = Result(
                                id=i.id,
                                student=i.student,
                                exam=i.exam,
                                total_question=i.total_question,
                                attempted=i.attempted,
                                total_marks=i.total_marks,
                                score=i.score,
                                result=i.result
                                )
                            create_archive_result.save(using="archive")
                            i.delete()
                        for i in personal_documents:
                            create_archive_personal_documents = PersonalDocuments(
                                id = i.id,
                                document=i.document,
                                document_name=i.document_name,
                                document_ID_no=i.document_ID_no,
                                student=i.student
                                )
                            create_archive_personal_documents.save(using="archive")
                            personal_document_images = PersonalDocumentsImages.objects.filter(document = i)
                            for j in personal_document_images:
                                create_archive_personal_documents_images = PersonalDocumentsImages(
                                    id=j.id,
                                    document = j.document,
                                    document_image=j.document_image
                                    )
                                create_archive_personal_documents_images.save(using="archive")
                                j.delete()
                            i.delete()
                        for i in student_documents:
                            create_archive_student_documents = StudentDocuments(
                                id=i.id,
                                student=i.student,
                                document=i.document,
                                document_name=i.document_name,
                                document_ID_no=i.document_ID_no,
                                document_image_front=i.document_image_front,
                                document_image_back=i.document_image_back
                                )
                            create_archive_student_documents.save(using="archive")
                            i.delete()

                        for i in university_examination:
                            create_archive_university_examination = UniversityExamination(
                                id=i.id,
                                student = i.student,
                                type = i.type,
                                amount = i.amount,
                                date = i.date,
                                examination = i.examination,
                                semyear = i.semyear,
                                paymentmode = i.paymentmode,
                                remarks = i.remarks
                                )
                            create_archive_university_examination.save(using="archive")
                            i.delete()

                        for i in result_uploaded:
                            create_archive_result_uploaded = ResultUploaded(
                                id=i.id,
                                student = i.student,
                                date = i.date,
                                examination = i.examination,
                                semyear = i.semyear,
                                uploaded = i.uploaded,
                                remarks = i.remarks
                                )
                            create_archive_result_uploaded.save(using="archive")
                            i.delete()
                        
                        getstudent.delete()
                        enrolled.delete()
                        return JsonResponse({'archived':'yes'})
                    else:
                        print("no")
                        getstudent.archive = False
                        getstudent.save()
                        
                        return JsonResponse({'archived':'yes'})
                except Student.DoesNotExist:
                    pass
        try:
            getstudent = Student.objects.get(enrollment_id = enroll_id)
            enroll = Enrolled.objects.get(student = getstudent.id)
            course = Course.objects.get(id = enroll.course.id)
            stream = Stream.objects.get(id = enroll.stream.id)
            # print(getstudent , enroll , course , stream)
            commonfees = PaymentReciept.objects.filter(Q(student = getstudent.id) & (Q(payment_for="Course Fees") | Q(payment_for="Tution Fees") | Q(payment_for="Re-Registration Fees") | Q(payment_for="Examination Fees")))
            additionalfees = PaymentReciept.objects.filter(Q(student = getstudent.id) & ~(Q(payment_for="Course Fees") | Q(payment_for="Tution Fees") | Q(payment_for="Re-Registration Fees") | Q(payment_for="Examination Fees")))
            # print("Common Fees : ",commonfees)
            # print("Additional Fees : ",additionalfees)
            obj = {
                "id":getstudent.id,
                "enrollment_id":getstudent.enrollment_id,
                "name":getstudent.name,
                "dateofbirth":getstudent.dateofbirth,
                "email":getstudent.email,
                "mobile":getstudent.mobile,
                "archive":getstudent.archive,
                "course_pattern":enroll.course_pattern,
                "session":enroll.session,
                "entry_mode":enroll.entry_mode,
                "total_semyear":enroll.total_semyear,
                "current_semyear":enroll.current_semyear,
                "course":course.name,
                "stream":stream.name,
                "commonfees":commonfees,
                "additionalfees":additionalfees
                
                
            }
            student.append(obj)
            
            found_student = "yes"
        except Student.DoesNotExist:
            pass
    params = {
        "display":display,
        "found_student":found_student,
        "student":student
    }
    return render(request,"super_admin/delete_archive.html",params)


@csrf_exempt
def PaymentSuccess(request):
    print("request")
    obj = {}
    if request.method == "POST":
        response = request.POST
        dumpdata = json.dumps(response)
        responsestring = json.loads(dumpdata)
        mihpayid = responsestring['mihpayid']
        mode = responsestring['mode']
        status = responsestring['status']
        unmappedstatus = responsestring['unmappedstatus']
        key = responsestring['key']
        txnid = responsestring['txnid']
        amount = responsestring['amount']
        cardCategory = responsestring['cardCategory']
        net_amount_debit = responsestring['net_amount_debit']
        addedon = responsestring['addedon']
        productinfo = responsestring['productinfo']
        firstname = responsestring['firstname']
        lastname = responsestring['lastname']

        email = responsestring['email']
        phone = responsestring['phone']
        payment_source = responsestring['payment_source']
        PG_TYPE = responsestring['PG_TYPE']
        bank_ref_num = responsestring['bank_ref_num']
        bankcode = responsestring['bankcode']
        cardnum = responsestring['cardnum']
        print(mihpayid)
        print(mode)
        print(status)
        print(unmappedstatus)
        print(key)
        print(txnid)
        print(amount)

        print(cardCategory)
        print(net_amount_debit)
        print(addedon)
        print(productinfo)
        print(firstname)
        print(lastname)

        print(email)
        print(phone)
        print(payment_source)
        print(PG_TYPE)
        print(bank_ref_num)
        print(bankcode)
        print(cardnum)
        try:
            getstudent = Student.objects.get(Q(email = email) & Q(mobile = phone))
            getenroll = Enrolled.objects.get(student = getstudent.id)
            getpaymentreciept = PaymentReciept.objects.filter(Q(student = getstudent.id) & Q(semyear = getenroll.current_semyear))
            print(getpaymentreciept)
            # add_payment_reciept = PaymentReciept(student = getstudent.id,fee_reciept_type="Course Fees",transaction_date= str(date.today()),cheque_no='',bank_name='',paidamount=net_amount_debit,pendingamount='0',transactionID = txnid,paymentmode='Online',remarks=status,session=getenroll.session,semyear=getenroll.current_semyear)
            add_payment_reciept = PaymentReciept(student = getstudent.id,payment_for="Course Fees",payment_type="Full Payment",fee_reciept_type="New",transaction_date= str(date.today()),cheque_no='',bank_name='',paidamount=net_amount_debit,pendingamount='0',transactionID = txnid,paymentmode='Online',remarks=status,session=getenroll.session,semyear=getenroll.current_semyear)
            add_payment_reciept.save()
    
            add_transactiondetails = TransactionDetails(transactionID=txnid,mihpayid=mihpayid,mode=mode,status=status,unmappedstatus=unmappedstatus,key=key,amount=amount,cardCategory=cardCategory,net_amount_debit=net_amount_debit,addedon=addedon,productinfo=productinfo,firstname=firstname,lastname=lastname,email=email,phone=phone,payment_source=payment_source,PG_TYPE=PG_TYPE,bank_ref_num=bank_ref_num,bankcode=bankcode,name_on_card='',cardnum=cardnum)
            add_transactiondetails.save()

            getlatestpaymentreciept = PaymentReciept.objects.latest('id')
            print("latest payment reciept :",getlatestpaymentreciept)
            paymentrecieptserializer = PaymentRecieptSerializer(getlatestpaymentreciept,many=False)
            view_student = Student.objects.get(id=getlatestpaymentreciept.student)
            view_enroll = Enrolled.objects.get(student=view_student.id)
            view_course = Course.objects.get(id = view_enroll.course)
            view_stream = Stream.objects.get(id = view_enroll.stream)
            temp = {
                "name":view_student.name,
                "email":view_student.email,
                "mobile":view_student.mobile,
                "enrollment_id":view_student.enrollment_id,
                "address":view_student.address,
                "country":view_student.country,
                "state":view_student.state,
                "city":view_student.city,
                "pincode":view_student.pincode,
                "course":view_course.name,
                "stream":view_stream.name
            }
            obj = {
                'view_reciept':paymentrecieptserializer.data,
                'personal':temp
                }
            
            
            
            
            
            res = sm(
                subject = "CIIS Payment Recieved",
                message = '''Thankyou For Making Payment of rs {} \n
                CIIS INDIA \n
                Click on the link below to download Invoice \n
                https://erp.ciisindia.in/getinvoice/{}'''.format(net_amount_debit,getlatestpaymentreciept.id),
                from_email = 'testmail@erp.ciisindia.in',
                recipient_list = [email],
                fail_silently=False,
                    )
            print(res)

        except Student.DoesNotExist:
            print("student Does not exist")
        # try:
        #     getstudenttransaction = SaveStudentTransaction.objects.get(txnid = txnid)
        #     getstudenttransaction.status = "success"
        #     getstudenttransaction.used = "yes"
        #     getstudenttransaction.archived = "yes"
        #     getstudenttransaction.save()
        # except SaveStudentTransaction.DoesNotExist:
        #     print("student transaction data not found")
        
    else:
        print("get")
    params = {
        "obj":obj
    }
    return render(request,"super_admin/payment_success.html",params)


def export_users_xls(request,student_ids):
    students = student_ids.split(',')
    print("students :",students)
    temp = []
    for i in students:
        getstudent = Student.objects.get(id = i)
        enrolled = Enrolled.objects.get(student= getstudent.id)
        course = Course.objects.get(id=enrolled.course.id)
        stream = Stream.objects.get(id = enrolled.stream.id)
        test_tuple = (
            getstudent.name,
            getstudent.email,
            getstudent.mobile,
            getstudent.enrollment_id,
            enrolled.course_pattern,
            enrolled.total_semyear,
            enrolled.current_semyear,
            course.name,
            stream.name
        )
        temp.append(test_tuple)
    # print("test_tuple :",temp)
    # for i in temp:
    #     print(i.values)
    # print("Temp : ",temp)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="DataSet.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Data Export') # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Name', 'Email Address', 'Mobile' , 'Enrollment ID','Course Pattern' , 'Total Semester', 'Current Semester' , 'Course Name', 'Stream Name']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) # at 0 row 0 column 

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = Student.objects.all().values_list('name', 'email' , 'mobile' , 'enrollment_id','dateofbirth')
    # print("rows :",rows)

    for row in temp:
        # print("row : ",row)
        # print("_______________________________________________________________________________________________________")
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)

    return response

def export_pdf(request,student_ids):
    students = student_ids.split(',')
    print("students :",students)
    temp = []
    for i in students:
        getstudent = Student.objects.get(id = i)
        enrolled = Enrolled.objects.get(student= getstudent.id)
        course = Course.objects.get(id=enrolled.course.id)
        stream = Stream.objects.get(id = enrolled.stream.id)
        paymentreciept = PaymentReciept.objects.filter(student=getstudent)
        payment_reciept_serializer = PaymentRecieptSerializer(paymentreciept,many=True)
        obj = {
            "student_id":getstudent.id,
            "name":getstudent.name,
            "email":getstudent.email,
            "mobile":getstudent.mobile,
            "enrollment_id":getstudent.enrollment_id,
            "course_pattern":enrolled.course_pattern,
            "total_semyear":enrolled.total_semyear,
            "current_semyear":enrolled.current_semyear,
            "course":course.name,
            "stream":stream.name,
            "paymentreciept":payment_reciept_serializer.data
        }
        temp.append(obj)
    print("temp :",temp)
    params = {
        "data":temp
    }
    return render(request,"super_admin/export_pdf.html",params)

## commented by Avani as function is incomplete - new function created for Change Course page
def changecourse (request):
    university_id = 0
    if request.method == "POST":
        enroll_id = request.POST.get('enroll_id')
        name = request.POST.get('name')
        course_data = []
        student_data = []
        print("enroll_id", enroll_id)
        if enroll_id != '':
            try:
                getStudent = Student.objects.get(enrollment_id = enroll_id)
                student_id = getStudent.id
                university_id = getStudent.university_id
                course_data = fetchCourseData(student_id, university_id)
                print("getStudent", getStudent)
                
                
            except Student.DoesNotExist:
                print("invalid enrollment id")
        elif name != '':
            try:
                getstudents = Student.objects.filter(
                            Q(name__icontains = name) |
                            Q(father_name__icontains = name) |
                            Q(mother_name__icontains = name)
                        )
                if len(getstudents) != 0:
                    print(getstudents)
                    if len(getstudents) == 1:
                        student_id = getstudents[0].id
                        university_id =getstudents[0].university_id
                        course_data = fetchCourseData(student_id, university_id)
                    if len(getstudents) > 1:
                        print("length", len(getstudents))
            except Student.DoesNotExist:
                print("invalid name")
            if len(getstudents) > 1:
                for i in getstudents:
                    obj = {
                        
                            "student_id":i.id,
                            "name":i.name,
                            "enrollment_id":i.enrollment_id,
                            
                        }
                    student_data.append(obj)
        print("coursedetails: ", course_data)
        print("studentdetails: ", student_data)
        try:
            get_all_course = Course.objects.filter(Q(university=university_id))
            print(get_all_course)
            courseserializer = CourseSerializer(get_all_course,many=True)
        except:
            courseserializer = ''
        try:
            sessionnames = SessionNames.objects.filter(Q(status=True))
            sessionserializer= SessionNames(sessionnames, many=True)
        except:
            sessionserializer = ''
        return JsonResponse({'coursedetails': course_data, 'studentdetails': student_data, 'all_course': courseserializer.data, 'sessionnames': sessionserializer})
    print(university_id)      
    params = {
        "all_course":Course.objects.filter(university_id = university_id),
        "sessionnames": SessionNames.objects.filter(status=True)
    }
    return render(request,"super_admin/editcourse.html", params)

## new function to only show quick registered students  added by Avani 15/07
@login_required(login_url='/login/')
def viewquickregister (request):
    params= {
                "students":Student.objects.filter(~Q(is_cancelled = True) & Q(is_quick_register=True)).order_by('-id'),
                "title" : 'List of Quick Registered Students'
                
            }
    return render(request,"super_admin/viewstudents.html", params)

## new function to only show registered students  added by Avani 15/07
@login_required(login_url='/login/')
def viewregister (request):
    params= {
                "students":Student.objects.filter(~Q(is_cancelled = True) & Q(is_quick_register=False)).order_by('-id'),
                "title" : 'List of Registered Students'
                
            }
    return render(request,"super_admin/viewstudents.html", params)

## new function to only payment status added by Avani 15/07
@login_required(login_url='/login/')
def viewpaymentstatus (request):
    student_list = []
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_data_entry:
            if request.method == "POST":
                payment_reciept_id = request.POST.get('payment_reciept_id')
                payment_reciept_status = request.POST.get('payment_reciept_status')
                if payment_reciept_id and payment_reciept_status:
                    get_payment_reciept = PaymentReciept.objects.get(id = payment_reciept_id)
                    get_payment_reciept.status = payment_reciept_status
                    total_fees = get_payment_reciept.semyearfees
                    paid_fees = get_payment_reciept.uncleared_amount
                    print(total_fees , paid_fees)
                    if int(paid_fees) > int(total_fees):
                        get_payment_reciept.advanceamount = int(paid_fees) - int(total_fees)
                        get_payment_reciept.paidamount = int(paid_fees)
                        get_payment_reciept.pendingamount = 0
                        get_payment_reciept.uncleared_amount = 0
                    elif int(paid_fees) < int(total_fees):
                        get_payment_reciept.advanceamount = 0
                        get_payment_reciept.paidamount = int(paid_fees)
                        get_payment_reciept.pendingamount = int(total_fees) - int(paid_fees)
                        get_payment_reciept.uncleared_amount = 0
                    elif int(paid_fees) == int(total_fees):
                        get_payment_reciept.advanceamount = 0
                        get_payment_reciept.paidamount = int(paid_fees)
                        get_payment_reciept.pendingamount = int(total_fees) - int(paid_fees)
                        get_payment_reciept.uncleared_amount = 0
                    # print("total fees :",get_payment_reciept.semyearfees)
                    # print("paid fees :",get_payment_reciept.paidamount)
                    # print("pending fees :",get_payment_reciept.pendingamount)
                    # print("uncleared amount :",get_payment_reciept.uncleared_amount)
                    get_payment_reciept.save()
                    updated_payment_data = []
                    all_pending_realisation = PaymentReciept.objects.filter(status = "Not Realised")
                    for i in all_pending_realisation:
                        try:
                            getstudent = Student.objects.get(id = i.student)
                            enrolled = Enrolled.objects.get(student = i.student)
                            course_name = Course.objects.get(id = enrolled.course)
                            stream_name = Stream.objects.get(id = enrolled.stream)
                            obj = {
                                "student_id":getstudent.id,
                                "name":getstudent.name,
                                "email":getstudent.email,
                                "mobile":getstudent.mobile,
                                "course":course_name.name,
                                "stream":stream_name.name,
                                "current_semester":enrolled.current_semyear,
                                "payment_reciept_id":i.id,                        
                                "payment_status":i.status,
                                "transactionID":i.transactionID,
                                "uncleared_amount":i.uncleared_amount
                            }
                            updated_payment_data.append(obj)
                        except Student.DoesNotExist:
                            pass
                    return JsonResponse({'data': updated_payment_data})

            all_pending_realisation = PaymentReciept.objects.filter(status = "Not Realised")
            for i in all_pending_realisation:
                try:
                    getstudent = Student.objects.get(id = i.student.id)
                    enrolled = Enrolled.objects.get(student = i.student)
                    course_name = Course.objects.get(id = enrolled.course.id)
                    stream_name = Stream.objects.get(id = enrolled.stream.id)
                    obj = {
                        "student_id":getstudent.id,
                        "name":getstudent.name,
                        "email":getstudent.email,
                        "mobile":getstudent.mobile,
                        "course":course_name.name,
                        "stream":stream_name.name,
                        "current_semester":enrolled.current_semyear,
                        "payment_reciept_id":i.id,                        
                        "payment_status":i.status,
                        "transactionID":i.transactionID,
                        "uncleared_amount":i.uncleared_amount
                    }
                    student_list.append(obj)
                except Student.DoesNotExist:
                    pass
    params= {
                "payment_student":student_list,
                
            }
    return render(request,"super_admin/viewpaymentstatus.html", params)

def fetchCourseData(student_id, university_id):

    try:
        coursedetails = Enrolled.objects.get(student_id = student_id)
        student = Student.objects.get(id = student_id)
        print(coursedetails)
        course = Course.objects.get(id=coursedetails.course_id)
        stream = Stream.objects.get(id = coursedetails.stream_id)
        ##added by Avani 14/08
        try:
            substream = SubStream.objects.get(id = coursedetails.substream_id)
            substreamname = substream.name
        except:
            substreamname = '-'
        ## end of add
        university = University.objects.get(id=university_id)
        course_data= {
                    
                        "course_pattern":coursedetails.course_pattern,
                        "total_semyear":coursedetails.total_semyear,
                        "current_semyear":coursedetails.current_semyear,
                        "course":course.name,
                        "stream":stream.name,
                        "substream": substreamname,##added by Avani 14/08
                        "course_duration": stream.sem,
                        "university":university.university_name,
                        "university_id":university.registrationID,
                        "enrollment_id": student.enrollment_id,
                        "session": coursedetails.session

        }
    except Enrolled.DoesNotExist:
        print("No course details found")
    
    return(course_data)

def savechangecourse(request):
    enroll_id = request.POST.get('enrollment_id')
    getstudent = Student.objects.get(enrollment_id = enroll_id)
    change_course = request.POST.get('change_course')
    change_stream = request.POST.get('change_stream')
    change_substream = request.POST.get('change_substream')##added by Avani 14/08
    change_study_pattern = request.POST.get('change_study_pattern')
    change_semyear = request.POST.get('change_semyear')
    full_course_fees = request.POST.get('full_course_fees')
    total_fees_paid = 0
    get_all_payment_reciepts = PaymentReciept.objects.filter(student=getstudent)
    for reciept in get_all_payment_reciepts:
        if reciept.payment_for == "Refund":
            print(reciept.payment_for)
            total_fees_paid -= int(reciept.paidamount)
        else:
            print(reciept.payment_for)
            total_fees_paid += int(reciept.paidamount)
            
    print("1total fees paid :",total_fees_paid)
    enrolled = Enrolled.objects.get(student=getstudent)
    change_session = request.POST.get('change_session')
    if change_session:
        print("2session :",change_session)
        if enrolled.session == change_session:
            pass
        else:
            enrolled.session = change_session
    if change_course and change_stream and change_study_pattern:
        try:
            getlatestreciept = PaymentReciept.objects.latest('id')
            tid = getlatestreciept.transactionID
            tranx = tid.replace("TXT445FE",'')
            transactionID =  str("TXT445FE") + str(int(tranx) + 1)
        except PaymentReciept.DoesNotExist:
            transactionID = "TXT445FE101"
                
        if change_study_pattern == "Full Course":
            if full_course_fees == '':
                calc = total_fees_paid
            else:
                calc = total_fees_paid - int(full_course_fees)
            total_year_fees = full_course_fees
            if calc >= 0:
                pending_amount = 0
                advance_amount = calc
            else:
                print("total fees is less than streamyear fees hence pending")
                pending_amount = int(full_course_fees) - total_fees_paid
                advance_amount = 0
                    
            enrolled.course = Course.objects.get(id=change_course)
            enrolled.stream = Stream.objects.get(id=change_stream)
            enrolled.course_pattern = change_study_pattern
            enrolled.total_semyear = 1
            enrolled.current_semyear = 1
                    
            get_all_payment_reciepts.delete()
            create_payment_reciept = PaymentReciept(
                student = getstudent,
                payment_for = "Course Fees",
                payment_categories = "New",
                payment_type = "Part Payment",
                fee_reciept_type = "New",
                transaction_date = date.today(),
                semyearfees = full_course_fees,
                paidamount = total_fees_paid,
                pendingamount=pending_amount,
                advanceamount=advance_amount,
                transactionID=transactionID,
                paymentmode = "Offline",
                status = "Realised",
                semyear=1,
                modified_by = request.user.id
            )
            create_payment_reciept.save()
        else:
            print('3total_fees_paid', total_fees_paid)
            print('4change_study_pattern', change_study_pattern)
            print('5change_stream', change_stream)
            print('6change_semyear', change_semyear)
            streamyearfees = CalculateFees(total_fees_paid,change_study_pattern,change_stream, change_substream, change_semyear)##modified by Avani 14/08
            print("7streamyearfees", streamyearfees.totalfees)
            if streamyearfees.totalfees == '':
                calc = total_fees_paid
            else:
                calc = total_fees_paid - int(streamyearfees.totalfees)
            total_year_fees = streamyearfees.totalfees
            if calc >= 0:
                pending_amount = 0
                advance_amount = calc
            else:
                print("total fees is less than streamyear fees hence pending")
                pending_amount = int(streamyearfees.totalfees) - total_fees_paid
                advance_amount = 0
            stream = Stream.objects.get(id=change_stream)
            get_all_student_fees = StudentFees.objects.filter(student=getstudent)
            get_all_student_fees.delete()
            if change_study_pattern == "Semester":
                getsemesterfees = SemesterFees.objects.filter(stream=stream)
                        
                for i in getsemesterfees:
                    addstudentfees = StudentFees(
                        student = getstudent,
                        studypattern="Semester",
                        stream=Stream.objects.get(id=stream.id),
                        tutionfees=i.tutionfees,
                        examinationfees=i.examinationfees,
                        bookfees=i.bookfees,
                        resittingfees=i.resittingfees,
                        entrancefees=i.entrancefees,
                        extrafees=i.extrafees,
                        discount=i.discount,
                        totalfees=i.totalfees,
                        sem=i.sem
                        )
                    addstudentfees.save()
                    total_semyear = int(stream.sem) * 2
            else:
                total_semyear = int(stream.sem)
                try: ## added by Avani as code breaking if no record in YearFees 17/07
                    getyearfees = YearFees.objects.filter(stream=stream)
                    print("8getyearfees", getyearfees)
                    for i in getyearfees:
                        addstudentfees = StudentFees(
                            student = getstudent,
                            studypattern="Annual",
                            stream=Stream.objects.get(id=stream.id),
                            tutionfees=i.tutionfees,
                            examinationfees=i.examinationfees,
                            bookfees=i.bookfees,
                            resittingfees=i.resittingfees,
                            entrancefees=i.entrancefees,
                            extrafees=i.extrafees,
                            discount=i.discount,
                            totalfees=i.totalfees,
                            sem=i.year
                            )
                        addstudentfees.save()
                except YearFees.DoesNotExist:
                    print("no record found")
            print("dffgggggfhf hfhgfjfjgj")
            enrolled.course = Course.objects.get(id=change_course)
            enrolled.stream = Stream.objects.get(id=change_stream)
            ##added by Avani 14/08
            if change_substream != '':
                enrolled.substream = SubStream.objects.get(id=change_substream)
            else:
                enrolled.substream = None
            ## end of add
            enrolled.course_pattern = change_study_pattern
            enrolled.total_semyear = total_semyear
            enrolled.current_semyear = change_semyear
            enrolled.save()
            get_all_payment_reciepts.delete()
            create_payment_reciept = PaymentReciept(
                student = getstudent,
                payment_for = "Course Fees",
                payment_categories = "New",
                payment_type = "Part Payment",
                fee_reciept_type = "New",
                transaction_date = date.today(),
                semyearfees = total_year_fees,
                paidamount = total_fees_paid,
                pendingamount=pending_amount,
                advanceamount=advance_amount,
                transactionID=transactionID,
                paymentmode = "Offline",
                status = "Realised",
                semyear=change_semyear,
                modified_by = request.user.id
            )
            create_payment_reciept.save()
    enrolled.save()        
    print("get_all_payment_reciepts :",get_all_payment_reciepts)
    return JsonResponse({'message': "Course changes done"})
                
def showstudentform(request, enroll_id): 
    print("enroll_id: ", enroll_id)
    try:
        getstudent = Student.objects.get(enrollment_id = enroll_id)
        try:
            country = Countries.objects.get(id = getstudent.country_id)
            try:
                state = States.objects.get(id = getstudent.state_id)
                try:
                    district = Cities.objects.get(id = getstudent.city_id)
                except Cities.DoesNotExist:
                    district=''
                    print("No District found")
            except States.DoesNotExist:
                state =''
                print("No state found")
        except Countries.DoesNotExist: 
            country= ''
            print("No Country found")
        addressdetails = {}
        addressdetails['country'] = country.name
        addressdetails['state'] = state.name
        addressdetails['district'] = district.name
        try:
            counselor_name = AdditionalEnrollmentDetails.objects.get(student_id = getstudent.id).counselor_name
        except AdditionalEnrollmentDetails.DoesNotExist:
            counselor_name = ''
        try:
            passport = StudentDocuments.objects.get(student_id = getstudent.id, document="Passport").document_ID_no
        except StudentDocuments.DoesNotExist:
            passport = ''
        try:
            aadhar = StudentDocuments.objects.get(student_id = getstudent.id, document="Aadhar Card").document_ID_no
        except StudentDocuments.DoesNotExist:
            aadhar = ''
        try:
            university = University.objects.get(id = getstudent.university_id).university_name
        except StudentDocuments.DoesNotExist:
            aadhar = ''
        try:
            coursedetails = Enrolled.objects.get(student_id= getstudent.id)
            try:
                coursename = Course.objects.get(id = coursedetails.course_id).name
            except:
                coursename = ''
            try:
                streamname = Stream.objects.get(id = coursedetails.stream_id).name 
            except:
                streamname = ''
        except:
            coursedetails = ''
        otherdetails = {}
        otherdetails['aadhar'] = aadhar
        otherdetails['passport'] = passport
        otherdetails['counselor'] = counselor_name
        otherdetails['university'] =university
        otherdetails['coursename'] = coursename
        otherdetails['streamname'] = streamname
        print("coursedtails", coursedetails)
        try:
            qualificationdetails = Qualification.objects.get(student_id = getstudent.id)
            print("qualificationdetails",qualificationdetails)
        except Qualification.DoesNotExist:
            qualificationdetails = ''
    except Student.DoesNotExist:
        getstudent= ''
        print("No student details found.")
   
   
    imagename = "/media/" + str(getstudent.image)
    print("imagename", imagename)
    params= {
        "student": getstudent,
        "addressdetails": addressdetails,
        "otherdetails": otherdetails,
        "coursedetails": coursedetails,
        "qualificationdetails": qualificationdetails,
        "image": imagename

    }
    return render(request,"super_admin/studentform.html", params)

def documentmanagement(request, enroll_id):
    try:
        getstudent = Student.objects.get(enrollment_id = enroll_id)
        qualificationdocs = {}
        try:
            qualificationdetails = Qualification.objects.get(student_id = getstudent.id)
            print("qualificationdetails",qualificationdetails)
            
            if qualificationdetails.secondary_document != '':
                qualificationdocs["secondary"]= '/media/' + str(qualificationdetails.secondary_document)
            else:
                qualificationdocs["secondary"]= ''
            if qualificationdetails.sr_document != '':
                qualificationdocs["sr"]= '/media/' + str(qualificationdetails.sr_document)
            else:
                qualificationdocs["sr"]= ''
            if qualificationdetails.under_document != '':
                qualificationdocs["under"]= '/media/' + str(qualificationdetails.under_document)
            else:
                qualificationdocs["under"]= ''
            if qualificationdetails.post_document != '':
                qualificationdocs["post"]= '/media/' + str(qualificationdetails.post_document)
            else:
                qualificationdocs['post'] = ''
            if qualificationdetails.mphil_document != '':
                qualificationdocs["mphil"]= '/media/' + str(qualificationdetails.mphil_document)
            else:
                qualificationdocs['mphil'] =''
                qualificationdocs['others'] = qualificationdetails.others
        except Qualification.DoesNotExist:
            qualificationdetails = ''
        try:
            personaldocuments = StudentDocuments.objects.filter(student_id = getstudent.id)
            print(personaldocuments)
        except StudentDocuments.DoesNotExist:
            personaldocuments = ''

    except Student.DoesNotExist:
        print("No student found")
    params = {
        "qualificationdocs":qualificationdocs,
        "personaldocs":personaldocuments
    }

    return render(request, "super_admin/documentmanagement.html", params)

def paymentmode(request):
    if request.user.is_superuser:
         if request.method == "POST":
                paymentmode = request.POST.get('paymentmode')
                try:
                    print("paymentmode", paymentmode)
                    name = PaymentModes.objects.get(name__iexact=paymentmode)
                    return JsonResponse({'msg': "Payment mode already exists"})
                except:
                    newpaymentmode = PaymentModes(name=paymentmode)
                    newpaymentmode.save()
                    return JsonResponse({'msg': "Payment mode added succesfully"})

    params={
       "all_payment_modes": PaymentModes.objects.all()

    }
    return render(request, "masters/paymentmode.html", params)

def editpaymentmode(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            name = request.POST.get('name')
            status = request.POST.get('status')
            print(id, status, name)
            try:
                editpaymentmode = PaymentModes.objects.get(id = id)
                editpaymentmode.name = name
                editpaymentmode.status = status
                editpaymentmode.save()
                msg ="Payment Mode updated successfully"
            except:
                msg="No Payment Mode found"
    return JsonResponse({'msg': msg})

def deletepaymentmode(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            try:
                deletepaymentmode = PaymentModes.objects.get(id = id)
                deletepaymentmode.delete()
                msg ="Payment Mode deleted successfully"
            except:
                msg="No Payment Mode found"
    return JsonResponse({'msg': msg})

def feereceiptoption(request):
    if request.user.is_superuser:
         if request.method == "POST":
                feereceipt = request.POST.get('feereceipt')
                try:
                    name = FeeReceiptOptions.objects.get(name__iexact=feereceipt)
                    return JsonResponse({'msg': "Fee Receipt Option already exists"})
                except:
                    newfeereceiptoption = FeeReceiptOptions(name=feereceipt)
                    newfeereceiptoption.save()
                    return JsonResponse({'msg': "Fee Receipt option added succesfully"})
    print(FeeReceiptOptions.objects.all())
    params={
       "all_feereceipt_options": FeeReceiptOptions.objects.all()

    }
    return render(request, "masters/feereceipt.html", params)

def editfeereceiptoption(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            name = request.POST.get('name')
            status = request.POST.get('status')
            print(id, status, name)
            try:
                editfeereceiptoption = FeeReceiptOptions.objects.get(id = id)
                editfeereceiptoption.name = name
                editfeereceiptoption.status = status
                editfeereceiptoption.save()
                msg ="Fee Receipt option updated successfully"
            except:
                msg="No Fee Receipt option found"
    return JsonResponse({'msg': msg})

def deletefeereceiptoption(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            try:
                deletefeereceiptoption = FeeReceiptOptions.objects.get(id = id)
                deletefeereceiptoption.delete()
                msg ="Fee Receipt option deleted successfully"
            except:
                msg="No Fee Receipt option found"
    return JsonResponse({'msg': msg})          
            

def bankname(request):
    if request.user.is_superuser:
         if request.method == "POST":
                bank = request.POST.get('bank')
                try:
                    name = BankNames.objects.get(name__iexact=bank)
                    return JsonResponse({'msg': "Bank Name already exists"})
                except:
                    newbank = BankNames(name=bank)
                    newbank.save()
                    return JsonResponse({'msg': "Bank Name added succesfully"})
    print(FeeReceiptOptions.objects.all())
    params={
       "all_banks": BankNames.objects.all()

    }
    return render(request, "masters/banks.html", params)

def editbankname(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            name = request.POST.get('name')
            status = request.POST.get('status')
            print(id, status, name)
            try:
                editbank = BankNames.objects.get(id = id)
                editbank.name = name
                editbank.status = status
                editbank.save()
                msg ="Bank name updated successfully"
            except:
                msg="No Bank name found"
    return JsonResponse({'msg': msg})

def deletebankname(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            try:
                deletebank = BankNames.objects.get(id = id)
                deletebank.delete()
                msg ="Bank Name deleted successfully"
            except:
                msg="No Bank name found"
    return JsonResponse({'msg': msg})   



def sessionname(request):
    if request.user.is_superuser:
         if request.method == "POST":
                session = request.POST.get('session')
                try:
                    name = SessionNames.objects.get(name__iexact=session)
                    return JsonResponse({'msg': "Session Name already exists"})
                except:
                    newsession = SessionNames(name=session)
                    newsession.save()
                    return JsonResponse({'msg': "Session Name added succesfully"})
    
    params={
       "all_sessions": SessionNames.objects.all()

    }
    return render(request, "masters/sessions.html", params)

def editsessionname(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            name = request.POST.get('name')
            status = request.POST.get('status')
            print(id, status, name)
            try:
                editsession = SessionNames.objects.get(id = id)
                editsession.name = name
                editsession.status = status
                editsession.save()
                msg ="Session Name updated successfully"
            except:
                msg="No Session name found"
    return JsonResponse({'msg': msg})

def deletesessionname(request):
    if request.user.is_superuser:
         if request.method == "POST":
            id = request.POST.get('id')
            try:
                deletebank = SessionNames.objects.get(id = id)
                deletebank.delete()
                msg ="Session Name deleted successfully"
            except:
                msg="No Session name found"
    return JsonResponse({'msg': msg})  

def addsubstream(request):##added by Avani 14/08
    if request.method == 'POST':
        stream = request.POST.get('stream')
        substream = request.POST.get('substream')
        print("stream substream", stream, substream)
        try:
            substream = SubStream.objects.get(name__iexact=substream)
            return JsonResponse({'msg': "Sub Stream already exists"})

        except:
           
            # getstream = Stream.objects.get(id = stream)
            newsubstream = SubStream(name = substream,stream_id=stream)
            newsubstream.save()
            return JsonResponse({'msg': 'Sub Stream added successfully'})
    
    universitydata = []
    university = University.objects.all()
    for i in university:
        courses = Course.objects.filter(university = i.id)
        obj = {
                    "university_name":i.university_name,
                    "course_name":courses,

                }
        universitydata.append(obj)
    params={
        "university": University.objects.all(),
        "univdata": universitydata
    }
    return render(request, 'super_admin/addsubstream.html', params)

def getstream(request):##added by Avani 14/08
    if request.user.is_superuser:
         if request.method == "POST":
             course = request.POST.get('coursename')
             stream = Stream.objects.filter(course_id= course)
             streamdata= StreamSerializer(stream,many=True)
    return JsonResponse({'stream':streamdata.data})

def updatesubstream(request):##added by Avani 14/08
    if request.method == 'POST':
        updatedlist = request.POST.get('updatesubstream')
        if updatedlist:
                jsondata = json.loads(updatedlist)
                data = jsondata['data']
                for i in data:
                    substream_id = i['substream_id']
                    substream_name = i['substream_name']
                    try:
                            modsubstream = SubStream.objects.get(id= substream_id)
                            if modsubstream.name == substream_name:
                                pass
                               
                            else:
                                modsubstream.name = substream_name
                                modsubstream.save()
                                msg = "Substream modified successfully"
                    except SubStream.DoesNotExist:
                            print("no substream")
                            msg="Substream does not exist"
    return JsonResponse({'msg':msg})

def deletesubstream(request):##added by Avani 14/08
    return render(request, 'super_admin/addsubstream.html')

def getsubstream(request):##added by Avani 14/08
    if request.method == "POST":
             streamid = request.POST.get('streamid')
             print('streamid', streamid)
             substream = SubStream.objects.filter(stream_id= streamid)
             print("Substream", substream)
             substreamdata= SubStreamSerializer(substream,many=True)
             print(substreamdata)
    return JsonResponse({'substream':substreamdata.data})














  