from rest_framework.decorators import api_view 
from rest_framework_simplejwt.tokens import RefreshToken
from .api_serializers import *
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes,parser_classes
from rest_framework.permissions import IsAuthenticated
from .models import *
import logging
from django.db import transaction
from django.db.models.query import QuerySet
import traceback
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import os
import json
from django.views.decorators.csrf import csrf_exempt
from datetime import date
from django.db.models import Q
import pandas as pd
from django.conf import settings
import openpyxl
from openpyxl.styles import Font

logger = logging.getLogger(__name__)
logger = logging.getLogger('student_registration')
handler = logging.FileHandler('student_registration.log')
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
logger.setLevel(logging.INFO)

def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token)
        }
    

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate
import logging

logger = logging.getLogger(__name__)

@api_view(['POST'])
def login_view(request):
    serializer = UserLoginSerializer(data=request.data)
    
    if serializer.is_valid():
        email = serializer.validated_data.get("email")
        password = serializer.validated_data.get("password")
        user = authenticate(email=email, password=password)

        if user is not None:
            token = get_tokens_for_user(user)

            if user.is_student:
                try:
                    student = Student.objects.get(email=email)
                    exams = StudentAppearingExam.objects.filter(
                        student_id__contains=[student.id]
                    ).select_related(
                        'exam__course__university', 
                        'exam__stream', 
                        'exam__subject', 
                        'exam__substream'
                    )
                except Student.DoesNotExist:
                    return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

                if not exams.exists():
                    return Response({
                        "message": "Login Successful",
                        "token": token,
                        "is_student": True,
                        "email": user.email,
                        "is_active": user.is_active,
                    }, status=status.HTTP_200_OK)

                exam_details = []
                examination_data = []
                university = None
                exam_progress = []

                for exam in exams:
                    try:
                        examination = exam.exam
                        university_obj = examination.course.university if examination.course and examination.course.university else None

                        if university_obj:
                            university = {
                                "id": university_obj.id,
                                "name": university_obj.university_name,
                                "city": university_obj.university_city,
                                "state": university_obj.university_state,
                                "pincode": university_obj.university_pincode,
                                "logo": request.build_absolute_uri(university_obj.university_logo.url) if university_obj.university_logo else None
                            }

                        exam_details.append({
                            "id": exam.id,
                            "exam_id": examination.id,
                            "examstartdate": exam.examstartdate,
                            "examenddate": exam.examenddate,
                            "examstarttime": exam.examstarttime,
                            "examendtime": exam.examendtime
                        })

                        examination_data.append({
                            "id": examination.id,
                            "course_id": examination.course.id,
                            "stream_id": examination.stream.id,
                            "subject_id": examination.subject.id,
                            "studypattern": examination.studypattern,
                            "semyear": examination.semyear,
                            "substream_id": examination.substream.id if examination.substream else None,
                            "course_name": examination.course.name,
                            "stream_name": examination.stream.name,
                            "subject_name": examination.subject.name,
                            "substream_name": examination.substream.name if examination.substream else None
                        })

                        # ExamSession time left
                        exam_session = ExamSession.objects.filter(student=student, exam=examination).first()
                        time_left_ms = exam_session.time_left_ms if exam_session else None

                        # Submitted answers
                        answered_questions_qs = SubmittedExamination.objects.filter(student=student, exam=examination)
                        answered_questions = list(answered_questions_qs.values_list('question', flat=True))

                        # Last answered question id
                        last_answered_question_id = None
                        if answered_questions:
                            try:
                                last_answered_question_id = max(int(q) for q in answered_questions if q.isdigit())
                            except Exception:
                                last_answered_question_id = None

                        exam_progress.append({
                            "exam_id": examination.id,
                            "time_left_ms": time_left_ms,
                            "answered_questions": answered_questions,
                            "last_answered_question_id": last_answered_question_id,
                        })

                    except Examination.DoesNotExist:
                        logger.error(f"Examination for exam_id {exam.id} does not exist. Skipping.")
                        continue

                # Fetch all Result records for this student filtered by student
                result_qs = Result.objects.filter(student=student)

                result_data = []
                for res in result_qs:
                    result_data.append({
                        "id": res.id,
                        "student_id": res.student.id,  # Added student_id here
                        "exam_id": res.exam.id if res.exam else None,
                        "total_question": res.total_question,
                        "attempted": res.attempted,
                        "total_marks": res.total_marks,
                        "score": res.score,
                        "result": res.result,
                        "percentage": res.percentage,
                        "examdetails_id": res.examdetails.id if res.examdetails else None,
                    })

                return Response({
                    "message": "Login Successful",
                    "is_student": True,
                    "token": token,
                    "student_id": student.id,
                    "student_name": student.name,
                    "university_logo": university["logo"] if university else None,
                    "exam_details": exam_details,
                    "examination_data": examination_data,
                    "university": university,
                    "exam_progress": exam_progress,
                    "result_data": result_data,  # Included result data here
                }, status=status.HTTP_200_OK)

            else:
                return Response({
                    "message": "Login Successful",
                    "token": token,
                    "is_student": False,
                    "email": user.email,
                    "is_active": user.is_active,
                    "user_id": user.id,
                }, status=status.HTTP_200_OK)

        else:
            return Response({"error": "Invalid Credentials"}, status=status.HTTP_400_BAD_REQUEST)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



  
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def add_university(request):
    user = request.user

    if request.method == 'GET':
        universities = University.objects.all()
        serializer = UniversitySerializer(universities, many=True)
        #logger.info(f"[{user.email}] fetched universities list.")
        return Response(serializer.data, status=status.HTTP_200_OK)

    if not (user.is_superuser or getattr(user, 'is_data_entry', False)):
        logger.warning(f"[{user.email}] Unauthorized university creation attempt.")
        return Response({"message": "You do not have permission to add a university."}, status=status.HTTP_403_FORBIDDEN)

    university_name = request.data.get("university_name", "").strip()
    university_address = request.data.get("university_address", "").strip()

    if not university_name or not university_address:
        logger.warning(f"[{user.email}] Missing university fields.")
        return Response({"message": "University name and address are required."}, status=status.HTTP_400_BAD_REQUEST)

    if University.objects.filter(university_name__iexact=university_name).exists():
        logger.warning(f"[{user.email}] Tried adding duplicate university: {university_name}")
        return Response({"message": "University name already exists."}, status=status.HTTP_406_NOT_ACCEPTABLE)

    lowercase = university_name.lower().replace(' ', '')
    today = str(date.today()).replace('-', '')
    registration_id = f"{lowercase}{today}UNI"

    data = {
        'university_name': university_name,
        'university_address': university_address,
        'university_city': request.data.get('university_city', ''),
        'university_state': request.data.get('university_state', ''),
        'university_pincode': request.data.get('university_pincode', ''),
        'registrationID': registration_id,
        'university_logo': request.FILES.get('university_logo', None)
    }

    serializer = UniversitySerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        #logger.info(f"[{user.email}] Successfully added university: {university_name}")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    logger.error(f"[{user.email}] Error validating university data: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def university_detail(request, university_id):
    user = request.user
    try:
        university = University.objects.get(pk=university_id)
    except University.DoesNotExist:
        logger.warning(f"[{user.email}] Tried accessing nonexistent university ID: {university_id}")
        return Response({"message": "University not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = UniversitySerializer(university)
        #logger.info(f"[{user.email}] viewed university ID: {university_id}")
        return Response(serializer.data)

    if request.method == 'PUT':
        if not (user.is_superuser or getattr(user, 'is_data_entry', False)):
            logger.warning(f"[{user.email}] Unauthorized university update attempt.")
            return Response({"message": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        serializer = UniversitySerializer(university, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            #logger.info(f"[{user.email}] updated university ID: {university_id}")
            return Response(serializer.data)
        logger.error(f"[{user.email}] error updating university ID: {university_id} | {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        if not user.is_superuser:
            logger.warning(f"[{user.email}] Unauthorized university deletion attempt.")
            return Response({"message": "Only superusers can delete universities."}, status=status.HTTP_403_FORBIDDEN)
        university.delete()
        #logger.info(f"[{user.email}] deleted university ID: {university_id}")
        return Response({"message": "University deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
  

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def create_user(request):
    user = request.user
    
    # Handle GET request
    if request.method == "GET":
        if user.is_superuser:
            users = User.objects.filter(Q(is_fee_clerk=True) | Q(is_data_entry=True))
            serializer = UserSerializers(users, many=True)
            #logger.info(f"[{user.email}] fetched all fee clerk and data entry users.")
            return Response(serializer.data, status=status.HTTP_200_OK)
        logger.warning(f"[{user.email}] attempted to fetch users without permission.")
        return Response({"message": "You do not have permission to view users."}, status=status.HTTP_403_FORBIDDEN)

    # Handle POST request
    if request.method == "POST":
        if user.is_superuser:
            # If is_superuser is not provided, default to False
            serializer = UserSerializers(data=request.data)
            request.data['is_superuser'] = True

            if serializer.is_valid():
                new_user = serializer.save()
                #logger.info(f"[{user.email}] created new user with email: {new_user.email}")
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            logger.error(f"[{user.email}] user creation failed: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        logger.warning(f"[{user.email}] attempted to create user without permission.")
        return Response({"message": "You do not have permission to create users."}, status=status.HTTP_403_FORBIDDEN)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_semester_fees(request):
    # if not request.user.is_superuser:
    #     logger.warning(f"Unauthorized access attempt by user {request.user.email}")
    #     return Response({"message": "You don't have permission to add"}, status=status.HTTP_403_FORBIDDEN)

    try:
        data = request.data if isinstance(request.data, list) else [request.data]
        responses = []

        for entry in data:
            try:
                required_fields = ['university_id', 'course_id', 'substream_id', 'sem']
                if not all(entry.get(field) for field in required_fields):
                    responses.append({"error": "Missing required fields."})
                    continue

                university = University.objects.filter(id=entry['university_id']).first()
                course = Course.objects.filter(id=entry['course_id'], university=university).first() if university else None
                stream = Stream.objects.filter(id=entry['stream_id'], course=course).first() if course else None
                substream = SubStream.objects.filter(id=entry['substream_id'], stream=stream).first() if stream else None

                if not all([university, course, stream, substream]):
                    responses.append({"error": "Invalid university, course, stream or substream reference."})
                    continue

                sem = entry['sem']
                if int(sem) > 8:
                    responses.append({"error": f"Semester '{sem}' cannot be more than 8"})
                    continue

                if SemesterFees.objects.filter(stream=stream, substream=substream, sem=sem).exists():
                    responses.append({"error": f"Semester fee for sem {sem} already exists"})
                    continue

                # Fee calculation
                tutionfees = float(entry.get('tutionfees', 0))
                examinationfees = float(entry.get('examinationfees', 0))
                bookfees = float(entry.get('bookfees', 0))
                resittingfees = float(entry.get('resittingfees', 0))
                entrancefees = float(entry.get('entrancefees', 0))
                extrafees = float(entry.get('extrafees', 0))
                discount = float(entry.get('discount', 0))

                totalfees = (
                    tutionfees + examinationfees + bookfees +
                    resittingfees + entrancefees + extrafees - discount
                )

                semester_fee = SemesterFees.objects.create(
                    stream=stream,
                    substream=substream,
                    sem=sem,
                    tutionfees=tutionfees,
                    examinationfees=examinationfees,
                    bookfees=bookfees,
                    resittingfees=resittingfees,
                    entrancefees=entrancefees,
                    extrafees=extrafees,
                    discount=discount,
                    totalfees=totalfees,
                    created_by=request.user.id,
                    modified_by=request.user.id
                )

                responses.append({
                    "message": f"Semester fee created for Semester {sem}",
                    "data": {
                        "stream": stream.id,
                        "substream": substream.id,
                        "sem": sem,
                        "totalfees": semester_fee.totalfees,
                    },
                })
            except Exception as entry_err:
                logger.error(f"Entry error: {entry_err}")
                responses.append({"error": str(entry_err)})

        return Response(responses, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Internal server error: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_year_fees(request):
    # if not request.user.is_superuser:
    #     logger.warning(f"Unauthorized access attempt by user {request.user.email}")
    #     return Response({"message": "You don't have permission to add"}, status=status.HTTP_403_FORBIDDEN)

    try:
        data = request.data if isinstance(request.data, list) else [request.data]
        responses = []

        for entry in data:
            try:
                required_fields = ['university_id', 'course_id', 'substream_id', 'year']
                if not all(entry.get(field) for field in required_fields):
                    responses.append({"error": "Missing required fields."})
                    continue

                university = University.objects.filter(id=entry['university_id']).first()
                course = Course.objects.filter(id=entry['course_id'], university=university).first() if university else None
                stream = Stream.objects.filter(id=entry['stream_id'], course=course).first() if course else None
                substream = SubStream.objects.filter(id=entry['substream_id'], stream=stream).first() if stream else None

                if not all([university, course, stream, substream]):
                    responses.append({"error": "Invalid university, course, stream or substream reference."})
                    continue

                year = entry['year']
                if int(year) > 4:
                    responses.append({"error": f"Year '{year}' cannot be greater than 4"})
                    continue

                if YearFees.objects.filter(stream=stream, substream=substream, year=year).exists():
                    responses.append({"error": f"Year fee for year {year} already exists"})
                    continue

                # Fee calculation
                totalfees = sum([
                    float(entry.get('tutionfees', 0)),
                    float(entry.get('examinationfees', 0)),
                    float(entry.get('bookfees', 0)),
                    float(entry.get('resittingfees', 0)),
                    float(entry.get('entrancefees', 0)),
                    float(entry.get('extrafees', 0)),
                ]) - float(entry.get('discount', 0))

                year_fee = YearFees.objects.create(
                    stream=stream,
                    substream=substream,
                    year=year,
                    tutionfees=entry.get('tutionfees', 0),
                    examinationfees=entry.get('examinationfees', 0),
                    bookfees=entry.get('bookfees', 0),
                    resittingfees=entry.get('resittingfees', 0),
                    entrancefees=entry.get('entrancefees', 0),
                    extrafees=entry.get('extrafees', 0),
                    discount=entry.get('discount', 0),
                    totalfees=totalfees,
                    created_by=request.user.id,
                    modified_by=request.user.id
                )

                responses.append({
                    "message": f"Year fee created for Year {year}",
                    "data": {
                        "stream": stream.id,
                        "substream": substream.id,
                        "year": year_fee.year,
                        "totalfees": year_fee.totalfees,
                    },
                })
            except Exception as entry_err:
                logger.error(f"Entry error: {entry_err}")
                responses.append({"error": str(entry_err)})

        return Response(responses, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Internal server error: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def payment_modes(request):
    # if not request.user.is_superuser:
    #     logger.warning(f"[{request.user.email}] Unauthorized access to payment_modes.")
    #     return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        modes = PaymentModes.objects.all()
        serializer = PaymentModesSerializer(modes, many=True)
        #logger.info("Payment modes list fetched.")
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = PaymentModesSerializer(data=request.data)
        if serializer.is_valid():
            name = serializer.validated_data.get('name')
            if PaymentModes.objects.filter(name__iexact=name).exists():
                logger.warning(f"Duplicate payment mode attempted: {name}")
                return Response({"error": "Mode already exists"}, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            #logger.info(f"Payment mode created: {name}")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.error(f"Payment mode creation failed: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def payment_mode_detail(request, id):
    try:
        mode = PaymentModes.objects.get(id=id)
    except PaymentModes.DoesNotExist:
        logger.warning(f"Payment mode with ID {id} not found.")
        return Response({"error": "PaymentMode not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = PaymentModesSerializer(mode)
        return Response(serializer.data)

    if request.method == 'PUT':
        serializer = PaymentModesSerializer(mode, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            #logger.info(f"Updated payment mode ID {id}.")
            return Response(serializer.data)
        logger.error(f"Failed to update payment mode ID {id}: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        mode.delete()
        #logger.info(f"Deleted payment mode ID {id}.")
        return Response({"message": "Payment mode deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def fee_receipt_options(request):
    # if not request.user.is_superuser:
    #     logger.warning(f"[{request.user.email}] Unauthorized access to fee_receipt_options.")
    #     return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        options = FeeReceiptOptions.objects.all()
        serializer = FeeReceiptOptionsSerializer(options, many=True)
        #logger.info("Fetched all fee receipt options.")
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = FeeReceiptOptionsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            #logger.info("Created new fee receipt option.")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.error(f"Failed to create fee receipt option: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def fee_receipt_option_detail(request, id):
    # if not request.user.is_superuser:
    #     logger.warning(f"[{request.user.email}] Unauthorized access to fee_receipt_option_detail.")
    #     return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    try:
        option = FeeReceiptOptions.objects.get(id=id)
    except FeeReceiptOptions.DoesNotExist:
        return Response({"error": "FeeReceiptOption not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = FeeReceiptOptionsSerializer(option)
        return Response(serializer.data)

    if request.method == 'PUT':
        serializer = FeeReceiptOptionsSerializer(option, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            #logger.info(f"Updated fee receipt option ID {id}.")
            return Response(serializer.data)
        logger.error(f"Failed to update fee receipt option ID {id}: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        option.delete()
        #logger.info(f"Deleted fee receipt option ID {id}.")
        return Response({"message": "Fee receipt option deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def bank_names(request):
    # if not request.user.is_superuser:
    #     logger.warning(f"[{request.user.email}] Unauthorized access to bank_names.")
    #     return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        banks = BankNames.objects.all()
        serializer = BankNamesSerializer(banks, many=True)
        #logger.info("Fetched all bank names.")
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = BankNamesSerializer(data=request.data)
        if serializer.is_valid():
            name = serializer.validated_data.get('name')
            if BankNames.objects.filter(name__iexact=name).exists():
                logger.warning(f"Attempted to create duplicate bank name: {name}")
                return Response({"error": "Bank name already exists"}, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            #logger.info(f"Created new bank name: {name}")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.error(f"Failed to create bank name: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
  
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def bank_name_detail(request, id):
    try:
        bank = BankNames.objects.get(id=id)
    except BankNames.DoesNotExist:
        logger.warning(f"Bank name with ID {id} not found.")
        return Response({"error": "BankName not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = BankNamesSerializer(bank)
        return Response(serializer.data)

    if request.method == 'PUT':
        serializer = BankNamesSerializer(bank, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            #logger.info(f"Updated bank name ID {id}.")
            return Response(serializer.data)
        logger.error(f"Failed to update bank name ID {id}: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        bank.delete()
        #logger.info(f"Deleted bank name ID {id}.")
        return Response({"message": "Bank name deleted successfully"}, status=status.HTTP_204_NO_CONTENT)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def session_names(request):
    """
    Handles GET and POST for SessionNames.
    Only superusers can perform these operations.
    """
    # if not request.user.is_superuser:
    #     logger.warning(f"[{request.user.email}] Unauthorized access to session_names.")
    #     return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        sessions = SessionNames.objects.all()
        serializer = SessionNamesSerializer(sessions, many=True)
        #logger.info("Fetched all session names.")
        return Response(serializer.data, status=status.HTTP_200_OK)

    if request.method == 'POST':
        serializer = SessionNamesSerializer(data=request.data)
        if serializer.is_valid():
            if SessionNames.objects.filter(name__iexact=serializer.validated_data['name']).exists():
                logger.warning("Attempt to create duplicate session name.")
                return Response({"error": "Session name already exists"}, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            #logger.info(f"Session name created: {serializer.validated_data['name']}")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        logger.error(f"Failed to create session name: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def session_name_detail(request, id):
    """
    Handles GET, PUT, and DELETE for a specific SessionName by ID.
    Only superusers can perform these operations.
    """
    # if not request.user.is_superuser:
    #     logger.warning(f"[{request.user.email}] Unauthorized access to session_name_detail.")
    #     return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    try:
        session = SessionNames.objects.get(id=id)
    except SessionNames.DoesNotExist:
        logger.warning(f"Session name ID {id} not found.")
        return Response({"error": "Session name not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = SessionNamesSerializer(session)
        return Response(serializer.data, status=status.HTTP_200_OK)

    if request.method == 'PUT':
        serializer = SessionNamesSerializer(session, data=request.data, partial=True)
        if serializer.is_valid():
            if 'name' in serializer.validated_data and SessionNames.objects.filter(
                name__iexact=serializer.validated_data['name']
            ).exclude(id=session.id).exists():
                logger.warning(f"Duplicate session name attempted in update: {serializer.validated_data['name']}")
                return Response({"error": "Session name already exists"}, status=status.HTTP_400_BAD_REQUEST)

            serializer.save()
            #logger.info(f"Updated session name ID {id}")
            return Response(serializer.data, status=status.HTTP_200_OK)
        logger.error(f"Failed to update session name ID {id}: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        session.delete()
        #logger.info(f"Deleted session name ID {id}")
        return Response({"message": "Session name deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user

    if not user.is_authenticated:
        logger.warning("Unauthorized password change attempt.")
        return Response(
            {"error": "Authentication required. User not found or anonymous."},
            status=status.HTTP_401_UNAUTHORIZED
        )

    serializer = ChangePasswordSerializer(data=request.data)

    if serializer.is_valid():
        user.set_password(serializer.validated_data['password'])
        user.save()
        #logger.info(f"[{user.email}] changed their password.")
        return Response({"message": "Password changed successfully."}, status=status.HTTP_200_OK)

    logger.error(f"[{user.email}] failed password change: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_courses_by_university(request):
    university_name = request.query_params.get('university')

    if not university_name:
        logger.warning("Missing university parameter in get_courses_by_university.")
        return Response({"error": "University name is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        university = University.objects.get(university_name=university_name)
    except University.DoesNotExist:
        logger.warning(f"University not found: {university_name}")
        return Response({"error": "University not found."}, status=status.HTTP_404_NOT_FOUND)

    courses = university.course_set.all()
    course_names = [course.name for course in courses]

    #logger.info(f"Courses fetched for university: {university_name}")
    return Response({
        "university": university_name,
        "courses": course_names
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stream_by_course_one(request):
    course_name = request.query_params.get('course')
    university_name = request.query_params.get('university')

    if not course_name or not university_name:
        logger.warning("Missing course or university in request.")
        return Response({"error": "Both 'course_name' and 'university_name' are required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        university = University.objects.get(university_name=university_name)
    except University.DoesNotExist:
        logger.error("University not found: '%s'", university_name)
        return Response({"error": f"University '{university_name}' not found."}, status=status.HTTP_404_NOT_FOUND)

    try:
        course = Course.objects.get(name=course_name, university=university)
    except Course.DoesNotExist:
        logger.error("Course '%s' not found for university '%s'", course_name, university_name)
        return Response({"error": f"Course '{course_name}' not found for university '{university_name}'."}, status=status.HTTP_404_NOT_FOUND)

    try:
        streams = Stream.objects.filter(course=course)
        stream_list = [
            {
                "stream_id": stream.id,
                "stream_name": stream.name,
                "semester": stream.sem,
                "year": stream.year
            }
            for stream in streams
        ]

        #logger.info("Streams fetched for course '%s' at university '%s'", course.name, university.university_name)

        return Response({
            "university_name": university.university_name,
            "course_name": course.name,
            "course_id": course.id,
            "streams": stream_list
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error("Exception occurred while fetching streams: %s", str(e))
        return Response({"error": "Unable to fetch streams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
  
  
from django.contrib.auth.hashers import make_password
from django.shortcuts import get_object_or_404


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_substreams_by_university_course_stream(request):
    university_name = request.query_params.get('university')
    course_name = request.query_params.get('course')
    stream_name = request.query_params.get('stream')

    if not university_name or not course_name or not stream_name:
        logger.warning("Missing one or more required parameters: university, course, stream.")
        return Response({"error": "University name, course name, and stream name are required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        university = University.objects.get(university_name=university_name)
    except University.DoesNotExist:
        logger.error("University not found: '%s'", university_name)
        return Response({"error": "University not found."}, status=status.HTTP_404_NOT_FOUND)

    try:
        course = Course.objects.get(university=university, name=course_name)
    except Course.DoesNotExist:
        logger.error("Course not found: '%s' under university '%s'", course_name, university_name)
        return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

    try:
        stream = Stream.objects.get(course=course, name=stream_name)
    except Stream.DoesNotExist:
        logger.error("Stream not found: '%s' in course '%s'", stream_name, course_name)
        return Response({"error": "Stream not found."}, status=status.HTTP_404_NOT_FOUND)

    substreams = SubStream.objects.filter(stream=stream)
    substream_names = [substream.name for substream in substreams]

    #logger.info("Fetched %d substreams for stream '%s'", len(substream_names), stream.name)
    return Response(substream_names, status=status.HTTP_200_OK)
        
from django.contrib.auth.hashers import make_password

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_by_enrollment_id(request):
    enrollment_id = request.query_params.get('enrollment_id')

    if not enrollment_id:
        logger.warning("Enrollment ID not provided in query parameters.")
        return Response({"error": "Enrollment ID is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student = Student.objects.get(enrollment_id=enrollment_id)
        serializer = StudentSearchSerializer(student)
        #logger.info(f"Student found with enrollment_id: {enrollment_id}")
        return Response({"data": serializer.data}, status=status.HTTP_200_OK)
    except Student.DoesNotExist:
        logger.warning(f"No student found for enrollment_id: {enrollment_id}")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_by_student_name(request):
    """
    Search for students by their name using a case-insensitive substring match.
    """
    name_query = request.query_params.get('name', '').strip()

    if not name_query:
        logger.warning("No name provided in student name search.")
        return Response({"error": "Name query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

    students = Student.objects.filter(name__icontains=name_query)

    if students.exists():
        serializer = StudentSearchSerializer(students, many=True)
        #logger.info(f"Found {students.count()} student(s) matching: '{name_query}'")
        return Response(serializer.data, status=status.HTTP_200_OK)
    else:
        #logger.info(f"No students found for name search: '{name_query}'")
        return Response({"message": "No students found."}, status=status.HTTP_404_NOT_FOUND)
  


@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def create_course(request):
    """
    POST: Create a new course for a given university.
    GET: Retrieve all courses or filter by university name.
    """
    university_name = (
        request.data.get('university_name') if request.method == 'POST'
        else request.query_params.get('university_name')
    )

    if request.method == 'POST':
        course_name = request.data.get('name')

        if not university_name or not course_name:
            return Response(
                {'error': 'Both university_name and name are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            university = University.objects.get(university_name=university_name)
        except University.DoesNotExist:
            return Response({'error': 'University not found.'}, status=status.HTTP_404_NOT_FOUND)

        if Course.objects.filter(university=university, name=course_name).exists():
            return Response(
                {'error': 'Course with this name already exists for the specified university.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = CourseSerializerTwo(data=request.data)
        if serializer.is_valid():
            serializer.save(university=university, created_by=request.user, modified_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # GET Request
    if university_name:
        try:
            university = University.objects.get(university_name=university_name)
            courses = Course.objects.filter(university=university)
        except University.DoesNotExist:
            return Response({'error': 'University not found.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        courses = Course.objects.all()

    serializer = CourseSerializerTwo(courses, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)
  
  
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_stream(request):
    """
    Create a new stream for a given course and university.
    """
    university_name = request.data.get('university_name')
    course_name = request.data.get('course_name')
    stream_name = request.data.get('stream_name')
    year = request.data.get('year') or datetime.now().year
    sem = request.data.get('sem')

    # Validate required fields
    if not university_name or not course_name or not stream_name:
        return Response(
            {'error': 'university_name, course_name, and stream_name are required.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not sem:
        return Response({'error': 'Semester is required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Validate University
    try:
        university = University.objects.get(university_name=university_name)
    except University.DoesNotExist:
        return Response({'error': 'University not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Validate Course
    try:
        course = Course.objects.get(name=course_name, university=university)
    except Course.DoesNotExist:
        return Response(
            {'error': 'Course not found for the specified university.'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Check for existing Stream
    if Stream.objects.filter(course=course, name=stream_name).exists():
        return Response(
            {'error': 'Stream with this name already exists for the specified course.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Prepare data for serializer
    data = request.data.copy()
    data['name'] = stream_name
    data['year'] = year

    serializer = StreamSerializer(data=data)
    if serializer.is_valid():
        serializer.save(course=course, created_by=request.user, modified_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
  
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_sub_stream(request):
    """
    Create a new substream for a specific stream, course, and university.
    """
    university_name = request.data.get('university_name')
    course_name = request.data.get('course_name')
    stream_name = request.data.get('stream_name')
    substream_name = request.data.get('substream_name')

    # Validate required fields
    missing_fields = []
    if not university_name: missing_fields.append('university_name')
    if not course_name: missing_fields.append('course_name')
    if not stream_name: missing_fields.append('stream_name')
    if not substream_name: missing_fields.append('substream_name')

    if missing_fields:
        return Response(
            {'error': f"The following fields are required: {', '.join(missing_fields)}"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        university = University.objects.get(university_name=university_name)
    except University.DoesNotExist:
        return Response({'error': 'University not found.'}, status=status.HTTP_404_NOT_FOUND)

    try:
        course = Course.objects.get(name=course_name, university=university)
    except Course.DoesNotExist:
        return Response({'error': 'Course not found for the specified university.'}, status=status.HTTP_404_NOT_FOUND)

    try:
        stream = Stream.objects.get(name=stream_name, course=course)
    except Stream.DoesNotExist:
        return Response({'error': 'Stream not found for the specified course.'}, status=status.HTTP_404_NOT_FOUND)

    if SubStream.objects.filter(stream=stream, name=substream_name).exists():
        return Response(
            {'error': 'SubStream with this name already exists for the specified stream.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    data = request.data.copy()
    data['name'] = substream_name

    serializer = SubStreamSerializer(data=data)
    if serializer.is_valid():
        serializer.save(stream=stream)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
  

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_course_details(request, student_id):
    """
    Retrieve a student's course details including university, course, stream, substream, and enrollment details.
    """
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} not found.")
        return Response({"error": f"Student with ID {student_id} not found."}, status=status.HTTP_404_NOT_FOUND)

    try:
        enrollment = Enrolled.objects.get(student=student)
    except Enrolled.DoesNotExist:
        logger.error(f"Enrollment not found for student ID {student_id}.")
        return Response({"error": "Enrollment details not found for the specified student."}, status=status.HTTP_404_NOT_FOUND)

    try:
        data = {
            "university_name": enrollment.course.university.university_name if enrollment.course and enrollment.course.university else None,
            "course_name": enrollment.course.name if enrollment.course else None,
            "stream_name": enrollment.stream.name if enrollment.stream else None,
            "substream_name": enrollment.substream.name if enrollment.substream else None,
            "study_pattern": enrollment.course_pattern or "",
            "session": enrollment.session or "",
            "semester": enrollment.current_semyear or "",
            "course_duration": enrollment.stream.sem if enrollment.stream else "",
        }

        return Response(data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Unexpected error while processing student ID {student_id}: {str(e)}")
        return Response({"error": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_student_course_details(request, student_id):
    """
    Update all fields for a student's course details.
    Requires all fields in the request data.
    """
    #logger.info("Update Student Course Details API called for student ID: %s with data: %s", student_id, request.data)

    # Validate if the student has an enrollment
    try:
        enrollment = Enrolled.objects.select_related(
            'course__university', 'stream', 'substream'
        ).get(student_id=student_id)
    except Enrolled.DoesNotExist:
        logger.error("Enrollment details not found for student ID: %s", student_id)
        return Response({"error": "Enrollment details not found for the specified student."}, status=status.HTTP_404_NOT_FOUND)

    # Extract and validate the input data
    required_fields = [
        "university_name", "course_name", "stream_name", "substream_name", 
        "study_pattern", "session", "semister", "course_duration"
    ]
    missing_fields = [field for field in required_fields if field not in request.data]
    
    if missing_fields:
        logger.error("Missing required fields: %s", missing_fields)
        return Response({"error": f"Missing required fields: {', '.join(missing_fields)}"}, status=status.HTTP_400_BAD_REQUEST)

    data = request.data

    # Validate university
    try:
        university = University.objects.get(university_name=data["university_name"])
    except University.DoesNotExist:
        logger.error("University '%s' not found.", data["university_name"])
        return Response({"error": "University not found."}, status=status.HTTP_404_NOT_FOUND)

    # Validate course
    try:
        course = Course.objects.get(name=data["course_name"], university=university)
    except Course.DoesNotExist:
        logger.error("Course '%s' not found for university '%s'.", data["course_name"], data["university_name"])
        return Response({"error": "Course not found for the specified university."}, status=status.HTTP_404_NOT_FOUND)

    # Validate stream
    try:
        stream = Stream.objects.get(name=data["stream_name"], course=course)
    except Stream.DoesNotExist:
        logger.error("Stream '%s' not found for course '%s'.", data["stream_name"], data["course_name"])
        return Response({"error": "Stream not found for the specified course."}, status=status.HTTP_404_NOT_FOUND)

    # Validate substream (mandatory)
    if not data["substream_name"]:
        logger.error("Substream name is required.")
        return Response({"error": "Substream name is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        substream = SubStream.objects.get(name=data["substream_name"], stream=stream)
    except SubStream.DoesNotExist:
        logger.error("Substream '%s' not found for stream '%s'.", data["substream_name"], data["stream_name"])
        return Response({"error": "Substream not found for the specified stream."}, status=status.HTTP_404_NOT_FOUND)

    # Update enrollment details
    enrollment.course = course
    enrollment.stream = stream
    enrollment.substream = substream
    enrollment.course_pattern = data["study_pattern"]
    enrollment.session = data["session"]
    enrollment.current_semyear = data["semister"]
    enrollment.stream.sem = data["course_duration"]  # Assuming this updates `course_duration` dynamically

    enrollment.save()

    #logger.info("Enrollment updated successfully for student ID: %s", student_id)

    # Prepare response data
    response_data = {
        "university_name": university.university_name,
        "course_name": course.name,
        "stream_name": stream.name,
        "substream_name": substream.name if substream else None,
        "study_pattern": enrollment.course_pattern,
        "session": enrollment.session,
        "semister": enrollment.current_semyear,
        "course_duration": enrollment.stream.sem,
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def universities_with_courses(request):
    try:
        # Fetch all universities
        universities = University.objects.all()

        # Manually match courses based on university_id
        result = {}
        for university in universities:
            # Fetch courses with id, name, and year for the current university
            courses = Course.objects.filter(university_id=university.id).values('id', 'name', 'year')

            # Format the course data as a list of dictionaries
            formatted_courses = []
            for course in courses:
                formatted_courses.append({
                    "id": course['id'],
                    "name":course['name'],
                    "year": course['year']
                })
            
            # Add to the result dictionary
            result[university.university_name] = formatted_courses

        #logger.info("Fetched universities and their courses successfully. Total universities: %d", len(universities))
        return Response(result, status=status.HTTP_200_OK)
    
    except Exception as e:
        logger.error("Error fetching universities and courses. Exception: %s", str(e))
        return Response({"message": "An error occurred while fetching data."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_course(request, course_id):
    try:
        # Fetch the course to update
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            logger.error("Update failed: Course with ID %d not found.", course_id)
            return Response({"message": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get data from the request
        data = request.data
        course_name = data.get('name', None)
        course_year = data.get('year', None)

        # Validate input
        if not course_name or not course_year:
            logger.warning("Invalid input for updating course ID %d. Data received: %s", course_id, data)
            return Response(
                {"message": "Both 'name' and 'year' fields are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update fields
        course.name = course_name
        course.year = course_year
        course.save()

        #logger.info("Course updated successfully. ID: %d, Name: %s, Year: %s", course_id, course_name, course_year)
        return Response(
            {
                "message": "Course updated successfully.",
                "course": {
                    "id": course.id,
                    "name": course.name,
                    "year": course.year
                }
            },
            status=status.HTTP_200_OK
        )

    except Exception as e:
        logger.error("Error updating course ID %d. Exception: %s", course_id, str(e))
        return Response({"message": "An error occurred while updating the course."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stream_by_course_two(request, course_id):
    #logger.info("Request to fetch streams for course ID %d received.", course_id)
    
    try:
        # Fetch the course object
        course = Course.objects.get(id=course_id)
        
        # Fetch streams related to this course
        streams = Stream.objects.filter(course=course)
        
        # Prepare the response data
        stream_list = []
        for stream in streams:
            stream_data = {
                "stream_name": stream.name,
                "semester": stream.sem,
                "year": stream.year
            }
            stream_list.append(stream_data)

        #logger.info("Successfully fetched %d streams for course '%s' (ID: %d).", len(stream_list), course.name, course_id)
        
        return Response({"course_name": course.name, "streams": stream_list}, status=status.HTTP_200_OK)

    except Course.DoesNotExist:
        logger.error("Stream fetch failed: Course with ID %d not found.", course_id)
        return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        logger.error("Error fetching streams for course ID %d. Exception: %s", course_id, str(e))
        return Response({"error": "Unable to fetch streams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_streams_by_course(request, course_id):
    #logger.info("Request to update streams for course ID %d received.", course_id)
    
    try:
        # Validate if course exists
        course = Course.objects.get(id=course_id)
        data = request.data

        if not isinstance(data, list):
            logger.warning("Invalid request format for updating streams. Expected a list, got: %s", type(data))
            return Response({"error": "Request data must be a list of streams."}, status=status.HTTP_400_BAD_REQUEST)

        updated_streams = []
        for stream_data in data:
            stream_id = stream_data.get('id')
            stream_name = stream_data.get('name')
            year = stream_data.get('year')

            if not stream_id:
                logger.error("Stream ID is missing in the update request for course ID %d.", course_id)
                return Response({"error": "Stream ID is required for updates."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                # Fetch the stream
                stream = Stream.objects.get(id=stream_id, course=course)

                # Update the fields
                if stream_name is not None:
                    stream.name = stream_name
                if year is not None:
                    stream.year = year
        
                stream.save()
                #logger.info("Stream updated: ID %d, Name '%s'", stream.id, stream.name)
                
                updated_streams.append({
                    "id": stream.id,
                    "name": stream.name,
                    "sem": stream.sem,
                    "year": stream.year,
                })

            except Stream.DoesNotExist:
                logger.error("Stream update failed: Stream ID %d not found for course '%s' (ID: %d).", stream_id, course.name, course_id)
                return Response({"error": f"Stream with ID {stream_id} not found for course {course.name}."}, status=status.HTTP_404_NOT_FOUND)

        #logger.info("Successfully updated %d streams for course '%s' (ID: %d).", len(updated_streams), course.name, course_id)
        return Response({
            "message": "Streams updated successfully.",
            "course_name": course.name,
            "updated_streams": updated_streams
        }, status=status.HTTP_200_OK)

    except Course.DoesNotExist:
        logger.error("Stream update failed: Course ID %d not found.", course_id)
        return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        logger.error("Error updating streams for course ID %d. Exception: %s", course_id, str(e))
        return Response({"error": "Unable to update streams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_substreams_by_stream(request, stream_id):
    """
    Update single or multiple substreams associated with a specific stream.
    """
    #logger.info("Request to update substreams for stream ID %d received.", stream_id)
    
    try:
        # Validate if the stream exists
        stream = Stream.objects.get(id=stream_id)
        data = request.data

        if not isinstance(data, list):
            logger.warning("Invalid request format for updating substreams. Expected a list, got: %s", type(data))
            return Response({"error": "Request data must be a list of substreams."}, status=status.HTTP_400_BAD_REQUEST)

        updated_substreams = []
        for substream_data in data:
            substream_id = substream_data.get('id')
            substream_name = substream_data.get('name')

            if not substream_id:
                logger.error("Substream ID is missing in the update request for stream ID %d.", stream_id)
                return Response({"error": "Substream ID is required for updates."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                # Fetch the substream
                substream = SubStream.objects.get(id=substream_id, stream=stream)

                # Update the name
                if substream_name:
                    substream.name = substream_name
                    substream.save()

                #logger.info("Substream updated: ID %d, Name '%s'", substream.id, substream.name)
                updated_substreams.append({
                    "id": substream.id,
                    "name": substream.name,
                    "stream_id": substream.stream.id,
                })

            except SubStream.DoesNotExist:
                logger.error(
                    "Substream update failed: Substream ID %d not found for stream '%s' (ID: %d).",
                    substream_id,
                    stream.name,
                    stream_id
                )
                return Response(
                    {"error": f"SubStream with ID {substream_id} not found for stream {stream.name}."},
                    status=status.HTTP_404_NOT_FOUND
                )

        #logger.info("Successfully updated %d substreams for stream '%s' (ID: %d).",len(updated_substreams),stream.name,stream_id)
        return Response({
            "message": "Substreams updated successfully.",
            "stream_name": stream.name,
            "updated_substreams": updated_substreams
        }, status=status.HTTP_200_OK)

    except Stream.DoesNotExist:
        logger.error("Substream update failed: Stream ID %d not found.", stream_id)
        return Response({"error": "Stream not found."}, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        logger.error("Error updating substreams for stream ID %d. Exception: %s", stream_id, str(e))
        return Response({"error": "Unable to update substreams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @api_view(['POST'])
# def quick_registration(request):
#   print('this function call')
#   # import pdb
#   # pdb.set_trace()
#   if not (request.user.is_superuser or request.user.is_data_entry):
#       logger.warning(f"Unauthorized access attempt by user {request.user.id}")
#       return Response({'message': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)

#   data = request.data
#   logger.debug(f"Request Data: {request.data}")

#   # Required Fields
#   required_fields = ['university', 'student_name', 'father_name', 'dob', 'email', 'mobile_number']
#   for field in required_fields:
#       if not data.get(field):
#           logger.error(f"Missing required field: {field}")
#           return Response({'message': f"'{field}' is required."}, status=status.HTTP_400_BAD_REQUEST)

#   try:
#       university_id = data.get('university')
#       student_email = data.get('email')
#       mobile_number = data.get('mobile_number')
#       payment_mode = data.get('payment_mode')
#       fees = data.get('fees')

#       # Fetch ForeignKey Objects
#       try:
#         university = University.objects.get(university_name=university_id)
#       except University.DoesNotExist:
#         logger.error(f"University with name '{university_id}' does not exist.")
#         return Response({'message': f"University '{university_id}' does not exist."}, status=status.HTTP_400_BAD_REQUEST)

#       # Check for duplicates
#       if Student.objects.filter(email=student_email).exists():
#           logger.error(f"Duplicate email: {student_email}")
#           return Response({'message': f"Email '{student_email}' is already registered."}, status=status.HTTP_400_BAD_REQUEST)

#       if Student.objects.filter(mobile=mobile_number).exists():
#           logger.error(f"Duplicate mobile number: {mobile_number}")
#           return Response({'message': f"Mobile number '{mobile_number}' is already registered."}, status=status.HTTP_400_BAD_REQUEST)

#       # Generate unique IDs
#       latest_student = Student.objects.latest('id') if Student.objects.exists() else None
#       enrollment_id = 5000 if not latest_student else int(latest_student.enrollment_id) + 1
#       registration_id = 250000 if not latest_student else int(latest_student.registration_id) + 1

#       with transaction.atomic():
#           # Create Student
#           student = Student(university=university,name=data.get('student_name'),father_name=data.get('father_name'),mother_name=data.get('mother_name'),dateofbirth=data.get('dob'),mobile=mobile_number,email=student_email,gender=data.get('gender'),alternateaddress=data.get('alternateaddress', ''),category=data.get('category', ''),enrollment_id=enrollment_id,registration_id=registration_id,is_enrolled=True,is_quick_register=True,created_by=request.user.id,modified_by=request.user.id,verified=True
#           )
          

#           # Create User for Student
#           user = User.objects.create(
#               email=student_email,
#               is_student=True,
#               password=make_password(student_email)
#           )
          

#           if payment_mode == "Cheque":
#             payment_status = "Not Realised"
#             uncleared_amount = fees
#             paidfees = 0
#           else:
#             payment_status = "Realised"
#             uncleared_amount = 0
#             paidfees = fees

#           # Validate Course, Stream, and SubStream dependencies
#           try:
#             course = Course.objects.get(name=data.get('course'), university=university)
#           except Course.DoesNotExist:
#             logger.error(f"Course '{data.get('course')}' does not belong to University '{university.university_name}'")
#             return Response({'message': f"Course '{data.get('course')}' does not belong to the specified university."}, status=status.HTTP_400_BAD_REQUEST)

#           try:
#             stream = Stream.objects.get(name=data.get('stream'), course=course)
#           except Stream.DoesNotExist:
#             logger.error(f"Stream '{data.get('stream')}' does not belong to Course '{course.name}'")
#             return Response({'message': f"Stream '{data.get('stream')}' does not belong to the specified course."}, status=status.HTTP_400_BAD_REQUEST)

#           try:
#             print(data.get('substream'),'subbbbbbb')
#             substream = SubStream.objects.get(name=data.get('substream'), stream=stream)
#           except SubStream.DoesNotExist:
#             logger.error(f"SubStream '{data.get('substream')}' does not belong to Stream '{stream.name}'")
#             return Response({'message': f"SubStream '{data.get('substream')}' does not belong to the specified stream."}, status=status.HTTP_400_BAD_REQUEST)

#           # Enroll Student in Course
#           student.save()
#           #logger.info(f"Student created: {student.name} (ID: {student.id})")
#           student.user = user
#           student.save()
#           #logger.info(f"User linked to student: {user.email}")
#           course_pattern=data.get('study_pattern')
#           course_pattern=course_pattern.capitalize()
#           print(course_pattern,'course_pattern')
#           enrollment = Enrolled(
#             student=student,course=course,stream=stream,substream=substream,course_pattern=course_pattern,session=data.get('session'),entry_mode=data.get('entry_mode'),total_semyear=stream.sem
#           )
#           enrollment.save()
#           #logger.info(f"Enrollment created for student: {student.name}")

#           latest_student = Student.objects.latest('id')
#           counselor_name=data.get('counselor_name')
#           university_enrollment_number=data.get('university_enrollment_number')
          
#           add_additional_details = AdditionalEnrollmentDetails(student=latest_student,counselor_name=counselor_name,university_enrollment_id=university_enrollment_number, reference_name='')
#           add_additional_details.save()
          
#           studypattern=data.get('studypattern')
#           fee_reciept_type = data.get('fee_reciept_type')
#           other_data = data.get('other_data')
#           transaction_date = data.get('transaction_date')
#           total_fees = data.get('total_fees')
#           bank_name=data.get('bank_name')
#           other_bank = data.get('other_bank')
#           cheque_no = data.get('cheque_no')
#           remarks = data.get('remarks')
#           session = data.get('session')
#           semyear = data.get('semyear')

#           if studypattern=="Semester":
#             try:
#               print('inside Try Semester ')
#               getsemesterfees = SemesterFees.objects.filter(stream=stream, substream=substream)
#               print('inside Try Semester ',getsemesterfees)
#               for i in getsemesterfees:
#                 print(i,'testststtss')
#                 addstudentfees = StudentFees(student = latest_student,studypattern="Semester",stream=stream,tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.sem, substream=substream)
#                 print('saved data before addstudent') 
#                 addstudentfees.save()  #Added By Ankit 13-12-24
#                 print('saved data addstudnet savedddddddd ') 

#             except SemesterFees.DoesNotExist:
#               logger.error(f"SemesterFees does not exists")
#               print('inside except semister not found')
#               pass
#           elif studypattern == "Annual":
#             try:
#               print("Annual try reached")
#               getyearfees = YearFees.objects.filter(stream=stream, substream=substream)
#               for i in getyearfees:
#                 addstudentfees = StudentFees(student = latest_student,studypattern="Annual",stream=Stream.objects.get(id=stream),tutionfees=i.tutionfees,examinationfees=i.examinationfees,bookfees=i.bookfees,resittingfees=i.resittingfees,entrancefees=i.entrancefees,extrafees=i.extrafees,discount=i.discount,totalfees=i.totalfees,sem=i.year, substream=i.substream)
#                 addstudentfees.save()
#             except YearFees.DoesNotExist:
#                 print("annual not reached")            
#           try:
#             getlatestreciept = PaymentReciept.objects.latest('id')
#             print(getlatestreciept,'iddddddddd')
#           except PaymentReciept.DoesNotExist:
#             print('inside PaymentReciept.DoesNotExist ')
#             getlatestreciept = "none"
#           if getlatestreciept == "none":
#               transactionID = "TXT445FE101"
#               print(transactionID,'  : IDDDDDDDDDDDDD ')
#           else:
#             tid = getlatestreciept.transactionID
#             tranx = tid.replace("TXT445FE",'')
#             transactionID =  str("TXT445FE") + str(int(tranx) + 1)
#           if fee_reciept_type == "Others":
#             reciept_type = other_data
#           else:
#             reciept_type = fee_reciept_type
#             print('reciept_type',reciept_type) 
#           if payment_mode == "Cheque":
#               print('inside payment_mode Cheque')
#               if studypattern == "Full Course":
#                 print('inside studypattern Full Course')
#                 if fees and fee_reciept_type and transaction_date and payment_mode:
#                   print('inside fees and fee_reciept_type and transaction_date and payment_mode ')

#                   pending_fees = int(total_fees) - int(fees)
#                   if pending_fees > 0:
#                       print("positive")
#                       pending_amount = pending_fees
#                       advance_amount = 0
#                   elif pending_fees == 0:
#                       print("no pending semyear clear")
#                       pending_amount = 0
#                       advance_amount = 0
#                   elif pending_fees < 0:
#                       print("negative hence advance payment")
#                       pending_amount = 0
#                       advance_amount = abs(pending_fees)
#                   if pending_fees == 0 | pending_fees < 0:
#                       paymenttype = "Full Payment"
#                   else:
#                       paymenttype = "Part Payment"
#                   if bank_name == "Others":
#                     print('inside bank_name == "Others"   ---------')
#                     add_payment_reciept = PaymentReciept(student = latest_student,payment_for="Course Fees",payment_categories = "New",payment_type=paymenttype,fee_reciept_type=reciept_type,transaction_date= transaction_date,cheque_no=cheque_no,bank_name=other_bank,semyearfees=total_fees,paidamount=0,pendingamount=total_fees,advanceamount = advance_amount,transactionID = transactionID,paymentmode=payment_mode,remarks=remarks,session=session,semyear="1",uncleared_amount=fees,status=payment_status)
#                     add_payment_reciept.save()
#                   else :
#                     print('inside else bank_name == "Others" ')
#                     add_payment_reciept = PaymentReciept(student = latest_student,payment_for="Course Fees",payment_categories = "New",payment_type=paymenttype,fee_reciept_type=reciept_type,transaction_date= transaction_date,cheque_no=cheque_no,bank_name=bank_name,semyearfees=total_fees,paidamount=fees,pendingamount=pending_amount,advanceamount = advance_amount,transactionID = transactionID,paymentmode=payment_mode,remarks=remarks,session=session,semyear=semyear,status=payment_status)
#                     add_payment_reciept.save()
#           else:
#               print('payment_mode == "Cheque" elseeeeee')
#               if studypattern == "Full Course":
#                 print('studypattern == "Full Course"')
#                 if fees and fee_reciept_type and transaction_date and payment_mode:
#                   print('fees and fee_reciept_type and transaction_date and payment_mode')  
#                   pending_fees = int(total_fees) - int(fees)
#                   if pending_fees > 0:
#                     print("positive")
#                     pending_amount = pending_fees
#                     advance_amount = 0
#                 elif pending_fees == 0:
#                     print("no pending semyear clear")
#                     pending_amount = 0
#                     advance_amount = 0
#                 elif pending_fees < 0:
#                     print("negative hence advance payment")
#                     pending_amount = 0
#                     advance_amount = abs(pending_fees)
#                 if pending_fees == 0 | pending_fees < 0:
#                     paymenttype = "Full Payment"
#                 else:
#                     paymenttype = "Part Payment"
#                 if bank_name == "Others":
#                     add_payment_reciept = PaymentReciept(student = latest_student,payment_for="Course Fees",payment_categories = "New",payment_type=paymenttype,fee_reciept_type=reciept_type,transaction_date= transaction_date,cheque_no=cheque_no,bank_name=other_bank,semyearfees=total_fees,paidamount=fees,pendingamount=pending_amount,advanceamount = advance_amount,transactionID = transactionID,paymentmode=payment_mode,remarks=remarks,session=session,semyear="1",status=payment_status)
#                     add_payment_reciept.save()
#                 else:
#                     print('inside else bank_name == "bank name" ')
#                     add_payment_reciept = PaymentReciept(student = latest_student,payment_for="Course Fees",payment_categories = "New",payment_type=paymenttype,fee_reciept_type=reciept_type,transaction_date= transaction_date,cheque_no=cheque_no,bank_name=bank_name,semyearfees=total_fees,paidamount=fees,pendingamount=pending_amount,advanceamount = advance_amount,transactionID = transactionID,paymentmode=payment_mode,remarks=remarks,session=session,semyear="1",status=payment_status)
#                     add_payment_reciept.save()
#               else:
#                 if fees and fee_reciept_type and transaction_date and payment_mode:
#                   pending_fees = int(total_fees) - int(fees)
#                   if pending_fees > 0:
#                       print("positive")
#                       pending_amount = pending_fees
#                       advance_amount = 0
#                   elif pending_fees == 0:
#                       print("no pending semyear clear")
#                       pending_amount = 0
#                       advance_amount = 0
#                   elif pending_fees < 0:
#                       print("negative hence advance payment")
#                       pending_amount = 0
#                       advance_amount = abs(pending_fees)
                      
#                   if pending_fees == 0:
#                       paymenttype = "Full Payment"
#                   else:
#                       paymenttype = "Part Payment"
#                   if bank_name == "Others":
#                       print('inside if else bank_name == "others" ')
#                       add_payment_reciept = PaymentReciept(student = latest_student,payment_for="Course Fees",payment_categories = "New",payment_type=paymenttype,fee_reciept_type=reciept_type,transaction_date= transaction_date,cheque_no=cheque_no,bank_name=other_bank,semyearfees=total_fees,paidamount=fees,pendingamount=pending_amount,advanceamount = advance_amount,transactionID = transactionID,paymentmode=payment_mode,remarks=remarks,session=session,semyear=semyear,status=payment_status)
#                       add_payment_reciept.save()
#                   else :
#                     print('inside if else bank_name == "others" ')
#                     add_payment_reciept = PaymentReciept(student = latest_student,payment_for="Course Fees",payment_categories = "New",payment_type=paymenttype,fee_reciept_type=reciept_type,transaction_date= transaction_date,cheque_no=cheque_no,bank_name=bank_name,semyearfees=total_fees,paidamount=fees,pendingamount=pending_amount,advanceamount = advance_amount,transactionID = transactionID,paymentmode=payment_mode,remarks=remarks,session=session,semyear=semyear,status=payment_status)
#                     add_payment_reciept.save()  
            
          
#       return Response({'message': f"Student '{student.name}' registered successfully."}, status=status.HTTP_201_CREATED)

#   except Exception as e:
#     logger.error(f"Unexpected error while fetching University: {str(e)}")
#     return Response({'message': 'Unexpected error occurred while fetching University.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])  # Specify that this view only accepts GET requests
@permission_classes([IsAuthenticated])
def get_sem_year_by_stream(request):
    course_id = request.GET.get('course')  # Use request.GET for query parameters
    university_id= request.GET.get('university')
    stream_id = request.GET.get('stream')

    try:
        university = University.objects.get(id=university_id)
        course = Course.objects.get(id=course_id, university=university)
        streams = Stream.objects.filter(id=stream_id, course=course)

        # Prepare the response data
        response_data = []
        for stream in streams:
            response_data.append({
                'year': stream.year,
                'stream_name': stream.name,
                'sem': stream.sem
            })

        # Return the response with the desired structure
        return Response({
            'university_name': university.university_name,
            'course_name': course.name,
            'streams': response_data
        }, status=200)

    except University.DoesNotExist:
        logger.error(f"University '{university_id}' not found.")
        return Response({'error': 'University not found'}, status=404)
    except Course.DoesNotExist:
        logger.error(f"Course '{course_id}' not found for university '{university_id}'.")
        return Response({'error': 'Course not found'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return Response({'error': 'An error occurred'}, status=500)

    except University.DoesNotExist:
        logger.error(f"University '{university_id}' not found.")
        return Response({'error': 'University not found'}, status=404)
    except Course.DoesNotExist:
        logger.error(f"Course '{course_id}' not found for university '{university_id}'.")
        return Response({'error': 'Course not found'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return Response({'error': 'An error occurred'}, status=500)
      
@api_view(['GET'])  # Specify that this view only accepts GET requests
@permission_classes([IsAuthenticated])
def get_sem_year_by_stream_byname(request):
    course_name = request.GET.get('course')  # Use request.GET for query parameters
    university_name = request.GET.get('university')
    stream_name = request.GET.get('stream')

    try:
        # Fetch the university by name
        university = University.objects.get(university_name=university_name)
        
        # Fetch the course by name and its associated university
        course = Course.objects.get(name=course_name, university=university)
        
        # Fetch the streams related to the course and with the given stream name
        streams = Stream.objects.filter(name=stream_name, course=course)

        # Prepare the response data
        response_data = []
        for stream in streams:
            response_data.append({
                'year': stream.year,
                'stream_name': stream.name,
                'sem': stream.sem
            })

        # Return the response with the desired structure
        return Response({
            'university_name': university.university_name,
            'course_name': course.name,
            'streams': response_data
        }, status=200)

    except University.DoesNotExist:
        logger.error(f"University '{university_name}' not found.")
        return Response({'error': 'University not found'}, status=404)
    except Course.DoesNotExist:
        logger.error(f"Course '{course_name}' not found for university '{university_name}'.")
        return Response({'error': 'Course not found'}, status=404)
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return Response({'error': 'An error occurred'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_fee_recipt_option(request):
    try:
        data = FeeReceiptOptions.objects.filter(status=True)
        serializer = FeeReceiptOptionsSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"error": str(e)},
              status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def bank_names_list_create(request):
    if request.method == 'GET':
        bank_names = BankNames.objects.filter(status=True)
        serializer = BankNamesSerializer(bank_names, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = BankNamesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def payment_modes_list_create(request):
    if request.method == 'GET':
        payment_modes = PaymentModes.objects.filter(status=True)
        serializer = PaymentModesSerializer(payment_modes, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = PaymentModesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def view_quick_registered_students(request):
    # Fetch all students who are quick registered
    students = Student.objects.filter(is_quick_register=True, is_cancelled=False).order_by('-id')
    serializer = Student_Quick_RegisteredSerializer(students, many=True)
    return Response(serializer.data)  

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_student(request, student_id):
    try:
        # Fetch the student by ID
        student = Student.objects.get(id=student_id)
        
        # Delete the student
        student.delete()
        
        return Response({"message": "Student deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    except Student.DoesNotExist:
        # If the student doesn't exist, return an error
        return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def student_registration(request):
    if not request.user.is_superuser:
        return Response({"error": "You are not authorized to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    data = request.data
    try:
        # Extract fields
        name = data.get('name')
        email = data.get('email')
        mobile = data.get('mobile_number')
        alternate_mobile = data.get('alternate_number')
        alternate_email = data.get('alternate_email')
        father_name = data.get('father_name')
        mother_name = data.get('mother_name')
        dob = data.get('dob')
        gender = data.get('gender')
        category = data.get('category')
        address = data.get('address')
        alternateaddress = data.get('alternateaddress')
        nationality = data.get('nationality')
        pincode = data.get('pincode')
        registration_number = data.get('registration_number')
        student_image = request.FILES.get('image')

        university_id = data.get('university')
        course_id = data.get('course')
        stream_id = data.get('stream')
        substream_id = data.get('substream')
        studypattern = data.get('studypattern', '').capitalize()
        semyear = data.get('semyear')
        session = data.get('session')
        entry_mode = data.get('entry_mode')

        country_id = data.get('country')
        state_id = data.get('state')
        city_id = data.get('city')

        # Validate uniqueness
        if Student.objects.filter(mobile=mobile).exists():
            return Response({"error": "Mobile number already registered."}, status=status.HTTP_400_BAD_REQUEST)
        if Student.objects.filter(email=email).exists():
            return Response({"error": "Email already registered."}, status=status.HTTP_400_BAD_REQUEST)

        # Foreign key validation
        try:
            university = University.objects.get(id=university_id)
            course = Course.objects.get(id=course_id, university=university)
            stream = Stream.objects.get(id=stream_id, course=course)
            substream = SubStream.objects.get(id=substream_id, stream=stream) if substream_id else None
            country = Countries.objects.get(id=country_id)
            state = States.objects.get(id=state_id, country_id=country_id)
            city = Cities.objects.get(id=city_id, state_id=state_id)
        except (University.DoesNotExist, Course.DoesNotExist, Stream.DoesNotExist,
                SubStream.DoesNotExist, Countries.DoesNotExist,
                States.DoesNotExist, Cities.DoesNotExist) as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Generate enrollment/registration IDs
        latest = Student.objects.order_by("-id").first()
        enrollment_id = int(latest.enrollment_id) + 1 if latest else 50000
        registration_id = int(latest.registration_id) + 1 if latest else 250000

        # Save Student
        student_data = {
            "name": name, "email": email, "mobile": mobile, "father_name": father_name,
            "mother_name": mother_name, "alternate_mobile1": alternate_mobile,
            "alternateemail": alternate_email, "enrollment_id": enrollment_id,
            "registration_id": registration_id, "dateofbirth": dob, "gender": gender,
            "category": category, "address": address, "alternateaddress": alternateaddress,
            "nationality": nationality, "pincode": pincode, "image": student_image,
            "university": university.id, "country": country.id, "state": state.id,
            "city": city.id, "verified": True, "is_enrolled": True,
            "created_by": request.user.id
        }
        student_serializer = StudentSerializer(data=student_data)
        if not student_serializer.is_valid():
            return Response({"error": "Validation failed", "details": student_serializer.errors}, status=400)
        student = student_serializer.save()

        if student:
          try:
              User.objects.create(email=email, is_student=True, password=make_password(mobile))
          except Exception as e:
              logger.warning("User creation failed: %s", e)

        # Save Enrollment
        total_semyear = int(stream.sem) * (2 if studypattern == "Semester" else 1)
        enrolled_data = {
            "student": student.id, "course": course.id, "stream": stream.id,
            "substream": substream.id if substream else None,
            "course_pattern": studypattern, "session": session,
            "entry_mode": entry_mode, "total_semyear": total_semyear,
            "current_semyear": semyear
        }
        enrolled_serializer = EnrolledSerializer(data=enrolled_data)
        if not enrolled_serializer.is_valid():
            return Response({"error": "Enrollment failed", "details": enrolled_serializer.errors}, status=400)
        enrolled_serializer.save()

        # Save Additional Enrollment Info
        additional_data = {
            "student": student.id,
            "counselor_name": data.get("counselor_name"),
            "university_enrollment_id": data.get("university_enrollment_number"),
            "reference_name": data.get("reference_name")
        }
        add_serializer = AdditionalEnrollmentDetailsSerializer(data=additional_data)
        if not add_serializer.is_valid():
            return Response({"error": "Additional Enrollment Failed", "details": add_serializer.errors}, status=400)
        add_serializer.save()

        # Qualifications
        others = []
        for key in data.keys():
            if key.startswith("qualifications[others]") and key.endswith("[year]"):
                try:
                    idx = key.split("[")[2].split("]")[0]
                    file_key = f"qualifications[others][{idx}][file]"
                    file_path = default_storage.save(f"University_Documents/{request.FILES[file_key].name}", request.FILES[file_key]) if file_key in request.FILES else None
                    others.append({
                        "year": data.get(f"qualifications[others][{idx}][year]"),
                        "board": data.get(f"qualifications[others][{idx}][board]"),
                        "doctype": data.get(f"qualifications[others][{idx}][doctype]"),
                        "file_path": file_path
                    })
                except Exception as e:
                    logger.error("Error in other qualifications: %s", e)

        qualification_data = {
            "student": student.id,
            "secondary_year": data.get("qualifications[secondary_year]"),
            "secondary_board": data.get("qualifications[secondary_board]"),
            "secondary_percentage": data.get("qualifications[secondary_percentage]"),
            "sr_year": data.get("qualifications[sr_year]"),
            "sr_board": data.get("qualifications[sr_board]"),
            "sr_percentage": data.get("qualifications[sr_percentage]"),
            "under_year": data.get("qualifications[under_year]"),
            "under_board": data.get("qualifications[under_board]"),
            "under_percentage": data.get("qualifications[under_percentage]"),
            "post_year": data.get("qualifications[post_year]"),
            "post_board": data.get("qualifications[post_board]"),
            "post_percentage": data.get("qualifications[post_percentage]"),
            "mphil_year": data.get("qualifications[mphil_year]"),
            "mphil_board": data.get("qualifications[mphil_board]"),
            "mphil_percentage": data.get("qualifications[mphil_percentage]"),
            "secondary_document": request.FILES.get("qualifications[secondary_document]"),
            "sr_document": request.FILES.get("qualifications[sr_document]"),
            "under_document": request.FILES.get("qualifications[under_document]"),
            "post_document": request.FILES.get("qualifications[post_document]"),
            "mphil_document": request.FILES.get("qualifications[mphil_document]"),
            "others": others
        }
        qualification_serializer = QualificationSerializer(data=qualification_data)
        if not qualification_serializer.is_valid():
            return Response({"error": "Qualification failed", "details": qualification_serializer.errors}, status=400)
        qualification_serializer.save()

        # Personal Documents
        index = 0
        documents = []
        while f'documents[{index}][document]' in data:
            doc_data = {
                "document": data.get(f'documents[{index}][document]'),
                "document_name": data.get(f'documents[{index}][document_name]'),
                "document_ID_no": data.get(f'documents[{index}][document_ID_no]'),
                "document_image_front": request.FILES.get(f'documents[{index}][document_image_front]'),
                "document_image_back": request.FILES.get(f'documents[{index}][document_image_back]')
            }
            doc_serializer = StudentDocumentsSerializer(data=doc_data)
            if doc_serializer.is_valid():
                validated = {
                    **doc_serializer.validated_data,
                    "student": student
                }
                documents.append(StudentDocuments(**validated))
            else:
                return Response({"error": "Invalid Document", "details": doc_serializer.errors}, status=400)
            index += 1

        if documents:
            StudentDocuments.objects.bulk_create(documents)

        # Assign Fees
        fees_model = SemesterFees if studypattern == "Semester" else YearFees
        fees = fees_model.objects.filter(stream=stream, substream=substream)
        for fee in fees:
            StudentFees.objects.create(
                student=student, stream=stream, substream=substream,
                studypattern=studypattern,
                tutionfees=fee.tutionfees, examinationfees=fee.examinationfees,
                totalfees=fee.totalfees, sem=fee.sem if studypattern == "Semester" else fee.year
            )

        # Save Payment
        try:
            latest_receipt = PaymentReciept.objects.latest("id")
            last_id = int(latest_receipt.transactionID.replace("TXT445FE", "")) if latest_receipt else 100
        except PaymentReciept.DoesNotExist:
            last_id = 100

        transaction_id = f"TXT445FE{last_id + 1}"
        payment_mode = data.get("payment_mode")
        fee_type = data.get("fee_reciept_type")
        other_type = data.get("other_data")
        receipt_type = other_type if fee_type == "Others" else fee_type
        total_fees = fees.last().totalfees if fees.exists() else 0
        paid = float(data.get("paidamount") or 0)
        cheque_no = data.get("cheque_no")
        bank_name = data.get("bank_name")
        other_bank = data.get("other_bank")
        remarks = data.get("remarks")
        transaction_date = data.get("transaction_date")
        session = data.get("session")

        pending = float(total_fees) - paid
        advance = abs(pending) if pending < 0 else 0
        pending = pending if pending > 0 else 0
        payment_type = "Full Payment" if pending == 0 else "Part Payment"
        bank_used = other_bank if bank_name == "Others" else bank_name
        payment_status = "Not Realised" if payment_mode == "Cheque" else "Realised"
        semyear_value = "1" if studypattern == "Full Course" else semyear
        uncleared_amount = paid if payment_mode == "Cheque" else None

        PaymentReciept.objects.create(
            student=student, payment_for=data.get("fee_recipt"),
            payment_categories="New", payment_type=payment_type,
            fee_reciept_type=receipt_type, transaction_date=transaction_date,
            cheque_no=cheque_no, bank_name=bank_used, semyearfees=total_fees,
            paidamount=0 if payment_mode == "Cheque" else paid,
            pendingamount=pending, advanceamount=advance,
            transactionID=transaction_id, paymentmode=payment_mode,
            remarks=remarks, session=session, semyear=semyear_value,
            uncleared_amount=uncleared_amount, status=payment_status
        )

        return Response({"message": "Student registered successfully."}, status=201)

    except Exception as e:
        logger.error("Unexpected error: %s", traceback.format_exc())
        return Response({"error": "An unexpected error occurred."}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_sem_fees(request):
    try:
        # Fetch query parameters
        stream_id = request.query_params.get('stream_id')
        substream_id = request.query_params.get('substream_id')

        if not stream_id:
            return Response({"error": "Stream ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch the stream object by ID
        stream = get_object_or_404(Stream, id=stream_id)

        # Fetch the substream object by ID (optional)
        substream = None
        if substream_id:
            substream = get_object_or_404(SubStream, id=substream_id, stream=stream)

        # Filter semester fees based on stream and substream
        semester_fees = SemesterFees.objects.filter(
            stream=stream, substream=substream
        ).values('sem', 'tutionfees', 'examinationfees', 'bookfees', 
                 'resittingfees', 'entrancefees', 'extrafees', 'discount', 'totalfees')

        # Format the response with 'sem' outside the object
        response_data = [
            {
                "sem": fee["sem"],
                "fees_details": {
                    "tutionfees": fee["tutionfees"],
                    "examinationfees": fee["examinationfees"],
                    "bookfees": fee["bookfees"],
                    "resittingfees": fee["resittingfees"],
                    "entrancefees": fee["entrancefees"],
                    "extrafees": fee["extrafees"],
                    "discount": fee["discount"],
                    "totalfees": fee["totalfees"]
                }
            }
            for fee in semester_fees
        ]

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_year_fees(request):
    try:
        stream_id = request.query_params.get('stream_id')
        substream_id = request.query_params.get('substream_id')

        if not stream_id:
            return Response({"error": "Stream ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        stream = get_object_or_404(Stream, id=stream_id)

        substream = None
        if substream_id:
            substream = get_object_or_404(SubStream, id=substream_id, stream=stream)

        year_fees = YearFees.objects.filter(
            stream=stream, substream=substream
        ).values('year', 'tutionfees', 'examinationfees', 'bookfees', 
                 'resittingfees', 'entrancefees', 'extrafees', 'discount', 'totalfees')

        response_data = [
            {
                "year": fee["year"],
                "fees_details": {
                    "tutionfees": fee["tutionfees"],
                    "examinationfees": fee["examinationfees"],
                    "bookfees": fee["bookfees"],
                    "resittingfees": fee["resittingfees"],
                    "entrancefees": fee["entrancefees"],
                    "extrafees": fee["extrafees"],
                    "discount": fee["discount"],
                    "totalfees": fee["totalfees"]
                }
            }
            for fee in year_fees
        ]

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#------------------------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_courses_by_university_with_id(request):
    university_id = request.query_params.get('university_id')

    if not university_id:
        return Response({"error": "University name is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        university = University.objects.get(id=university_id)
    except University.DoesNotExist:
        return Response({"error": "University not found."}, status=status.HTTP_404_NOT_FOUND)

    # Fetch all courses with id and name
    courses = university.course_set.all()
    course_details = [
        {"course_id": course.id, "name": course.name}
        for course in courses
    ]

    # Return both university and course details
    return Response({
        "university_id": university.id,
        "university_name": university.university_name,
        "courses": course_details
    }, status=status.HTTP_200_OK)
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stream_by_course_with_id(request):
    # Get course_id and university_id from query parameters
    course_id = request.query_params.get('course_id')
    university_id = request.query_params.get('university_id')

    #logger.info("Request to fetch streams for course ID '%s' at university ID '%s' received.", course_id, university_id)
    
    # Check if both course_id and university_id are provided
    if not course_id or not university_id:
        return Response({"error": "Both 'course_id' and 'university_id' are required."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Fetch the university object
        university = University.objects.get(id=university_id)

        # Fetch the course object related to the given university
        course = Course.objects.get(id=course_id, university=university)
        
        # Fetch streams related to this course
        streams = Stream.objects.filter(course=course)
        
        # Prepare the response data
        stream_list = [
            {
                "stream_id": stream.id,
                "stream_name": stream.name,
                "semester": stream.sem,
                "year": stream.year
            }
            for stream in streams
        ]

        #logger.info("Successfully fetched %d streams for course '%s' (ID: %d) at university '%s' (ID: %d).",len(stream_list), course.name, course.id, university.university_name, university.id)
        
        return Response({
            "university_id": university.id,
            "university_name": university.university_name,
            "course_id": course.id,
            "course_name": course.name,
            "streams": stream_list
        }, status=status.HTTP_200_OK)

    except University.DoesNotExist:
        logger.error("Stream fetch failed: University with ID '%s' not found.", university_id)
        return Response({"error": f"University with ID '{university_id}' not found."}, status=status.HTTP_404_NOT_FOUND)

    except Course.DoesNotExist:
        logger.error("Stream fetch failed: Course with ID '%s' not found for university ID '%s'.", course_id, university_id)
        return Response({"error": f"Course with ID '{course_id}' not found for university ID '{university_id}'."}, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
 
        return Response({"error": "Unable to fetch streams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_substreams_by_university_course_stream_with_id(request):
    # Get IDs from query parameters
    university_id = request.query_params.get('university_id')
    course_id = request.query_params.get('course_id')
    stream_id = request.query_params.get('stream_id')

    #logger.info("Request to fetch substreams for university ID '%s', course ID '%s', stream ID '%s' received.", university_id, course_id, stream_id)    
    # Check if all required parameters are provided
    if not university_id or not course_id or not stream_id:
        return Response(
            {"error": "University ID, course ID, and stream ID are required."},
            status=status.HTTP_400_BAD_REQUEST
        )    
    try:
        # Fetch the university object
        university = University.objects.get(id=university_id)

        # Fetch the course object related to the given university
        course = Course.objects.get(id=course_id, university=university)

        # Fetch the stream object related to the given course
        stream = Stream.objects.get(id=stream_id, course=course)

        # Fetch substreams related to the stream
        substreams = SubStream.objects.filter(stream=stream)
        
        # Prepare the response data
        substream_list = [
            {
                "substream_id": substream.id,
                "substream_name": substream.name
            }
            for substream in substreams
        ]
        #logger.info("Successfully fetched %d substreams for stream '%s' (ID: %d), course '%s' (ID: %d), university '%s' (ID: %d).", len(substream_list), stream.name, stream.id, course.name, course.id, university.university_name, university.id)

        return Response({
            "university_id": university.id,
            "university_name": university.university_name,
            "course_id": course.id,
            "course_name": course.name,
            "stream_id": stream.id,
            "stream_name": stream.name,
            "substreams": substream_list
        }, status=status.HTTP_200_OK)

    except University.DoesNotExist:
        logger.error("Substream fetch failed: University with ID '%s' not found.", university_id)
        return Response({"error": f"University with ID '{university_id}' not found."}, status=status.HTTP_404_NOT_FOUND)

    except Course.DoesNotExist:
        logger.error("Substream fetch failed: Course with ID '%s' not found for university ID '%s'.", course_id, university_id)
        return Response({"error": f"Course with ID '{course_id}' not found for university ID '{university_id}'."}, status=status.HTTP_404_NOT_FOUND)

    except Stream.DoesNotExist:
        logger.error("Substream fetch failed: Stream with ID '%s' not found for course ID '%s' and university ID '%s'.", 
                     stream_id, course_id, university_id)
        return Response(
            {"error": f"Stream with ID '{stream_id}' not found for course ID '{course_id}' and university ID '{university_id}'."},
            status=status.HTTP_404_NOT_FOUND
        )

    except Exception as e:
        logger.error("Error fetching substreams for university ID '%s', course ID '%s', stream ID '%s'. Exception: %s", 
                     university_id, course_id, stream_id, str(e))
        return Response({"error": "Unable to fetch substreams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_country(request):
    try:
        from .models import Countries  # ✅ Make sure this matches your actual model
        data = Countries.objects.all()
        serializer = CountrySerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_country: {str(e)}")
        return Response(
            {"error": "An internal server error occurred."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_states(request):
    try:
        country_input = request.query_params.get('country_id')

        if not country_input:
            return Response(
                {"error": "country_id is required as a query parameter."},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import Countries  

        if country_input.isdigit():
            data = States.objects.filter(country_id=int(country_input))
        else:
            country = Countries.objects.filter(name__iexact=country_input.strip()).first()
            if not country:
                return Response({"error": f"No country found with name '{country_input}'"}, status=404)
            data = States.objects.filter(country_id=country.id)

        serializer = StateSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error in get_states: {str(e)}")
        return Response({"error": "An internal server error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_cities(request):
    try:
        state_input = request.query_params.get('state_id')

        if not state_input:
            return Response(
                {"error": "state_id is required as a query parameter."},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import States  # Ensure States model is imported

        if state_input.isdigit():
            data = Cities.objects.filter(state_id=int(state_input))
        else:
            state = States.objects.filter(name__iexact=state_input.strip()).first()
            if not state:
                return Response({"error": f"No state found with name '{state_input}'"}, status=404)
            data = Cities.objects.filter(state_id=state.id)

        serializer = CitySerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error in get_cities: {str(e)}")
        return Response({"error": "An internal server error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_details(request, enrollment_id):
    if not enrollment_id:
        logger.error("Missing required parameter: enrollment_id in request URL")
        return Response(
            {"error": "Missing required parameter: enrollment_id"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        student = Student.objects.get(enrollment_id=enrollment_id)

        # === STUDENT DATA ===
        student_data = {
            "id": student.id,
            "name": student.name,
            "father_name": student.father_name,
            "mother_name": student.mother_name,
            "dateofbirth": student.dateofbirth,
            "country": student.country.name if student.country else None,
            "state": student.state.name if student.state else None,
            "city": student.city.name if student.city else None,
            "country_id": student.country.id if student.country else None,
            "state_id": student.state.id if student.state else None,
            "city_id": student.city.id if student.city else None,
            "image": student.image.url if student.image else None,
            "mobile": student.mobile,
            "nationality": student.nationality,
            "pincode": student.pincode,
            "alternate_mobile1": student.alternate_mobile1,
            "email": student.email,
            "alternateemail": student.alternateemail,
            "gender": student.gender,
            "category": student.category,
            "address": student.address,
            "alternateaddress": student.alternateaddress,
            "student_remarks": student.student_remarks,
            "registration_id": student.registration_id,
            "enrollment_id": student.enrollment_id,
            "university": student.university.university_name if student.university else None,
            "university_id": student.university.id if student.university else None,
        }

        # === ENROLLMENT DATA ===
        enrollment = Enrolled.objects.filter(student=student).first()
        enrollment_data = {
            "course": enrollment.course.name if enrollment and enrollment.course else None,
            "stream": enrollment.stream.name if enrollment and enrollment.stream else None,
            "substream": enrollment.substream.name if enrollment and enrollment.substream else None,
            "course_id": enrollment.course.id if enrollment and enrollment.course else None,
            "stream_id": enrollment.stream.id if enrollment and enrollment.stream else None,
            "substream_id": enrollment.substream.id if enrollment and enrollment.substream else None,
            "studypattern": enrollment.course_pattern if enrollment else None,
            "session": enrollment.session if enrollment else None,
            "entry_mode": enrollment.entry_mode if enrollment else None,
            "total_semyear": enrollment.total_semyear if enrollment else None,
            "semyear": enrollment.current_semyear if enrollment else None,
        } if enrollment else {}

        # === ADDITIONAL ENROLLMENT DETAILS ===
        additional_details = AdditionalEnrollmentDetails.objects.filter(student=student).first()
        additional_details_data = {
            "counselor_name": additional_details.counselor_name if additional_details else None,
            "reference_name": additional_details.reference_name if additional_details else None,
            "university_enroll_number": additional_details.university_enrollment_id if additional_details else None,
        }

        # === DOCUMENTS ===
        student_documents = StudentDocuments.objects.filter(student=student)
        student_documents_data = StudentDocumentsSerializerGET(student_documents, many=True).data

        # === QUALIFICATION ===
        qualifications = Qualification.objects.filter(student=student).first()
        qualification_data = QualificationSerializer(qualifications).data if qualifications else None

        # === PAYMENT RECEIPT ===
        latest_payment = PaymentReciept.objects.filter(student=student).last()
        due_amount = None
        payment_data = {}

        if latest_payment:
            try:
                semyearfees = float(latest_payment.semyearfees) if latest_payment.semyearfees else 0.0
                paidamount = float(latest_payment.paidamount) if latest_payment.paidamount else 0.0
                due_amount = semyearfees - paidamount

                payment_data = {
                    "Exam_Fee": latest_payment.payment_for,
                    "transaction_date": latest_payment.transaction_date,
                    "total_fees": latest_payment.semyearfees,
                    "Due": due_amount,
                    "cheque_no": latest_payment.cheque_no,
                    "bank_name": latest_payment.bank_name,
                    "paymentmode": latest_payment.paymentmode,
                    "remarks": latest_payment.remarks,
                    "paidamount": latest_payment.paidamount,
                    "fee_recipt_type": latest_payment.fee_reciept_type,
                }

            except Exception as pe:
                logger.warning(f"Could not calculate payment due for enrollment_id {enrollment_id}: {pe}")
                payment_data = {}

        # === FINAL RESPONSE ===
        response_data = {
            **student_data,
            **enrollment_data,
            **additional_details_data,
            "documents": student_documents_data,
            "qualifications": qualification_data,
            "payment_details": payment_data,
        }

        return Response(response_data, status=status.HTTP_200_OK)

    except Student.DoesNotExist:
        logger.error(f"Student with enrollment_id {enrollment_id} not found.")
        return Response(
            {"error": "Student not found."},
            status=status.HTTP_404_NOT_FOUND
        )

    except Exception as e:
        logger.error(f"Unexpected error while retrieving student details for enrollment_id {enrollment_id}: {str(e)}")
        return Response(
            {"error": "An unexpected error occurred while retrieving student details."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

       
        
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_student_details(request, enrollment_id):

    try:
        student = Student.objects.get(enrollment_id=enrollment_id)
        for field in ['name', 'father_name', 'dateofbirth', 'mobile', 'email', 'gender', 'category', 'address']:
            if field in request.data:
                setattr(student, field, request.data[field])
        
        if 'country' in request.data:
            try:
                student.country = Countries.objects.get(id=request.data['country'])
            except Countries.DoesNotExist:
                return Response({"error": f"Country with id {request.data['country']} does not exist."}, status=400)

        if 'state' in request.data:
            try:
                student.state = States.objects.get(id=request.data['state'])
            except States.DoesNotExist:
                return Response({"error": f"State with id {request.data['state']} does not exist."}, status=400)  
        if 'city' in request.data:
            try:
                student.city = Cities.objects.get(id=request.data['city'])
            except Cities.DoesNotExist:
                return Response({"error": f"City with id {request.data['city']} does not exist."}, status=400)
        
        if 'image' in request.data:
            student.image = request.data['image']
        student.save()

        # Update enrollment details
        enrollment = Enrolled.objects.filter(student=student).first()
        if enrollment:
            for field, data_field in [
                ('course', 'course'),
                ('stream', 'stream'),
                ('substream', 'substream'),
                ('course_pattern', 'studypattern'),
                ('session', 'session'),
                ('entry_mode', 'entry_mode'),
                ('total_semyear', 'total_semyear'),
                ('current_semyear', 'semyear')
            ]:
                if data_field in request.data:
                    if data_field == 'course':
                        enrollment.course = Course.objects.get(id=request.data[data_field])
                    elif data_field == 'stream':
                        enrollment.stream = Stream.objects.get(id=request.data[data_field])
                    elif data_field == 'substream':
                      substream_id = request.data[data_field]
                      if substream_id and str(substream_id).isdigit():
                          enrollment.substream = SubStream.objects.get(id=substream_id)
                      else:
                          enrollment.substream = None
                    else:
                        setattr(enrollment, field, request.data[data_field])
            enrollment.save()
        # Update or create payment receipt
        payment_data = request.data.get('payment_details', {})
        print(payment_data,'payment_data')
        if payment_data:
            latest_payment = PaymentReciept.objects.filter(student=student).last()
            if latest_payment:
                for field in ['remarks', 'transaction_date', 'bank_name', 'cheque_no', 'paymentmode']:
                    if field in payment_data:
                        setattr(latest_payment, field, payment_data[field])
                latest_payment.save()
            else:
                PaymentReciept.objects.create(
                    student=student,
                    remarks=payment_data.get('remarks'),
                    transaction_date=payment_data.get('transaction_date'),
                    bank_name=payment_data.get('bank_name'),
                    cheque_no=payment_data.get('cheque_no'),
                    paymentmode=payment_data.get('paymentmode'),
                    transactionID=payment_data.get('transaction_id')
                )

        # Update additional enrollment details
        additional_details, created = AdditionalEnrollmentDetails.objects.get_or_create(student=student)
        for field, data_field in [
            ('counselor_name', 'counselor_name'),
            ('reference_name', 'reference_name'),
            ('university_enrollment_id', 'university_enroll_number')
        ]:
            if data_field in request.data:
                setattr(additional_details, field, request.data[data_field])
        additional_details.save()

        # Update or add new documents
        documents_data = request.data.get('documents', [])
        for document_data in documents_data:
            document_id = document_data.get('id')
            if document_id:
                try:
                    # Update existing document using its ID
                    document = StudentDocuments.objects.get(id=document_id, student=student)
                    for field in ['document', 'document_name', 'document_ID_no', 'document_image_front', 'document_image_back']:
                        if field in document_data:
                            setattr(document, field, document_data[field])
                    document.save()
                except StudentDocuments.DoesNotExist:
                    return Response(
                        {"error": f"Document with id {document_id} not found for this student."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Check if a document with the same attributes already exists for the student
                existing_document = StudentDocuments.objects.filter(
                    student=student,
                    document=document_data.get('document'),
                    document_name=document_data.get('document_name'),
                    document_ID_no=document_data.get('document_ID_no')
                ).first()

                if existing_document:
                    # Update existing document
                    for field in ['document_image_front', 'document_image_back']:
                        if field in document_data:
                            setattr(existing_document, field, document_data[field])
                    existing_document.save()
                else:
                    # Create a new document if no match is found
                    StudentDocuments.objects.create(
                        student=student,
                        document=document_data.get('document'),
                        document_name=document_data.get('document_name'),
                        document_ID_no=document_data.get('document_ID_no'),
                        document_image_front=document_data.get('document_image_front'),
                        document_image_back=document_data.get('document_image_back'),
                    )

        # Update qualifications
        qualifications_data = request.data.get('qualifications', {})
        qualifications, created = Qualification.objects.get_or_create(student=student)

        for field in [
            'secondary_year', 'sr_year', 'under_year', 'post_year', 'mphil_year', 
            'secondary_board', 'sr_board', 'under_board', 'post_board', 'mphil_board',
            'secondary_percentage', 'sr_percentage', 'under_percentage', 'post_percentage', 'mphil_percentage',
            'secondary_document', 'sr_document', 'under_document', 'post_document', 'mphil_document'
        ]:
            if field in qualifications_data:
                setattr(qualifications, field, qualifications_data[field])

        others = qualifications_data.get('others', [])
        qualifications.others = others if others is not None else []
        qualifications.save()

        # Get updated student details after the update
        student_serializer = StudentSerializer(student)

        # Fetch enrollment details
        enrollment_data = {
            "course": enrollment.course.name if enrollment.course else None,
            "stream": enrollment.stream.name if enrollment.stream else None,
            "substream": enrollment.substream.name if enrollment.substream else None,
            "studypattern": enrollment.course_pattern,
            "session": enrollment.session,
            "entry_mode": enrollment.entry_mode,
            "total_semyear": enrollment.total_semyear,
            "semyear": enrollment.current_semyear,
        }

        # Fetch additional enrollment details
        additional_details_data = {
            "counselor_name": additional_details.counselor_name if additional_details else None,
            "reference_name": additional_details.reference_name if additional_details else None,
            "university_enroll_number": additional_details.university_enrollment_id if additional_details else None,
        }

        # Fetch student documents
        student_documents = StudentDocuments.objects.filter(student=student)
        student_documents_data = StudentDocumentsSerializerGET(student_documents, many=True).data

        # Fetch qualifications
        qualification_data = QualificationSerializer(qualifications).data if qualifications else None

        # Fetch latest payment details
        latest_payment = PaymentReciept.objects.filter(student=student).last()
        print(latest_payment,'latest_payment')
        payment_data = {
            "transaction_date": latest_payment.transaction_date if latest_payment else None,
            "cheque_no": latest_payment.cheque_no if latest_payment else None,
            "bank_name": latest_payment.bank_name if latest_payment else None,
            "paymentmode": latest_payment.paymentmode if latest_payment else None,
            "remarks": latest_payment.remarks if latest_payment else None,
        }  

        # Prepare response data
        response_data = {
            'message': 'Student details updated successfully.',
            'student': student_serializer.data,
            'enrollment': enrollment_data,
            'additional_details': additional_details_data,
            'documents': student_documents_data,
            'qualifications': qualification_data,
            'payment_details': payment_data
        }

        return Response(response_data, status=status.HTTP_200_OK)

    except Student.DoesNotExist:
        return Response(
            {"error": "Student not found."},
            status=status.HTTP_404_NOT_FOUND
        )
    except Course.DoesNotExist:
        return Response(
            {"error": "Course not found."},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Stream.DoesNotExist:
        return Response(
            {"error": "Stream not found."},
            status=status.HTTP_400_BAD_REQUEST
        )
    except SubStream.DoesNotExist:
        return Response(
            {"error": "SubStream not found."},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_registration_list(request):
  students = Student.objects.filter(is_quick_register=False, is_cancelled=False).order_by('-id')
  serializer = Student_Quick_RegisteredSerializer(students, many=True)
  return Response(serializer.data)  


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def quick_registration(request):
    if not request.user.is_superuser:
        return Response({"error": "You are not authorized to perform this action."}, status=403)

    data = request.data
    image = request.FILES.get('image')

    try:
        name = data.get('name')
        email = data.get('email')
        mobile = data.get('mobile_number')
        university_id = data.get('university')
        course_id = data.get('course')
        stream_id = data.get('stream')
        substream_id = data.get('substream')
        studypattern = data.get('studypattern', '').capitalize()
        semyear = data.get('semyear')
        session = data.get('session')
        entry_mode = data.get('entry_mode')

        if Student.objects.filter(mobile=mobile).exists():
            return Response({"error": "Mobile number already registered."}, status=400)
        if Student.objects.filter(email=email).exists():
            return Response({"error": "Email already registered."}, status=400)

        try:
            university = University.objects.get(id=university_id)
            course = Course.objects.get(id=course_id, university=university)
            stream = Stream.objects.get(id=stream_id, course=course)
            substream = SubStream.objects.get(id=substream_id, stream=stream) if substream_id else None
        except Exception as e:
            return Response({"error": f"Invalid foreign key: {e}"}, status=400)

        latest = Student.objects.order_by("-id").first()
        enrollment_id = int(latest.enrollment_id) + 1 if latest else 50000
        registration_id = int(latest.registration_id) + 1 if latest else 250000

        with transaction.atomic():
            student_serializer = StudentSerializer(data={
                "name": name,
                "email": email,
                "mobile": mobile,
                "enrollment_id": enrollment_id,
                "registration_id": registration_id,
                "dob": data.get('dob'),
                "university": university.id,
                "is_quick_register": True,
                "image": image
            })
            student_serializer.is_valid(raise_exception=True)
            student = student_serializer.save()
            if student:
              try:
                  User.objects.create(email=email, is_student=True, password=make_password(mobile))
              except Exception as e:
                  logger.warning(f"User creation failed: {e}")

            enrolled_serializer = EnrolledSerializer(data={
                "student": student.id,
                "course": course.id,
                "stream": stream.id,
                "substream": substream.id if substream else None,
                "course_pattern": studypattern,
                "session": session,
                "entry_mode": entry_mode,
                "total_semyear": int(stream.sem) * (2 if studypattern == "Semester" else 1),
                "current_semyear": semyear
            })
            enrolled_serializer.is_valid(raise_exception=True)
            enrolled_serializer.save()

            add_serializer = AdditionalEnrollmentDetailsSerializer(data={
                "student": student.id,
                "counselor_name": data.get("counselor_name"),
                "university_enrollment_id": data.get("university_enrollment_number"),
                "reference_name": data.get("reference_name")
            })
            add_serializer.is_valid(raise_exception=True)
            add_serializer.save()

            fees_model = SemesterFees if studypattern == "Semester" else YearFees
            fees = fees_model.objects.filter(stream=stream, substream=substream)
            for fee in fees:
                StudentFees.objects.create(
                    student=student, stream=stream, substream=substream,
                    studypattern=studypattern, tutionfees=fee.tutionfees,
                    examinationfees=fee.examinationfees, totalfees=fee.totalfees,
                    sem=fee.sem if studypattern == "Semester" else fee.year
                )

            try:
                latest_receipt = PaymentReciept.objects.latest('id')
                last_id = int(latest_receipt.transactionID.replace("TXT445FE", ""))
                transactionID = f"TXT445FE{last_id + 1}"
            except PaymentReciept.DoesNotExist:
                transactionID = "TXT445FE101"

            total_fees = float(fees.last().totalfees) if fees.exists() else 0
            paid = float(data.get("paidamount") or 0)
            pending = total_fees - paid
            advance = abs(pending) if pending < 0 else 0
            pending = pending if pending > 0 else 0

            PaymentReciept.objects.create(
                student=student,
                payment_for=data.get("fee_recipt"),
                payment_categories="New",
                payment_type="Full Payment" if pending == 0 else "Part Payment",
                fee_reciept_type=data.get("other_data") if data.get("fee_reciept_type") == "Others" else data.get("fee_reciept_type"),
                transaction_date=data.get("transaction_date"),
                cheque_no=data.get("cheque_no"),
                bank_name=data.get("other_bank") if data.get("bank_name") == "Others" else data.get("bank_name"),
                semyearfees=total_fees,
                paidamount=0 if data.get("payment_mode") == "Cheque" else paid,
                pendingamount=pending,
                advanceamount=advance,
                transactionID=transactionID,
                paymentmode=data.get("payment_mode"),
                remarks=data.get("remarks"),
                session=session,
                semyear="1" if studypattern == "Full Course" else semyear,
                uncleared_amount=paid if data.get("payment_mode") == "Cheque" else None,
                status="Not Realised" if data.get("payment_mode") == "Cheque" else "Realised"
            )

        return Response({"message": "Student registered successfully."}, status=201)

    except Exception as e:
        logger.exception("Unexpected error during quick registration.")
        return Response({"error": str(e)}, status=500)





# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def create_subject(request):
#     #logger.info("Received request to create subject.")
    
#     # Extract input data from the request
#     subject_data = {
#         'name': request.data.get('name'),
#         'code': request.data.get('code'),
#         'stream_id': request.data.get('stream_id'),  # Ensure you're passing the ID here
#         'substream_id': request.data.get('substream_id'),  # This may be null or missing
#         'studypattern': request.data.get('studypattern'),
#         'semyear': request.data.get('semyear'),
#     }
#     logger.debug(f"Extracted subject data: {subject_data}")

#     # Validate required fields
#     if not subject_data['name'] or not subject_data['code']:
#         logger.warning("Subject name and code are required but missing.")
#         return Response({'error': 'Subject name and code are required.'}, status=status.HTTP_400_BAD_REQUEST)

#     # Get the stream object
#     try:
#         stream = Stream.objects.get(id=subject_data['stream_id'])
#         #logger.info(f"Stream with id {subject_data['stream_id']} found.")
#     except Stream.DoesNotExist:
#         logger.error(f"Stream with id {subject_data['stream_id']} not found.")
#         return Response({'error': 'Stream not found.'}, status=status.HTTP_404_NOT_FOUND)

#     # Get the substream object (optional)
#     substream = None
#     if subject_data['substream_id']:
#         try:
#             substream = SubStream.objects.get(id=subject_data['substream_id'], stream=stream)
#             #logger.info(f"Substream with id {subject_data['substream_id']} found for the stream.")
#         except SubStream.DoesNotExist:
#             logger.error(f"Substream with id {subject_data['substream_id']} not found for stream with id {subject_data['stream_id']}.")
#             return Response({'error': 'SubStream not found for the specified stream.'}, status=status.HTTP_404_NOT_FOUND)
#     else:
#         #logger.info("No substream ID provided, substream will be set to null.")

#     # Check for duplicate Subject
#     if Subject.objects.filter(
#         name=subject_data['name'],
#         code=subject_data['code'],
#         stream=stream,
#         substream=substream
#     ).exists():
#         logger.warning(f"Duplicate subject found: {subject_data['name']} with code {subject_data['code']} in the specified stream/substream.")
#         return Response(
#             {'error': 'Subject with this name and code already exists for the specified stream/substream.'},
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # Pass stream and substream ID to the serializer context
#     subject_data['stream'] = subject_data['stream_id']  # Pass stream_id instead of the stream object
#     subject_data['substream'] = subject_data['substream_id']  # Pass substream_id instead of the substream object

#     serializer = SubjectSerializer(data=subject_data)
#     if serializer.is_valid():
#         serializer.save()
#         #logger.info(f"Subject '{subject_data['name']}' with code '{subject_data['code']}' successfully created.")
#         return Response(serializer.data, status=status.HTTP_201_CREATED)

#     # Handle validation errors
#     logger.error(f"Validation failed: {serializer.errors}")
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_subject(request):
    #logger.info("Received request to create subject.")
    
    # Extract input data from the request
    subject_data = {
        'name': request.data.get('name'),
        'code': request.data.get('code'),
        'stream_id': request.data.get('stream_id'),
        'substream_id': request.data.get('substream_id'),
        'studypattern': request.data.get('studypattern'),
        'semyear': request.data.get('semyear'),
    }
    logger.debug(f"Extracted subject data: {subject_data}")

    # Validate required fields
    if not subject_data['name'] or not subject_data['code']:
        logger.warning("Subject name and code are required but missing.")
        return Response({'error': 'Subject name and code are required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Get the stream object
    try:
        stream = Stream.objects.get(id=subject_data['stream_id'])
        #logger.info(f"Stream with id {subject_data['stream_id']} found.")
    except Stream.DoesNotExist:
        logger.error(f"Stream with id {subject_data['stream_id']} not found.")
        return Response({'error': 'Stream not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Get the substream object (optional)
    substream = None
    if subject_data['substream_id']:
        try:
            substream = SubStream.objects.get(id=subject_data['substream_id'], stream=stream)
            #logger.info(f"Substream with id {subject_data['substream_id']} found for the stream.")
        except SubStream.DoesNotExist:
            logger.error(f"Substream with id {subject_data['substream_id']} not found for stream with id {subject_data['stream_id']}.")
            return Response({'error': 'SubStream not found for the specified stream.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        pass
        #logger.info("No substream ID provided, substream will be set to null.")

    # ✅ Improved duplicate check logic
    duplicate_subject = Subject.objects.filter(
        name__iexact=subject_data['name'],
        code__iexact=subject_data['code'],
        stream=stream,
        substream=substream,
        studypattern__iexact=subject_data['studypattern'],
        semyear__iexact=subject_data['semyear']
    ).first()

    if duplicate_subject:
        logger.warning(f"Subject already exists: {subject_data['name']} ({subject_data['code']})")
        return Response({'msg': 'Subject already exists'}, status=status.HTTP_200_OK)

    # Prepare data for serializer
    subject_data['stream'] = subject_data['stream_id']
    subject_data['substream'] = subject_data['substream_id']

    serializer = SubjectSerializer(data=subject_data)
    if serializer.is_valid():
        serializer.save()
        #logger.info(f"Subject '{subject_data['name']}' with code '{subject_data['code']}' successfully created.")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    logger.error(f"Validation failed: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from django.http import HttpResponse,FileResponse
from io import BytesIO

def validate_row(row):
    errors = []
    if not row['University']:
        errors.append("University not found")
    if not row['Course']:
        errors.append("Course not found")
    if not row['Stream']:
        errors.append("Stream not found")
    if row['email'] and Student.objects.filter(email=row['email']).exists():
        errors.append("Email already exists")
    if row['mobile'] and Student.objects.filter(mobile=row['mobile']).exists():
        errors.append("Mobile number already exists")
    return errors

from openpyxl import load_workbook
from datetime import datetime

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def bulk_student_upload(request):
#     if 'upload_file' not in request.FILES:
#         return Response({"error": "No file uploaded."}, status=400)

#     file = request.FILES['upload_file']
#     try:
#         workbook = load_workbook(file)
#         sheet = workbook.active
#     except Exception as e:
#         return Response({"error": f"Invalid Excel file: {str(e)}"}, status=400)

#     errors = []
#     successes = []
#     row_number = 1

#     for row in sheet.iter_rows(min_row=2, values_only=True):
#       row_number += 1
#       if all(value is None for value in row):
#           continue  # Skip empty rows

#       try:
#           # Unpack row data
#           (
#               name, date_of_birth, mobile_number, email, university_name, course_name, stream_name,
#               substream_name, current_semyear, admission_type, session, course_pattern,
#               total_semyear, country_name, state_name, city_name
#           ) = row

#           # Process date_of_birth
#           if isinstance(date_of_birth, datetime):
#               date_of_birth = date_of_birth.date()  # Convert to date object if it's already datetime
#           else:
#               date_of_birth = datetime.strptime(date_of_birth, "%d-%m-%Y").date()  # Parse string

#           print(f"Processing Row {row_number}: {row}")

#           # Validate required fields
#           if not all([name, date_of_birth, mobile_number, email, university_name, course_name, stream_name]):
#               errors.append(f"Row {row_number}: Missing required fields.")
#               continue

#           # Fetch related models
#           university = University.objects.filter(university_name=university_name).first()
#           course = Course.objects.filter(name=course_name).first()
#           stream = Stream.objects.filter(name=stream_name).first()
#           substream = SubStream.objects.filter(name=substream_name).first() if substream_name else None
#           country = Countries.objects.filter(name=country_name).first()
#           state = States.objects.filter(name=state_name).first()
#           city = Cities.objects.filter(name=city_name).first()

#           print(f"Lookups for Row {row_number}: University={university}, Course={course}, Stream={stream}, "
#                 f"SubStream={substream}, Country={country}, State={state}, City={city}")

#           if not all([university, course, stream, country, state, city]):
#               errors.append(f"Row {row_number}: Invalid data for required fields.")
#               continue

#           # Generate enrollment and registration IDs
#           try:
#               last_student = Student.objects.latest('id')
#               enrollment_id = int(last_student.enrollment_id) + 1
#               registration_id = int(last_student.registration_id) + 1
#           except Student.DoesNotExist:
#               enrollment_id = 50000
#               registration_id = 250000

#           # Create student
#           user = User(
#               email=email,
#               is_student=True,
#               password=make_password(email),
#           )
#           user.save()

#           student = Student(
#               name=name,
#               dateofbirth=date_of_birth,
#               mobile=mobile_number,
#               email=email,
#               university=university,
#               enrollment_id=enrollment_id,
#               registration_id=registration_id,
#               country=country,
#               state=state,
#               city=city,
#               student_remarks='Bulk Data Upload',
#               verified=True,
#               is_enrolled=True,
#               user=user  # Associate the student with the created user

#           )
#           student.save()

#           # Create enrollment details
#           enrollment = Enrolled(
#               student=student,
#               course=course,
#               stream=stream,
#               substream=substream,
#               current_semyear=current_semyear,
#               total_semyear=total_semyear,
#               course_pattern=course_pattern.capitalize(),
#               session=session,
#               entry_mode=admission_type,
#           )
#           enrollment.save()

#           successes.append(f"Row {row_number}: Student data added successfully.")
#       except Exception as e:
#           errors.append(f"Row {row_number}: {str(e)}")
#     return Response({"success": successes, "errors": errors})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_student_upload(request):
    if 'upload_file' not in request.FILES:
        return Response({"errors": ["No file uploaded."]}, status=400)

    file = request.FILES['upload_file']
    try:
        workbook = load_workbook(file)
        sheet = workbook.active
    except Exception as e:
        logger.error(f"Invalid Excel file: {str(e)}")
        return Response({"errors": [f"Invalid Excel file: {str(e)}"]}, status=400)

    errors = []
    successes = []
    row_number = 1

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_number += 1
        if all(value is None for value in row):
            continue

        try:
            (
                name, date_of_birth, mobile_number, email, university_name, course_name, stream_name,
                substream_name, current_semyear, admission_type, session, course_pattern,
                total_semyear, country_name, state_name, city_name
            ) = row

            if isinstance(date_of_birth, datetime):
                date_of_birth = date_of_birth.date()
            else:
                date_of_birth = datetime.strptime(date_of_birth, "%d-%m-%Y").date()

            #logger.info(f"Processing Row {row_number}: {row}")

            required_fields = [name, date_of_birth, mobile_number, email, university_name, course_name, stream_name]
            if not all(required_fields):
                missing = [field_name for field_name, value in zip(
                    ["Name", "Date of Birth", "Mobile Number", "Email", "University Name", "Course Name", "Stream Name"],
                    required_fields
                ) if not value]
                error_msg = f"Row {row_number}: Missing required fields: {', '.join(missing)}."
                logger.warning(error_msg)
                errors.append(error_msg)
                continue

            mobile_number_str = str(mobile_number).strip()
            email = str(email).strip().lower()

            # Check if user with same email already exists
            if User.objects.filter(email=email).exists():
                msg = f"Row {row_number}: Email '{email}' already exists. Skipping entry."
                logger.warning(msg)
                errors.append(msg)
                continue

            # Check if student already exists (based on mobile or email)
            if Student.objects.filter(email=email).exists():
                msg = f"Row {row_number}: Student with email '{email}' already exists. Skipping entry."
                logger.warning(msg)
                errors.append(msg)
                continue

            if Student.objects.filter(mobile=mobile_number_str).exists():
                msg = f"Row {row_number}: Student with mobile '{mobile_number_str}' already exists. Skipping entry."
                logger.warning(msg)
                errors.append(msg)
                continue

            university = University.objects.filter(university_name=university_name).first()
            course = Course.objects.filter(name=course_name, university=university).first()
            stream = Stream.objects.filter(name=stream_name, course=course).first()
            substream = SubStream.objects.filter(name=substream_name, stream=stream).first() if substream_name else None
            country = Countries.objects.filter(name=country_name).first()
            state = States.objects.filter(name=state_name).first()
            city = Cities.objects.filter(name=city_name).first()

            missing_refs = []
            if not university: missing_refs.append("University")
            if not course: missing_refs.append("Course")
            if not stream: missing_refs.append("Stream")
            if not country: missing_refs.append("Country")
            if not state: missing_refs.append("State")
            if not city: missing_refs.append("City")

            if missing_refs:
                msg = f"Row {row_number}: Invalid references for fields: {', '.join(missing_refs)}."
                logger.warning(msg)
                errors.append(msg)
                continue

            try:
                last_student = Student.objects.latest('id')
                enrollment_id = int(last_student.enrollment_id) + 1
                registration_id = int(last_student.registration_id) + 1
            except Student.DoesNotExist:
                enrollment_id = 50000
                registration_id = 250000

            # Create User
            user = User.objects.create(
                email=email,
                mobile=mobile_number_str,
                is_student=True,
                password=make_password(mobile_number_str)
            )

            # Create Student
            student = Student.objects.create(
                name=name,
                dateofbirth=date_of_birth,
                mobile=mobile_number_str,
                email=email,
                university=university,
                enrollment_id=enrollment_id,
                registration_id=registration_id,
                country=country,
                state=state,
                city=city,
                student_remarks="Bulk Data Upload",
                verified=True,
                is_enrolled=True,
                user=user
            )

            # Enroll student
            Enrolled.objects.create(
                student=student,
                course=course,
                stream=stream,
                substream=substream,
                current_semyear=current_semyear,
                total_semyear=total_semyear,
                course_pattern=(course_pattern or "").capitalize(),
                session=session,
                entry_mode=admission_type,
            )

            msg = f"Row {row_number}: Student '{name}' added successfully."
            #logger.info(msg)
            successes.append(msg)

        except Exception as e:
            error_msg = f"Row {row_number}: Unexpected error: {str(e)}"
            logger.error(error_msg, exc_info=True)
            errors.append(error_msg)

    return Response({
        "success": successes,
        "errors": errors
    }, status=200)

#-----------------------------------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def fetch_subject(request):
    try:
        stream_id = request.query_params.get('stream')
        substream_id = request.query_params.get('substream')  # Optional
        study_pattern = request.query_params.get('study_pattern')
        semyear = request.query_params.get('semyear')
        
        # Initialize the queryset for subjects
        subject_query = Subject.objects.filter(
            stream_id=stream_id,  # Filtering by stream
            studypattern=study_pattern,  # Filtering by study pattern
            semyear=semyear  # Filtering by semester/year
        )

        # If substream is provided, filter by substream
        if substream_id:
            subject_query = subject_query.filter(substream_id=substream_id)
        else:
            # If substream is not provided, include subjects that don't have a substream
            subject_query = subject_query.filter(substream_id__isnull=True)

        # Check if subjects exist for the given parameters
        if not subject_query.exists():
            logger.warning("No subjects found for the given parameters.")
            return Response({"message": "No subjects found."}, status=status.HTTP_404_NOT_FOUND)

        # Serialize the subject data
        subject_data = SubjectSerializer(subject_query, many=True)

        # Return the subjects data
        return Response({"data": subject_data.data}, status=status.HTTP_200_OK)

    except Exception as e:
        # Log the error if an exception occurs
        logger.error("An error occurred while fetching subjects: %s", str(e), exc_info=True)
        return Response({"message": "An error occurred while processing your request."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
def validate_excel_file(file_path):
    required_columns = [
        "COURSE", "STREAM", "SUBSTREAM", "SESSION", "MODE", "YEAR/SEMESTER",
        "SUBJECT NAME", "TYPE OF EXAM", "QUESTION TYPE", "DIFFICULTY LEVEL",
        "QUESTION", "OPTION 1", "OPTION 2", "OPTION 3", "OPTION 4",
        "ANSWER", "MARKS", "EXAM DURATION", "PASSING MARKS"
    ]
    try:
        data = pd.read_excel(file_path)
        missing_columns = [col for col in required_columns if col not in data.columns]
        if missing_columns:
            error_message = f"Missing columns in Excel file: {', '.join(missing_columns)}"
            logger.error(error_message)
            return None, error_message
        return data, None
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
        return None, "Invalid Excel file"

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def bulk_exam_upload(request):
    try:
        if 'file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded"}, status=400)

        file = request.FILES['file']
        temp_file_path = os.path.join(settings.MEDIA_ROOT, 'temp', file.name)
        os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)
        with open(temp_file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        # Validate the uploaded Excel file
        data, error = validate_excel_file(temp_file_path)
        if error:
            return Response({"status": "error", "message": error}, status=400)

        university_id = request.data.get("university")
        try:
            university = University.objects.get(id=university_id)
        except University.DoesNotExist:
            return Response({"status": "error", "message": f"University with ID {university_id} not found"}, status=400)

        errors = []
        success_rows = []

        with transaction.atomic():
            for index, row in data.iterrows():
                try:
                    course_name = str(row["COURSE"]).strip()
                    stream_name = str(row["STREAM"]).strip()
                    subject_name = str(row["SUBJECT NAME"]).strip()
                    substream_name = str(row["SUBSTREAM"]).strip() if pd.notna(row["SUBSTREAM"]) and str(row["SUBSTREAM"]).strip().lower() != 'nan' else None
                    exam_semyear = str(row["YEAR/SEMESTER"]).strip()

                    course = Course.objects.get(name=course_name, university=university)
                    stream = Stream.objects.get(name=stream_name, course=course)

                    substream = None
                    if substream_name:
                        substream = SubStream.objects.get(name=substream_name, stream=stream)

                    subject = Subject.objects.filter(
                        name=subject_name,
                        stream=stream,
                        substream=substream,
                        semyear=exam_semyear
                    ).first()

                    if not subject:
                        raise Subject.DoesNotExist(
                            f"Subject '{subject_name}' with sem/year '{exam_semyear}' not found in row {index + 2}"
                        )

                    exam = Examination.objects.filter(
                        university=university,
                        course=course,
                        stream=stream,
                        substream=substream,
                        subject=subject,
                        examtype=str(row["TYPE OF EXAM"]).strip(),
                        session=str(row["SESSION"]).strip(),
                        studypattern=str(row["MODE"]).strip()
                    ).first()

                    if exam:
                        exam.totalquestions = (int(exam.totalquestions) if exam.totalquestions else 0) + 1
                        exam.totalmarks = (int(exam.totalmarks) if exam.totalmarks else 0) + int(row["MARKS"])
                        exam.save()
                    else:
                        exam = Examination.objects.create(
                            university=university,
                            course=course,
                            stream=stream,
                            substream=substream,
                            subject=subject,
                            examtype=str(row["TYPE OF EXAM"]).strip(),
                            examduration=str(row["EXAM DURATION"]).strip(),
                            studypattern=str(row["MODE"]).strip(),
                            semyear=exam_semyear,
                            session=str(row["SESSION"]).strip(),
                            totalquestions=1,
                            totalmarks=row["MARKS"],
                            passingmarks=row["PASSING MARKS"],
                            created_by="Bulk File Uploaded",
                            active=True,
                            archive=False,
                        )

                    # Create Questions
                    Questions.objects.create(
                        exam=exam,
                        question=str(row["QUESTION"]).strip(),
                        type=str(row["QUESTION TYPE"]).strip(),
                        marks=int(row["MARKS"]),
                        option1=str(row["OPTION 1"]).strip(),
                        option2=str(row["OPTION 2"]).strip(),
                        option3=str(row["OPTION 3"]).strip(),
                        option4=str(row["OPTION 4"]).strip(),
                        answer=str(row["ANSWER"]).strip().lower(),
                        difficultylevel=str(row["DIFFICULTY LEVEL"]).strip(),
                    )

                    success_rows.append(f"Row {index + 2}: Question added successfully.")

                except (Course.DoesNotExist, Stream.DoesNotExist, SubStream.DoesNotExist, Subject.DoesNotExist) as e:
                    error_message = f"{str(e)} in row {index + 2}"
                    logger.error(error_message)
                    errors.append(error_message)
                    continue
                except Exception as e:
                    error_message = f"Row {index + 2}: Unexpected error: {str(e)}"
                    logger.error(error_message, exc_info=True)
                    errors.append(error_message)
                    continue

        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)

        return Response({
            "status": "success",
            "message": "Data processed successfully",
            "success_logs": success_rows,
            "error_logs": errors
        }, status=200)

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}", exc_info=True)
        return Response({"status": "error", "message": "Internal server error"}, status=500)

logger = logging.getLogger('student_registration')
@api_view(['GET'])
def filter_questions(request):
    try:
        exam_id = request.query_params.get('exam_id')

        if not exam_id:
            return Response({"error": "Missing exam_id in request"}, status=status.HTTP_400_BAD_REQUEST)

        questions_queryset = Questions.objects.filter(exam__id=exam_id).distinct()
        serializer = QuestionsSerializer(questions_queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error occurred while filtering questions: {str(e)}")
        return Response({"error": "Something went wrong!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# @api_view(['POST'])
# def fetch_exam(request):
#     try:
#         university = request.data.get("university")
#         course = request.data.get("course")
#         stream = request.data.get("stream") or request.data.get("Stream")  # handle both 'stream' and 'Stream'
#         session = request.data.get("session")
#         studypattern = request.data.get("studypattern")
#         semyear = request.data.get("semyear")
#         substream = request.data.get("substream")  # May be "" or actual ID

#         if substream == "":
#             substream = None

#         # === BUILD EXAM FILTERS (dynamic Q) ===
#         exam_filters = Q(university_id=university) & Q(course_id=course) & Q(stream_id=stream) & Q(active=True)

#         if session not in ["", None, "0"]:
#             exam_filters &= Q(session=session)

#         if studypattern not in ["", None, "0"]:
#             exam_filters &= Q(studypattern=studypattern)

#         if semyear not in ["", None, "0"]:
#             exam_filters &= Q(semyear=semyear)

#         if substream is None:
#             exam_filters &= (Q(substream__isnull=True) | Q(substream_id__isnull=True))
#         else:
#             exam_filters &= Q(substream_id=substream)

#         examinations = Examination.objects.filter(exam_filters).select_related("subject")

#         if not examinations.exists():
#             return Response({"message": "No examinations found matching the criteria."}, status=404)

#         # === BUILD STUDENT FILTERS (dynamic Q) ===
#         student_filters = Q(course_id=course) & Q(stream_id=stream)

#         if session not in ["", None, "0"]:
#             student_filters &= Q(session=session)

#         if studypattern not in ["", None, "0"]:
#             student_filters &= Q(course_pattern=studypattern)

#         if semyear not in ["", None, "0"]:
#             student_filters &= Q(current_semyear=semyear)

#         if substream is None:
#             student_filters &= (Q(substream__isnull=True) | Q(substream_id__isnull=True))
#         else:
#             student_filters &= Q(substream_id=substream)

#         enrolled_students = Enrolled.objects.filter(student_filters).select_related('student')

#         students_list = []
#         for enrollment in enrolled_students:
#             try:
#                 student_obj = Student.objects.get(id=enrollment.student.id)
#                 students_list.append(student_obj)
#             except Student.DoesNotExist:
#                 continue

#         # === SERIALIZE DATA ===
#         student_serializer = StudentSerializer(students_list, many=True)
#         exam_serializer = ExaminationSubjectSerializer(examinations, many=True)

#         return Response({
#             "studentdata": student_serializer.data,
#             "exams": exam_serializer.data
#         }, status=200)

#     except Exception as e:
#         logger.error(f"Error in fetch_exam API: {str(e)}")
#         return Response({"error": "An error occurred while processing the request."}, status=500)


# @api_view(['POST'])
# def fetch_exam(request):
#     try:
#         university = request.data.get("university")
#         course = request.data.get("course")
#         stream = request.data.get("stream") or request.data.get("Stream")  # Allow both
#         session = request.data.get("session")
#         studypattern = request.data.get("studypattern")
#         semyear = request.data.get("semyear")
#         substream = request.data.get("substream")

#         # Normalize empty substream
#         if not substream or substream in ["", "null", "None", None, 0, "0"]:
#             substream = None

#         # === EXAM FILTERS ===
#         exam_filters = Q(university_id=university) & Q(course_id=course) & Q(stream_id=stream) & Q(active=True)

#         if session:
#             exam_filters &= Q(session=session)

#         if studypattern:
#             exam_filters &= Q(studypattern=studypattern)

#         if semyear:
#             exam_filters &= Q(semyear=semyear)

#         if substream is None:
#             exam_filters &= Q(substream__isnull=True)
#         else:
#             exam_filters &= Q(substream_id=substream)

#         examinations = Examination.objects.filter(exam_filters).select_related("subject")

#         if not examinations.exists():
#             return Response({"message": "No examinations found matching the criteria."}, status=404)

#         # === STUDENT FILTERS ===
#         student_filters = Q(course_id=course) & Q(stream_id=stream)

#         if session:
#             student_filters &= Q(session=session)

#         if studypattern:
#             student_filters &= Q(course_pattern=studypattern)

#         if semyear:
#             student_filters &= Q(current_semyear=semyear)

#         if substream is None:
#             student_filters &= Q(substream__isnull=True)
#         else:
#             student_filters &= Q(substream_id=substream)

#         enrolled_students = Enrolled.objects.filter(student_filters).select_related('student')

#         students_list = [enr.student for enr in enrolled_students if enr.student]

#         # === SERIALIZE ===
#         student_serializer = StudentSerializer(students_list, many=True)
#         exam_serializer = ExaminationSubjectSerializer(examinations, many=True)

#         return Response({
#             "studentdata": student_serializer.data,
#             "exams": exam_serializer.data
#         }, status=200)

#     except Exception as e:
#         logger.error(f"Error in fetch_exam API: {str(e)}")
#         return Response({"error": "An error occurred while processing the request."}, status=500)


@api_view(['POST'])
def fetch_exam(request):
    try:
        university = request.data.get("university")
        course = request.data.get("course")
        stream = request.data.get("stream") or request.data.get("Stream")
        session = request.data.get("session")
        studypattern = request.data.get("studypattern")
        semyear = request.data.get("semyear")
        substream = request.data.get("substream")

        # Treat blank/invalid substream as no filter
        apply_substream_filter = False
        if substream and str(substream).strip().isdigit():
            apply_substream_filter = True

        # === BASE EXAM FILTERS ===
        exam_filters = Q(university_id=university) & Q(course_id=course) & Q(stream_id=stream) & Q(active=True)

        if session:
            exam_filters &= Q(session=session)
        if studypattern:
            exam_filters &= Q(studypattern=studypattern)
        if semyear:
            exam_filters &= Q(semyear=semyear)
        if apply_substream_filter:
            exam_filters &= Q(substream_id=substream)

        examinations = Examination.objects.filter(exam_filters).select_related("subject")

        if not examinations.exists():
            return Response({"message": "No examinations found matching the criteria."}, status=404)

        # === STUDENT FILTERS ===
        student_filters = Q(course_id=course) & Q(stream_id=stream)

        if session:
            student_filters &= Q(session=session)
        if studypattern:
            student_filters &= Q(course_pattern=studypattern)
        if semyear:
            student_filters &= Q(current_semyear=semyear)
        if apply_substream_filter:
            student_filters &= Q(substream_id=substream)

        enrolled_students = Enrolled.objects.filter(student_filters).select_related('student')
        students_list = [e.student for e in enrolled_students if e.student]

        student_serializer = StudentSerializer(students_list, many=True)
        exam_serializer = ExaminationSubjectSerializer(examinations, many=True)

        return Response({
            "studentdata": student_serializer.data,
            "exams": exam_serializer.data
        }, status=200)

    except Exception as e:
        logger.error(f"Error in fetch_exam API: {str(e)}", exc_info=True)
        return Response({"error": "An error occurred while processing the request."}, status=500)


      
logger = logging.getLogger('student_registration')
@api_view(['POST'])
def view_assigned_students(request):
    try:
        # Extract parameters
        params = {
            "university": request.data.get("university"),
            "course": request.data.get("course"),
            "stream": request.data.get("stream"),
            "session": request.data.get("session"),
            "studypattern": request.data.get("studypattern"),
            "semyear": request.data.get("semyear"),
            "subject": request.data.get("subject"),
        }

        # Handle substream separately
        substream = request.data.get("substream")
        if substream not in [None, "", "0", 0]:
            try:
                params["substream"] = int(substream)  # Safely cast to int
            except (ValueError, TypeError):
                logger.error(f"Invalid substream value: {substream}")
                return Response({'message': 'Invalid substream value provided.'}, status=400)

        # Remove None, empty or invalid entries
        filters = {key: value for key, value in params.items() if value not in [None, '', '0']}

        exams = Examination.objects.filter(**filters)

        if not exams.exists():
            return Response({'message': 'No results found', 'students': []}, status=404)

        formatted_data = []
        for exam in exams:
            student_data = StudentAppearingExam.objects.filter(exam=exam.id)
            if not student_data.exists():
                continue

            serializer = StudentAppearingExamSerializer(student_data, many=True)
            for record in serializer.data:
                for student_id in record['student_id']:
                    has_appeared = ExamSession.objects.filter(student_id=student_id, exam=exam.id).exists()
                    submitted_exam = Result.objects.filter(student_id=student_id, exam=exam).exists()

                    if has_appeared and submitted_exam:
                        status = "Appeared"
                    elif has_appeared:
                        status = "Ongoing Exam"
                    else:
                        status = "Not Appeared"

                    student = Student.objects.filter(id=student_id).first()
                    if student:
                        formatted_data.append({
                            'student_id': student_id,
                            'student_name': student.name,
                            'student_email': student.email,
                            'student_mobile': student.mobile,
                            'exam_id': record['exam'],
                            'examstarttime': record['examstarttime'],
                            'examendtime': record['examendtime'],
                            'examstartdate': record['examstartdate'],
                            'examenddate': record['examenddate'],
                            'id': record['id'],
                            'status': status,
                        })

        if not formatted_data:
            return Response({'message': 'No students assigned', 'students': []}, status=404)

        return Response({'message': 'Students fetched successfully', 'students': formatted_data}, status=200)

    except Exception as e:
        logger.error(f"Error fetching assigned students: {e}", exc_info=True)
        return Response({'message': 'An error occurred while fetching assigned students'}, status=500)

from django.core.mail import send_mail, EmailMessage
from threading import Thread

logger = logging.getLogger('student_registration')
def send_exam_email(subject, message, recipient_list):
    """
    Sends an email asynchronously.
    """
    try:
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipient_list,
        )
        email.send(fail_silently=False)
        #logger.info(f"Email sent successfully to {recipient_list}")
    except Exception as e:
        logger.error(f"Failed to send email to {recipient_list}. Error: {e}")

@api_view(['POST'])
def resend_exam_email(request):
    data = request.data
    student_id = data.get("studentid")
    exam_id = data.get("examid")
    exam_start_date = data.get("examstartdate")
    exam_end_date = data.get("examenddate")
    exam_start_time = data.get("examstarttime")
    exam_end_time = data.get("examendtime")

    if not all([student_id, exam_id, exam_start_date, exam_end_date, exam_start_time, exam_end_time]):
        logger.error("Required fields are missing in the request.")
        return Response(
            {"error": "Student ID, Exam ID, Exam Start/End Dates, and Start/End Times are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        student = Student.objects.get(id=int(student_id))
        exam = Examination.objects.get(id=int(exam_id))
        subject = Subject.objects.get(id=exam.subject_id)

        university_name = student.university.university_name if student.university else "Your University"

        email_subject = f"{university_name} Examination - Exam Details"
        login_url = f"{settings.DOMAIN_NAME}"
        email_message = f"""Dear {student.name},

Your examination details are as follows:

Subject: {subject.name}
Exam Start Date: {exam_start_date}
Exam End Date: {exam_end_date}
Exam Start Time: {exam_start_time}
Exam End Time: {exam_end_time}

Please log in using the credentials below:
Email: {student.email}
Password: {student.mobile}

Click here to log in:
{login_url}

Best regards,
{university_name} Examination Team
"""

        Thread(target=send_exam_email, args=(email_subject, email_message, [student.email])).start()

        return Response({"message": "Email resent successfully."}, status=status.HTTP_200_OK)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} does not exist.")
        return Response({"error": "Invalid student ID."}, status=status.HTTP_400_BAD_REQUEST)
    except Examination.DoesNotExist:
        logger.error(f"Examination with ID {exam_id} does not exist.")
        return Response({"error": "Invalid exam ID."}, status=status.HTTP_400_BAD_REQUEST)
    except Subject.DoesNotExist:
        logger.error(f"Subject for exam ID {exam_id} does not exist.")
        return Response({"error": "Invalid subject for the given exam."}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.exception(f"Error occurred while resending email: {e}")
        return Response({"error": "An error occurred while resending the email."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


        
@api_view(['POST'])
def save_exam_details(request):
    try:
        examdata = request.data.get("examsdata")
        studentdata = request.data.get("studentdata")

        if not examdata:
            return Response({'message': 'Exam data is required'}, status=400)
        if not studentdata:
            return Response({'message': 'Student data is required'}, status=400)

        messages = []
        errors = []

        for exam in examdata:
            exam_id = exam.get("examination_id")
            start_date = exam.get("start_date")
            end_date = exam.get("end_date")
            start_time = exam.get("start_time")
            end_time = exam.get("end_time")

            if not all([exam_id, start_date, end_date, start_time, end_time]):
                errors.append("Incomplete exam details provided.")
                continue

            try:
                exam_instance = Examination.objects.get(id=exam_id)

                existing_exam, created = StudentAppearingExam.objects.get_or_create(
                    exam=exam_instance,
                    examstartdate=start_date,
                    examenddate=end_date,
                    examstarttime=start_time,
                    examendtime=end_time,
                    defaults={"student_id": []},
                )

                current_students = existing_exam.student_id or []
                new_student_ids = [student['id'] for student in studentdata if student['id'] not in current_students]

                if new_student_ids:
                    current_students.extend(new_student_ids)
                    existing_exam.student_id = current_students
                    existing_exam.save()

                    for student_id in new_student_ids:
                        try:
                            student_instance = Student.objects.get(id=student_id)
                            messages.append(f"Exam details saved successfully for {student_instance.name}")

                            email_subject = "Examination Details"
                            login_url = f"{settings.DOMAIN_NAME}"
                            email_message = f"""
                            Dear {student_instance.name},

                            You are invited to take the {exam_instance.subject.name} {exam_instance.studypattern} {exam_instance.semyear} Test.
                            Exam Link: {login_url}
                            User ID: {student_instance.email}
                            Password: {student_instance.mobile}

                            The exam is available from {start_date} to {end_date} between {start_time} and {end_time}.
                            """

                            Thread(
                                target=send_exam_email,
                                args=(email_subject, email_message, [student_instance.email])
                            ).start()

                        except Student.DoesNotExist:
                            errors.append(f"Student with ID {student_id} not found.")
                        except Exception as email_error:
                            errors.append(f"Failed to send email to student ID {student_id}: {str(email_error)}")

            except Examination.DoesNotExist:
                errors.append(f"Invalid exam ID: {exam_id}")
            except Exception as e:
                errors.append(f"Failed to save exam details for exam ID {exam_id}: {str(e)}")

        return Response({'messages': messages, 'errors': errors}, status=200)

    except Exception as e:
        return Response({'message': 'An unexpected error occurred.', 'error': str(e)}, status=500)

      
      
@api_view(['GET'])
def view_set_examination(request):
    try:
        # Extract query parameters
        query_params = {
            'university': request.query_params.get('university'),
            'course': request.query_params.get('course'),
            'stream': request.query_params.get('stream'),
            'substream': request.query_params.get('substream'),
            'session': request.query_params.get('session'),
            'studypattern': request.query_params.get('studypattern'),
            'semyear': request.query_params.get('semyear')
        }

        # Initialize the examinations queryset
        examinations = Examination.objects.all()

        # Check for empty or invalid parameters before filtering
        for param, value in query_params.items():
            if value is not None and value != '':  # Only filter if the parameter is provided and non-empty
                # Dynamically filter the queryset based on the query parameters
                # Ensure the value is a number for parameters like course, stream, university
                if param in ['university', 'course', 'stream', 'substream'] and value.isdigit():
                    examinations = examinations.filter(**{param: value})
                else:
                    # Handle case where value is invalid (non-numeric when a number is expected)
                    if param in ['university', 'course', 'stream', 'substream']:
                        logger.error(f"Invalid {param}: {value}")
                        return Response({"error": f"Invalid {param} parameter value."}, status=status.HTTP_400_BAD_REQUEST)

        # If `substream` is passed, apply that filter too
        if query_params['substream'] is not None and query_params['substream'] != '':
            examinations = examinations.filter(substream=query_params['substream'])

        # If no substream is provided, show examinations for the same course, stream, and university, but no substream
        if query_params['substream'] is None or query_params['substream'] == '':
            examinations = examinations.filter(substream__isnull=True)

        # Serialize the filtered examination data
        serializer = ExaminationSerializer(examinations, many=True)


        # Return the serialized data in the response
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    except Exception as e:
        # Log the error if an exception occurs
        logger.error(f"Error occurred while fetching examinations: {str(e)}")
        
        # Return a structured error response
        return Response(
            {"error": "An error occurred while processing your request."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# @api_view(['GET'])
# def view_set_examination(request):
#     try:
#         # Extract query parameters
#         university = request.query_params.get('university')
#         course = request.query_params.get('course')
#         stream = request.query_params.get('stream')
#         substream = request.query_params.get('substream')
#         session = request.query_params.get('session')
#         studypattern = request.query_params.get('studypattern')
#         semyear = request.query_params.get('semyear')

#         # Validate semyear presence
#         if not semyear or semyear.strip() == '':
#             return Response({"error": "semyear parameter is required."}, status=400)
#         semyear = semyear.strip()

#         # Initialize queryset
#         examinations = Examination.objects.all()

#         # Filter by semyear (case insensitive exact match)
#         examinations = examinations.filter(semyear__iexact=semyear)

#         # Filter other numeric parameters if valid
#         for param in ['university', 'course', 'stream', 'substream']:
#             value = locals()[param]
#             if value and value.strip() != '':
#                 if not value.isdigit():
#                     return Response({"error": f"Invalid {param} parameter value."}, status=400)
#                 examinations = examinations.filter(**{param: value.strip()})

#         # Filter session and studypattern if provided
#         if session and session.strip() != '':
#             examinations = examinations.filter(session=session.strip())
#         if studypattern and studypattern.strip() != '':
#             examinations = examinations.filter(studypattern=studypattern.strip())

#         # If substream is empty or None, filter exams with substream NULL
#         if not substream or substream.strip() == '':
#             examinations = examinations.filter(substream__isnull=True)

#         serializer = ExaminationSerializer(examinations, many=True)
#         return Response(serializer.data, status=200)

#     except Exception as e:
#         logger.error(f"Error occurred while fetching examinations: {str(e)}")
#         return Response({"error": "An error occurred while processing your request."}, status=500)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_exam_for_subject(request):
    try:
        # Extract request data
        data = request.data
        university_id = data.get('university')
        course_id = data.get('course')
        stream_id = data.get('stream')
        substream_id = data.get('substream')
        subject_id = data.get('subject')
        examtype = data.get('examtype')
        examduration = data.get('examduration')
        session = data.get('session')
        studypattern = data.get('studypattern')
        semyear = data.get('semyear')
        totalmarks = data.get('totalmarks')
        passingmarks = data.get('passingmarks')
        created_by = data.get('created_by')

        # Validate dependencies
        university = University.objects.filter(id=university_id).first()
        if not university:
            logger.error(f"Invalid university ID: {university_id}")
            return Response({"error": "Invalid university ID."}, status=status.HTTP_400_BAD_REQUEST)

        course = Course.objects.filter(id=course_id, university=university).first()
        if not course:
            logger.error(f"Invalid course ID: {course_id} for university ID: {university_id}")
            return Response({"error": "Invalid course for the given university."}, status=status.HTTP_400_BAD_REQUEST)

        stream = Stream.objects.filter(id=stream_id, course=course).first()
        if not stream:
            logger.error(f"Invalid stream ID: {stream_id} for course ID: {course_id}")
            return Response({"error": "Invalid stream for the given course."}, status=status.HTTP_400_BAD_REQUEST)

        substream = None
        if substream_id:
            substream = SubStream.objects.filter(id=substream_id, stream=stream).first()
            if not substream:
                logger.error(f"Invalid substream ID: {substream_id} for stream ID: {stream_id}")
                return Response({"error": "Invalid substream for the given stream."}, status=status.HTTP_400_BAD_REQUEST)

        subject = Subject.objects.filter(id=subject_id, stream=stream, substream=substream).first()
        if not subject:
            logger.error(f"Invalid subject ID: {subject_id} for stream ID: {stream_id} and substream ID: {substream_id}")
            return Response({"error": "Invalid subject for the given stream and substream."}, status=status.HTTP_400_BAD_REQUEST)

        # Create Examination
        examination = Examination.objects.create(
            university=university,
            course=course,
            stream=stream,
            substream=substream,
            subject=subject,
            examtype=examtype,
            examduration=examduration,
            studypattern=studypattern,
            semyear=semyear,
            session=session,
            totalmarks=totalmarks,
            passingmarks=passingmarks,
            created_by=created_by
        )

        # Process Excel File
        excel_file = request.FILES.get('file')
        if not excel_file:
            logger.error("No Excel file uploaded.")
            return Response({"error": "Questions Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            data = pd.read_excel(excel_file)
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return Response({"error": "Invalid Excel file."}, status=status.HTTP_400_BAD_REQUEST)

        total_questions = 0
        error_count = 0
        error_messages = []
        required_columns = ['Question', 'Question Type', 'Marks', 'Option 1', 'Option 2', 'Option 3', 'Option 4', 'Answer', 'Difficulty Level']
        missing_columns = [col for col in required_columns if col not in data.columns]

        if missing_columns:
            logger.error(f"Missing columns in Excel: {', '.join(missing_columns)}")
            return Response({"error": f"Missing columns in Excel: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

        for idx, row in data.iterrows():
            try:
                # Check for missing mandatory fields
                if any(pd.isna(row.get(col)) for col in required_columns):
                    error_messages.append(f"Row {idx + 1}: Missing mandatory fields.")
                    error_count += 1
                    continue

                Questions.objects.create(
                    exam=examination,
                    question=row['Question'],
                    type=row['Question Type'],
                    marks=int(row['Marks']),
                    option1=row['Option 1'],
                    option2=row['Option 2'],
                    option3=row['Option 3'],
                    option4=row['Option 4'],
                    answer=row['Answer'].lower(),
                    difficultylevel=row['Difficulty Level']
                )
                total_questions += 1
            except Exception as e:
                logger.error(f"Error saving question at row {idx + 1}: {str(e)}")
                error_messages.append(f"Row {idx + 1}: {str(e)}")
                error_count += 1

        examination.totalquestions = total_questions
        examination.save()

        response_data = {
            "status": "success",
            "message": "Data processed successfully.",
            "total_success": total_questions,
            "total_errors": error_count,
            "errors": error_messages
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return Response({"error": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def delete_exam_for_student(request):
    try:
        if not (request.user.is_superuser or getattr(request.user, 'is_data_entry', False)):
            logger.warning(f"User {request.user} attempted unauthorized access to recall_exam_for_student.")
            return Response({"error": "You do not have permission to perform this action."}, status=403)
        
        student_id = request.data.get("studentid")
        exam_id = request.data.get("examid")
        
        if not student_id or not exam_id:
            logger.error("Missing required fields: studentid or examid.")
            return Response({"error": "Both studentid and examid are required."}, status=400)
        
        get_exam_data = StudentAppearingExam.objects.filter(id=exam_id).first()
        student = Student.objects.filter(id=student_id).first()
        
        if not get_exam_data:
            logger.error(f"Exam record with id {exam_id} not found.")
            return Response({"error": "Exam record not found."}, status=404)
        
        if not student:
            logger.error(f"Student record with id {student_id} not found.")
            return Response({"error": "Student record not found."}, status=404)
        
        list_of_students = get_exam_data.student_id
        
        if isinstance(list_of_students, list) and int(student_id) in list_of_students:
            university_name = student.university.university_name if student.university else "Your University"
            
            if len(list_of_students) > 1:
                list_of_students.remove(int(student_id))
                get_exam_data.student_id = list_of_students
                get_exam_data.save()
                
                email_subject = f"{university_name} Examination - Recall of Exam"
                email_message = f"An exam incorrectly assigned to you has been recalled by {university_name}."
                Thread(target=send_exam_email, args=(email_subject, email_message, [student.email])).start()
                
                return Response({"message": "Exam recalled for the student."}, status=200)
            else:
                get_exam_data.delete()
                
                email_subject = f"{university_name} Examination - Recall of Exam"
                email_message = f"An exam incorrectly assigned to you has been recalled by {university_name}."
                Thread(target=send_exam_email, args=(email_subject, email_message, [student.email])).start()
                
                return Response({"message": "Exam recalled and record deleted as only one student was present."}, status=200)
        else:
            logger.warning(f"Student ID {student_id} not found in the student list for exam {exam_id}.")
            return Response({"error": "Student ID not found in the list."}, status=404)
    except Exception as e:
        logger.error(f"Unexpected error in recall_exam_for_student: {str(e)}")
        return Response({"error": "An unexpected error occurred."}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reassign_student(request):
    if not (request.user.is_superuser or (hasattr(request.user, 'is_data_entry') and request.user.is_data_entry)):
        logger.warning(f"Unauthorized access attempt by user {request.user.id}")
        return Response({"message": "You do not have permission to perform this action."}, status=status.HTTP_403_FORBIDDEN)

    data = request.data
    student_id = data.get("studentid")
    exam_id = data.get("examid")
    exam_start_time = data.get("examstarttime")
    exam_end_time = data.get("examendtime")
    exam_start_date = data.get("examstartdate")
    exam_end_date = data.get("examenddate")

    try:
        exam = Examination.objects.get(id=int(exam_id))
        subject = Subject.objects.get(id=exam.subject_id)
    except Examination.DoesNotExist:
        logger.error(f"Examination with ID {exam_id} does not exist.")
        return Response({"message": "Invalid exam ID."}, status=status.HTTP_400_BAD_REQUEST)
    except Subject.DoesNotExist:
        logger.error(f"Subject for exam ID {exam_id} does not exist.")
        return Response({"message": "Invalid subject for the given exam."}, status=status.HTTP_400_BAD_REQUEST)
    except ValueError as e:
        logger.error(f"Invalid exam ID: {exam_id}. Error: {e}")
        return Response({"message": "Exam ID must be a valid integer."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        new_exam = StudentAppearingExam(
            exam=exam,
            examstartdate=exam_start_date,
            examenddate=exam_end_date,
            examstarttime=exam_start_time,
            examendtime=exam_end_time,
            student_id=[int(student_id)]
        )
        new_exam.save()

        student = Student.objects.get(id=student_id)
        university_name = student.university.university_name if student.university else "Your University"

        email_subject = f"{university_name} Examination - Reassign of Exam"
        login_url = f"{settings.DOMAIN_NAME}"
        email_message = f"""Dear {student.name},

Your exam for {subject.name} has been reassigned by {university_name}. Please find the new details:

Exam Start Date: {exam_start_date}
Exam End Date: {exam_end_date}
Exam Start Time: {exam_start_time}
Exam End Time: {exam_end_time}

Enter Email ID & Password in the link below:

Email: {student.email}
Password: {student.mobile}

Click on the link to Login:
{login_url}
"""
        Thread(target=send_exam_email, args=(email_subject, email_message, [student.email])).start()

        return Response({"message": "Exam reassigned successfully."}, status=status.HTTP_200_OK)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} does not exist.")
        return Response({"message": "Invalid student ID."}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.exception(f"Error occurred while reassigning exam: {e}")
        return Response({"message": "An error occurred while reassigning the exam."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_course_duration(request):
    try:
        # Extract query parameters
        university_id = request.query_params.get('university')
        course_id = request.query_params.get('course')
        stream_id = request.query_params.get('stream')

        # Validate and fetch University
        university = University.objects.filter(id=university_id).first()
        if not university:
            logger.error(f"Invalid university ID: {university_id}")
            return Response({"error": "Invalid university ID."}, status=status.HTTP_400_BAD_REQUEST)

        # Validate and fetch Course
        course = Course.objects.filter(id=course_id, university=university).first()
        if not course:
            logger.error(f"Invalid course ID: {course_id} for university ID: {university_id}")
            return Response({"error": "Invalid course for the given university."}, status=status.HTTP_400_BAD_REQUEST)

        # Validate and fetch Stream
        stream = Stream.objects.filter(id=stream_id, course=course).first()
        if not stream:
            logger.error(f"Invalid stream ID: {stream_id} for course ID: {course_id}")
            return Response({"error": "Invalid stream for the given course."}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch semester information from the Stream model
        sem = stream.sem
        #logger.info(f"Semester retrieved for stream ID {stream_id}: {sem}")

        # Return semester in the response
        return Response({"stream": stream.name, "sem": sem}, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception("An unexpected error occurred in get_course_duration.")
        return Response({"error": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      

# @api_view(['GET'])
# def student_login(request):
#     try:
#         email = request.query_params.get('email')
#         password = request.query_params.get('password')

#         if not email or not password:
#             logger.warning("Email or password not provided")
#             return Response({"error": "Email and password are required."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             student = Student.objects.get(email=email)
#         except Student.DoesNotExist:
#             logger.error(f"Student with email {email} not found")
#             return Response({"error": "Invalid email or password."}, status=status.HTTP_404_NOT_FOUND)

#         if password != student.mobile:
#             logger.warning(f"Invalid password attempt for student {email}")
#             return Response({"error": "Invalid email or password."}, status=status.HTTP_401_UNAUTHORIZED)

#         # Check if the student is in any StudentAppearingExam
#         exams = StudentAppearingExam.objects.filter(student_id__contains=[student.id])

#         if not exams.exists():
#             #logger.info(f"No exams found for student {email} (ID: {student.id})")
#             return Response({"message": "No exams found for this student."}, status=status.HTTP_200_OK)

#         # Prepare the exam-related data
#         exam_details = []
#         examination_data = []

#         for exam in exams:
#             exam_details.append({
#                 "exam_id": exam.exam.id,
#                 "examstartdate": exam.examstartdate,
#                 "examenddate": exam.examenddate,
#                 "examstarttime": exam.examstarttime,
#                 "examendtime": exam.examendtime
#             })

#             # Fetch details from the Examination table
#             examination = exam.exam
#             examination_data.append({
#                 "id": examination.id,
#                 "course_id": examination.course.id,
#                 "stream_id": examination.stream.id,
#                 "subject_id": examination.subject.id,
#                 "studypattern":examination.studypattern,
#                 "semyear":examination.semyear,
#                 "substream_id": examination.substream.id if examination.substream else None,
#                 "course_name": examination.course.name,
#                 "stream_name": examination.stream.name,
#                 "subject_name": examination.subject.name,
#                 "substream_name": examination.substream.name if examination.substream else None
                
#             })

#         #logger.info(f"Successfully fetched exam details for student {email} (ID: {student.id})") 

#         return Response({
#             "student_id": student.id,
#             "exam_details": exam_details,
#             "examination_data": examination_data
#         }, status=status.HTTP_200_OK)

#     except Exception as e:
#         logger.exception("An error occurred while processing the student login.")
#         return Response({"error": "An internal server error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def student_login(request):
    try:
        email = request.data.get('email')
        password = request.data.get('password')

        if not email or not password:
            logger.error("Email or password not provided")
            return Response({"error": "Email and password are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(email=email)
        except Student.DoesNotExist:
            logger.error(f"Student with email {email} not found")
            return Response({"error": "Invalid email or password."}, status=status.HTTP_404_NOT_FOUND)

        # Authenticate student (assuming password is stored as mobile number)
        if password != student.mobile:
            logger.error(f"Invalid password attempt for student {email}")
            return Response({"error": "Invalid email or password."}, status=status.HTTP_401_UNAUTHORIZED)

        # Generate JWT Token
        try:
            token = get_tokens_for_user(student)  # Ensure this function is working
        except Exception as e:
            logger.error(f"Error generating token for student {email}: {str(e)}")
            return Response({"error": "An error occurred while generating the token."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Fetch student exams by checking if student.id is in the student_id JSON list
        try:
            exams = StudentAppearingExam.objects.filter(student_id__contains=[student.id])
        except Exception as e:
            logger.error(f"Error fetching exams for student {email}: {str(e)}")
            return Response({"error": "An error occurred while fetching exams."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if not exams.exists():
            return Response({
                "message": "No exams found for this student.",
                "token": token
            }, status=status.HTTP_200_OK)

        # Prepare exam-related data
        exam_details = []
        examination_data = []

        for exam in exams:
            try:
                exam_details.append({
                    "exam_id": exam.exam.id,
                    "examstartdate": exam.examstartdate,
                    "examenddate": exam.examenddate,
                    "examstarttime": exam.examstarttime,
                    "examendtime": exam.examendtime
                })
            except Exception as e:
                logger.error(f"Error fetching exam details for exam {exam.id}: {str(e)}")
                continue  # Skip this exam if it's missing details

            try:
                examination = exam.exam
                examination_data.append({
                    "id": examination.id,
                    "course_id": examination.course.id,
                    "stream_id": examination.stream.id,
                    "subject_id": examination.subject.id,
                    "studypattern": examination.studypattern,
                    "semyear": examination.semyear,
                    "substream_id": examination.substream.id if examination.substream else None,
                    "course_name": examination.course.name,
                    "stream_name": examination.stream.name,
                    "subject_name": examination.subject.name,
                    "substream_name": examination.substream.name if examination.substream else None
                })
            except Exception as e:
                logger.error(f"Error fetching examination data for exam {exam.id}: {str(e)}")
                continue  # Skip this examination if it's missing related data

        return Response({
            "message": "Login Successful",
            "token": token,
            "student_id": student.id,
            "exam_details": exam_details,
            "examination_data": examination_data
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"An error occurred while processing the student login: {str(e)}")
        return Response({"error": f"An internal server error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_excel_for_set_exam_for_subject(request):
    try:
        # Define the path to the predefined Excel file
        file_name = "SubjectWiseQuestionTemplate.xlsx"
        file_path = os.path.join('media', 'Download_Excel_Format', file_name)

        # Check if the file exists
        if not os.path.exists(file_path):
            logger.error("File not found: %s", file_path)
            return Response(
                {"error": "The requested Excel file does not exist."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Serve the file as a response
        response = FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=file_name,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        return response

    except Exception as e:
        # Log any unexpected errors
        logger.error("Error while serving Excel file: %s", str(e))
        return Response(
            {"error": "An error occurred while processing the request."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_subjects(request):
    try:
        stream = request.query_params.get('stream')
        substream = request.query_params.get('substream')  # Can be "", None, or valid ID
        semyear = request.query_params.get('semyear')

        if not stream:
            return Response({'error': "Stream parameter is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not semyear:
            return Response({'error': "Semester/Year parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Base filters
        filters = {'stream__id': stream, 'semyear': semyear}

        # Filter by substream only if passed and valid
        if substream not in [None, '', 'null']:
            filters['substream__id'] = substream
        # else: no substream filter applied to include all with/without substream

        subjects = Subject.objects.filter(**filters).values(
            'id', 'name', 'code', 'stream', 'substream', 'studypattern', 'semyear'
        )

        if not subjects.exists():
            return Response({'message': "No subjects found."}, status=status.HTTP_404_NOT_FOUND)

        return Response({'subjects': list(subjects)}, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception(f"Error in get_all_subjects: {str(e)}")
        return Response({'error': "Internal server error."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def fetch_questions_based_on_exam_id(request):
    """
    Fetch questions for a given exam ID.
    """
    exam_id = request.query_params.get('exam_id')
    if not exam_id:
        logger.error("Exam ID is missing in the request.")
        return Response({"error": "Exam ID is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        questions = Questions.objects.filter(exam_id=exam_id)
        if not questions.exists():
            #logger.info(f"No questions found for Exam ID: {exam_id}")
            return Response({"message": "No questions found for the given Exam ID."}, status=status.HTTP_404_NOT_FOUND)
        
        # Prepare response data
        data = []
        for question in questions:
            data.append({
                "id": question.id,
                "exam_id": question.exam_id,
                "question": question.question,
                "image": question.image,
                "type": question.type,
                "marks": question.marks,
                "options": [
                    question.option1, question.option2, question.option3,
                    question.option4
                ],
                "short_answer": question.shortanswer,
                "answer": question.answer,
                "difficulty_level": question.difficultylevel
            })
        
        #logger.info(f"Successfully fetched {len(data)} questions for Exam ID: {exam_id}")
        return Response({"questions": data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error fetching questions for Exam ID {exam_id}: {str(e)}")
        return Response({"error": "An error occurred while fetching questions."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

   
@api_view(['POST'])
def save_all_questions_answers(request):
    try:
        data = request.data
        student_id = data.get("student_id")
        exam_id = data.get("exam_id")
        questions = data.get("questions")

        #logger.info(f"Received submission for student_id={student_id}, exam_id={exam_id}")

        # Validate presence of required fields
        if not student_id or not exam_id or not questions:
            logger.error("Missing required fields: student_id, exam_id, or questions.")
            return Response({
                "error": "Missing required fields: student_id, exam_id, or questions."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Fetch related instances
        try:
            student_instance = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            logger.error(f"Student with id {student_id} not found.")
            return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            exam_instance = Examination.objects.get(id=exam_id)
        except Examination.DoesNotExist:
            logger.error(f"Examination with id {exam_id} not found.")
            return Response({"error": "Examination not found."}, status=status.HTTP_404_NOT_FOUND)

        saved_answers = []

        for question in questions:
            question_id = question.get("id")
            submitted_answer = question.get("submitted_answer")
            
            print('submitted_answer',submitted_answer)

            if not question_id or submitted_answer is None:
                logger.warning(f"Skipping question with incomplete data: {question}")
                continue

            try:
                question_instance = Questions.objects.get(id=question_id)
            except Questions.DoesNotExist:
                logger.warning(f"Question with ID {question_id} does not exist. Skipping.")
                continue

            # Get correct answer and marks
            correct_answer = question_instance.answer.lower().strip() if question_instance.answer else ""
            
            print('submitted_answer correct_answer ',correct_answer,submitted_answer)
            
            submitted_value = submitted_answer.lower().strip()
            marks = question_instance.marks or "0"

            # Evaluate
            if correct_answer == submitted_value:
                result = "Right"
                marks_obtained = marks
            else:
                result = "Wrong"
                marks_obtained = "0"

            # Save submission
            SubmittedExamination.objects.create(
                student=student_instance,
                exam=exam_instance,
                question=str(question_id),  # Question ID stored as CharField
                type=question_instance.type,
                marks=marks,
                marks_obtained=marks_obtained,
                submitted_answer=submitted_answer,
                answer=correct_answer,
                result=result,
            )

            #logger.info(f"Saved answer for question_id={question_id}: {result}, Marks={marks_obtained}")

            saved_answers.append({
                "question_id": question_id,
                "result": result,
                "marks_obtained": marks_obtained,
            })

        #logger.info(f"Successfully saved {len(saved_answers)} answers for student {student_id}.")

        return Response({
            "message": "Submitted answers saved successfully.",
            "saved_answers": saved_answers
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.exception(f"Unexpected error while saving answers: {str(e)}")
        return Response({
            "error": "An internal server error occurred."
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_result_to_show_based_on_subject(request):
    try:
        # Extract query parameters
        query_params = {
            'university': request.query_params.get('university'),
            'course': request.query_params.get('course'),
            'stream': request.query_params.get('stream'),
            'substream': request.query_params.get('substream'),
            'session': request.query_params.get('session'),
            'studypattern': request.query_params.get('studypattern'),
            'semyear': request.query_params.get('semyear')
        }

        # Filter examinations based on the provided query parameters
        examinations = Examination.objects.all()
        for param, value in query_params.items():
            if value:
                examinations = examinations.filter(**{param: value})

        # Construct response data
        response_data = []
        for exam in examinations:
            # Fetch all student-exam records related to this exam
            student_exam_entries = StudentAppearingExam.objects.filter(exam=exam)

            # Extract student IDs from all records
            student_ids = []
            for entry in student_exam_entries:
                if entry.student_id:  # Ensure it's not empty
                    student_ids.extend(entry.student_id)

            # Remove duplicates if necessary
            student_ids = list(set(student_ids))

            # Use the first StudentAppearingExam entry to get exam timing details
            student_exam_details = student_exam_entries.first()

            response_data.append({
                "id": exam.id,
                "subject": exam.subject.id if exam.subject else None,
                "subject_name": exam.subject.name if exam.subject else None,
                "exam_start_date": student_exam_details.examstartdate if student_exam_details else None,
                "exam_end_date": student_exam_details.examenddate if student_exam_details else None,
                "exam_start_time": student_exam_details.examstarttime if student_exam_details else None,
                "exam_end_time": student_exam_details.examendtime if student_exam_details else None,
                "total_questions": exam.totalquestions,
                "total_marks": exam.totalmarks,
                "student_ids": student_ids  # Include student IDs in the response
            })

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error occurred while fetching examinations: {str(e)}")
        return Response(
            {"error": "An error occurred while processing your request."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


from openpyxl import Workbook
from openpyxl.styles import Font

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_to_excel(request):
    """
    Export data to Excel based on filters provided in the request.
    """
    try:
        # Extract filters from POST data
        university = request.data.get("university")
        course = request.data.get("course")
        stream = request.data.get("stream")
        session = request.data.get("session")
        studypattern = request.data.get("studypattern")
        semyear = request.data.get("semyear")
        substream = request.data.get("substream")

        #logger.info(f"Received export request with filters: {request.data}")

        # Validate required filters
        if not university or not course or not stream:
            return Response({'message': 'Missing required fields: university, course, stream'}, status=400)

        # Build query filters dynamically
        filters = Q(university=university) & Q(course=course) & Q(stream=stream)
        if session:
            filters &= Q(session=session)
        if studypattern:
            filters &= Q(studypattern=studypattern)
        if semyear:
            filters &= Q(semyear=semyear)
        if substream:
            filters &= Q(substream=substream)

        # Fetch relevant data
        fetch_exams = Examination.objects.filter(filters)
        if not fetch_exams.exists():
            logger.warning("No exams found for the given filters.")
            return Response({'message': 'No results found'}, status=404)

        formatted_data = []
        for exam in fetch_exams:
            subject = Subject.objects.filter(id=exam.subject_id).first()
            student_data = StudentAppearingExam.objects.filter(exam=exam.id)

            for student_record in student_data:
                for student_id in student_record.student_id:
                    student = Student.objects.filter(id=student_id).first()
                    status = "Not Appeared"
                    result_data = Result.objects.filter(student_id=student_id, exam=exam.id).first()

                    if result_data:
                        status = "Appeared"
                        marks_obtained = result_data.score
                        result = result_data.result
                    else:
                        marks_obtained = ''
                        result = ''

                    formatted_data.append({
                        'University': university,
                        'Course': course,
                        'Stream': stream,
                        'Substream': substream if substream else '',
                        'Session': session,
                        'Study Pattern': studypattern,
                        'Semester/Year': semyear,
                        'Subject Name': subject.name if subject else '',
                        'Student Name': student.name if student else '',
                        'Marks Obtained': marks_obtained,
                        'Result': result,
                        'Status': status
                    })

        if not formatted_data:
            logger.warning("No student data found to export.")
            return Response({'message': 'No data available for export'}, status=400)

        # Generate Excel file
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Exported Data"

        # Write headers
        headers = list(formatted_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Write data rows
        for row_num, row_data in enumerate(formatted_data, start=2):
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=row_num, column=col_num).value = row_data[header]

        # Create response with HttpResponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'
        workbook.save(response)

        #logger.info("Excel file generated and sent successfully.")
        return response

    except Exception as e:
        logger.exception(f"Error occurred while exporting data: {str(e)}")
        return Response({'message': 'An internal server error occurred.'}, status=500)

from django.db.models import Sum, Count
@api_view(['POST'])
def generate_result(request):
    try:
        # Extract student_id and exam_id from the request
        student_id = request.data.get('student_id')
        exam_id = request.data.get('exam_id')

        # Validate input
        if not student_id or not exam_id:
            logger.error("Student ID or Exam ID is missing.")
            return Response({"error": "Student ID and Exam ID are required."}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch submitted examinations for the student and exam
        submitted_exams = SubmittedExamination.objects.filter(student_id=student_id, exam_id=exam_id)
        if not submitted_exams.exists():
            logger.error(f"No submitted examinations found for student {student_id} and exam {exam_id}.")
            return Response({"error": "No submitted examinations found."}, status=status.HTTP_404_NOT_FOUND)

        # Calculate total marks obtained and attempted questions
        total_marks_obtained = submitted_exams.aggregate(total_marks=Sum('marks_obtained'))['total_marks'] or 0
        attempted_questions = submitted_exams.exclude(submitted_answer__isnull=True).exclude(submitted_answer="").count()

        # Get the total questions count
        total_questions = submitted_exams.count()

        # Fetch examination details
        exam = Examination.objects.get(id=exam_id)
        total_marks = int(exam.totalmarks)
        passing_marks = int(exam.passingmarks) if exam.passingmarks else 0

        # Calculate percentage
        percentage = (int(total_marks_obtained) / total_marks) * 100 if total_marks > 0 else 0

        # Determine result
        result_status = "Pass" if total_marks_obtained >= passing_marks else "Fail"

        # Create or update the Result entry
        result, created = Result.objects.update_or_create(
            student_id=student_id,
            exam_id=exam_id,
            defaults={
                "total_question": total_questions,
                "attempted": attempted_questions,
                "total_marks": total_marks,
                "score": int(total_marks_obtained),
                "percentage": round(percentage, 2),
                "result": result_status,
                "created_by": request.user.username if request.user else None,
                "modified_by": request.user.username if request.user else None,
            },
        )

        #logger.info(f"Result {'created' if created else 'updated'} successfully for student {student_id} and exam {exam_id}.")
        return Response({
            "message": "Result generated successfully.",
            "result_id": result.id,
            "percentage": round(percentage, 2),
            "result_status": result_status
        }, status=status.HTTP_200_OK)

    except Examination.DoesNotExist:
        logger.error(f"Examination with ID {exam_id} does not exist.")
        return Response({"error": "Examination not found."}, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        logger.exception("An error occurred while generating the result.")
        return Response({"error": "An unexpected error occurred.", "details": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def show_result(request):
    try:
        # Get query parameters
        student_id = request.query_params.get('student_id')
        exam_id = request.query_params.get('exam_id')

        # Validate query parameters
        if not student_id or not exam_id:
            logger.error("Student ID or Exam ID is missing.")
            return Response({"error": "Student ID and Exam ID are required."}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch the result based on student_id and exam_id
        result = Result.objects.filter(student_id=student_id, exam_id=exam_id).select_related('student', 'exam').first()
        if not result:
            logger.error(f"No result found for student ID {student_id} and exam ID {exam_id}.")
            return Response({"error": "No result found."}, status=status.HTTP_404_NOT_FOUND)

        # Fetch student details
        student = result.student
        response_data = {
            "student_name": student.name,
            "email": student.email,
            "enrollment_id": student.enrollment_id,
            "exam_id": result.exam.id,
            "total_questions": result.total_question,
            "total_marks": result.total_marks,
            "score": result.score,
            "percentage": result.percentage,
            "result_status": result.result,
        }

        #logger.info(f"Result fetched successfully for student {student_id} and exam {exam_id}.")
        return Response({"message": "Result fetched successfully.", "data": response_data}, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception("An error occurred while fetching the result.")
        return Response({"error": "An unexpected error occurred.", "details": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_university(request, university_id):
    try:
        university = University.objects.get(id=university_id)
        if Course.objects.filter(university=university).exists():
            return Response({"message": "Cannot delete university as it has associated courses."}, status=status.HTTP_400_BAD_REQUEST)
        university_name = university.university_name
        university.delete()
        #logger.info(f"University '{university_name}' with ID {university_id} deleted successfully.")
        return Response({"message": f"University '{university_name}' deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    except University.DoesNotExist:
        logger.warning(f"Attempt to delete non-existent university with ID {university_id}.")
        return Response({"message": "University not found."}, status=status.HTTP_404_NOT_FOUND)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_course(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
        if Stream.objects.filter(course=course).exists():
            return Response({"message": "Cannot delete course as it has associated streams."}, status=status.HTTP_400_BAD_REQUEST)
        course_name = course.name
        course.delete()
        #logger.info(f"Course '{course_name}' with ID {course_id} deleted successfully.")
        return Response({"message": f"Course '{course_name}' deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    except Course.DoesNotExist:
        logger.warning(f"Attempt to delete non-existent course with ID {course_id}.")
        return Response({"message": "Course not found."}, status=status.HTTP_404_NOT_FOUND)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_stream(request, stream_id):
    try:
        stream = Stream.objects.get(id=stream_id)
        if SubStream.objects.filter(stream=stream).exists() or Subject.objects.filter(stream=stream).exists():
            return Response({"message": "Cannot delete stream as it has associated substreams or subjects."}, status=status.HTTP_400_BAD_REQUEST)
        stream_name = stream.name
        stream.delete()
        #logger.info(f"Stream '{stream_name}' with ID {stream_id} deleted successfully.")
        return Response({"message": f"Stream '{stream_name}' deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    except Stream.DoesNotExist:
        logger.warning(f"Attempt to delete non-existent stream with ID {stream_id}.")
        return Response({"message": "Stream not found."}, status=status.HTTP_404_NOT_FOUND)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_substream(request, substream_id):
    try:
        substream = SubStream.objects.get(id=substream_id)
        if Subject.objects.filter(substream=substream).exists():
            return Response({"message": "Cannot delete substream as it has associated subjects."}, status=status.HTTP_400_BAD_REQUEST)
        substream_name = substream.name
        substream.delete()
        #logger.info(f"SubStream '{substream_name}' with ID {substream_id} deleted successfully.")
        return Response({"message": f"SubStream '{substream_name}' deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    except SubStream.DoesNotExist:
        logger.warning(f"Attempt to delete non-existent substream with ID {substream_id}.")
        return Response({"message": "SubStream not found."}, status=status.HTTP_404_NOT_FOUND)
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_subject(request, subject_id):
    try:
        subject = Subject.objects.get(id=subject_id)
        subject_name = subject.name
        subject.delete()
        #logger.info(f"Subject '{subject_name}' with ID {subject_id} deleted successfully.")
        return Response({"message": f"Subject '{subject_name}' deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    except Subject.DoesNotExist:
        logger.warning(f"Attempt to delete non-existent subject with ID {subject_id}.")
        return Response({"message": "Subject not found."}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_substreams_with_id_by_university_course_stream(request):
    university_name = request.query_params.get('university')
    course_name = request.query_params.get('course')
    stream_name = request.query_params.get('stream')
    if not university_name or not course_name or not stream_name:
        return Response({"error": "University name, course name, and stream name are required."}, status=status.HTTP_400_BAD_REQUEST)
    try:
        university = University.objects.get(university_name=university_name)
    except University.DoesNotExist:
        return Response({"error": "University not found."}, status=status.HTTP_404_NOT_FOUND)
    try:
        course = Course.objects.get(university=university, name=course_name)
    except Course.DoesNotExist:
        return Response({"error": "Course not found."}, status=status.HTTP_404_NOT_FOUND)
    try:
        stream = Stream.objects.get(course=course, name=stream_name)
    except Stream.DoesNotExist:
        return Response({"error": "Stream not found."}, status=status.HTTP_404_NOT_FOUND)
    substreams = SubStream.objects.filter(stream=stream)
    substream_list = [{"id": substream.id, "name": substream.name} for substream in substreams]
    return Response(substream_list, status=status.HTTP_200_OK)
  
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_subjects_by_stream(request, stream_id):
    try:
        subjects = Subject.objects.filter(stream_id=stream_id)
        if not subjects.exists():
            return Response({"status": "error", "message": "No subjects found for the given stream."}, status=status.HTTP_404_NOT_FOUND)
        subject_data = [
            {   "id":subject.id,
                "studypattern": subject.studypattern,
                "semyear": subject.semyear,
                "name": subject.name,
                "code": subject.code
            }
            for subject in subjects
        ]
        return Response({"status": "success", "data": subject_data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error fetching subjects for stream_id {stream_id}: {str(e)}")
        return Response({"status": "error", "message": "Something went wrong!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
  

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_of_all_registered_student(request):
    try:
        all_students = Student.objects.filter(archive=False).order_by('-id')

        student_data = []
        for student in all_students:
            # Get enrollment details
            enrolled = Enrolled.objects.filter(student=student).first()

            if enrolled:
                current_semyear = enrolled.current_semyear
                entry_mode = enrolled.entry_mode if enrolled.entry_mode else ""
                study_pattern_mode = enrolled.course_pattern if enrolled.course_pattern else ""
            else:
                current_semyear = ""
                entry_mode = ""
                study_pattern_mode = ""

            # Determine Source (is_quick_register condition)
            is_quick_register = 'SR' if not student.is_quick_register else 'QR'

            student_data.append({
                "id":student.id,
                "current_semyear": current_semyear,
                "enrollment_id": student.enrollment_id,
                "student_name": student.name,
                "university_name": student.university.university_name if student.university else "",
                "source": is_quick_register,
                "study_pattern_mode": study_pattern_mode,
                "entry_mode": entry_mode,
                "enrollment_date": student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else ""
            })
        
        return Response({"status": "success", "data": student_data}, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error fetching students: {str(e)}")
        return Response({"status": "error", "message": "Something went wrong!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def get_student_enroll_to_next_year(request, id):
    try:
        student = Student.objects.get(id=id)
        enrolled = Enrolled.objects.filter(student=student).first()

        if not enrolled:
            return Response({"status": "error", "message": "Student enrollment not found!"}, status=status.HTTP_404_NOT_FOUND)

        if request.method == "GET":
            if int(enrolled.current_semyear) == int(enrolled.total_semyear):
                return Response({"status": "error", "message": "Student is in Final Semester / Year!"}, status=status.HTTP_400_BAD_REQUEST)
            elif int(enrolled.current_semyear) < int(enrolled.total_semyear):
                student_data = {
                    "id": student.id,
                    "total_semyear": enrolled.total_semyear,
                    "course": enrolled.course.name if enrolled.course else "",  # Course Name
                    "stream": enrolled.stream.name if enrolled.stream else "",  # Stream Name
                    "current_semyear": enrolled.current_semyear,
                    "next_semyear": int(enrolled.current_semyear) + 1,
                }
                #logger.info(f"GET Request - Student ID: {id}, Response: {student_data}")
                return Response({"status": "success", "data": student_data}, status=status.HTTP_200_OK)

        if request.method == "POST":
            if int(enrolled.current_semyear) == int(enrolled.total_semyear):
                return Response({"status": "error", "message": "Student is already in Final Semester / Year!"}, status=status.HTTP_400_BAD_REQUEST)

            new_semyear = int(enrolled.current_semyear) + 1
            enrolled.current_semyear = str(new_semyear)
            enrolled.save()

            response_data = {
                "id": student.id,
                "message": f"Student has been enrolled to Semester/Year {new_semyear}",
                "updated_current_semyear": new_semyear,
                "total_semyear": enrolled.total_semyear,
                "course": enrolled.course.name if enrolled.course else "",  # Course Name
                "stream": enrolled.stream.name if enrolled.stream else "",  # Stream Name
            }

            #logger.info(f"POST Request - Student ID: {id}, New Semester/Year: {new_semyear}, Response: {response_data}")
            return Response({"status": "success", "data": response_data}, status=status.HTTP_200_OK)

    except Student.DoesNotExist:
        logger.error(f"Student with ID {id} not found!")
        return Response({"status": "error", "message": "Student not Found!"}, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        logger.error(f"Error processing request for student {id}: {str(e)}")
        return Response({"status": "error", "message": "Something went wrong!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_multiple_subjects(request):
    try:
        subjects_data = request.data.get("subjects", [])

        if not subjects_data:
            logger.error("No subjects data provided for update.")
            return Response({"status": "error", "message": "No subjects data provided."}, status=status.HTTP_400_BAD_REQUEST)

        updated_subjects = []
        errors = []

        for subject_data in subjects_data:
            subject_id = subject_data.get("id")
            if not subject_id:
                errors.append({"id": None, "message": "Subject ID is required."})
                continue

            subject = Subject.objects.filter(id=subject_id).first()
            if not subject:
                errors.append({"id": subject_id, "message": "Subject not found."})
                continue

            fields_updated = []

            if 'studypattern' in subject_data:
                subject.studypattern = subject_data['studypattern']
                fields_updated.append("Study Pattern")

            if 'semyear' in subject_data:
                subject.semyear = subject_data['semyear']
                fields_updated.append("Semester/Year")

            if 'name' in subject_data:
                subject.name = subject_data['name']
                fields_updated.append("Subject Name")

            if 'code' in subject_data:
                subject.code = subject_data['code']
                fields_updated.append("Subject Code")

            if not fields_updated:
                errors.append({"id": subject_id, "message": "No valid fields provided for update."})
                continue

            subject.save()
            updated_subjects.append({"id": subject.id, "updated_fields": fields_updated})
            #logger.info(f"Updated Subject ID {subject.id}. Fields updated: {', '.join(fields_updated)}")

        response_data = {"status": "success", "updated_subjects": updated_subjects}
        if errors:
            response_data["errors"] = errors

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error updating subjects: {str(e)}")
        return Response({"status": "error", "message": "Something went wrong!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def register_cancel_student(request, id):
    try:
        student = Student.objects.get(id=id)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {id} not found.")
        return Response({"status": "error", "message": "Student not found!"}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == "GET":
        student_data = {
            'id': student.id,
            'name': student.name,
            'mobile': student.mobile,
            'email': student.email,
            'enrolled': student.is_enrolled,
            'active': student.is_enrolled,
        }
        return Response({"status": "success", "data": student_data}, status=status.HTTP_200_OK)
    
    if request.method == "POST":
        cancel_status = request.data.get('cancel_status')
        if cancel_status == "in-active":
            student.archive = True
            student.is_enrolled = False
            student.is_cancelled = True
            student.save()
            #logger.info(f"Student with ID {id} has been cancelled.")
            return Response({"status": "success", "message": "Student registration cancelled successfully."}, status=status.HTTP_200_OK)
        else:
            logger.error(f"Invalid cancel_status received: {cancel_status}")
            return Response({"status": "error", "message": "Invalid cancel_status value."}, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({"status": "error", "message": "Invalid request method."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(["POST", "GET"])
@permission_classes([IsAuthenticated])
def registered_new_university_enrollment_number(request):
    try:
        if request.method == "POST":
            data = request.data
            student_id = data.get("student_id")
            course_id = data.get("course_id")
            enrollment_id = data.get("enrollment_id")
            enrollment_type = data.get("type")

            if not student_id or not course_id or not enrollment_id:
                logger.error("Missing required fields: student_id, course_id, or enrollment_id")
                return Response({"status": "error", "message": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST)

            if enrollment_type != "new":
                logger.error("Invalid enrollment type. Expected 'new'")
                return Response({"status": "error", "message": "Invalid enrollment type. Expected 'new'"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate Student existence
            try:
                student = Student.objects.get(id=student_id)
            except Student.DoesNotExist:
                logger.error(f"Student with ID {student_id} does not exist")
                return Response({"status": "error", "message": "Invalid student ID"}, status=status.HTTP_400_BAD_REQUEST)

            # Check for duplicate enrollment ID
            if UniversityEnrollment.objects.filter(enrollment_id=enrollment_id).exists():
                logger.error(f"Enrollment ID {enrollment_id} already exists")
                return Response({"status": "error", "message": "Enrollment ID already exists"}, status=status.HTTP_400_BAD_REQUEST)

            # Create new enrollment
            enrollment = UniversityEnrollment.objects.create(
                student=student,
                type="new",
                course_id=course_id,
                course_name=student.course.name if hasattr(student, "course") else "",
                enrollment_id=enrollment_id,
            )

            return Response({
                "status": "success",
                "message": "Enrollment registered successfully",
                "data": {
                    "id": enrollment.id,
                    "student_id": student_id,
                    "type": "new",
                    "course_id": course_id,
                    "course_name": enrollment.course_name,
                    "enrollment_id": enrollment_id
                }
            }, status=status.HTTP_201_CREATED)

        elif request.method == "GET":
            student_id = request.query_params.get("student_id")

            if not student_id:
                logger.error("Missing required field: student_id")
                return Response({"status": "error", "message": "student_id is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate Student existence
            try:
                student = Student.objects.get(id=student_id)
            except Student.DoesNotExist:
                logger.error(f"Student with ID {student_id} does not exist")
                return Response({"status": "error", "message": "Invalid student ID"}, status=status.HTTP_400_BAD_REQUEST)

            enrollments = UniversityEnrollment.objects.filter(student=student)

            if not enrollments.exists():
                return Response({"status": "error", "message": "No enrollment records found"}, status=status.HTTP_404_NOT_FOUND)

            enrollment_data = [
                {
                    "id": enrollment.id,
                    "student_id": enrollment.student.id,
                    "student_name": enrollment.student.name,
                    "type": enrollment.type,
                    "course_id": enrollment.course_id,
                    "course_name": enrollment.course_name,
                    "enrollment_id": enrollment.enrollment_id,
                }
                for enrollment in enrollments
            ]

            return Response({
                "status": "success",
                "message": "Enrollment records fetched successfully",
                "data": enrollment_data
            }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return Response({"status": "error", "message": "Internal Server Error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST", "GET"])
@permission_classes([IsAuthenticated])
def registered_old_university_enrollment_number(request):
    try:
        if request.method == "POST":
            data = request.data
            student_id = data.get("student_id")
            course_id = data.get("course_id")
            enrollment_id = data.get("enrollment_id")
            enrollment_type = data.get("type")

            if not student_id or not course_id or not enrollment_id:
                logger.error("Missing required fields: student_id, course_id, or enrollment_id")
                return Response({"status": "error", "message": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST)

            if enrollment_type != "old":
                logger.error("Invalid enrollment type. Expected 'old'")
                return Response({"status": "error", "message": "Invalid enrollment type. Expected 'old'"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate Student existence
            try:
                student = Student.objects.get(id=student_id)
            except Student.DoesNotExist:
                logger.error(f"Student with ID {student_id} does not exist")
                return Response({"status": "error", "message": "Invalid student ID"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate Course existence
            try:
                course = Course.objects.get(id=course_id)
            except Course.DoesNotExist:
                logger.error(f"Course with ID {course_id} does not exist")
                return Response({"status": "error", "message": "Invalid course ID"}, status=status.HTTP_400_BAD_REQUEST)

            # Check for duplicate enrollment ID
            if UniversityEnrollment.objects.filter(enrollment_id=enrollment_id).exists():
                logger.error(f"Enrollment ID {enrollment_id} already exists")
                return Response({"status": "error", "message": "Enrollment ID already exists"}, status=status.HTTP_400_BAD_REQUEST)

            # Create new enrollment
            enrollment = UniversityEnrollment.objects.create(
                student=student,
                type="old",
                course_id=course_id,
                course_name=course.name,
                enrollment_id=enrollment_id,
            )

            return Response({
                "status": "success",
                "message": "Enrollment registered successfully",
                "data": {
                    "id": enrollment.id,
                    "student_id": student_id,
                    "type": "old",
                    "course_id": course_id,
                    "course_name": course.name,
                    "enrollment_id": enrollment_id
                }
            }, status=status.HTTP_201_CREATED)

        elif request.method == "GET":
            student_id = request.query_params.get("student_id")

            if not student_id:
                logger.error("Missing required field: student_id")
                return Response({"status": "error", "message": "student_id is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate Student existence
            try:
                student = Student.objects.get(id=student_id)
            except Student.DoesNotExist:
                logger.error(f"Student with ID {student_id} does not exist")
                return Response({"status": "error", "message": "Invalid student ID"}, status=status.HTTP_400_BAD_REQUEST)

            enrollments = UniversityEnrollment.objects.filter(student=student, type="old")

            if not enrollments.exists():
                return Response({"status": "error", "message": "No old enrollment records found"}, status=status.HTTP_404_NOT_FOUND)

            enrollment_data = [
                {
                    "id": enrollment.id,
                    "student_id": enrollment.student.id,
                    "student_name": enrollment.student.name,
                    "type": enrollment.type,
                    "course_id": enrollment.course_id,
                    "course_name": enrollment.course_name,
                    "enrollment_id": enrollment.enrollment_id,
                }
                for enrollment in enrollments
            ]

            return Response({
                "status": "success",
                "message": "Old enrollment records fetched successfully",
                "data": enrollment_data
            }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return Response({"status": "error", "message": "Internal Server Error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET", "POST"])
def courier_api(request):
    if request.method == "GET":
        student_id = request.query_params.get("student_id")
        if not student_id:
            return Response(
                {"status": "error", "message": "student_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            student = get_object_or_404(Student, id=student_id)  # Cleaner lookup
            couriers = Courier.objects.filter(student=student).values(
                "id", "article_name", "courier_from", "courier_to", "booking_date", 
                "courier_company", "tracking_id", "remarks"
            )

            return Response({
                "status": "success",
                "message": "Courier records retrieved successfully",
                "data": list(couriers)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}")
            return Response(
                {"status": "error", "message": "Internal Server Error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    elif request.method == "POST":
        try:
            data = request.data
            student_id = data.get("student_id")
            article_name = data.get("article_name")
            courier_from = data.get("courier_from")
            courier_to = data.get("courier_to")
            booking_date = data.get("booking_date")
            courier_company = data.get("courier_company")
            tracking_id = data.get("tracking_id")
            remarks = data.get("remarks")

            # Validate student_id
            if not student_id:
                return Response(
                    {"status": "error", "message": "student_id is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            student = get_object_or_404(Student, id=student_id)  # Cleaner lookup

            # Create the courier record
            courier = Courier.objects.create(
                student=student,
                article_name=article_name,
                courier_from=courier_from,
                courier_to=courier_to,
                booking_date=booking_date,
                courier_company=courier_company,
                tracking_id=tracking_id,
                remarks=remarks
            )

            return Response({
                "status": "success",
                "message": "Courier record created successfully",
                "data": {
                    "id": courier.id,
                    "student_id": student_id,
                    "article_name": article_name,
                    "courier_from": courier_from,
                    "courier_to": courier_to,
                    "booking_date": booking_date,
                    "courier_company": courier_company,
                    "tracking_id": tracking_id,
                    "remarks": remarks
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}")
            return Response(
                {"status": "error", "message": "Internal Server Error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(['GET'])
def get_additional_fees(request):
    student_id = request.query_params.get("student_id")
    if not student_id:
        logger.error("Student ID is required.")
        return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student = Student.objects.get(id=student_id)
        fees = PaymentReciept.objects.filter(student=student)
        
        response_data = []
        for fee in fees:
            response_data.append({
                "student_id": student.id,
                "fees_id":fee.id,
                "semyear": fee.semyear,
                "payment_for": fee.payment_for,
                "payment_type": fee.payment_type,
                "session": fee.session,
                "transaction_date": fee.transaction_date,
                "paymentmode": fee.paymentmode,
                "cheque_no": fee.cheque_no,
                "paidamount": fee.paidamount,
                "pendingamount": fee.pendingamount,
            })
        
        return Response(response_data, status=status.HTTP_200_OK)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} does not exist.")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return Response({"error": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_additional_fees(request):
    student_id = request.data.get('student_id')
    if not student_id:
        logger.error("Student ID is required.")
        return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} does not exist.")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
    
    extrafees_feesfor = request.data.get('extrafees_feesfor')
    extrafees_amount = request.data.get('extrafees_amount')
    extrafees_feestype = request.data.get('extrafees_feestype')
    extrafees_semyear = request.data.get('extrafees_semyear')
    extrafees_transactiondate = request.data.get('extrafees_transactiondate')
    extrafees_paymentmode = request.data.get('extrafees_paymentmode')
    extrafees_chequeno = request.data.get('extrafees_chequeno')
    extrafees_bankname = request.data.get('extrafees_bankname')
    extrafees_remarks = request.data.get('extrafees_remarks')
    
    if not all([extrafees_feesfor, extrafees_amount, extrafees_feestype, extrafees_semyear]):
        logger.error("Missing required fee details.")
        return Response({"error": "Missing required fee details."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        getlatestreciept = PaymentReciept.objects.latest('id')
        tid = getlatestreciept.transactionID.replace("TXT445FE", '')
        transactionID = f"TXT445FE{int(tid) + 1}"
    except PaymentReciept.DoesNotExist:
        transactionID = "TXT445FE101"
    
    try:
        add_payment_reciept = PaymentReciept(
            student=student,
            payment_for=extrafees_feesfor,
            payment_type=extrafees_feestype,
            fee_reciept_type="",
            transaction_date=extrafees_transactiondate,
            cheque_no=extrafees_chequeno,
            bank_name=extrafees_bankname,
            paidamount=extrafees_amount,
            pendingamount="0",
            transactionID=transactionID,
            paymentmode=extrafees_paymentmode,
            remarks=extrafees_remarks,
            session="",
            semyear=extrafees_semyear
        )
        add_payment_reciept.save()
        
        response_data = {
            "student_id": student.id,
            "semyear": add_payment_reciept.semyear,
            "payment_for": add_payment_reciept.payment_for,
            "payment_type": add_payment_reciept.payment_type,
            "transaction_date": add_payment_reciept.transaction_date,
            "paymentmode": add_payment_reciept.paymentmode,
            "cheque_no": add_payment_reciept.cheque_no,
            "paidamount": add_payment_reciept.paidamount,
            "pendingamount": add_payment_reciept.pendingamount,
            "transactionID": add_payment_reciept.transactionID,
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    except Exception as e:
        logger.error(f"Error saving payment receipt: {str(e)}")
        return Response({"error": "An error occurred while saving the payment receipt."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT'])
def update_additional_fees(request):
    fees_id = request.data.get('fees_id')
    student_id = request.data.get('student_id')
    
    if not fees_id or not student_id:
        logger.error("Fees ID and Student ID are required.")
        return Response({"error": "Fees ID and Student ID are required."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        payment_receipt = PaymentReciept.objects.get(id=fees_id, student_id=student_id)
    except PaymentReciept.DoesNotExist:
        logger.error("No payment receipt found for the given Fees ID and Student ID.")
        return Response({"error": "No payment receipt found."}, status=status.HTTP_404_NOT_FOUND)
    
    # Update fields if provided in request
    for field in ['payment_for', 'payment_type', 'transaction_date', 'cheque_no', 'bank_name', 
                  'paidamount', 'pendingamount', 'paymentmode', 'remarks', 'semyear', 'status']:
        if request.data.get(field) is not None:
            setattr(payment_receipt, field, request.data[field])
    
    payment_receipt.save()
    
    response_data = {
        "fees_id": payment_receipt.id,
        "student_id": payment_receipt.student.id,
        "semyear": payment_receipt.semyear,
        "payment_for": payment_receipt.payment_for,
        "payment_type": payment_receipt.payment_type,
        "transaction_date": payment_receipt.transaction_date,
        "paymentmode": payment_receipt.paymentmode,
        "cheque_no": payment_receipt.cheque_no,
        "paidamount": payment_receipt.paidamount,
        "pendingamount": payment_receipt.pendingamount,
        "transactionID": payment_receipt.transactionID,
        "status": payment_receipt.status,
    }
    
    return Response(response_data, status=status.HTTP_200_OK)
  
@api_view(['POST', 'GET'])
def result_uploaded_view(request):
    if request.method == 'POST':
        student_id = request.data.get('student_id')
        date = request.data.get('date')
        examination = request.data.get('examination')
        semyear = request.data.get('semyear')
        uploaded = request.data.get('uploaded')
        remarks = request.data.get('remarks')
        
        if not student_id:
            logger.error("Student ID is required.")
            return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            logger.error(f"Student with ID {student_id} not found.")
            return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
        
        result_uploaded = ResultUploaded(
            student=student,
            date=date,
            examination=examination,
            semyear=semyear,
            uploaded=uploaded,
            remarks=remarks
        )
        result_uploaded.save()
        print('data save')
        response_data = {
            "id": result_uploaded.id,
            "student_id": result_uploaded.student.id,
            "date": result_uploaded.date,
            "examination": result_uploaded.examination,
            "semyear": result_uploaded.semyear,
            "uploaded": result_uploaded.uploaded,
            "remarks": result_uploaded.remarks
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)

    elif request.method == 'GET':
        student_id = request.GET.get('student_id')
        print('inside get',student_id)

        if not student_id:
            return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = ResultUploaded.objects.filter(student__id=student_id)

        if not queryset.exists():
            return Response({"error": "No results found for this student."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ResultUploadedSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
def update_result_uploaded(request, result_id):
    try:
        result_uploaded = ResultUploaded.objects.get(id=result_id)
    except ResultUploaded.DoesNotExist:
        logger.error(f"Result with ID {result_id} not found.")
        return Response({"error": "Result not found."}, status=status.HTTP_404_NOT_FOUND)
    
    result_uploaded.date = request.data.get('date', result_uploaded.date)
    result_uploaded.examination = request.data.get('examination', result_uploaded.examination)
    result_uploaded.semyear = request.data.get('semyear', result_uploaded.semyear)
    result_uploaded.uploaded = request.data.get('uploaded', result_uploaded.uploaded)
    result_uploaded.remarks = request.data.get('remarks', result_uploaded.remarks)
    
    result_uploaded.save()
    
    response_data = {
        "id": result_uploaded.id,
        "student_id": result_uploaded.student.id,
        "date": result_uploaded.date,
        "examination": result_uploaded.examination,
        "semyear": result_uploaded.semyear,
        "uploaded": result_uploaded.uploaded,
        "remarks": result_uploaded.remarks
    }
    
    return Response(response_data, status=status.HTTP_200_OK)
  
@api_view(['POST'])
def create_university_examination(request):
    student_id = request.data.get('student_id')
    exam_type = request.data.get('type')
    amount = request.data.get('amount')
    date = request.data.get('date')
    examination = request.data.get('examination')
    semyear = request.data.get('semyear')
    paymentmode = request.data.get('paymentmode')
    remarks = request.data.get('remarks')
    
    if not student_id:
        logger.error("Student ID is required.")
        return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} not found.")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
    
    university_exam = UniversityExamination(
        student=student,
        type=exam_type,
        amount=amount,
        date=date,
        examination=examination,
        semyear=semyear,
        paymentmode=paymentmode,
        remarks=remarks
    )
    university_exam.save()
    
    response_data = {
        "id": university_exam.id,
        "student_id": university_exam.student.id,
        "type": university_exam.type,
        "amount": university_exam.amount,
        "date": university_exam.date,
        "examination": university_exam.examination,
        "semyear": university_exam.semyear,
        "paymentmode": university_exam.paymentmode,
        "remarks": university_exam.remarks
    }
    
    return Response(response_data, status=status.HTTP_201_CREATED)
  
@api_view(['POST'])
def create_university_reregistration(request):
    student_id = request.data.get('student_id')
    type = request.data.get('type')
    amount = request.data.get('amount')
    date = request.data.get('date')
    examination = request.data.get('examination')
    semyear = request.data.get('semyear')
    paymentmode = request.data.get('paymentmode')
    remarks = request.data.get('remarks')
    
    if not student_id:
        logger.error("Student ID is required.")
        return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} not found.")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
    
    university_exam = UniversityExamination(
        student=student,
        type=type,
        amount=amount,
        date=date,
        examination=examination,
        semyear=semyear,
        paymentmode=paymentmode,
        remarks=remarks
    )
    university_exam.save()
    
    response_data = {
        "id": university_exam.id,
        "student_id": university_exam.student.id,
        "type": university_exam.type,
        "amount": university_exam.amount,
        "date": university_exam.date,
        "examination": university_exam.examination,
        "semyear": university_exam.semyear,
        "paymentmode": university_exam.paymentmode,
        "remarks": university_exam.remarks
    }
    
    return Response(response_data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
def get_university_reregistration(request):
    student_id = request.query_params.get("student_id")
    
    if not student_id:
        logger.error("Student ID is required.")
        return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        student = Student.objects.get(id=student_id)
        exams = UniversityExamination.objects.filter(student=student)
        
        response_data = [
            {
                "id": exam.id,
                "student_id": exam.student.id,
                "type": exam.type,
                "amount": exam.amount,
                "date": exam.date,
                "examination": exam.examination,
                "semyear": exam.semyear,
                "paymentmode": exam.paymentmode,
                "remarks": exam.remarks
            } for exam in exams
        ]
        
        return Response(response_data, status=status.HTTP_200_OK)
    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} not found.")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return Response({"error": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
@api_view(['GET'])
def get_paid_fees(request):
    student_id = request.query_params.get("student_id")

    if not student_id:
        logger.error("Student ID is required.")
        return Response({"error": "Student ID is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        getstudent = Student.objects.get(id=student_id)
        university = getstudent.university  # Fetch the related university

        response_data = {
            "name": getstudent.name,
            "dateofbirth": getstudent.dateofbirth,
            "email": getstudent.email,
            "mobile": getstudent.mobile,
            "address": getstudent.address,
            "city": getstudent.city.name if getstudent.city else None,
            "state": getstudent.state.name if getstudent.state else None,
            "pincode": getstudent.pincode,
            "country": getstudent.country.name if getstudent.country else None,
            "enrollment_id": getstudent.enrollment_id,
            "university_name": university.university_name,
            "university_address": university.university_address,
            "university_city": university.university_city,
            "university_state": university.university_state,
            "university_pincode": university.university_pincode,
            "university_logo": str(university.university_logo.url) if university.university_logo else None,
            "registrationID": university.registrationID
        }

        return Response(response_data, status=status.HTTP_200_OK)

    except Student.DoesNotExist:
        logger.error(f"Student with ID {student_id} not found.")
        return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return Response({"error": "An unexpected error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def save_single_question_answer(request):
    try:
        data = request.data

        # Convert student_id and exam_id to int for filtering
        try:
            student_id = int(data.get("student_id"))
            exam_id = int(data.get("exam_id"))
        except (TypeError, ValueError):
            return Response({"error": "Invalid student_id or exam_id. Must be integer."}, status=status.HTTP_400_BAD_REQUEST)

        question_id = data.get("question_id")
        submitted_answer = data.get("submitted_answer")
        exam_details_id = data.get("exam_details_id")  # NEW: get exam_details_id from request

        if not student_id or not exam_id or not question_id or submitted_answer is None:
            return Response({"error": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST)

        # Validate question exists
        try:
            question_instance = Questions.objects.get(id=question_id)
        except Questions.DoesNotExist:
            return Response({"error": f"Question ID {question_id} does not exist"}, status=status.HTTP_404_NOT_FOUND)

        # Validate exam_details_id
        try:
            exam_details_instance = StudentAppearingExam.objects.get(id=int(exam_details_id))
        except (StudentAppearingExam.DoesNotExist, ValueError, TypeError):
            return Response({"error": "Invalid or missing exam_details_id"}, status=status.HTTP_400_BAD_REQUEST)

        correct_answer = question_instance.answer.lower()
        marks = question_instance.marks

        if submitted_answer.lower() == correct_answer:
            result = "Right"
            marks_obtained = marks
        else:
            result = "Wrong"
            marks_obtained = "0"

        submission, created = SubmittedExamination.objects.update_or_create(
            student_id=student_id,
            exam_id=exam_id,
            question=str(question_id),
            examdetails=exam_details_instance,  # use passed exam_details_id here
            defaults={
                "type": question_instance.type,
                "marks": marks,
                "marks_obtained": marks_obtained,
                "submitted_answer": submitted_answer,
                "answer": correct_answer,
                "result": result,
            }
        )

        return Response({
            "message": "Answer updated successfully." if not created else "Answer submitted successfully.",
            "question_id": question_id,
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception(f"An error occurred while saving answer: {str(e)}")
        return Response({"error": "An internal server error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_of_all_cancelled_student(request):
    try:
        students = Student.objects.filter(is_cancelled=True).order_by('-id')

        student_data = []
        for student in students:
            enrolled = Enrolled.objects.filter(student=student).first()

            student_data.append({
                "id": student.id,
                "mobile": student.mobile,
                "name": student.name,
                "registration_id": student.registration_id,
                "email": student.email,
                "current_semyear": enrolled.current_semyear if enrolled else "",
                "enrollment_id": student.enrollment_id,
                "student_name": student.name,
                "university_name": student.university.university_name if student.university else "",
                "source": 'QR' if student.is_quick_register else 'SR',
                "study_pattern_mode": enrolled.course_pattern if enrolled and enrolled.course_pattern else "",
                "entry_mode": enrolled.entry_mode if enrolled and enrolled.entry_mode else "",
                "enrollment_date": student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else ""
            })

        return Response({"status": "success", "data": student_data}, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error fetching cancelled students: {str(e)}")
        return Response({"status": "error", "message": "Something went wrong!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def document_management(request, enrollment_id):
    if not enrollment_id:
        logger.error("Missing required parameter: enrollment_id in request URL")
        return Response(
            {"error": "Missing required parameter: enrollment_id"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Fetch the student by enrollment_id
        student = Student.objects.get(enrollment_id=enrollment_id)
        #logger.info(f"Fetching documents for student: {student.name} (Enrollment ID: {student.enrollment_id})")

        # Fetch student documents
        student_documents = StudentDocuments.objects.filter(student=student)
        student_documents_data = StudentDocumentsSerializerGET(student_documents, many=True).data

        # Fetch student qualifications
        qualifications = Qualification.objects.filter(student=student).first()
        qualification_data = QualificationSerializer(qualifications).data if qualifications else None

        # Prepare response data
        response_data = {
            "image": student.image.url if student.image else None,
            "documents": student_documents_data,
            "qualifications": qualification_data
        }

        return Response(response_data, status=status.HTTP_200_OK)

    except Student.DoesNotExist:
        logger.error(f"Student with enrollment_id {enrollment_id} not found.")
        return Response(
            {"error": "Student not found."},
            status=status.HTTP_404_NOT_FOUND
        )

    except Exception as e:
        logger.error(f"Unexpected error while retrieving documents for student {enrollment_id}: {str(e)}")
        return Response(
            {"error": "An unexpected error occurred while retrieving student documents."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
        


@api_view(["POST"])
def save_exam_timer(request):
    student_id = request.data.get("student_id")
    exam_id = request.data.get("exam_id")
    time_left_ms_raw = request.data.get("time_left_ms")
    exam_details_id = request.data.get("exam_details_id")

    if not all([student_id, exam_id, time_left_ms_raw, exam_details_id]):
        return Response({"error": "Missing data"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        time_left_ms = int(time_left_ms_raw)
    except (TypeError, ValueError):
        return Response({"error": "Invalid time_left_ms, must be integer"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        examdetails_instance = StudentAppearingExam.objects.get(id=int(exam_details_id))
    except (StudentAppearingExam.DoesNotExist, ValueError):
        return Response({"error": "Invalid exam_details_id"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student_obj = Student.objects.get(id=student_id)
        exam_obj = Examination.objects.get(id=exam_id)
    except (Student.DoesNotExist, Examination.DoesNotExist):
        return Response({"error": "Student or Exam not found"}, status=status.HTTP_404_NOT_FOUND)

    try:
        session = ExamSession.objects.get(
            student=student_obj,
            exam=exam_obj,
            examdetails=examdetails_instance
        )
        session.time_left_ms = time_left_ms
        session.save()
        created = False
    except ExamSession.DoesNotExist:
        duration_minutes = getattr(exam_obj, "examduration", None)
        if duration_minutes is not None:
            init_time_left_ms = int(duration_minutes) * 60 * 1000
        else:
            init_time_left_ms = time_left_ms

        session = ExamSession.objects.create(
            student=student_obj,
            exam=exam_obj,
            examdetails=examdetails_instance,
            time_left_ms=init_time_left_ms
        )
        created = True

    return Response({"status": "saved", "created": created, "time_left_ms": session.time_left_ms}, status=status.HTTP_200_OK)



@api_view(["GET"])
def get_exam_timer(request):
    student_id = request.query_params.get("student_id")
    exam_id = request.query_params.get("exam_id")
    exam_details_id = request.query_params.get("exam_details_id")

    if not all([student_id, exam_id, exam_details_id]):
        return Response({"error": "Missing data"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        session = ExamSession.objects.get(
            student_id=student_id,
            exam_id=exam_id,
            examdetails_id=exam_details_id
        )
        return Response({"time_left_ms": session.time_left_ms})
    except ExamSession.DoesNotExist:
        return Response({"time_left_ms": None})


@api_view(['POST'])
def save_result_after_exam(request):
    """
    Save the result after an exam is finished.
    Use provided exam_details_id instead of latest.
    """
    try:
        data = request.data
        student_id = data.get("student_id")
        exam_id = data.get("exam_id")
        exam_details_id = data.get("exam_details_id")  # <-- new param
        created_by = data.get("created_by", None)
        modified_by = data.get("modified_by", None)

        if not student_id or not exam_id:
            return Response({"error": "Missing required fields: student_id or exam_id"}, 
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            student_id_int = int(student_id)
            exam_id_int = int(exam_id)
        except (TypeError, ValueError):
            return Response({"error": "Invalid student_id or exam_id"}, status=status.HTTP_400_BAD_REQUEST)

        examdetails_instance = None
        if exam_details_id:
            try:
                exam_details_id_int = int(exam_details_id)
                examdetails_instance = StudentAppearingExam.objects.get(id=exam_details_id_int)
            except (ValueError, StudentAppearingExam.DoesNotExist):
                return Response({"error": "Invalid exam_details_id"}, status=status.HTTP_400_BAD_REQUEST)

        if not examdetails_instance:
            # fallback: get latest
            examdetails_instance = StudentAppearingExam.objects.filter(
                student_id__contains=[student_id_int],
                exam_id=exam_id_int
            ).order_by('-id').first()

        if not examdetails_instance:
            return Response({"error": "No StudentAppearingExam record found for given student and exam."},
                            status=status.HTTP_404_NOT_FOUND)


        questions = Questions.objects.filter(exam_id=exam_id_int)

        for question in questions:
            submitted_answer = SubmittedExamination.objects.filter(
                student_id=student_id_int,
                exam_id=exam_id_int,
                question=str(question.id),
                examdetails=examdetails_instance
            ).order_by('-id').first()

            if not submitted_answer:
                SubmittedExamination.objects.create(
                    student_id=student_id_int,
                    exam_id=exam_id_int,
                    question=str(question.id),
                    type=question.type,
                    marks=question.marks,
                    marks_obtained="0",
                    submitted_answer="NA",
                    answer=question.answer,
                    result="Wrong",
                    examdetails=examdetails_instance
                )

        total_questions = questions.count()
        attempted = SubmittedExamination.objects.filter(
            student_id=student_id_int, 
            exam_id=exam_id_int,
            examdetails=examdetails_instance
        ).filter(
                ~Q(submitted_answer="NA"),
                ~Q(submitted_answer="")
            ).count()
        score = SubmittedExamination.objects.filter(
            student_id=student_id_int, 
            exam_id=exam_id_int, 
            result="Right",
            examdetails=examdetails_instance
        ).count()
        total_marks = questions.aggregate(Sum('marks'))['marks__sum'] or 0

        exam = Examination.objects.get(id=exam_id_int)
        passing_marks = int(exam.passingmarks or 50)
        percentage = (score / total_questions) * 100 if total_questions > 0 else 0
        result_status = "Passed" if percentage >= passing_marks else "Failed"

        result_obj, created = Result.objects.update_or_create(
            student_id=student_id_int,
            exam_id=exam_id_int,
            examdetails=examdetails_instance,
            defaults={
                'total_question': str(total_questions),
                'attempted': str(attempted),
                'total_marks': str(total_marks),
                'score': str(score),
                'result': result_status,
                'percentage': percentage,
                'created_by': created_by,
                'modified_by': modified_by,
            }
        )

        return Response({
            "message": "Exam results saved successfully",
            "result_id": result_obj.id
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception(f"Error saving exam result: {str(e)}")
        return Response({"error": "Internal server error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def check_exam_result(request):
    """
    Check if a result has been created for a student and exam.
    """
    try:
        student_id = request.query_params.get("student_id")
        exam_id = request.query_params.get("exam_id")
        
        print(student_id," --->student_id",exam_id,'" --->exam_id"')

        if not student_id or not exam_id:
            return Response({"error": "Missing required fields: student_id or exam_id"}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the result exists
        result = Result.objects.filter(student_id=student_id, exam_id=exam_id).exists()

        return Response({
            "has_result": result
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception(f"An error occurred while checking exam result: {str(e)}")
        return Response({"error": "An internal server error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
@api_view(['GET'])
def check_exam_result(request):
    """
    Check if a result has been created for a student, exam, and specific examdetails (StudentAppearingExam).
    """
    try:
        student_id = request.query_params.get("student_id")
        exam_id = request.query_params.get("exam_id")
        examdetails_id = request.query_params.get("examdetails_id")

        # Validate presence
        if not student_id or not exam_id or not examdetails_id:
            return Response({"error": "Missing required fields: student_id, exam_id or examdetails_id"}, status=status.HTTP_400_BAD_REQUEST)

        # Convert to integers safely
        try:
            student_id_int = int(student_id)
            exam_id_int = int(exam_id)
            examdetails_id_int = int(examdetails_id)
        except ValueError:
            return Response({"error": "student_id, exam_id and examdetails_id must be integers"}, status=status.HTTP_400_BAD_REQUEST)

        # Filter with typed parameters
        result_exists = Result.objects.filter(
            student_id=student_id_int,
            exam_id=exam_id_int,
            examdetails_id=examdetails_id_int
        ).exists()

        return Response({"has_result": result_exists}, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception(f"Error checking exam result: {str(e)}")
        return Response({"error": "An internal server error occurred."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)  


@api_view(['GET'])
def check_exam_availability(request):
    """
    Check if the current time is between exam start and end time, and if the exam date is valid for the student.
    """
    try:
        student_id = request.query_params.get("student_id")
        exam_id = request.query_params.get("exam_id")

        if not student_id or not exam_id:
            return Response({"error": "Missing student_id or exam_id."}, status=400)

        # Get the exam and student appearing data
        student_exam = StudentAppearingExam.objects.filter(exam_id=exam_id, student_id__contains=[student_id]).first()

        if not student_exam:
            return Response({"error": "Student is not enrolled for this exam."}, status=404)

        # Get exam start/end date and time
        exam = Examination.objects.get(id=exam_id)
        print('examexam',exam)
        print('examstartdate',student_exam.examstartdate)
        print('examenddate',student_exam.examenddate)
        print('examstarttime',student_exam.examstarttime)
        print('examendtime',student_exam.examendtime)

        start_datetime = datetime.combine(student_exam.examstartdate, student_exam.examstarttime)
        end_datetime = datetime.combine(student_exam.examenddate, student_exam.examendtime)

        now = datetime.now()

        # Check if the current date and time are within the exam window
        is_within_exam_window = now >= start_datetime and now <= end_datetime

        if is_within_exam_window:
            return Response({"can_start": True, "message": "You can start the exam."})
        else:
            return Response({"can_start": False, "message": "The exam is not available at this time."})

    except Exception as e:
        return Response({"error": f"An error occurred: {str(e)}"}, status=500)
      

@api_view(['POST'])
def create_category(request):
    serializer = CategoriesSerializer(data=request.data)
    if serializer.is_valid():
        try:
            serializer.save()
            #logger.info(f"Category created successfully: {serializer.data}")
            return Response({
                'message': 'Category created successfully',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error(f"Error saving category: {str(e)}", exc_info=True)
            return Response({'error': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        logger.warning(f"Category creation failed due to validation errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
def update_category(request, pk):
    try:
        category = Categories.objects.get(pk=pk)
    except Categories.DoesNotExist:
        logger.warning(f"Category update failed: Category with ID {pk} not found")
        return Response({'error': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = CategoriesSerializer(category, data=request.data, partial=(request.method == 'PATCH'))
    if serializer.is_valid():
        try:
            serializer.save()
            #logger.info(f"Category updated successfully (ID: {pk}): {serializer.data}")
            return Response({
                'message': 'Category updated successfully',
                'data': serializer.data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error updating category ID {pk}: {str(e)}", exc_info=True)
            return Response({'error': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        logger.warning(f"Category update failed for ID {pk} due to validation errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
      
@api_view(['GET'])
def list_categories(request):
    try:
        categories = Categories.objects.all().order_by('-created_at')  # Optional: sorted by latest
        serializer = CategoriesSerializer(categories, many=True)
        #logger.info(f"{len(categories)} categories fetched successfully.")
        return Response({'data': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error fetching categories: {str(e)}", exc_info=True)
        return Response({'error': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
      
@api_view(['POST'])
def create_source(request):
    serializer = SourceSerializer(data=request.data)
    if serializer.is_valid():
        try:
            serializer.save()
            #logger.info(f"Source created: {serializer.data}")
            return Response({'message': 'Source created successfully', 'data': serializer.data}, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error(f"Error saving source: {str(e)}", exc_info=True)
            return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        logger.warning(f"Validation failed: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
def update_source(request, pk):
    try:
        source = Source.objects.get(pk=pk)
    except Source.DoesNotExist:
        logger.warning(f"Source with ID {pk} not found")
        return Response({'error': 'Source not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = SourceSerializer(source, data=request.data, partial=(request.method == 'PATCH'))
    if serializer.is_valid():
        try:
            serializer.save()
            #logger.info(f"Source updated (ID {pk}): {serializer.data}")
            return Response({'message': 'Source updated successfully', 'data': serializer.data})
        except Exception as e:
            logger.error(f"Error updating source (ID {pk}): {str(e)}", exc_info=True)
            return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        logger.warning(f"Validation failed during update for ID {pk}: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_sources(request):
    try:
        sources = Source.objects.all()
        serializer = SourceSerializer(sources, many=True)
        #logger.info(f"{len(sources)} sources fetched.")
        return Response({'data': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error fetching sources: {str(e)}", exc_info=True)
        return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
@api_view(['GET'])
def list_statuses(request):
    try:
        statuses = Status.objects.all().order_by('-id')
        serializer = StatusSerializer(statuses, many=True)
        return Response({'data': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error listing statuses: {e}")
        return Response({'error': 'Something went wrong'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_status(request):
    serializer = StatusSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Status created successfully', 'data': serializer.data}, status=status.HTTP_201_CREATED)
    logger.warning(f"Create status validation error: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
def update_status(request, pk):
    try:
        status_instance = Status.objects.get(pk=pk)
    except Status.DoesNotExist:
        logger.warning(f"Status not found with id: {pk}")
        return Response({'error': 'Status not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = StatusSerializer(status_instance, data=request.data, partial=(request.method == 'PATCH'))
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Status updated successfully', 'data': serializer.data}, status=status.HTTP_200_OK)
    logger.warning(f"Update status validation error: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_common_lead_labels(request):
    try:
        labels = Common_Lead_Label.objects.all().order_by('-id')
        serializer = CommonLeadLabelSerializer(labels, many=True)
        return Response({'data': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error listing common lead labels: {e}")
        return Response({'error': 'Something went wrong'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_common_lead_label(request):
    serializer = CommonLeadLabelSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Common Lead Label created successfully', 'data': serializer.data}, status=status.HTTP_201_CREATED)
    logger.warning(f"Validation error while creating label: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
def update_common_lead_label(request, pk):
    try:
        label = Common_Lead_Label.objects.get(pk=pk)
    except Common_Lead_Label.DoesNotExist:
        logger.warning(f"Label not found with id: {pk}")
        return Response({'error': 'Common Lead Label not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = CommonLeadLabelSerializer(label, data=request.data, partial=(request.method == 'PATCH'))
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Common Lead Label updated successfully', 'data': serializer.data}, status=status.HTTP_200_OK)
    logger.warning(f"Validation error while updating label: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
  

@api_view(['GET'])
def list_colors(request):
    try:
        colors = Color.objects.all().order_by('-id')
        serializer = ColorSerializer(colors, many=True)
        return Response({'data': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error listing colors: {e}")
        return Response({'error': 'Something went wrong'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def create_color(request):
    serializer = ColorSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Color created successfully', 'data': serializer.data}, status=status.HTTP_201_CREATED)
    logger.warning(f"Create color validation error: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'PATCH'])
def update_color(request, pk):
    try:
        color_instance = Color.objects.get(pk=pk)
    except Color.DoesNotExist:
        logger.warning(f"Color not found with id: {pk}")
        return Response({'error': 'Color not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ColorSerializer(color_instance, data=request.data, partial=(request.method == 'PATCH'))
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Color updated successfully', 'data': serializer.data}, status=status.HTTP_200_OK)
    logger.warning(f"Update color validation error: {serializer.errors}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
  
@api_view(['GET'])
def sync_answers(request):
    student_id = request.query_params.get("student_id")
    exam_id = request.query_params.get("exam_id")
    exam_details_id = request.query_params.get("exam_details_id")  # new param

    if not student_id or not exam_id or not exam_details_id:
        return Response({"error": "Missing student_id, exam_id, or exam_details_id"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student_obj = Student.objects.get(id=student_id)
        exam_obj = Examination.objects.get(id=exam_id)
        exam_details_obj = StudentAppearingExam.objects.get(id=exam_details_id)
    except (Student.DoesNotExist, Examination.DoesNotExist, StudentAppearingExam.DoesNotExist):
        return Response({"error": "Student, Exam, or Exam Details not found"}, status=status.HTTP_404_NOT_FOUND)

    submitted_answers = SubmittedExamination.objects.filter(
        student=student_obj,
        exam=exam_obj,
        examdetails=exam_details_obj
    )

    answer_dict = {}
    for ans in submitted_answers:
        try:
            question_id_int = int(ans.question)
            answer_dict[question_id_int] = ans.submitted_answer
        except (ValueError, TypeError):
            continue

    return Response({"answers": answer_dict}, status=status.HTTP_200_OK)


@api_view(['POST'])
def registered_save_enrollment_to_next_semyear(request):
    student_id = request.data.get('student_id')
    current_semyear = request.data.get('current_semyear')
    total_semyear = request.data.get('total_semyear')
    next_semyear = request.data.get('next_semyear')

    if not all([student_id, current_semyear, total_semyear, next_semyear]):
        logger.warning("Missing required fields")
        return Response({"status": "error", "message": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        current = int(current_semyear)
        total = int(total_semyear)
        next_s = int(next_semyear)
    except ValueError:
        logger.error("Invalid data types for semester fields")
        return Response({"status": "error", "message": "Invalid semester values"}, status=status.HTTP_400_BAD_REQUEST)

    if total <= current:
        logger.warning(f"Next semester not allowed. Current: {current}, Total: {total}")
        return Response({"status": "error", "message": "Cannot enroll beyond total sem/year"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        enrolled = Enrolled.objects.get(student=student_id)
        enrolled.current_semyear = next_s
        enrolled.save()
        #logger.info(f"Student {student_id} promoted to sem/year {next_s}")
        return Response({"status": "success", "message": "Student enrolled to next semester/year"}, status=status.HTTP_200_OK)
    except Enrolled.DoesNotExist:
        logger.error(f"No enrollment found for student ID {student_id}")
        return Response({"status": "error", "message": "Enrollment not found"}, status=status.HTTP_404_NOT_FOUND)
      
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_exam_data_to_excel(request):
    user = request.user
    #logger.info(f"[EXPORT] Request received by user: {user.username} (ID: {user.id})")

    if not (user.is_superuser or getattr(user, 'is_data_entry', False)):
        logger.warning(f"[EXPORT] Unauthorized access attempt by user: {user.username}")
        return Response({"message": "Unauthorized access"}, status=status.HTTP_403_FORBIDDEN)

    data = request.data
    university = data.get("university")
    course = data.get("course")
    stream = data.get('stream')
    session = data.get('session')
    studypattern = data.get('studypattern')
    semyear = data.get('semyear')
    substream = data.get('substream')

    logger.debug(f"[EXPORT] Filters received: university={university}, course={course}, stream={stream}, session={session}, studypattern={studypattern}, semyear={semyear}, substream={substream}")

    try:
        filters = Q(university=university, course=course, stream=stream)
        if session not in ['0', '', None]:
            filters &= Q(session=session)
        if studypattern not in ['0', '', None]:
            filters &= Q(studypattern=studypattern)
        if semyear not in ['0', '', None]:
            filters &= Q(semyear=semyear)
        if substream not in ['0', '', None]:
            filters &= Q(substream=substream)

        exams = Examination.objects.filter(filters)
        if not exams.exists():
            #logger.info("[EXPORT] No matching exams found for the given filters.")
            return Response({"message": "No results found"}, status=status.HTTP_400_BAD_REQUEST)

        formatted_data = []
        for exam in exams:
            logger.debug(f"[EXPORT] Processing exam: ID={exam.id}, subject={exam.subject_id}")
            subject = Subject.objects.filter(id=exam.subject_id).first()
            student_exam_qs = StudentAppearingExam.objects.filter(exam=exam.id)

            if student_exam_qs.exists():
                serializer = StudentAppearingExamSerializer(student_exam_qs, many=True)
                for record in serializer.data:
                    for student_id in record['student_id']:
                        logger.debug(f"[EXPORT] Processing student ID={student_id} for exam ID={exam.id}")
                        has_appeared = ExamSession.objects.filter(
                            student_id=student_id,
                            exam=exam.id,
                            examdetails_id=record['id']
                        )
                        submitted_exam = Result.objects.filter(
                            student_id=student_id,
                            exam=exam,
                            examdetails_id=record['id']
                        )

                        status_str = "Not Appeared"
                        if has_appeared.exists() and submitted_exam.exists():
                            status_str = "Appeared"
                        elif has_appeared.exists():
                            status_str = "Ongoing Exam"

                        student = Student.objects.filter(id=student_id).first()
                        university_obj = University.objects.filter(id=university).first()
                        course_obj = Course.objects.filter(id=course).first()
                        stream_obj = Stream.objects.filter(id=stream).first()
                        substream_obj = SubStream.objects.filter(id=substream).first() if substream else None
                        enrolled = Enrolled.objects.filter(student_id=student_id).first()

                        if not enrolled:
                            logger.warning(f"[EXPORT] No enrollment found for student ID={student_id}. Skipping.")
                            continue

                        course_duration = int(enrolled.total_semyear) if enrolled.course_pattern == 'Annual' else int(enrolled.total_semyear) / 2

                        marks_obtained = ''
                        attempted_questions = ''
                        result_status = ''
                        if status_str == "Appeared":
                            sub_exam = submitted_exam.first()
                            if sub_exam:
                                marks_obtained = sub_exam.score
                                attempted_questions = sub_exam.attempted
                                result_status = sub_exam.result

                        formatted_data.append({
                            'University': university_obj.university_name if university_obj else '',
                            'Course': course_obj.name if course_obj else '',
                            'Stream': stream_obj.name if stream_obj else '',
                            'Substream': substream_obj.name if substream_obj else '',
                            'Session': session,
                            'Study Pattern': studypattern,
                            'Semester/Year': semyear,
                            'Course Duration': course_duration,
                            'Subject Name': subject.name if subject else '',
                            'Student Name': student.name if student else '',
                            'Student Email ID': student.email if student else '',
                            'Student Mobile Number': student.mobile if student else '',
                            'Enrollment ID': student.enrollment_id if student else '',
                            'Exam Start Time': record['examstarttime'],
                            'Exam End Time': record['examendtime'],
                            'Exam Start Date': record['examstartdate'],
                            'Exam End Date': record['examenddate'],
                            'Status': status_str,
                            'Total Marks': exam.totalmarks,
                            'Passing Marks': exam.passingmarks,
                            'Marks Obtained': marks_obtained,
                            'Questions Attempted': attempted_questions,
                            'Result': result_status,
                        })

        if not formatted_data:
            #logger.info("[EXPORT] No student data available to export.")
            return Response({"message": "No students assigned"}, status=status.HTTP_400_BAD_REQUEST)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Exam Data"

        headers = list(formatted_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header.upper()
            cell.font = Font(bold=True)

        for row_num, row_data in enumerate(formatted_data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num).value = value.upper() if isinstance(value, str) else value

        #logger.info(f"[EXPORT] Excel file prepared with {len(formatted_data)} records.")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = 'attachment; filename="exam_data_export.xlsx"'
        workbook.save(response)

        #logger.info("[EXPORT] Excel file successfully sent as response.")
        return response

    except Exception as e:
        logger.exception(f"[EXPORT] An unexpected error occurred: {str(e)}")
        return Response({"message": "An error occurred during export."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

import zipfile
import io

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def view_all_assigned_students_api(request):
    print('inside view_all_assigned_students_api')
    user = request.user
    #logger.info(f"[VIEW-STUDENTS] API accessed by user: {user.username} (ID: {user.id})")

    if not (user.is_superuser or getattr(user, 'is_data_entry', False)):
        logger.warning(f"[VIEW-STUDENTS] Unauthorized access attempt by {user.username}")
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

    try:
        data = request.data
        university_id = int(data.get("university"))
        course_id = int(data.get("course"))
        stream_id = int(data.get("stream") or data.get("Stream"))
        session = data.get("session", "").strip()
        studypattern = data.get("studypattern", "").upper().strip()
        semyear = data.get("semyear", "").strip()
        substream_id = data.get("substream")

        # ✅ Ensure semyear is mandatory
        if not semyear:
            logger.warning("[VIEW-STUDENTS] Missing semyear in request.")
            return Response({'error': 'Semester/Year (semyear) is required.'}, status=400)

        # Prepare filter for matching results via exams
        exam_filter = {
            'course_id': course_id,
            'stream_id': stream_id,
            'semyear__iexact': semyear
        }

        if substream_id and str(substream_id).lower() != 'null':
            try:
                exam_filter['substream_id'] = int(substream_id)
            except ValueError:
                logger.error("[VIEW-STUDENTS] Invalid substream_id format.")
                return Response({'error': 'Invalid substream_id'}, status=400)

        if session:
            exam_filter['session__iexact'] = session

        if studypattern:
            exam_filter['studypattern__iexact'] = studypattern

        logger.debug(f"[VIEW-STUDENTS] Exam filter: {exam_filter}")

        # Get results that match the exam filters
        matched_results = Result.objects.filter(
            exam__in=Examination.objects.filter(**exam_filter),
            student__university_id=university_id
        ).select_related('student', 'exam')

        # Build unique student response from results
        student_data = []
        seen_ids = set()

        for r in matched_results:
            if r.student and r.student.id not in seen_ids:
                student = r.student
                seen_ids.add(student.id)
                student_data.append({
                    "student_id": student.id,
                    "name": student.name,
                    "email": student.email,
                    "mobile": student.mobile,
                    "enrollment_id": student.enrollment_id,
                    "current_semyear": "",  # optional: could be pulled from Enrolled
                    "session": r.exam.session,
                    "entry_mode": "",       # optional: could be pulled from Enrolled
                    "course_pattern": r.exam.studypattern,
                })

        #logger.info(f"[VIEW-STUDENTS] Found {len(student_data)} students with result records")
        return Response({'message': 'Students fetched successfully', 'students': student_data}, status=200)

    except Exception as e:
        logger.exception(f"[VIEW-STUDENTS] Unexpected error: {str(e)}")
        return Response({'error': 'Internal server error'}, status=500)

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def fetch_complete_student_data_api(request):
#     user = request.user
#     #logger.info(f"[EXPORT-ZIP] Started by {user.username} (ID={user.id})")

#     if not (user.is_superuser or getattr(user, 'is_data_entry', False)):
#         logger.warning(f"[EXPORT-ZIP] Unauthorized access attempt by {user.username}")
#         return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

#     try:
#         data = request.data

#         # FIX: ensure student_ids is parsed correctly
#         raw_ids = data.get('student_ids', [])
#         if isinstance(raw_ids, str):
#             try:
#                 student_ids = json.loads(raw_ids)
#             except json.JSONDecodeError:
#                 return Response({'error': 'Invalid student_ids format'}, status=400)
#         else:
#             student_ids = raw_ids

#         course_id = data.get("course")
#         stream_id = data.get("stream")
#         substream_id = data.get("substream")
#         session = data.get("session")
#         semyear = data.get("semyear")
#         studypattern = data.get("studypattern")

#         course = Course.objects.get(id=course_id)
#         stream = Stream.objects.get(id=stream_id)
#         substream = SubStream.objects.get(id=substream_id) if substream_id else None

#         results = Result.objects.filter(
#             student_id__in=student_ids,
#             exam__semyear=str(semyear)  # cast to string to match field type
#         ).select_related('exam__subject', 'student').order_by('student_id', 'exam_id')


#         student_subject_map = {}
#         student_name_map = {}
#         all_exam_ids, all_examdetail_ids, all_student_ids = set(), set(), set()

#         for r in results:
#             sid = str(r.student_id)
#             subject_name = r.exam.subject.name.strip() if r.exam and r.exam.subject else f"Exam_{r.exam_id}"
#             student_name = r.student.name.strip() if r.student else f"Student_{sid}"
#             student_subject_map.setdefault(sid, {}).setdefault(subject_name, []).append(r)
#             student_name_map[sid] = student_name
#             all_exam_ids.add(r.exam_id)
#             all_examdetail_ids.add(r.examdetails_id)
#             all_student_ids.add(r.student_id)

#         students_without_data = []
#         for sid in student_ids:
#             if int(sid) not in all_student_ids:
#                 try:
#                     student = Student.objects.get(id=sid)
#                     students_without_data.append(student.name.strip())
#                 except Student.DoesNotExist:
#                     students_without_data.append(f"Student_{sid}")

#         questions = Questions.objects.filter(exam_id__in=all_exam_ids)
#         questions_map = {}
#         for q in questions:
#             questions_map.setdefault(q.exam_id, []).append(q)

#         exam_sessions = ExamSession.objects.filter(
#             examdetails__in=all_examdetail_ids,
#             student__in=all_student_ids
#         )
#         exam_time_map = {
#             (es.examdetails_id, es.student_id): f"{int(es.time_left_ms / 60000)} mins"  # Convert ms to minutes
#             for es in exam_sessions
#         }

#         submitted_answers = SubmittedExamination.objects.filter(
#             student_id__in=all_student_ids,
#             examdetails_id__in=all_examdetail_ids
#         )

#         submitted_answer_map = {
#             (int(sa.question), sa.student_id, sa.examdetails_id): sa
#             for sa in submitted_answers if sa.question.isdigit()
#         }

#         zip_buffer = io.BytesIO()
#         with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
#             for sid, subjects in student_subject_map.items():
#                 student_name = student_name_map.get(sid, f"Student_{sid}")
#                 safe_name = student_name.replace(" ", "_").replace("/", "_")
#                 excel_buffer = io.BytesIO()

#                 with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
#                     for subject_name, result_list in subjects.items():
#                         any_result = result_list[0]

#                         exam_month = (
#                             any_result.examdetails.examstartdate.strftime("%B %Y")
#                             if any_result.examdetails and any_result.examdetails.examstartdate
#                             else ""
#                         )
#                         time_key = (any_result.examdetails_id, any_result.student_id)
#                         time_appeared = f"{exam_time_map[time_key]} mins" if time_key in exam_time_map else "Not Recorded"
#                         duration = f"{any_result.exam.examduration} mins" if any_result.exam else ""

#                         metadata = [
#                             ['Student Name', student_name],
#                             ['Course', course.name],
#                             ['Stream', stream.name],
#                         ]
#                         if substream:
#                             metadata.append(['Substream', substream.name])
#                         metadata += [
#                             ['Session', session],
#                             ['Study Pattern', studypattern],
#                             ['Semester/Year', any_result.exam.semyear or semyear],
#                             ['Subject Name', subject_name],
#                             ['Exam Month', exam_month],
#                             ['Time Appeared', time_appeared],
#                             ['Duration', duration]
#                         ]
#                         meta_df = pd.DataFrame(metadata, columns=["", ""])

#                         exam_questions = questions_map.get(any_result.exam_id, [])
#                         question_rows = []
#                         for idx, q in enumerate(exam_questions, start=1):
#                             key = (q.id, any_result.student_id, any_result.examdetails_id)
#                             sa = submitted_answer_map.get(key)
#                             question_rows.append({
#                                 "SR. NO.": idx,
#                                 "QUESTION": q.question,
#                                 "OPTION 1": q.option1,
#                                 "OPTION 2": q.option2,
#                                 "OPTION 3": q.option3,
#                                 "OPTION 4": q.option4,
#                                 "MARKED ANSWER": sa.submitted_answer if sa else "",
#                                 "CORRECT ANSWER": sa.answer if sa else "",
#                                 "OBTAINED MARKS": sa.marks_obtained if sa else "",
#                                 "MAX MARKS": sa.marks if sa else ""
#                             })

#                         data_df = pd.DataFrame(question_rows)
#                         sheet_name = subject_name[:31]
#                         meta_df.to_excel(writer, sheet_name=sheet_name, startrow=0, header=False, index=False)
#                         data_df.to_excel(writer, sheet_name=sheet_name, startrow=len(metadata) + 2, index=False)

#                         worksheet = writer.sheets[sheet_name]
#                         workbook = writer.book
#                         bold_border = workbook.add_format({'bold': True, 'border': 1})
#                         border_only = workbook.add_format({'border': 1})
#                         header_format = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1})
#                         right_align_bold_border = workbook.add_format({'bold': True, 'border': 1, 'align': 'right'})
#                         left_align_border = workbook.add_format({'border': 1, 'align': 'left'})

#                         for i in range(len(metadata)):
#                             worksheet.write(i, 0, metadata[i][0], bold_border)
#                             worksheet.write(i, 1, metadata[i][1], border_only)

#                         for col_num, value in enumerate(data_df.columns.values):
#                             worksheet.write(len(metadata) + 2, col_num, value, header_format)

#                         worksheet.set_column('A:A', 25)
#                         worksheet.set_column('B:B', 80)
#                         worksheet.set_column('C:G', 25)
#                         worksheet.set_column('H:J', 18)

#                         total_obtained = sum(
#                             float(sa.marks_obtained)
#                             for sa in submitted_answer_map.values()
#                             if sa.student_id == any_result.student_id and
#                                sa.examdetails_id == any_result.examdetails_id and
#                                sa.question.isdigit() and
#                                int(sa.question) in [q.id for q in exam_questions] and
#                                sa.marks_obtained
#                         )

#                         total_max = sum(
#                             float(sa.marks)
#                             for sa in submitted_answer_map.values()
#                             if sa.student_id == any_result.student_id and
#                                sa.examdetails_id == any_result.examdetails_id and
#                                sa.question.isdigit() and
#                                int(sa.question) in [q.id for q in exam_questions] and
#                                sa.marks
#                         )

#                         summary_row = len(metadata) + 2 + len(data_df) + 2
                        
#                         try:
#                             worksheet.merge_range(summary_row, 0, summary_row, 7, "MARKS OBTAINED", right_align_bold_border)
#                         except Exception as merge_err:
#                             logger.warning(f"Skipping merge range due to overlap: {merge_err}")
#                             worksheet.write(summary_row, 7, "MARKS OBTAINED", right_align_bold_border)

#                         worksheet.write(summary_row, 8, total_obtained, left_align_border)
#                         worksheet.write(summary_row, 9, total_max, left_align_border)

                        
                        
#                 zip_file.writestr(f"{safe_name}.xlsx", excel_buffer.getvalue())

#         if not student_subject_map:
#             logger.warning("[EXPORT-ZIP] No data found for selected students.")
#             return Response({'error': 'No data found for selected students'}, status=404)

#         zip_buffer.seek(0)
#         filename_parts = [course.name.replace(" ", "_"), stream.name.replace(" ", "_")]
#         if substream:
#             filename_parts.append(substream.name.replace(" ", "_"))
#         filename = "_".join(filename_parts).lower() + ".zip"

#         response = HttpResponse(zip_buffer, content_type='application/zip')
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         if students_without_data:
#             response['X-Missing-Students'] = ', '.join(students_without_data[:10])
#             response['X-Missing-Count'] = str(len(students_without_data))

#         #logger.info(f"[EXPORT-ZIP] Successfully generated ZIP for {len(student_subject_map)} students")
#         return response

#     except Exception as e:
#         logger.exception(f"[EXPORT-ZIP] Internal error: {str(e)}")
#         return Response({'error': 'Internal Server Error'}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def fetch_complete_student_data_api(request):
    user = request.user
    if not (user.is_superuser or getattr(user, 'is_data_entry', False)):
        return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

    try:
        data = request.data

        raw_ids = data.get('student_ids', [])
        if isinstance(raw_ids, str):
            try:
                student_ids = json.loads(raw_ids)
            except json.JSONDecodeError:
                return Response({'error': 'Invalid student_ids format'}, status=400)
        else:
            student_ids = raw_ids

        course_id = data.get("course")
        stream_id = data.get("stream")
        substream_id = data.get("substream")
        session = data.get("session")
        semyear = data.get("semyear")
        studypattern = data.get("studypattern")

        course = Course.objects.get(id=course_id)
        stream = Stream.objects.get(id=stream_id)
        substream = SubStream.objects.get(id=substream_id) if substream_id else None

        results = Result.objects.filter(
            student_id__in=student_ids,
            exam__semyear=str(semyear)
        ).select_related('exam__subject', 'student').order_by('student_id', 'exam_id')

        student_subject_map = {}
        student_name_map = {}
        all_exam_ids, all_examdetail_ids, all_student_ids = set(), set(), set()

        for r in results:
            sid = str(r.student_id)
            if not r.exam or not r.exam.subject:
                continue
            subject_id = r.exam.subject.id
            subject_name = r.exam.subject.name.strip()
            subject_code = r.exam.subject_code if hasattr(r.exam, 'subject_code') else f"EXAM{r.exam.id}"
            subject_key = (subject_id, subject_name, subject_code)

            student_subject_map.setdefault(sid, {}).setdefault(subject_key, []).append(r)
            student_name_map[sid] = r.student.name.strip()
            all_exam_ids.add(r.exam_id)
            all_examdetail_ids.add(r.examdetails_id)
            all_student_ids.add(r.student_id)

        students_without_data = []
        for sid in student_ids:
            if int(sid) not in all_student_ids:
                try:
                    student = Student.objects.get(id=sid)
                    students_without_data.append(student.name.strip())
                except Student.DoesNotExist:
                    students_without_data.append(f"Student_{sid}")

        questions = Questions.objects.filter(exam_id__in=all_exam_ids)
        questions_map = {}
        for q in questions:
            questions_map.setdefault(q.exam_id, []).append(q)

        exam_sessions = ExamSession.objects.filter(
            examdetails__in=all_examdetail_ids,
            student__in=all_student_ids
        )
        exam_time_map = {
            (es.examdetails_id, es.student_id): f"{int(es.time_left_ms / 60000)} mins"
            for es in exam_sessions
        }

        submitted_answers = SubmittedExamination.objects.filter(
            student_id__in=all_student_ids,
            examdetails_id__in=all_examdetail_ids
        )
        submitted_answer_map = {
            (int(sa.question), sa.student_id, sa.examdetails_id): sa
            for sa in submitted_answers if sa.question.isdigit()
        }

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for sid, subjects in student_subject_map.items():
                student_name = student_name_map.get(sid, f"Student_{sid}")
                safe_name = student_name.replace(" ", "_").replace("/", "_")
                excel_buffer = io.BytesIO()

                with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                    for (subject_id, subject_name, subject_code), result_list in subjects.items():
                        any_result = result_list[0]

                        exam_month = (
                            any_result.examdetails.examstartdate.strftime("%B %Y")
                            if any_result.examdetails and any_result.examdetails.examstartdate
                            else ""
                        )
                        time_key = (any_result.examdetails_id, any_result.student_id)
                        time_appeared = exam_time_map.get(time_key, "Not Recorded")
                        duration = f"{any_result.exam.examduration} mins" if any_result.exam else ""

                        metadata = [
                            ['Student Name', student_name],
                            ['Full Subject Name', subject_name],
                            ['Course', course.name],
                            ['Stream', stream.name],
                        ]
                        if substream:
                            metadata.append(['Substream', substream.name])
                        metadata += [
                            ['Session', session],
                            ['Study Pattern', studypattern],
                            ['Semester/Year', any_result.exam.semyear or semyear],
                            ['Exam Month', exam_month],
                            ['Time Appeared', time_appeared],
                            ['Duration', duration]
                        ]
                        meta_df = pd.DataFrame(metadata, columns=["", ""])

                        exam_questions = questions_map.get(any_result.exam_id, [])
                        question_rows = []
                        for idx, q in enumerate(exam_questions, start=1):
                            key = (q.id, any_result.student_id, any_result.examdetails_id)
                            sa = submitted_answer_map.get(key)
                            question_rows.append({
                                "SR. NO.": idx,
                                "QUESTION": q.question,
                                "OPTION 1": q.option1,
                                "OPTION 2": q.option2,
                                "OPTION 3": q.option3,
                                "OPTION 4": q.option4,
                                "MARKED ANSWER": sa.submitted_answer if sa else "",
                                "CORRECT ANSWER": sa.answer if sa else "",
                                "OBTAINED MARKS": sa.marks_obtained if sa else "",
                                "MAX MARKS": sa.marks if sa else ""
                            })

                        data_df = pd.DataFrame(question_rows)

                        # Generate Unique Sheet Name using Subject Code
                        sheet_base_name = subject_name.replace(" ", "_")[:20]  # Trim to avoid overflow
                        sheet_name = f"{subject_code}_{sheet_base_name}"[:31]  # Ensure <=31 chars

                        meta_df.to_excel(writer, sheet_name=sheet_name, startrow=0, header=False, index=False)
                        data_df.to_excel(writer, sheet_name=sheet_name, startrow=len(metadata) + 2, index=False)

                        worksheet = writer.sheets[sheet_name]
                        workbook = writer.book
                        bold_border = workbook.add_format({'bold': True, 'border': 1})
                        border_only = workbook.add_format({'border': 1})
                        header_format = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1})
                        right_align_bold_border = workbook.add_format({'bold': True, 'border': 1, 'align': 'right'})
                        left_align_border = workbook.add_format({'border': 1, 'align': 'left'})

                        for i in range(len(metadata)):
                            worksheet.write(i, 0, metadata[i][0], bold_border)
                            worksheet.write(i, 1, metadata[i][1], border_only)

                        for col_num, value in enumerate(data_df.columns.values):
                            worksheet.write(len(metadata) + 2, col_num, value, header_format)

                        worksheet.set_column('A:A', 25)
                        worksheet.set_column('B:B', 80)
                        worksheet.set_column('C:G', 25)
                        worksheet.set_column('H:J', 18)

                        total_obtained = sum(
                            float(sa.marks_obtained)
                            for sa in submitted_answer_map.values()
                            if sa.student_id == any_result.student_id and
                               sa.examdetails_id == any_result.examdetails_id and
                               sa.question.isdigit() and
                               int(sa.question) in [q.id for q in exam_questions] and
                               sa.marks_obtained
                        )

                        total_max = sum(
                            float(sa.marks)
                            for sa in submitted_answer_map.values()
                            if sa.student_id == any_result.student_id and
                               sa.examdetails_id == any_result.examdetails_id and
                               sa.question.isdigit() and
                               int(sa.question) in [q.id for q in exam_questions] and
                               sa.marks
                        )

                        summary_row = len(metadata) + 2 + len(data_df) + 2
                        try:
                            worksheet.merge_range(summary_row, 0, summary_row, 7, "MARKS OBTAINED", right_align_bold_border)
                        except Exception:
                            worksheet.write(summary_row, 7, "MARKS OBTAINED", right_align_bold_border)

                        worksheet.write(summary_row, 8, total_obtained, left_align_border)
                        worksheet.write(summary_row, 9, total_max, left_align_border)

                zip_file.writestr(f"{safe_name}.xlsx", excel_buffer.getvalue())

        if not student_subject_map:
            return Response({'error': 'No data found for selected students'}, status=404)

        zip_buffer.seek(0)
        filename_parts = [course.name.replace(" ", "_"), stream.name.replace(" ", "_")]
        if substream:
            filename_parts.append(substream.name.replace(" ", "_"))
        filename = "_".join(filename_parts).lower() + ".zip"

        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        if students_without_data:
            response['X-Missing-Students'] = ', '.join(students_without_data[:10])
            response['X-Missing-Count'] = str(len(students_without_data))

        return response

    except Exception as e:
        logger.exception(f"[EXPORT-ZIP] Internal error: {str(e)}")
        return Response({'error': 'Internal Server Error'}, status=500)

